#!/usr/bin/env python3
"""授权曲库 Excel 导入

列名自动识别（复用 monitor 的别名模糊匹配思路）。
标准字段：歌曲ID / 歌名 / 艺人 / 专辑 / 版本 / 版权公司 / 词作者 / 曲作者
"""

import json
import os
import re

from . import db

# 列别名表：key=标准字段，value=可能的表头写法（小写、去符号后比对）
COLUMN_ALIASES = {
    'song_id': ['歌曲id', '歌曲i d', 'id', 'songid', 'musicid', '曲目id'],
    'song_name': ['歌名', '歌曲名', '歌曲名称', '曲名', '作品名', '作品名称', '歌曲',
                  'songname', 'song', 'title', 'name', '曲目', '曲目名称'],
    'artist': ['歌手', '歌手名', '歌手名称', '演唱', '演唱者', '演唱人', '艺人', '表演者',
               'artist', 'singer', 'performer', '主唱'],
    'album': ['专辑', '专辑名', '专辑名称', 'album'],
    'version': ['版本', 'version', 'ver'],
    'copyright_company': ['版权公司', '版权方', '版权', '公司', 'copyright', 'company', '厂牌'],
    'lyricist': ['词作者', '作词', '作词人', '词', '填词', 'lyricist', 'lyric', '词作'],
    'composer': ['曲作者', '作曲', '作曲人', '曲', 'composer', 'compose', '曲作'],
}

REQUIRED = ['song_name']


def _norm_header(h):
    if h is None:
        return ''
    return re.sub(r'[\s\u3000:：()（）\[\]【】/、_-]', '', str(h)).strip().lower()


def detect_columns(headers):
    mapping = {}
    used = set()
    normed = [_norm_header(h) for h in headers]
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            a = _norm_header(alias)
            for i, h in enumerate(normed):
                if i in used or not h:
                    continue
                if h == a:
                    mapping[field] = i
                    used.add(i)
                    break
            if field in mapping:
                break
    for field, aliases in COLUMN_ALIASES.items():
        if field in mapping:
            continue
        for alias in aliases:
            a = _norm_header(alias)
            for i, h in enumerate(normed):
                if i in used or not h:
                    continue
                if a in h or h in a:
                    mapping[field] = i
                    used.add(i)
                    break
            if field in mapping:
                break
    unmatched = [headers[i] for i in range(len(headers)) if i not in used and normed[i]]
    return mapping, unmatched


def read_excel(path, sheet=None, header_row=None, max_scan=20):
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.csv', '.txt', '.tsv'):
        return _read_csv(path)

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not all_rows:
        raise ValueError('Excel 为空')

    best_i, best_map, best_hit = 0, {}, -1
    scan_to = min(max_scan, len(all_rows))
    rng = [header_row] if header_row is not None else range(scan_to)
    for i in rng:
        m, _ = detect_columns(all_rows[i])
        if len(m) > best_hit:
            best_i, best_map, best_hit = i, m, len(m)

    headers = [('' if c is None else str(c).strip()) for c in all_rows[best_i]]
    rows = all_rows[best_i + 1:]
    meta = {'sheet': ws.title, 'sheets': wb.sheetnames if hasattr(wb, 'sheetnames') else [],
            'header_row': best_i, 'total_rows': len(rows)}
    return headers, rows, meta


def _read_csv(path):
    import csv
    encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']
    for enc in encodings:
        try:
            with open(path, 'r', encoding=enc, newline='') as f:
                sample = f.read(4096)
                f.seek(0)
                delim = '\t' if sample.count('\t') > sample.count(',') else ','
                rows = list(csv.reader(f, delimiter=delim))
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError('CSV 编码无法识别（试过 utf-8 / gbk / gb18030）')
    if not rows:
        raise ValueError('CSV 为空')
    return rows[0], rows[1:], {'sheet': 'csv', 'sheets': [], 'header_row': 0,
                               'total_rows': len(rows) - 1}


def preview(path, limit=10):
    headers, rows, meta = read_excel(path)
    mapping, unmatched = detect_columns(headers)
    samples = []
    for r in rows[:limit]:
        samples.append({f: _cell(r, mapping.get(f)) for f in mapping})
    missing = [f for f in REQUIRED if f not in mapping]
    return {
        'file': os.path.basename(path),
        'headers': headers,
        'mapping': {f: headers[i] for f, i in mapping.items()},
        'unmatched_headers': unmatched,
        'missing_required': missing,
        'total_rows': meta['total_rows'],
        'header_row': meta['header_row'],
        'sheet': meta['sheet'],
        'samples': samples,
    }


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ''
    v = row[idx]
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def import_excel(path, batch=None, mapping_override=None, dry_run=False):
    db.init_db()
    headers, rows, meta = read_excel(path)
    mapping, _ = detect_columns(headers)
    if mapping_override:
        for field, header_name in mapping_override.items():
            if header_name in headers:
                mapping[field] = headers.index(header_name)

    missing = [f for f in REQUIRED if f not in mapping]
    if missing:
        raise ValueError(f'必需列未识别到：{missing}；实际表头={headers}')

    batch = batch or f'{os.path.basename(path)}@{db.now_str()}'
    stat = {'total': 0, 'inserted': 0, 'skipped_empty': 0,
            'dup_in_file': 0, 'batch': batch}
    seen_keys = set()
    before = db.get_conn().execute('SELECT COUNT(*) c FROM catalog').fetchone()['c']

    for i, row in enumerate(rows):
        name = _cell(row, mapping.get('song_name'))
        if not name:
            stat['skipped_empty'] += 1
            continue
        artist = _cell(row, mapping.get('artist'))
        version = _cell(row, mapping.get('version'))
        key = db._catalog_key(name, artist, version)
        if key in seen_keys:
            stat['dup_in_file'] += 1
            continue
        seen_keys.add(key)
        stat['total'] += 1
        if dry_run:
            continue
        raw = {headers[j]: _cell(row, j) for j in range(min(len(headers), len(row)))}
        db.upsert_catalog(
            song_id=_cell(row, mapping.get('song_id')),
            song_name=name,
            artist=artist,
            album=_cell(row, mapping.get('album')),
            version=version,
            copyright_company=_cell(row, mapping.get('copyright_company')),
            lyricist=_cell(row, mapping.get('lyricist')),
            composer=_cell(row, mapping.get('composer')),
            batch=batch,
            source_row=meta['header_row'] + 2 + i,
            raw=raw,
        )

    after = db.get_conn().execute('SELECT COUNT(*) c FROM catalog').fetchone()['c']
    stat['inserted'] = after - before
    stat['catalog_total'] = after
    return stat


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print('usage: python -m evidence.importer <excel_path> [--preview|--import]')
        sys.exit(1)
    p = os.path.expanduser(sys.argv[1])
    mode = sys.argv[2] if len(sys.argv) > 2 else '--preview'
    if mode == '--import':
        print(json.dumps(import_excel(p), ensure_ascii=False, indent=2))
    else:
        print(json.dumps(preview(p), ensure_ascii=False, indent=2))
