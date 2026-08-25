# -*- coding: utf-8 -*-
"""
batch_v2 — 2万首级别批量任务系统（SQLite + 后台 worker + 断点续跑）

定位：解决原 /api/batch_search 一锅端跑几千首会超时/丢结果/不能关浏览器的痛点。
- 任务持久化到 SQLite（~/.musicfinder/batch_v2.db），崩溃后自动续跑
- 后台 ThreadPoolExecutor 处理，可中途取消/查询进度/增量导出
- 范围精简：仅 3 平台（QQ / 酷狗 / 网易云）+ 2 字段（收藏量 + 单曲链接）
- 两阶段：QQ/酷狗 走 12 路快速池；网易云因官方强限流 + 405 熔断，单独 1 路错峰慢跑（断点续跑）

设计：单一全局引擎在 register(app) 时挂载；QQ/酷狗走共享池，网易云走独立慢速池。
复用 app.py 的 search_all + search_netease + _pick_best_for_batch（register 时 app 已加载完）。
"""

import os
import re
import io
import sys
import json
import time
import random
import sqlite3
import threading
import concurrent.futures
from datetime import datetime
from flask import request, jsonify

# ─────────── 常量 ───────────
TASK_PLATFORMS = ['qq', 'kugou', 'netease']
TASK_PLAT_NAMES = {'qq': 'QQ音乐', 'kugou': '酷狗', 'netease': '网易云'}
DB_PATH = os.path.expanduser('~/.musicfinder/batch_v2.db')
WORKER_POOL_SIZE = 12         # QQ/酷狗 快速池（同时处理几首歌）。过大易触发平台限流
V2_SEARCH_TIMEOUT = 25        # 单首搜索超时秒数（避开慢响应阻塞 worker）
ITEM_MAX_ATTEMPTS = 2         # 单首失败重试次数（fail-fast，节省时间）
ITEM_RETRY_BASE_SEC = 0.6     # 退避基准
ITEM_STAGGER_SEC = 1.0        # 错峰（每首启动前随机 0~该值，避免突发限流）
# 网易云独立池：405「操作频繁」实测是**账号级**限流（带 Cookie 才触发），匿名请求不受限。
NETEASE_POOL_SIZE = 6         # 网易云独立池（匿名请求不受账号级限流，可适度并发）
NETEASE_PACE_SEC = 0.3        # 两次网易云请求之间至少间隔（秒），避开「操作频繁」
NETEASE_COOLDOWN = 900        # 网易云 405 熔断后的冷却时长（秒），与 app.py 一致
NETEASE_MAX_ATTEMPTS = 3      # 单首网易云重试上限（仍失败则留空，不阻塞整体）
TASK_AUTORUN_ON_STARTUP = True # 启动时自动续跑中断的任务

# ─── 平台限流闸门 ───
# 实测（2026-08-07，99 首×6 轮压测）：
#   · 酷狗收藏量接口 mget_collect 跑到约第 400 首开始返回空
#   · QQ 搜索跑到约第 600 首整个返回 {"message":"query error"} + 空列表
# 两者都是 HTTP 200 的「软封」，不抛异常。必须靠连续空结果来判定。
PLAT_EMPTY_STREAK_LIMIT = 8    # 同一平台连续 N 首拿不到 → 判定被限流
PLAT_COOLDOWN_SEC = 300        # 熔断后暂停该平台多久，然后放 1 首出去试水
PLAT_PROBE_INTERVAL = 120      # 试水失败后，隔多久再试一次（酷狗软限流常只持续 1~2 分钟，过长的冷却会拖慢 2 万首整体进度）
PLAT_MAX_TRIES = 5             # 单首歌在同一平台最多真跑几次，超了就认「查无此歌」
# 酷狗是 HTTP 200 返回 0 条的典型「软限流」——代码分不清是「冷门歌真没有」还是
# 「被风控挡了」。一旦撞上限流就被误判成「查无此歌」会静默丢数据，所以对酷狗的
# 「空结果」给足重试耐心：靠补跑线程跨很多个 45s 周期反复试，平台恢复的那一下
# 就能补上。明确返回错误 / 其它平台仍用 PLAT_MAX_TRIES。
KUGOU_EMPTY_PATIENCE = 60
SWEEP_INTERVAL_SEC = 45        # 补跑扫描间隔（平台恢复后自动把留空的歌捡回来）
SWEEP_BATCH = 200              # 空闲时每轮最多捡多少首，避免瞬间打爆刚恢复的平台
SWEEP_BATCH_BUSY = 20          # 还有歌没跑过第一遍时，补跑降为涓流（前台优先）
SWEEP_QUEUE_CEIL = 120         # 工作队列已堆这么多任务时，这轮补跑直接跳过
FINALIZE_STABLE_SWEEPS = 8     # 连续多少轮补跑无进展才收尾（NULL收藏量填0），避免把"平台还在恢复"误判成"查无"
ABSOLUTE_MAX_TRIES = 20        # 已放弃(真查无/长限流)的歌，平台恢复后再给的总重试上限（仅补跑阶段用）
GAP_MAX_TRIES = 8             # 缺口行(兄弟平台已确认有这首歌)的弱势平台：需连续 GAP_MAX_TRIES 次「干净空」才肯判未收录，
ORPHAN_MAX_TRIES = 4          # v4.25.15 残缺行兜底：有收藏量但没链接的行，sweeper 最多重审 4 次，仍拒收则放弃（保留残缺）
                             # 否则留 _done=0 让 sweeper 借兄弟精确名反复重搜（跨平台回填）。防限流窗口把缺口「假完成」吃掉。
STAGE1_MAX_WAIT_SEC = 0        # v4.27.12 起不再睡等：平台冷却时 _v2_search_fast 立即返 blocked，item 交 sweeper 重试（见 _process_item_safe）

# ─── 收藏量批量合并 ───
# QQ 的 _fetch_qq_details 和酷狗的 _fetch_kugou_collection_counts 本来就收列表、
# 一次请求查一批。之前每首歌单独调一次 = 白白多打了 N-1 倍请求，也正是酷狗
# 每跑 400 首就触发人机验证的主因。改成攒够一批再发：
#   每首请求数 4 → 约 2.05，收藏量接口的压力直接降到 1/40。
FAV_BATCH_SIZE = 40            # 攒够多少首发一次
FAV_FLUSH_SEC = 2.5            # 攒不满也最多等这么久（避免尾巴卡住）

# ─── 自适应节流 ───
# 实测的核心结论：吞吐从来不是瓶颈（12 并发能跑 630 首/分），限流才是。
# 全速跑 400~600 首必被封，而且一封就是几小时——2 万首这么跑永远跑不完。
# 所以主动降速：撞墙就腰斩，平稳一段再慢慢加回来（AIMD，和 TCP 拥塞控制一个思路）。
STAGE1_RATE_START = 120        # 起步速度（首/分钟）
STAGE1_RATE_MIN = 15
STAGE1_RATE_MAX = 300
RATE_RECOVER_AFTER = 240       # 平稳这么多秒没撞墙，就上调 25%

# 平台内部 ID 的「库列名」与「搜索结果字段名」映射。
# 收藏量接口认的 ID 跟链接里的 ID 不是一个东西：
#   QQ   链接带 mid（字符串），收藏量接口要 songId（数字）
#   酷狗 链接带 hash，收藏量接口要 album_audio_id（即 _mixsongid）
# 所以必须单独存一份，否则补跑只能靠重新搜索来重新拿 ID。
_ID_COL = {'qq': 'qq_songid', 'kugou': 'kugou_mixsongid'}
_ID_FIELD = {'qq': '_songid', 'kugou': '_mixsongid'}

# 进程退出（重启/热重载/关机）时线程池会抛这个。它不是"这首歌查不到"，
# 而是"我被打断了"，必须区别对待，否则中断一次就永久丢一批歌。
_SHUTDOWN_HINTS = ('interpreter shutdown', 'cannot schedule new futures')


def _is_shutdown_err(e):
    s = str(e)
    return isinstance(e, RuntimeError) and any(h in s for h in _SHUTDOWN_HINTS)


def _plat_id(platform, row):
    """从搜索结果里取出该平台的收藏量查询 ID，取不到返回 ''。"""
    if not row:
        return ''
    v = row.get(_ID_FIELD[platform])
    return str(v).strip() if v not in (None, '') else ''


def _canon_url(platform, sid):
    """用平台内部 ID 重建规范播放页链接（兜底，避免「有匹配却无链接」的残缺行）。"""
    if not sid:
        return ''
    if platform == 'qq':
        return f'https://y.qq.com/n/ryqq/songDetail/{sid}'
    if platform == 'kugou':
        return f'https://www.kugou.com/song/#hash={sid}'
    if platform == 'netease':
        return f'https://music.163.com/#/song?id={sid}'
    return ''


# ─────────── 自适应节流器 ───────────
class _Pacer:
    """全局发车节流：控制阶段1 每分钟放行多少首歌。

    撞到平台限流 → 速度腰斩；平稳跑一段 → 慢慢加回来。
    这样 2 万首的任务会自己找到「这个 IP 今天能承受的最快速度」，
    不需要人盯着调参数。
    """

    def __init__(self, rate=STAGE1_RATE_START):
        self._lock = threading.Lock()
        self._rate = float(rate)
        self._next_slot = time.time()
        self._last_penalty = 0.0

    @property
    def rate(self):
        return round(self._rate, 1)

    def acquire(self):
        """排队领一个发车名额，必要时阻塞。"""
        with self._lock:
            self._maybe_recover()
            interval = 60.0 / max(1e-6, self._rate)
            now = time.time()
            slot = max(now, self._next_slot)
            self._next_slot = slot + interval
        wait = slot - time.time()
        if wait > 0:
            time.sleep(wait)

    def penalize(self, why=''):
        with self._lock:
            old = self._rate
            self._rate = max(STAGE1_RATE_MIN, self._rate / 2)
            self._last_penalty = time.time()
            # 撞墙时队列里可能已经排了一堆车，把发车时刻拉回当下重新按新速率排
            self._next_slot = max(time.time(), self._next_slot - 30)
        if round(old, 1) != round(self._rate, 1):
            print(f'[batch_v2] 撞到限流({why}) → 降速 {old:.0f} → {self._rate:.0f} 首/分')

    def _maybe_recover(self):
        if self._rate >= STAGE1_RATE_MAX:
            return
        anchor = self._last_penalty or 0
        if anchor and time.time() - anchor >= RATE_RECOVER_AFTER:
            old = self._rate
            self._rate = min(STAGE1_RATE_MAX, self._rate * 1.25)
            self._last_penalty = time.time()
            print(f'[batch_v2] 平稳运行，提速 {old:.0f} → {self._rate:.0f} 首/分')


# ─────────── 平台健康闸门 ───────────
class _PlatGate:
    """按平台跟踪「连续空结果」，达到阈值就熔断该平台一段时间。

    平台软限流的特征是 HTTP 200 + 空列表（QQ 会带 message='query error'），
    不会抛异常也不会返回错误码。不主动检测的话，2 万首里被限流的那几千首
    会被静默写成空白，用户拿到表才发现——那时候已经晚了。

    实测（2026-08-07）平台软封可以持续 3 小时以上，所以不能简单「冷却 N 秒后全量恢复」
    —— 那样一恢复就 12 并发怼上去，只会立刻被再封一轮。这里用「冷却 → 放 1 首试水 →
    成了才全开」的探针模式。
    """

    def __init__(self, pacer=None):
        self._lock = threading.Lock()
        self._streak = {}      # 连续空结果计数
        self._until = {}       # 熔断到期时间戳
        self._probing = {}     # 该平台是否已放出试水请求（同一时刻只放 1 个）
        self._pacer = pacer    # 熔断时顺带通知节流器降速

    def is_cooling(self, code):
        """只读版：纯粹问「这个平台还在冷却吗」，不会消耗试水名额。

        blocked() 有副作用（冷却到期时会放行一个探针），所以轮询等待必须用这个。
        """
        with self._lock:
            return time.time() < self._until.get(code, 0)

    def blocked(self, code):
        """True = 现在别打这个平台。冷却到期后只放行 1 个试水请求。"""
        with self._lock:
            until = self._until.get(code, 0)
            if not until:
                return False
            if time.time() < until:
                return True
            # 冷却到期：放 1 首出去探路，其余继续挡着
            if self._probing.get(code):
                return True
            self._probing[code] = True
            print(f'[batch_v2] 平台 {code} 冷却结束，放 1 首试水')
            return False

    def report(self, code, ok):
        """ok=True 表示这次拿到了东西（清零连击）；False 表示空结果。"""
        with self._lock:
            probing = self._probing.pop(code, False)
            if ok:
                self._streak[code] = 0
                if self._until.get(code):
                    self._until[code] = 0
                    print(f'[batch_v2] 平台 {code} 已恢复，全速续跑')
                return
            if probing or self._until.get(code):
                # 试水没过 → 继续压着，隔一段再试
                self._until[code] = time.time() + PLAT_PROBE_INTERVAL
                return
            n = self._streak.get(code, 0) + 1
            self._streak[code] = n
            tripped = n >= PLAT_EMPTY_STREAK_LIMIT
            if tripped:
                self._until[code] = time.time() + PLAT_COOLDOWN_SEC
                self._streak[code] = 0
                print(f'[batch_v2] 平台 {code} 连续 {PLAT_EMPTY_STREAK_LIMIT} 次空结果 '
                      f'→ 判定被限流，暂停 {PLAT_COOLDOWN_SEC}s')
        if tripped and self._pacer:
            self._pacer.penalize(code)      # 锁外调用，别把 pacer 的锁套进来

    def snapshot(self):
        """返回 {平台: 还要冷却几秒}，只含仍在熔断中的。"""
        with self._lock:
            now = time.time()
            return {k: max(0, int(v - now)) for k, v in self._until.items() if v > now}


# ─────────── 全局引擎单例 ───────────
_ENGINE = None
_ENGINE_LOCK = threading.Lock()


def get_engine(app_module, start_workers=True):
    """延迟创建引擎单例。register(app) 末尾调用一次即可。"""
    global _ENGINE
    with _ENGINE_LOCK:
        if _ENGINE is None:
            _ENGINE = _BatchV2Engine(app_module, start_workers=start_workers)
        return _ENGINE


# ─────────── 引擎主体 ───────────
class _BatchV2Engine:
    def __init__(self, app_module, start_workers=True):
        """start_workers=False 时只建对象、不起后台线程。

        Flask 开了 use_reloader 之后，reloader 的父进程也会把 app.py 整个跑一遍，
        于是引擎被创建两份、补跑线程和续跑任务全部翻倍 —— 对外请求量直接 ×2，
        本来能撑 600 首的额度 300 首就打光了。只有真正提供服务的那个进程该起线程。
        """
        self.app_module = app_module
        self._db_path = DB_PATH
        self._db_lock = threading.Lock()        # SQLite 写入串行化（多线程安全）
        self._init_db()
        self.executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=WORKER_POOL_SIZE,
            thread_name_prefix='batchv2',
        )
        self.ne_executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=NETEASE_POOL_SIZE,
            thread_name_prefix='batchv2ne',
        )
        self.pacer = _Pacer()
        self.gate = _PlatGate(self.pacer)
        # 收藏量合批队列：搜索线程只管把命中的歌丢进来，由 flusher 攒批统一查
        self._fav_lock = threading.Lock()
        self._fav_q = {'qq': [], 'kugou': []}   # [(入队时间, item_id, 结果dict), ...]
        self._fav_inflight = set()              # 正在等收藏量的 (平台, item_id)，补跑要避开
        self._finalize_state = {}               # task_id -> {'last_fav':int|None,'stable':int} 收尾稳定性跟踪
        if not start_workers:
            print('[batch_v2] 非服务进程，后台线程不启动（避免 reloader 双开导致请求翻倍）')
            return
        self._fav_flusher = threading.Thread(target=self._fav_flush_loop, daemon=True,
                                             name='batchv2fav')
        self._fav_flusher.start()
        # 补跑线程：平台从限流中恢复后，把之前留空的歌自动捡回来重查
        self._sweeper = threading.Thread(target=self._sweeper_loop, daemon=True,
                                         name='batchv2sweeper')
        self._sweeper.start()
        if TASK_AUTORUN_ON_STARTUP:
            self._autorun_resume()
            self._autofix_v4253()

    # ─── DB ───
    def _connect(self):
        """开 SQLite 连接，遇瞬时锁错自动重试。

        为什么必须重试：worker 池 + 网易云慢速池 + 收藏量合批线程 + 补跑线程 + 仪表盘轮询
        一起抢这个 db 文件时，OS 偶尔会抛 `unable to open database file` 或
        `database is locked`。这是 macOS WAL/SHM 的瞬时竞争，不是真错。
        不重试的话 row 会被 _mark_item_failed 永久打入 failed —— 2 万首规模下
        哪怕只有 0.01% 概率，每跑 1 万首也会有 1 首默默丢，用户完全感知不到。

        行为：3 次重试，间隔 0.2s/0.4s/0.8s 退避。只对瞬时 OperationalError 重试，
        其他错误（语法错、磁盘满、权限错）原样上抛，不掩盖真问题。
        """
        last_err = None
        for attempt in range(3):
            try:
                conn = sqlite3.connect(self._db_path, timeout=30, isolation_level=None)
                conn.execute('PRAGMA journal_mode=WAL')
                conn.execute('PRAGMA synchronous=NORMAL')
                conn.row_factory = sqlite3.Row
                return conn
            except sqlite3.OperationalError as e:
                msg = str(e)
                if 'unable to open database file' in msg or 'database is locked' in msg:
                    last_err = e
                    time.sleep(0.2 * (2 ** attempt))
                    continue
                raise
        # 三次都败，原始异常上抛（外层 _mark_item_failed 会兜底）
        raise last_err

    def _init_db(self):
        with self._db_lock, self._connect() as c:
            c.executescript('''
                CREATE TABLE IF NOT EXISTS tasks (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    name        TEXT NOT NULL,
                    status      TEXT NOT NULL DEFAULT 'pending',
                    total       INTEGER NOT NULL DEFAULT 0,
                    done        INTEGER NOT NULL DEFAULT 0,
                    failed      INTEGER NOT NULL DEFAULT 0,
                    options_json TEXT,
                    error       TEXT,
                    created_at  REAL NOT NULL,
                    started_at  REAL,
                    finished_at REAL
                );
                CREATE TABLE IF NOT EXISTS task_items (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id       INTEGER NOT NULL,
                    idx           INTEGER NOT NULL,
                    song_name     TEXT NOT NULL,
                    performer     TEXT DEFAULT '',
                    lyricist      TEXT DEFAULT '',
                    composer      TEXT DEFAULT '',
                    status        TEXT NOT NULL DEFAULT 'pending',
                    attempts      INTEGER NOT NULL DEFAULT 0,
                    last_error    TEXT,
                    qq_url        TEXT DEFAULT '',
                    qq_favorites  INTEGER,
                    kugou_url        TEXT DEFAULT '',
                    kugou_favorites  INTEGER,
                    netease_url      TEXT DEFAULT '',
                    netease_favorites INTEGER,
                    started_at   REAL,
                    finished_at  REAL
                );
                CREATE INDEX IF NOT EXISTS idx_items_task ON task_items(task_id, status);
            ''')
            # 兼容旧库：补每平台独立的完成标记。
            # 为什么每个平台要单独记：平台被限流时是「HTTP 200 + 空结果」，不报错。
            # 只用一个总的 status 的话，被限流那段时间的歌会被当成「查过了、没有」，
            # 悄悄写成空白列。分平台记录 + 平台恢复后自动补跑，才不会漏。
            for _sql in (
                'ALTER TABLE task_items ADD COLUMN netease_done INTEGER NOT NULL DEFAULT 0',
                "ALTER TABLE task_items ADD COLUMN netease_error TEXT DEFAULT ''",
                'ALTER TABLE task_items ADD COLUMN qq_done INTEGER NOT NULL DEFAULT 0',
                'ALTER TABLE task_items ADD COLUMN kugou_done INTEGER NOT NULL DEFAULT 0',
                'ALTER TABLE task_items ADD COLUMN qq_tries INTEGER NOT NULL DEFAULT 0',
                'ALTER TABLE task_items ADD COLUMN kugou_tries INTEGER NOT NULL DEFAULT 0',
                # 平台内部 ID：收藏量接口认的不是链接里的那个 ID
                #   QQ   收藏量要数字 songId（链接里只有 mid，反解不出来）
                #   酷狗 收藏量要 album_audio_id（链接里只有 hash，同样反解不出来）
                # 搜索时顺手存下来，补跑「只缺收藏量」的歌就不用再搜一遍 —— 这一步
                # 能省掉 600+ 次无谓搜索，对限流额度是实打实的节省。
                "ALTER TABLE task_items ADD COLUMN qq_songid TEXT DEFAULT ''",
                "ALTER TABLE task_items ADD COLUMN kugou_mixsongid TEXT DEFAULT ''",
                # 匹配质量标签。平台上遍地是蹭原唱名的 UGC 翻唱号（「周杰伦./Asasblue」），
                # 它们的收藏量跟原版能差 5 个数量级。必须把「这条数据到底匹配得准不准」
                # 如实写进结果，否则用户拿翻唱的数字当原版用，决策直接跑偏。
                "ALTER TABLE task_items ADD COLUMN qq_match TEXT DEFAULT ''",
                "ALTER TABLE task_items ADD COLUMN kugou_match TEXT DEFAULT ''",
                "ALTER TABLE task_items ADD COLUMN netease_match TEXT DEFAULT ''",
                # 未收录「2 次确认」计数：区分「平台限流」还是「真的搜不到」。
                #   连续 2 次「平台不在限流 + 返回空」→ 判定确认未收录（留 NULL，不填 0）。
                #   被风控挡回 → 重置计数，等平台恢复再做第 2 次确认。
                "ALTER TABLE task_items ADD COLUMN qq_confirms INTEGER NOT NULL DEFAULT 0",
                "ALTER TABLE task_items ADD COLUMN kugou_confirms INTEGER NOT NULL DEFAULT 0",
                "ALTER TABLE task_items ADD COLUMN netease_confirms INTEGER NOT NULL DEFAULT 0",
                "ALTER TABLE task_items ADD COLUMN netease_tries INTEGER NOT NULL DEFAULT 0",
                # 任务级收尾标记：所有歌都 resolved（有数据 or 确认未收录）后，
                # status='completed' 才是真话（每一首都对应了数据，未收录的留 NULL 不填 0）。
                "ALTER TABLE tasks ADD COLUMN finalized INTEGER NOT NULL DEFAULT 0",
                # v4.25.16「最近查到的歌」实时刷新：last_touched_at 记录这行被
                # 真正处理过的时刻（主批量完成 / 补跑命中 / 网易云补回任一都刷），
                # 与首轮 finished_at 解耦，避免补跑阶段 UI 面板定格不动。
                "ALTER TABLE task_items ADD COLUMN last_touched_at REAL",
            ):
                try:
                    c.execute(_sql)
                except Exception:
                    pass
            # 回填：新加的 *_done 列默认 0，会让补跑线程把历史上早就查全的歌
            # 全部重跑一遍（1386 条老数据 → 白白再打 2700 次请求，还会立刻把平台打封）。
            # 数据齐全（链接 + 收藏量都有）的直接认定已完成。
            for _p in ('qq', 'kugou'):
                try:
                    c.execute(
                        f"UPDATE task_items SET {_p}_done=1 "
                        f"WHERE {_p}_done=0 AND COALESCE({_p}_url,'')<>'' AND {_p}_favorites IS NOT NULL"
                    )
                except Exception:
                    pass
            # 重置「旧代码已放弃 / 已标未收录」但还没做 2 次确认的歌，让新代码重新走
            # 「2 次干净确认」流程，真正区分是平台限流还是真的搜不到。
            # 已确认（confirms>=2）的不再动，避免每次重启都把确认过的歌重搜一遍。
            # 只动「链接空 + 收藏量空」的歌（有链接的歌是另一回事，不在此列）。
            for _p in ('qq', 'kugou', 'netease'):
                try:
                    c.execute(
                        f"UPDATE task_items SET {_p}_done=0, {_p}_confirms=0 "
                        f"WHERE {_p}_done=1 AND {_p}_url='' AND {_p}_favorites IS NULL "
                        f"AND COALESCE({_p}_confirms,0) < 2"
                    )
                except Exception:
                    pass

    # ─── 任务创建 ───
    def submit(self, name, songs):
        """songs: [{song_name, performer, lyricist, composer}, ...]"""
        name = (name or '批量任务').strip()[:100]
        if not songs:
            raise ValueError('songs 不能为空')
        now = time.time()
        with self._db_lock, self._connect() as c:
            c.execute('BEGIN')
            try:
                c.execute(
                    'INSERT INTO tasks (name, status, total, options_json, created_at) '
                    'VALUES (?, ?, ?, ?, ?)',
                    (name, 'pending', len(songs), json.dumps({'version': 'v2'}), now),
                )
                task_id = c.execute('SELECT last_insert_rowid()').fetchone()[0]
                c.executemany(
                    'INSERT INTO task_items (task_id, idx, song_name, performer, lyricist, composer) '
                    'VALUES (?, ?, ?, ?, ?, ?)',
                    [(task_id, i, s['song_name'], s.get('performer', ''),
                      s.get('lyricist', ''), s.get('composer', ''))
                     for i, s in enumerate(songs)],
                )
                c.execute('COMMIT')
            except Exception:
                c.execute('ROLLBACK')
                raise
        # 触发后台处理
        self._kickoff_task(task_id)
        return task_id

    def _kickoff_task(self, task_id):
        """把任务里所有项重新分发：QQ/酷狗 走快速池，网易云 走慢速池。"""
        with self._db_lock, self._connect() as c:
            row = c.execute(
                "SELECT status FROM tasks WHERE id=?", (task_id,)
            ).fetchone()
            if row is None:
                return
            if row['status'] == 'cancelled':
                return
            # status='completed' 只代表阶段1（QQ/酷狗）跑完了，网易云第二阶段可能还欠着。
            # 这种任务不能直接 return，否则重启后那批网易云字段会永远留空。
            ne_only_mode = (row['status'] == 'completed')
            if ne_only_mode:
                fast_ids = []
            else:
                # 标记 running（仅当还是 pending 或上轮中断的 running）
                c.execute(
                    "UPDATE tasks SET status='running', started_at=COALESCE(started_at, ?) "
                    "WHERE id=? AND status IN ('pending','running')",
                    (time.time(), task_id),
                )
                # 残留的 running item 改回 pending（崩溃后接续）
                c.execute(
                    "UPDATE task_items SET status='pending' "
                    "WHERE task_id=? AND status='running'",
                    (task_id,),
                )
                # 计数器按实际行数重算：退回队列的歌不能还挂在 failed 上，
                # 否则进度条永远显示 100% / 有失败，用户会以为跑完了。
                c.execute(
                    "UPDATE tasks SET "
                    "  done  =(SELECT COUNT(*) FROM task_items WHERE task_id=tasks.id AND status='done'),"
                    "  failed=(SELECT COUNT(*) FROM task_items WHERE task_id=tasks.id AND status='failed') "
                    "WHERE id=?",
                    (task_id,),
                )
                fast_ids = [
                    r[0] for r in c.execute(
                        "SELECT id FROM task_items WHERE task_id=? AND status='pending' ORDER BY idx",
                        (task_id,),
                    ).fetchall()
                ]
            # 阶段1 已完、但网易云还没跑的项，交给慢速池续跑
            ne_ids = [
                r[0] for r in c.execute(
                    "SELECT id FROM task_items WHERE task_id=? AND status='done' AND netease_done=0 ORDER BY idx",
                    (task_id,),
                ).fetchall()
            ]
        for iid in fast_ids:
            self.executor.submit(self._process_item_safe, task_id, iid)
        for iid in ne_ids:
            self.ne_executor.submit(self._process_netease_safe, task_id, iid)

    # ─── 单首处理 ───
    def _process_item_safe(self, task_id, item_id):
        """包装层：错峰 + 异常吞掉 + 计数器更新。"""
        time.sleep(random.uniform(0, ITEM_STAGGER_SEC))
        # v4.27.12 修复：不再让工人「原地睡等平台恢复」。
        # 旧逻辑（STAGE1_MAX_WAIT_SEC=3600）在 QQ+酷狗 双冷却时让【所有】工人线程
        # 睡等最长 1 小时，等于把整个 executor 冻死——表现就是「上传 2000 首→点搜索
        # →进度条卡死→实际没查」。而 _v2_search_fast 在平台冷却时会【立即】返回
        # state='blocked'（不轰接口、不耗配额），item 随后被标记待 sweeper 补跑重试，
        # 根本不需要睡等。睡等只会白白冻结队列 1 小时，毫无收益。
        # 所以这里跳过睡等，直接领发车名额去跑；冷却的平台自然走 blocked→sweeper 重试。
        self.pacer.acquire()          # 领发车名额（撞过限流的话这里会自动变慢）
        try:
            ok = self._process_item(task_id, item_id)
        except Exception as e:
            if _is_shutdown_err(e):
                # 进程正在退出（重启/热重载/关机）导致的，不是这首歌的问题。
                # 标 failed 会让它永久出局（补跑线程只捞 status='done' 的），
                # 2 万首跑一夜中途重启一次就会静默丢几十首 —— 必须退回 pending。
                self._requeue_item(item_id)
                return
            # 兜底：理论上 _process_item 已捕获；这是双重保险
            self._mark_item_failed(item_id, 'unhandled: ' + repr(e)[:200])
            ok = False
        # 计数 + 任务完成判定
        self._on_item_finished(task_id, ok)

    def _process_item(self, task_id, item_id, sweep=False):
        """阶段1（QQ/酷狗）。sweep=True 表示这是补跑，只补还没拿到的那个平台。"""
        with self._db_lock, self._connect() as c:
            row = c.execute(
                "SELECT * FROM task_items WHERE id=?", (item_id,)
            ).fetchone()
            if not row:
                return False
            if not sweep and row['status'] != 'pending':
                return row['status'] == 'done'
            item = dict(row)

        # 只查还没搞定的平台（补跑时能省掉一半请求，也避免重复打已限流的平台）。
        # 补跑(sweep)模式下：首轮已「放弃」(p_done=1 且链接空) 的歌，只要还没撞到
        # ABSOLUTE_MAX_TRIES 总上限，就再给它一次机会——平台从限流恢复后这些歌
        # 大多能补回来，否则它们会永久成空格、导出一片空白列。
        need = [p for p in ('qq', 'kugou')
                if (not item.get(p + '_done'))
                or (item.get(p + '_url') in (None, '')
                    and (item.get(p + '_confirms') or 0) < 2
                    and (item.get(p + '_tries') or 0) < ABSOLUTE_MAX_TRIES)]

        # ── 快速通道：链接和平台 ID 都齐了，只差收藏量 ──
        # 这种歌不用再搜一遍，直接丢进收藏量合批队列。补跑场景下 600+ 首都属于
        # 这一类，省下来的搜索请求全是限流额度。
        fav_only = []
        for p in list(need):
            if item.get(p + '_favorites') is not None:
                continue
            pid = (item.get(_ID_COL[p]) or '').strip()
            if pid and (item.get(p + '_url') or ''):
                fav_only.append(p)
                need.remove(p)
        for p in fav_only:
            self._enqueue_fav(p, item_id, {_ID_FIELD[p]: item.get(_ID_COL[p])})

        if not need:
            # v4.25.16：补跑阶段「这歌所有平台都已 resolved」也算被处理过，刷 last_touched_at
            # 让 UI「最近查到的歌」面板能看到后台仍在动（否则 sweeper 静默无事可做时面板定格）。
            try:
                with self._db_lock, self._connect() as c:
                    c.execute("UPDATE task_items SET last_touched_at=? WHERE id=?",
                              (time.time(), item_id))
            except Exception:
                pass
            if self._netease_needs_work(item):
                self.ne_executor.submit(self._process_netease_safe, task_id, item_id)
            return True

        try:
            # 已存库的兄弟平台链接作为「已知命中」，供跨平台回填借名（不重搜、只借精确歌名+歌手）
            known_hits = {}
            if item.get('qq_url'):
                known_hits['qq'] = {'song_name': item['song_name'],
                                    'performer': item['performer']}
            if item.get('kugou_url'):
                known_hits['kugou'] = {'song_name': item['song_name'],
                                       'performer': item['performer']}
            res = self._v2_search_fast(item['song_name'], item['performer'], need,
                                       item.get('lyricist', '') or '', item.get('composer', '') or '',
                                       known_hits)
            last_err = None
        except Exception as e:
            res = {p: {'state': 'error', 'picked': None} for p in need}
            last_err = f'{type(e).__name__}: {e}'[:200]

        def _fav(b):
            if not b:
                return None
            v = b.get('collection_count')
            try:
                return int(v) if v is not None else None
            except Exception:
                return None

        def _url(b):
            if not b:
                return ''
            return (b.get('song_url') or b.get('link') or '') or ''

        updates = {'finished_at': time.time(), 'last_touched_at': time.time()}
        fav_targets = []                       # [(平台, 命中的结果dict), ...] 等会儿丢进合批队列
        if not sweep:
            # 阶段1 首轮：无论平台答没答，这首歌的「第一遍」就算走完了，
            # 状态置 done 是为了放行网易云。QQ/酷狗 缺的部分靠 tries + 补跑线程兜。
            updates.update({'status': 'done', 'attempts': (item.get('attempts') or 0) + 1,
                            'last_error': last_err, 'started_at': item.get('started_at') or time.time()})

        for p in need:
            # 缺口行识别：另一平台已确认有这首歌(QQ/酷狗 兄弟链接在库) → 本次弱势平台的
            # 「搜不到」极可能是限流误判，留可重试、不急判未收录（跨平台回填借名重搜）。
            brother_has = bool((p == 'kugou' and (item.get('qq_url') or '')) or
                              (p == 'qq' and (item.get('kugou_url') or '')))
            r = res.get(p) or {'state': 'error', 'picked': None}
            st = r['state']
            if st == 'blocked':
                continue                      # 熔断跳过，不算一次尝试
            # 被风控挡回（HTTP 200 空列表 + 平台正在熔断）：这次的「空」不可信，
            # 既不算尝试，也不累计确认次数——之前的「空结果」要作废重来。
            throttled = st in ('empty', 'error') and self.gate.is_cooling(p)
            tries = (item.get(p + '_tries') or 0) + (0 if throttled else 1)
            updates[p + '_tries'] = tries
            best = (r['picked'] or [None])[0]

            if st == 'searched' and best:
                _label = str((r.get('picked') or ['', ''])[1] or '')[:60]
                # v4.25.2 写库铁律：只有「精准匹配 / 四维全中 / 多艺人合唱」才落库。
                # 「存疑(疑似翻唱) / 其他:xxx / 近似」等标签一律视为未收录，
                # 绝不把别人（尤其大艺人同名）的收藏量算到用户头上。
                _is_match = (_label.startswith('精准匹配')
                             or _label.startswith('匹配('))
                if _is_match:
                    # 链接立刻落库（就算收藏量后面查不到，链接也是用户要的一半数据）
                    # v4.25.19 防御：候选可能只带平台内部 ID、没带 song_url（限流降级/
                    # 历史搜索结构差异），此时用存的 ID 兜底重建规范链接，
                    # 绝不留下「有匹配却无链接」的残缺行（数据完整性优先）。
                    _u = _url(best)
                    if not _u:
                        _sid = (best.get('_songid') or best.get('_songmid')
                                or best.get('_mixsongid') or best.get('_hash') or '')
                        if _sid:
                            _u = _canon_url(p, str(_sid))
                    updates[p + '_url'] = _u
                    updates[p + '_match'] = _label
                    pid = _plat_id(p, best)
                    if pid:
                        updates[_ID_COL[p]] = pid    # 存平台内部 ID，补跑时免重搜
                    fv = _fav(best)
                    if fv is not None:
                        updates[p + '_favorites'] = fv
                        updates[p + '_done'] = 1
                    else:
                        fav_targets.append((p, best))   # 排队等合批查收藏量
                else:
                    # 非匹配标签（如酷狗返回了结果但歌名/歌手严格匹配未中）。
                    # 缺口行特例：兄弟平台(QQ)已确认有这首歌，本次酷狗严格匹配未中多半是
                    # 「限流期拿到了错的候选」或「歌名在酷狗里的写法不同」。不急判未收录，
                    # 留 _done=0 让 sweeper 借兄弟精确歌名+歌手重搜（跨平台回填）。
                    # 带重试上限 GAP_MAX_TRIES 防死循环：到了仍非匹配才放弃（留 NULL，不填 0）。
                    if brother_has and tries < GAP_MAX_TRIES:
                        # 缺口行：留可重试，sweeper 下次借名重搜。不置 done / 不写脏 label。
                        pass
                    else:
                        updates[p + '_done'] = 1
                        updates[p + '_confirms'] = 2
                        updates[p + '_match'] = _label
            elif st == 'ok':
                # 平台显式答了「没有这首歌」→ 强确认未收录。
                # 实事求是：url 留空、收藏量留 NULL，绝不填 0。
                updates[p + '_done'] = 1
                updates[p + '_confirms'] = 2
            elif throttled:
                # 被风控挡回 → 之前的「空结果」不可信，重置确认计数，原样挂着等平台恢复。
                updates[p + '_confirms'] = 0
            else:
                # empty / error 且未被熔断：分不清「冷门歌真没有」还是「软限流挡了」。
                # 实事求是原则：先不填任何值，累计「干净空结果」确认次数，
                # 必须经过 2 次确认才能判定为「真的搜不到」（否则留 NULL 等补跑）。
                #   缺口行(兄弟平台已有这首歌)：放宽到 GAP_MAX_TRIES 次干净空才肯判未收录，
                #   给足重试机会（限流窗口一过，跨平台回填借名就能补回来），绝不「假完成」。
                #   撞到总尝试上限 ABSOLUTE_MAX_TRIES 仍未确认 → 放弃，留 NULL
                #     （标记「未确认/可能限流」，导出也是 null，界面单独提示）。
                confirms = (item.get(p + '_confirms') or 0) + 1
                updates[p + '_confirms'] = confirms
                _cap = GAP_MAX_TRIES if brother_has else 2
                if confirms >= _cap:
                    updates[p + '_done'] = 1   # 达到确认阈值 → 确认未收录（NULL，不填 0）
                elif tries >= ABSOLUTE_MAX_TRIES:
                    updates[p + '_done'] = 1   # 总尝试到顶仍未确认 → 放弃，留 NULL（未确认）
                # 否则不置 done → 补跑线程下次再确认

        with self._db_lock, self._connect() as c:
            sets = ', '.join(f"{k}=?" for k in updates.keys())
            c.execute(f"UPDATE task_items SET {sets} WHERE id=?",
                      (*updates.values(), item_id))

        for p, best in fav_targets:
            self._enqueue_fav(p, item_id, best)

        if self._netease_needs_work(item):
            self.ne_executor.submit(self._process_netease_safe, task_id, item_id)
        return True

    # ─── 收藏量合批 ───
    def _enqueue_fav(self, platform, item_id, row):
        with self._fav_lock:
            # ⚠️ 必须按 (平台, 歌) 去重，不能只按歌。
            # 同一首歌 QQ 和酷狗的收藏量是两笔独立请求，只按 item_id 去重会把
            # 后入队的那个平台整个吞掉 —— 表现就是「酷狗收藏量永远是空的」。
            if (platform, item_id) in self._fav_inflight:
                return                  # 已在队列/请求中，别重复排队（补跑线程会反复扫到同一首）
            self._fav_q[platform].append((time.time(), item_id, row))
            self._fav_inflight.add((platform, item_id))

    def _fav_flush_loop(self):
        while True:
            time.sleep(0.5)
            for p in ('qq', 'kugou'):
                try:
                    self._fav_flush(p)
                except Exception as e:
                    print(f'[batch_v2] 收藏量合批({p})异常:', repr(e)[:180])

    def _fav_flush(self, platform):
        fav_code = platform + ':fav'
        if self.gate.is_cooling(fav_code):
            return                      # 该通道在冷却，队列先囤着，恢复了一起发
        with self._fav_lock:
            q = self._fav_q[platform]
            if not q:
                return
            ready = len(q) >= FAV_BATCH_SIZE or (time.time() - q[0][0]) >= FAV_FLUSH_SEC
            if not ready:
                return
            batch, self._fav_q[platform] = q[:FAV_BATCH_SIZE], q[FAV_BATCH_SIZE:]

        if self.gate.blocked(fav_code):       # 冷却刚到期、试水名额被别人占了
            with self._fav_lock:
                self._fav_q[platform] = batch + self._fav_q[platform]
            return

        rows = [b[2] for b in batch]
        app = self.app_module
        try:
            if platform == 'qq':
                app._fetch_qq_details(rows, app.get_cookie_string('qq'), only_fav=True)
            else:
                app._fetch_kugou_collection_counts(rows)
        except Exception as e:
            print(f'[batch_v2] 收藏量批量请求({platform})失败:', repr(e)[:150])

        got = sum(1 for r in rows if r.get('collection_count') is not None)
        self.gate.report(fav_code, got > 0)

        if got == 0:
            # 整批颗粒无收 = 接口被挡了，退回队列等下一轮（不消耗 tries）
            with self._fav_lock:
                self._fav_q[platform] = batch + self._fav_q[platform]
            return

        with self._db_lock, self._connect() as c:
            for _ts, item_id, row in batch:
                cnt = row.get('collection_count')
                if cnt is None:
                    continue              # 这首没查到，留给补跑
                c.execute(
                    f"UPDATE task_items SET {platform}_favorites=?, {platform}_done=1 WHERE id=?",
                    (cnt, item_id),
                )
        with self._fav_lock:
            # 整批（不论查到没查到）都从 inflight 摘掉：
            # 查到的已落库；没查到的交给补跑线程下一轮重来。
            self._fav_inflight -= {(platform, b[1]) for b in batch}

    def _mark_item_failed(self, item_id, err):
        with self._db_lock, self._connect() as c:
            c.execute(
                "UPDATE task_items SET status='failed', last_error=?, finished_at=? WHERE id=?",
                (err[:200], time.time(), item_id),
            )

    def _requeue_item(self, item_id):
        """把因为进程退出而中断的歌退回队列，下次启动 _autorun_resume 会续跑。"""
        try:
            with self._db_lock, self._connect() as c:
                c.execute(
                    "UPDATE task_items SET status='pending', last_error='中断，已重新排队' "
                    "WHERE id=? AND status<>'done'", (item_id,)
                )
        except Exception:
            pass          # 解释器都在关了，写不进去也别再抛异常

    # ─── 补跑线程：平台恢复后把留空的歌捡回来 ───
    def _sweeper_loop(self):
        """常驻后台。每 SWEEP_INTERVAL_SEC 扫一次「首轮跑完但某个平台没拿到数据」的歌。

        为什么必须有它：QQ/酷狗 的软限流是 HTTP 200 + 空列表，一封就是几小时。
        2 万首跑下来必然会撞上，撞上的那几千首如果不回头补，导出的表就是一片空白列。
        """
        while True:
            time.sleep(SWEEP_INTERVAL_SEC)
            try:
                self._sweep_once()
            except Exception as e:
                print('[batch_v2] 补跑线程异常:', repr(e)[:200])
            # 收尾：把「已跑干净、连续多轮补跑无新进展」的任务标记 completed，
            # 剩余 NULL 收藏量全部填 0（真没收录=0 收藏量），网页上的「已完成」才是真话。
            try:
                with self._db_lock, self._connect() as c:
                    tids = [r[0] for r in c.execute(
                        "SELECT id FROM tasks WHERE status='running'").fetchall()]
                for tid in tids:
                    self._maybe_finalize(tid)
            except Exception as e:
                print('[batch_v2] 收尾检查异常:', repr(e)[:200])
            # v4.25.15 残缺行兜底：有收藏量但没链接（旧两阶段写库遗留），重搜补回链接。
            try:
                self._sweep_orphans()
            except Exception as e:
                print('[batch_v2] 残缺行兜底异常:', repr(e)[:200])

    def _sweep_once(self):
        # 两个平台都还在熔断里就别扫了，省得白跑。
        # ⚠️ 这里必须用只读的 is_cooling —— 用 blocked() 会把「冷却到期后的试水名额」
        # 在这儿就消耗掉，真正干活的线程再问就永远是 True，探针永远不会被归还，
        # 平台就再也恢复不了了。
        # 搜索通道和收藏量通道是两个独立的限流桶，要分开判断：
        #   · 搜索活着 → 捞「还没搜到」的（受 tries 上限约束）
        #   · 收藏量活着 → 捞「链接和 ID 都有、只差收藏量」的（不受 tries 约束，
        #     因为它只花一次合批请求，且 tries 是给「搜不到」计数的，跟这儿无关）
        cond_list = []
        for p in ('qq', 'kugou'):
            parts = []
            if not self.gate.is_cooling(p):
                # 只捡还没 resolved 的歌（含「待第 2 次确认未收录」的）。
                # 已确认未收录(p_done=1 且 confirms>=2)的不再动，避免反复重搜。
                parts.append(f"ti.{p}_done=0")
            if not self.gate.is_cooling(p + ':fav'):
                parts.append(f"(ti.{p}_favorites IS NULL AND ti.{_ID_COL[p]} <> '')")
            if parts:
                cond_list.append(' OR '.join(parts))
        # 网易云是独立慢速池（_process_netease 内部错峰），只要它还没 resolved 也纳入补跑，
        # 否则「仅网易云非精准」的行即便被复位也永远不会被重抓（v4.25.2）。
        cond_list.append("ti.netease_done=0")
        if not cond_list:
            return

        # ── 前台优先 ──
        # 补跑和首轮共用同一条节流线（pacer）和同一个线程池。历史积压动辄上千首，
        # 一轮捞 200 首就能把配额吃干净，新提交的任务几分钟一首都跑不出来。
        # 所以：① 队列已经堆高了就整轮跳过；② 还有歌没跑过第一遍时，补跑降为涓流。
        try:
            if self.executor._work_queue.qsize() > SWEEP_QUEUE_CEIL:
                return
        except Exception:
            pass
        with self._db_lock, self._connect() as c:
            busy = c.execute(
                "SELECT 1 FROM task_items ti JOIN tasks t ON t.id=ti.task_id "
                "WHERE ti.status='pending' AND t.status='running' LIMIT 1"
            ).fetchone() is not None
        limit = SWEEP_BATCH_BUSY if busy else SWEEP_BATCH

        running_conds = ' OR '.join(cond_list)
        # ⚠️ v4.27.12 修复：已完成任务只补网易云（第二阶段），不再扫 QQ/酷狗。
        # 旧逻辑对 status IN ('running','completed') 一视同仁，导致 8 天前 19590 首的
        # 旧任务（#26）被无限反复扫掠 QQ/酷狗 残缺行（日志 4358 次"补跑捡回"），
        # 霸占共享 executor 队列，把新提交的批量任务饿死在 pending——
        # 表现就是「上传 2000 首→点搜索→进度条卡死→实际没查」。
        # 已完成任务的 QQ/酷狗 残空已由 _maybe_finalize 判定为「未收录/放弃」并留 NULL，
        # 不应再自动回头扫；用户要补可手动点「补跑空结果」。
        completed_conds = "ti.netease_done=0"
        sql = f'''
            SELECT ti.id AS iid, ti.task_id AS tid
            FROM task_items ti
            JOIN tasks t ON t.id = ti.task_id
            WHERE ti.status = 'done'
              AND (
                (t.status = 'running' AND ({running_conds}))
                OR (t.status = 'completed' AND {completed_conds})
              )
              AND NOT EXISTS (
                  SELECT 1 FROM task_items p
                  WHERE p.task_id = ti.task_id AND p.status = 'pending'
              )
            ORDER BY ti.id
            LIMIT ?
        '''
        with self._db_lock, self._connect() as c:
            rows = c.execute(sql, (limit,)).fetchall()
        if not rows:
            return
        # 已经在收藏量队列里排队的别动，否则会白白再搜一次。
        # inflight 现在是 (平台, item_id)，这里保守地按歌粒度跳过：
        # 最多延后一个 sweep 周期，等它 flush 完下轮自然会被捡回来。
        with self._fav_lock:
            inflight = {iid for _p, iid in self._fav_inflight}
        rows = [r for r in rows if r['iid'] not in inflight]
        if not rows:
            return
        _live = [p for p in ('qq', 'kugou')
                 if not self.gate.is_cooling(p) or not self.gate.is_cooling(p + ':fav')]
        print(f'[batch_v2] 补跑：捡回 {len(rows)} 首（可用通道 {"/".join(_live) or "无"}'
              f'{"，涓流模式-前台优先" if busy else ""}）')
        for r in rows:
            self.executor.submit(self._sweep_item_safe, r['tid'], r['iid'])

    def _sweep_item_safe(self, task_id, item_id):
        """补跑单首：只补缺的平台，不动任务进度计数（那是首轮的事）。"""
        self.pacer.acquire()          # 补跑和首轮共用同一条节流线，别把刚恢复的平台又打死
        try:
            self._process_item(task_id, item_id, sweep=True)
        except Exception:
            pass

    def _sweep_orphans(self):
        """v4.25.15 残缺行兜底：有收藏量但没链接（p_done=1 & p_favorites>0 & p_url=''）。

        这些行是旧两阶段写库遗留：_pick_best_for_batch 把变体歌手（dollyy99 / 分手厨房张铎曦）
        判「存疑翻唱」拒收 → url 空，但收藏量已先写入。普通补跑按 p_done=0 不会捞到它们。
        现在 _pick_best_for_batch 已支持变体归一（_alias_variants），重搜多半能接受并填回 url。
        复用搜索通道（受 _PlatGate 限流/熔断保护），不新增额度压力；每平台最多重审
        ORPHAN_MAX_TRIES 次，仍拒收则放弃（保留残缺，不写脏数据）。
        """
        targets = []
        with self._db_lock, self._connect() as c:
            for p in ('qq', 'kugou'):
                if self.gate.is_cooling(p):
                    continue          # 平台在冷却就别凑热闹
                rows = c.execute(f'''
                    SELECT ti.id AS iid, ti.task_id AS tid
                    FROM task_items ti JOIN tasks t ON t.id=ti.task_id
                    WHERE ti.status='done'
                      AND t.status='running'
                      AND ti.{p}_done=1
                      AND ti.{p}_favorites IS NOT NULL
                      AND COALESCE(ti.{p}_url,'')=''
                      AND COALESCE(ti.{p}_tries,0) < ?
                    ORDER BY ti.id LIMIT ?
                ''', (ORPHAN_MAX_TRIES, SWEEP_BATCH)).fetchall()
                targets.extend((r['tid'], r['iid'], p) for r in rows)
        if not targets:
            return
        # 前台优先：队列堆积高时整轮跳过
        try:
            if self.executor._work_queue.qsize() > SWEEP_QUEUE_CEIL:
                return
        except Exception:
            pass
        with self._fav_lock:
            inflight = {iid for _p, iid in self._fav_inflight}
        targets = [t for t in targets if t[1] not in inflight]
        if not targets:
            return
        print(f'[batch_v2] 残缺行兜底：捡回 {len(targets)} 首')
        for tid, iid, p in targets:
            self.executor.submit(self._sweep_orphan_item_safe, tid, iid, p)

    def _sweep_orphan_item_safe(self, task_id, item_id, platform):
        self.pacer.acquire()
        try:
            self._sweep_orphan_item(task_id, item_id, platform)
        except Exception:
            pass

    def _sweep_orphan_item(self, task_id, item_id, platform):
        """残缺行重审单首：重搜该平台，接受即填 url（保留收藏量），拒收累计 tries。"""
        with self._db_lock, self._connect() as c:
            row = c.execute("SELECT * FROM task_items WHERE id=?", (item_id,)).fetchone()
            if not row:
                return
            item = dict(row)
        if self.gate.blocked(platform):
            return
        song = (item.get('song_name') or '').strip()
        perf = (item.get('performer') or '').strip()
        lyr = (item.get('lyricist') or '').strip()
        comp = (item.get('composer') or '').strip()
        try:
            if platform == 'qq':
                res = self.app_module.search_qq(f'{song} {perf}'.strip(), 30)
            else:
                res = self.app_module.search_kugou(f'{song} {perf}'.strip(), 30, light=True)
        except Exception:
            return
        if not res:
            return
        best, label = self.app_module._pick_best_for_batch(
            res, platform, song, perf, lyr, comp, self._enrich_for_4d)
        with self._db_lock, self._connect() as c:
            tries = (item.get(platform + '_tries') or 0) + 1
            if best and (label or '').startswith(('精准匹配', '匹配(')):
                url = (best.get('song_url') or best.get('link') or '').strip()
                if url:
                    c.execute(
                        f"UPDATE task_items SET {platform}_url=?, {platform}_match=?, "
                        f"{platform}_done=1, {platform}_tries=? WHERE id=?",
                        (url, label[:60], tries, item_id))
                    print(f'[batch_v2][残缺补] ✅ {platform} 《{song}》- {perf} 补回链接')
                    return
            # 仍拒收：到上限(ORPHAN_MAX_TRIES)则清掉旧 bug 留下的假收藏量（导出表干净），
            # 未到上限保留，下个 sweep 周期再借名重搜。
            if tries >= ORPHAN_MAX_TRIES:
                c.execute(
                    f"UPDATE task_items SET {platform}_favorites=NULL, "
                    f"{platform}_match=?, {platform}_tries=? WHERE id=?",
                    (label[:60] if label else '', tries, item_id))
                print(f'[batch_v2][残缺补] 🧹 {platform} 《{song}》- {perf} '
                      f'重试{ORPHAN_MAX_TRIES}次仍拒收，清假收藏量')
            else:
                c.execute(f"UPDATE task_items SET {platform}_tries=? WHERE id=?", (tries, item_id))

    def _enrich_for_4d(self, result, code):
        """四维校验补抓：给单首候选补词曲作者（best-effort，失败/缺词曲返回 None）。

        v4.25.2：对「歌手名相似但未精确(asc 70~99)」的最佳候选，批量查歌在挑出版本后
        会调用它补抓词曲详情，词曲均命中曲库即升「精准匹配(四项全中)」。
        只在 QQ / 网易云 启用 —— 酷狗词曲接口（歌词 API）最不稳且最易限流，
        本次四维校验跳过它，靠歌手名归一兜底即可。
        """
        try:
            time.sleep(0.35)  # 细节接口也留点节流余量，别把搜索额度打爆
            if code == 'qq':
                self.app_module._fetch_qq_details([result], skip_comments=True)
            elif code == 'netease':
                self.app_module._fetch_netease_details([result])
            else:
                return None
            if result.get('lyricist') or result.get('composer'):
                return result
        except Exception:
            return None
        return None

    def _v2_search_fast(self, song_name, performer, need=('qq', 'kugou'),
                        lyricist='', composer='', known_hits=None):
        """阶段1：QQ/酷狗 精简搜索。

        刻意绕开 search_all —— 后者会给整页 30 条结果统一补「词曲作者/评论数/在听人数/
        唱片公司/Discogs 厂牌」，单首歌要发几十甚至上百个请求。2 万首规模下这既慢
        （实测单首 6~25s）又必然触发平台风控。
        这里只做三步：搜列表(light) → 挑出命中的那一首 → 只给这一首补收藏量。

        返回 {平台: {'state': ..., 'picked': ...}}，state 取值：
          ok       平台答了，但这平台没有这首歌（picked 为空）→ 直接收工
          searched 搜到了歌，链接可以落库；收藏量交给合批 flusher 后补
          empty    搜索返回空列表（多半是被限流了）
          blocked  该平台正在熔断冷却中，本次直接跳过（不计入尝试次数）
          error    抛异常/超时
        """
        app = self.app_module
        # 多艺人组合（如"zss,御鹿神谷,贰啾啾啾啾"）时只用第一个艺人名，
        # 避免关键词过长导致 QQ 搜索分词失败、搜不到目标歌曲
        if performer:
            _first_artist = re.split(r'[/&,，、;；|]+', performer)[0].strip()
            kw = f'{song_name} {_first_artist}'.strip() if _first_artist else song_name
        else:
            kw = song_name

        def _one(code):
            if self.gate.blocked(code):
                return {'state': 'blocked', 'picked': None}
            try:
                if code == 'qq':
                    res = app.search_qq(kw, 30)              # search_qq 本身不抓详情
                else:
                    res = app.search_kugou(kw, 30, light=True)
            except Exception as e:
                self.gate.report(code, False)
                return {'state': 'error', 'picked': None, 'err': repr(e)[:120]}
            if not res:
                # 搜索列表为空。单看一首分不清是「冷门歌真没有」还是「被风控了」，
                # 交给 _PlatGate 用连续空结果去判断。
                self.gate.report(code, False)
                return {'state': 'empty', 'picked': None}

            self.gate.report(code, True)                      # 搜索通道是活的
            picked = app._pick_best_for_batch(
                res, code, song_name, performer, lyricist, composer, self._enrich_for_4d)
            best = picked[0] if picked else None
            if not best:
                return {'state': 'ok', 'picked': picked}      # 平台答了，就是没这首

            # 收藏量不在这儿查 —— 交给 flusher 攒批（见 FAV_BATCH_SIZE 注释）。
            # 这里只负责把链接落库，done 由 flusher 拿到收藏量后再置。
            return {'state': 'searched', 'picked': picked}

        out = {}
        local_ex = concurrent.futures.ThreadPoolExecutor(max_workers=2)
        try:
            futs = {p: local_ex.submit(_one, p) for p in need}
            deadline = time.time() + V2_SEARCH_TIMEOUT
            for p, f in futs.items():
                try:
                    out[p] = f.result(timeout=max(0.1, deadline - time.time()))
                except Exception as e:
                    out[p] = {'state': 'error', 'picked': None, 'err': repr(e)[:120]}
        finally:
            local_ex.shutdown(wait=False)
        # ── 跨平台回填（v4.25.3）：QQ 与酷狗同属腾讯系，发行强耦合 ──
        # 任一方「本次或已存库」命中(拿到链接) 而另一方「本次」未命中时，用命中方的
        # 精确歌名+歌手重搜另一方，把腾讯系该有的歌补回来。网易云独立（可能独家/
        # 可能全网），不参与回填，避免凭空造数据。
        if ('qq' in need or 'kugou' in need) and \
                ('qq' in out or 'kugou' in out or (known_hits or {})):
            self._cross_backfill(out, song_name, performer, lyricist, composer, known_hits)
        return out

    def _cross_backfill(self, out, song_name, performer, lyricist, composer, known_hits=None):
        """腾讯系耦合回填：QQ/酷狗 任一方有命中(本次或已存库)而另一方「本次」未命中时，
        用命中方的精确歌名+歌手重搜未命中方，把腾讯系该有的歌补回来。

        known_hits：来自已存库兄弟平台的 {'song_name','performer'}（补跑时兄弟平台
        已 done、本次没重搜，但能借它的精确名去搜另一侧）。
        """
        app = self.app_module
        kh = known_hits or {}

        def _hit(o):
            # 统一返回「命中的结果 dict」或 None（无命中）。
            # 之前写成 `o and o.get('state')=='searched' and picked[0]`，当酷狗搜到结果
            # 但严格匹配未中(state='ok')时返回 False（falsy 但非 None），导致上层
            # `kg_out_miss = _hit(...) is None` 误判成 False → 跨平台回填（酷狗←QQ）永远
            # 不触发（v4.25.3/4.25.4 期间正向回填 0 次的真因）。
            if not o or o.get('state') != 'searched':
                return None
            return (o.get('picked') or [None])[0]

        def _refetch(miss_code, seed):
            # 主尝试：歌名+歌手（和原始 kw 一致）
            # 多艺人组合时只用第一个艺人名，避免关键词过长导致搜索分词失败
            _seed_perf = seed.get('performer', '')
            if _seed_perf:
                _first = re.split(r'[/&,，、;；|]+', _seed_perf)[0].strip()
                kw2 = f"{seed.get('song_name', '')} {_first}".strip() if _first else song_name
            else:
                kw2 = seed.get('song_name', '') or song_name

            def _try(kw):
                try:
                    if miss_code == 'qq':
                        res = app.search_qq(kw, 30)
                    else:
                        res = app.search_kugou(kw, 30, light=True)
                except Exception:
                    return None
                if not res:
                    return None
                return app._pick_best_for_batch(
                    res, miss_code, song_name, performer, lyricist, composer, self._enrich_for_4d)

            p = _try(kw2)
            if p and p[0]:
                return p
            # fallback（v4.25.5 关键修复）：酷狗/QQ 搜索口对「歌名+歌手」长 query 极易跑偏
            # （组合歌手名、带后缀歌手），返回一堆不相关的歌；但只用「歌名」单搜反而精准命中。
            # 实测 task26 上千首「QQ有/酷狗空」缺口歌，酷狗里其实都有，只是带歌手搜不到。
            # 仅歌名搜时用库里 performer 做严格匹配（_pick_best_for_batch 已保护，同名不同歌手会判未收录），
            # 不会误配。
            sn = (seed.get('song_name') or '').strip()
            if sn and sn != kw2:
                p2 = _try(sn)
                if p2 and p2[0]:
                    return p2
            return None

        # 任意来源的命中（本次 out 或 已存库 known_hits）
        qq_hit = _hit(out.get('qq')) or kh.get('qq')
        kg_hit = _hit(out.get('kugou')) or kh.get('kugou')
        qq_out_miss = _hit(out.get('qq')) is None
        kg_out_miss = _hit(out.get('kugou')) is None
        if qq_hit and kg_out_miss and not self.gate.blocked('kugou'):
            p = _refetch('kugou', qq_hit)
            if p and p[0]:
                out['kugou'] = {'state': 'searched', 'picked': p, 'cross_backfilled': True}
                print(f'[batch_v2][回填] 酷狗←QQ命中「{song_name}」重搜补回')
        if kg_hit and qq_out_miss and not self.gate.blocked('qq'):
            p = _refetch('qq', kg_hit)
            if p and p[0]:
                out['qq'] = {'state': 'searched', 'picked': p, 'cross_backfilled': True}
                print(f'[batch_v2][回填] QQ←酷狗命中「{song_name}」重搜补回')

    # ─── 网易云慢速池（独立 1 路，尊重 405 熔断 + 错峰）───
    def _netease_needs_work(self, item):
        """网易云是否还需要补跑：没拿到链接 / 还没确认未收录 / 没撞总上限。"""
        if item.get('netease_done') and (item.get('netease_confirms') or 0) >= 2:
            return False                      # 已确认未收录，别再跑
        if item.get('netease_url') not in (None, ''):
            return False                      # 已有链接（收藏量随链接一次存好）
        if (item.get('netease_tries') or 0) >= ABSOLUTE_MAX_TRIES:
            return False                      # 已撞总上限，放弃
        return True

    def _process_netease_safe(self, task_id, item_id):
        try:
            self._process_netease(task_id, item_id)
        except Exception as e:
            if _is_shutdown_err(e):
                return        # 进程在退出，netease_done 保持 0，下次启动自动续跑
            self._mark_netease_done(item_id, 'ne-unhandled: ' + repr(e)[:180])

    def _process_netease(self, task_id, item_id):
        app = self.app_module
        row = self._get_item(item_id)
        if not row:
            return
        if row.get('netease_done') and (row.get('netease_confirms') or 0) >= 2:
            return                      # 已确认未收录，别再跑
        # 任务已取消则不再补网易云
        with self._db_lock, self._connect() as c:
            t = c.execute("SELECT status FROM tasks WHERE id=?", (task_id,)).fetchone()
            if t and t['status'] == 'cancelled':
                return
        song_name = row['song_name']; performer = row['performer']
        if performer:
            _first_artist = re.split(r'[/&,，、;；|]+', performer)[0].strip()
            kw = f'{song_name} {_first_artist}'.strip() if _first_artist else song_name
        else:
            kw = song_name
        blocked_seen = False
        for attempt in range(1, NETEASE_MAX_ATTEMPTS + 1):
            # 尊重熔断：网易云刚被限流时，等冷却结束再打（别越打封越死）
            blocked = getattr(app, '_NETEASE_BLOCKED_AT', 0.0)
            wait = (blocked + NETEASE_COOLDOWN) - time.time()
            if wait > 0:
                blocked_seen = True
                time.sleep(min(wait, NETEASE_COOLDOWN) + 2)
            # 错峰，避开「操作频繁」
            time.sleep(NETEASE_PACE_SEC)
            try:
                # light=True：只要列表，跳过 _fetch_netease_details（那会对整页 30 首各发
                # 评论/专辑/歌词 3 个请求 = 90 个无用请求，既慢又极易触发限流）
                res = app.search_netease(kw, 30, light=True)
            except Exception:
                res = None
            if res:
                ne = app._pick_best_for_batch(
                    res, 'netease', song_name, performer,
                    row.get('lyricist', '') or '', row.get('composer', '') or '', self._enrich_for_4d)
                _best = ne[0] if ne else None
                _mlabel = str((ne or ['', ''])[1] or '')[:60] if ne else ''
                # v4.24 兜底防线：_pick_best_for_batch 的 label 是「其他:xxx/近似/低相关/存疑」
                # 等明确"非匹配"标签时，**不**把它的 URL/收藏量写进库——这是上一版的脏数据源
                # （典型案例：搜「在你说再见之前」时 _song_name_score 错把中间子串「你说」打 90 分，
                # pick 选上了枯木逢春的《你说》（id 1420082472，78万收藏），但实际搜的歌网易云根本没收）。
                # 这些情况等同"未收录"：url 清空、fav 留 NULL，netease_confirms=2 锁死防再跑。
                _label = _mlabel or ''
                _is_fallback = (
                    not _best
                    or _label.startswith('低相关')
                    or _label.startswith('其他:')
                    or _label.startswith('近似')
                    or _label.startswith('存疑')
                )
                if _is_fallback:
                    with self._db_lock, self._connect() as c:
                        c.execute(
                            "UPDATE task_items SET netease_url='', netease_favorites=NULL, "
                            "netease_done=1, netease_confirms=2, netease_error=?, netease_match=?, "
                            "last_touched_at=? WHERE id=?",
                            (f'netease: 兜底标签({_label})不写库', _mlabel, time.time(), item_id),
                        )
                    return
                # 只给命中的那一首取红心数（eapi 加密接口，免登录）
                if _best and _best.get('_song_id'):
                    try:
                        _red = app._netease_eapi_red_count([_best.get('_song_id')]) or {}
                        _info = _red.get(str(_best.get('_song_id'))) or {}
                        if _info.get('count') is not None:
                            _best['collection_count'] = _info.get('count')
                    except Exception:
                        pass
                def _fav(b):
                    if not b: return None
                    v = b.get('collection_count')
                    try: return int(v) if v is not None else None
                    except Exception: return None
                def _url(b):
                    if not b: return ''
                    return (b.get('song_url') or b.get('link') or '') or ''
                err = None
                with self._db_lock, self._connect() as c:
                    c.execute(
                        "UPDATE task_items SET netease_url=?, netease_favorites=?, "
                        "netease_done=1, netease_confirms=2, netease_error=?, netease_match=?, "
                        "last_touched_at=? WHERE id=?",
                        (_url(ne[0] if ne else None), _fav(ne[0] if ne else None), err, _mlabel,
                         time.time(), item_id),
                    )
                return
            # 空结果：可能又被限流，下一轮等冷却
        # 本轮多次重试后仍空。判断是否撞了限流：
        if blocked_seen:
            # 这一轮在限流窗口里 → 重置确认计数，留待后续补跑再做第 2 次确认
            with self._db_lock, self._connect() as c:
                c.execute("UPDATE task_items SET netease_done=0, netease_confirms=0, "
                          "last_touched_at=? WHERE id=?", (time.time(), item_id,))
            tries = (row.get('netease_tries') or 0) + 1
            if tries < ABSOLUTE_MAX_TRIES:
                with self._db_lock, self._connect() as c:
                    c.execute("UPDATE task_items SET netease_tries=? WHERE id=?", (tries, item_id))
                self.ne_executor.submit(self._process_netease_safe, task_id, item_id)
            else:
                # 总上限到了仍确认不了 → 放弃，留 NULL（未确认/可能限流）
                with self._db_lock, self._connect() as c:
                    c.execute("UPDATE task_items SET netease_done=1, netease_confirms=0, "
                              "last_touched_at=? WHERE id=?",
                              (time.time(), item_id,))
            return
        # 没撞限流、纯空结果 → 第 1 次干净确认
        confirms = (row.get('netease_confirms') or 0) + 1
        if confirms >= 2:
            # 两次干净空结果 → 确认未收录，收藏量留 NULL 不填 0
            with self._db_lock, self._connect() as c:
                c.execute(
                    "UPDATE task_items SET netease_done=1, netease_confirms=2, "
                    "netease_error='netease: 两次空结果确认未收录', last_touched_at=? WHERE id=?",
                    (time.time(), item_id,))
        else:
            # 还需第 2 次确认 → 自我重投一次（受 ABSOLUTE 总上限约束）
            tries = (row.get('netease_tries') or 0) + 1
            with self._db_lock, self._connect() as c:
                c.execute("UPDATE task_items SET netease_done=0, netease_tries=?, "
                          "last_touched_at=? WHERE id=?", (tries, time.time(), item_id))
            if tries < ABSOLUTE_MAX_TRIES:
                self.ne_executor.submit(self._process_netease_safe, task_id, item_id)
            else:
                with self._db_lock, self._connect() as c:
                    c.execute("UPDATE task_items SET netease_done=1, last_touched_at=? WHERE id=?",
                              (time.time(), item_id,))

    def _mark_netease_done(self, item_id, err):
        with self._db_lock, self._connect() as c:
            c.execute(
                "UPDATE task_items SET netease_done=1, netease_error=?, last_touched_at=? WHERE id=?",
                (err, time.time(), item_id),
            )

    def _task_cancelled(self, task_id):
        try:
            with self._db_lock, self._connect() as c:
                r = c.execute("SELECT status FROM tasks WHERE id=?", (task_id,)).fetchone()
            return bool(r) and r['status'] == 'cancelled'
        except Exception:
            return False

    def _get_item(self, item_id):
        with self._db_lock, self._connect() as c:
            row = c.execute("SELECT * FROM task_items WHERE id=?", (item_id,)).fetchone()
            return dict(row) if row else None

    def _on_item_finished(self, task_id, ok):
        """每首完成后更新任务计数器；done+failed==total 时标记完成。"""
        with self._db_lock, self._connect() as c:
            if ok:
                c.execute("UPDATE tasks SET done=done+1 WHERE id=?", (task_id,))
            else:
                c.execute("UPDATE tasks SET failed=failed+1 WHERE id=?", (task_id,))
            row = c.execute(
                "SELECT done, failed, total, status FROM tasks WHERE id=?", (task_id,)
            ).fetchone()
            # ⚠️ 不再在这里 flip 成 completed：第一遍搜完 done 就到 total，但 QQ/酷狗
            # 收藏量往往还没补完、还有歌在补跑线程里。提前 completed 会让网页「假完成」
            # （用户上次看到的就是这个）。真正的 completed 由 _maybe_finalize 在补跑稳定、
            # 收尾把剩余 NULL 收藏量填 0 之后才标记。

    # ─── 查询 ───
    def get_status(self, task_id):
        with self._db_lock, self._connect() as c:
            t = c.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()
            if not t:
                return None
            errs = c.execute(
                "SELECT idx, song_name, performer, last_error FROM task_items "
                "WHERE task_id=? AND status='failed' ORDER BY idx LIMIT 30",
                (task_id,),
            ).fetchall()
            ne = c.execute(
                "SELECT COUNT(*) AS tot, COALESCE(SUM(netease_done),0) AS done "
                "FROM task_items WHERE task_id=?",
                (task_id,),
            ).fetchone()
            # 三平台各自的实收率：用户最关心的其实是「有多少首真的拿到了收藏量」，
            # 而不是「跑完了多少首」。被限流留空的会在这里露馅。
            cov = c.execute(
                """SELECT COUNT(*) AS tot,
                          SUM(CASE WHEN COALESCE(qq_url,'')<>'' THEN 1 ELSE 0 END) AS qq_u,
                          SUM(CASE WHEN qq_favorites      IS NOT NULL THEN 1 ELSE 0 END) AS qq_f,
                          -- 真·缺收藏量：仅「搜到链接 + done=1 + fav 还没拉到」，排除 pending/confirmed/unknown
                          SUM(CASE WHEN COALESCE(qq_url,'')<>'' AND qq_favorites IS NULL AND qq_done=1 THEN 1 ELSE 0 END) AS qq_fn,
                          SUM(CASE WHEN COALESCE(qq_url,'')='' AND qq_favorites IS NULL AND COALESCE(qq_confirms,0)>=2 THEN 1 ELSE 0 END) AS qq_cn,
                          SUM(CASE WHEN COALESCE(qq_url,'')='' AND qq_favorites IS NULL AND qq_done=1 AND COALESCE(qq_confirms,0)<2 THEN 1 ELSE 0 END) AS qq_un,
                          SUM(CASE WHEN COALESCE(kugou_url,'')<>'' THEN 1 ELSE 0 END) AS kg_u,
                          SUM(CASE WHEN kugou_favorites   IS NOT NULL THEN 1 ELSE 0 END) AS kg_f,
                          SUM(CASE WHEN COALESCE(kugou_url,'')<>'' AND kugou_favorites IS NULL AND kugou_done=1 THEN 1 ELSE 0 END) AS kg_fn,
                          SUM(CASE WHEN COALESCE(kugou_url,'')='' AND kugou_favorites IS NULL AND COALESCE(kugou_confirms,0)>=2 THEN 1 ELSE 0 END) AS kg_cn,
                          SUM(CASE WHEN COALESCE(kugou_url,'')='' AND kugou_favorites IS NULL AND kugou_done=1 AND COALESCE(kugou_confirms,0)<2 THEN 1 ELSE 0 END) AS kg_un,
                          SUM(CASE WHEN COALESCE(netease_url,'')<>'' THEN 1 ELSE 0 END) AS ne_u,
                          SUM(CASE WHEN netease_favorites IS NOT NULL THEN 1 ELSE 0 END) AS ne_f,
                          SUM(CASE WHEN COALESCE(netease_url,'')<>'' AND netease_favorites IS NULL AND netease_done=1 THEN 1 ELSE 0 END) AS ne_fn,
                          SUM(CASE WHEN COALESCE(netease_url,'')='' AND netease_favorites IS NULL AND COALESCE(netease_confirms,0)>=2 THEN 1 ELSE 0 END) AS ne_cn,
                          SUM(CASE WHEN COALESCE(netease_url,'')='' AND netease_favorites IS NULL AND netease_done=1 AND COALESCE(netease_confirms,0)<2 THEN 1 ELSE 0 END) AS ne_un,
                          SUM(CASE WHEN qq_done=0    THEN 1 ELSE 0 END) AS qq_pend,
                          SUM(CASE WHEN kugou_done=0 THEN 1 ELSE 0 END) AS kg_pend,
                          -- 全平台已 resolved（任一平台未 resolved 都算 pending）
                          SUM(CASE WHEN qq_done=1 AND kugou_done=1 AND netease_done=1 THEN 1 ELSE 0 END) AS fully_resolved,
                          -- 任一平台还在 pending 或 unknown 状态（连 2 次确认都还没完成的）
                          SUM(CASE WHEN qq_done=0 OR kugou_done=0 OR netease_done=0
                                    OR (COALESCE(qq_url,'')='' AND COALESCE(qq_confirms,0)<2)
                                    OR (COALESCE(kugou_url,'')='' AND COALESCE(kugou_confirms,0)<2)
                                    OR (COALESCE(netease_url,'')='' AND COALESCE(netease_confirms,0)<2)
                                  THEN 1 ELSE 0 END) AS still_unresolved
                   FROM task_items WHERE task_id=?""",
                (task_id,),
            ).fetchone()
        # 最近已拿到收藏量的歌（让用户实时看到「数据一条条进来」）。
        # 注意：收藏量由异步工作者在搜索完成后补拉，刚搜完的歌这一刻 fav 还是 NULL，
        # 所以只挑「至少在一个平台拿到收藏量」的，避免面板全是空值误导用户。
        recent = c.execute(
            "SELECT song_name, performer, qq_url, qq_favorites, kugou_url, "
            "kugou_favorites, netease_url, netease_favorites, qq_match, "
            "kugou_match, netease_match, last_touched_at, finished_at FROM task_items "
            "WHERE task_id=? AND status='done' AND ("
            "qq_favorites IS NOT NULL OR kugou_favorites IS NOT NULL "
            "OR netease_favorites IS NOT NULL) "
            "ORDER BY COALESCE(last_touched_at, finished_at, 0) DESC, idx DESC LIMIT 15",
            (task_id,),
        ).fetchall()
        st = _row_to_task_status(t, errs, ne['tot'], ne['done'],
                                  fully_resolved=cov['fully_resolved'] or 0,
                                  still_unresolved=cov['still_unresolved'] or 0)
        # 「下一首」：队列里等待处理的下一首（worker 正在并发跑它后面的歌）。
        # 用 status='pending' 的最小 idx —— 一目了然告诉用户「马上要查这首」。
        # 注意多 worker 并发时只是「队列顺序」不是「当前正在处理」，但用户视角看
        # 这个比"看 worker 内部状态"直观 100 倍。
        nxt = c.execute(
            "SELECT idx, song_name, performer FROM task_items "
            "WHERE task_id=? AND status='pending' ORDER BY idx ASC LIMIT 1",
            (task_id,),
        ).fetchone()
        st['processing'] = dict(nxt) if nxt else None
        # 「正在跑」：5 个被 worker 最近 touch 过、还在 pending 的歌。
        # 多 worker 并发时最真实的"此刻正忙"。
        ongoing = c.execute(
            "SELECT idx, song_name, performer FROM task_items "
            "WHERE task_id=? AND status='pending' "
            "ORDER BY COALESCE(started_at, 0) DESC, idx ASC LIMIT 5",
            (task_id,),
        ).fetchall()
        st['ongoing'] = [dict(o) for o in ongoing]
        st['recent_rows'] = [dict(r) for r in recent]
        st['finalized'] = bool(t['finalized'])
        # 「已完成」按「全平台都已 resolved」算（3 平台都拿到数据 OR 都 2 次确认留空）。
        # 这是用户最关心的「每首歌都对应了真实数据」的真正完成度，避免「100% 完成、
        # 但还有 1 万首没 2 次确认」的假完成。
        st['coverage'] = {
            'total': cov['tot'] or 0,
            'fully_resolved': cov['fully_resolved'] or 0,
            'still_unresolved': cov['still_unresolved'] or 0,
            'qq': {'url': cov['qq_u'] or 0, 'fav': cov['qq_f'] or 0,
                   'favnull': cov['qq_fn'] or 0,
                   'confirmed_null': cov['qq_cn'] or 0,
                   'unknown_null': cov['qq_un'] or 0,
                   'pending': cov['qq_pend'] or 0},
            'kugou': {'url': cov['kg_u'] or 0, 'fav': cov['kg_f'] or 0,
                      'favnull': cov['kg_fn'] or 0,
                      'confirmed_null': cov['kg_cn'] or 0,
                      'unknown_null': cov['kg_un'] or 0,
                      'pending': cov['kg_pend'] or 0},
            'netease': {'url': cov['ne_u'] or 0, 'fav': cov['ne_f'] or 0,
                        'favnull': cov['ne_fn'] or 0,
                        'confirmed_null': cov['ne_cn'] or 0,
                        'unknown_null': cov['ne_un'] or 0,
                        'pending': (ne['tot'] or 0) - (ne['done'] or 0)},
        }
        st['throttled'] = self.gate.snapshot()
        st['rate_per_min'] = self.pacer.rate
        with self._fav_lock:
            st['fav_queued'] = sum(len(v) for v in self._fav_q.values())
        return st

    def list_tasks(self, limit=50):
        with self._db_lock, self._connect() as c:
            rows = c.execute(
                "SELECT * FROM tasks ORDER BY id DESC LIMIT ?", (limit,)
            ).fetchall()
            return [_row_to_task_summary(dict(r)) for r in rows]

    def get_results_incremental(self, task_id, since_idx=-1, limit=200, include='done'):
        """增量拉取已完成结果 — 前端实时追加歌曲到表格用。
        Returns: {rows, max_idx, done, total, task_status, finished}
        Raises KeyError 当任务不存在。
        """
        status_filt = "AND status IN ('done','failed')" if include == 'all' else "AND status='done'"
        with self._db_lock, self._connect() as c:
            t = c.execute(
                "SELECT id, status, total, done, failed FROM tasks WHERE id=?", (task_id,)
            ).fetchone()
            if not t:
                raise KeyError(task_id)
            rows = c.execute(
                "SELECT idx, song_name, performer, "
                "qq_url, qq_favorites, qq_match, "
                "kugou_url, kugou_favorites, kugou_match, "
                "netease_url, netease_favorites, netease_match, "
                "started_at, finished_at, last_touched_at "
                "FROM task_items WHERE task_id=? "
                f"{status_filt} AND idx>? "
                "ORDER BY idx ASC LIMIT ?",
                (task_id, since_idx, limit),
            ).fetchall()
            cnt = c.execute(
                "SELECT COUNT(*) AS n FROM task_items "
                "WHERE task_id=? AND status='done'",
                (task_id,),
            ).fetchone()
            max_idx_row = c.execute(
                "SELECT MAX(idx) AS m FROM task_items WHERE task_id=?",
                (task_id,),
            ).fetchone()
        return {
            'rows': [dict(r) for r in rows],
            'max_idx': (max_idx_row['m'] if max_idx_row else 0) or 0,
            'done': cnt['n'] if cnt else 0,
            'total': t['total'] or 0,
            'task_status': t['status'],
            'finished': t['status'] in ('completed', 'cancelled', 'failed'),
        }

    def cancel(self, task_id):
        with self._db_lock, self._connect() as c:
            c.execute(
                "UPDATE tasks SET status='cancelled', finished_at=COALESCE(finished_at, ?) "
                "WHERE id=? AND status IN ('pending','running')",
                (time.time(), task_id),
            )
            c.execute(
                "UPDATE task_items SET status='pending' "
                "WHERE task_id=? AND status='running'", (task_id,)
            )
            # 注：已经在 executor 队列里飞着的 future 仍会跑完，但不再 kickoff 新的。
            # 已 pending 的 item 也保留 pending（重新 kickoff 会重启它们），所以改为删：
            # 这里保守做法：保留 pending，下次启动时 _autorun_resume 会续跑 → 但用户已取消，
            # 所以我们要明确把 pending 改完。如果用户取消后想重启，可以手动调 resume 接口。
            # 简化：取消时把所有未完成项标 cancelled（不再处理）。
            c.execute(
                "UPDATE task_items SET status='cancelled' "
                "WHERE task_id=? AND status='pending'",
                (task_id,),
            )
            # 已完成 / 失败 的不动
            # 重新结算 total（这里 total 仍是原值，done+failed 不变）

    def retry_failed(self, task_id, only_transient=True):
        """把任务的 failed 行重新入 pending 队列，并 kickoff 起来。

        only_transient=True（默认）：只重试「瞬时错」行（last_error 含 database/locked
        关键字）。这种重试 99% 能成功，是给 SQLite 锁竞争误伤的兜底。
        only_transient=False：所有 failed 都重试（含真接口错/网络错）。**慎用**，
        那种重试大概率还会失败，会再烧一轮平台配额。

        返回：(requeued_count, transient_count, all_failed_count)
        """
        transient_pattern = "%unable to open database%"
        locked_pattern = "%database is locked%"
        with self._db_lock, self._connect() as c:
            # ① 先看下总量
            all_failed = c.execute(
                "SELECT COUNT(*) FROM task_items WHERE task_id=? AND status='failed'",
                (task_id,),
            ).fetchone()[0]
            transient_failed = c.execute(
                "SELECT COUNT(*) FROM task_items WHERE task_id=? AND status='failed' "
                "AND (last_error LIKE ? OR last_error LIKE ?)",
                (task_id, transient_pattern, locked_pattern),
            ).fetchone()[0]
            if all_failed == 0:
                return (0, transient_failed, 0)

            if only_transient:
                rows = c.execute(
                    "SELECT id, idx FROM task_items WHERE task_id=? AND status='failed' "
                    "AND (last_error LIKE ? OR last_error LIKE ?) "
                    "ORDER BY idx",
                    (task_id, transient_pattern, locked_pattern),
                ).fetchall()
            else:
                rows = c.execute(
                    "SELECT id, idx FROM task_items WHERE task_id=? AND status='failed' "
                    "ORDER BY idx",
                    (task_id,),
                ).fetchall()
            ids = [r[0] for r in rows]
            if not ids:
                return (0, transient_failed, all_failed)

            # ② 重排队（清 last_error / started_at / finished_at，留 attempts 历史）
            placeholders = ','.join('?' * len(ids))
            c.execute(
                f"UPDATE task_items SET status='pending', last_error=NULL, "
                f"started_at=NULL, finished_at=NULL "
                f"WHERE id IN ({placeholders})",
                ids,
            )
            # ③ 任务整体拉回 running（已 completed 的也要拉回，否则 _kickoff_task 不会处理）
            c.execute(
                "UPDATE tasks SET "
                "  done  =(SELECT COUNT(*) FROM task_items WHERE task_id=tasks.id AND status='done'),"
                "  failed=(SELECT COUNT(*) FROM task_items WHERE task_id=tasks.id AND status='failed'),"
                "  status='running', started_at=COALESCE(started_at, ?), finished_at=NULL "
                "WHERE id=?",
                (time.time(), task_id),
            )
        # ④ 把刚入队的歌扔进快速线程池
        for iid in ids:
            self.executor.submit(self._process_item_safe, task_id, iid)
        return (len(ids), transient_failed, all_failed)

    def retry_empty(self, task_id, platforms, force=False):
        """把 task_items 那些「done 但指定平台链接为空」的行清空对应列，
        status 退回 pending，让 worker 重跑那一列。

        force=False（默认·安全模式）：只重搜「可疑空」——
          该平台从未完成(done=0) 或 只确认<2次(存疑/可能限流)。
          已 2 次确认未收录(done=1 & confirms>=2)的歌平台确实没有，重搜必仍空，
          直接跳过——避免反复重试同一批已确认数据（cosplay，浪费配额）。
        force=True（换 Cookie 重捞专用）：重搜所有空结果（含已确认未收录），
          用于用户更换了平台 Cookie 后想用新权限把之前搜不到的歌再捞一遍。

        返回：{requeued, by_platform, remaining_empty, mode}
        设计上不重置整行（保留已跑成功的平台），只对缺的列发起单独的"补跑"。
        """
        plat_cols = {
            'qq': ('qq_url', 'qq_match'),
            'kugou': ('kugou_url', 'kugou_match'),
            'netease': ('netease_url', 'netease_match'),
        }
        target = [(plat_cols[p][0], plat_cols[p][1]) for p in platforms if p in plat_cols]
        target_cols = [t[0] for t in target]
        cols = {p: plat_cols[p][0] for p in platforms if p in plat_cols}
        if not target_cols:
            return {'requeued': 0, 'by_platform': {}, 'remaining_empty': {}, 'mode': 'force' if force else 'safe'}

        by_platform = {p: 0 for p in platforms}
        # ① 找出「值得重试」的空结果：
        #    force=False → 跳过「已确认未收录」(match 含'未收录'/'低相关'，平台确实没有)，只重搜可能查到的；
        #    force=True  → 所有空都重搜（换 Cookie 重捞）。
        if force:
            conds = ' OR '.join(f'({c} IS NULL OR {c} = "")' for (c, _) in target)
        else:
            # 只重试「可能查到」的：残缺行(搜到匹配但 url 没存) / 真缺口(从未搜过) / 限流。
            # 跳过「已确认未收录」(match 含'未收录'/'低相关'——重搜必仍空，避免 cosplay 烧配额)。
            parts = []
            for url_c, match_c in target:
                parts.append(
                    f'( ({url_c} IS NULL OR {url_c} = "") AND '
                    f'( {match_c} IS NULL OR ({match_c} NOT LIKE "%未收录%" AND {match_c} NOT LIKE "%低相关%") ) )'
                )
            conds = ' OR '.join(parts)
        with self._db_lock, self._connect() as c:
            rows = c.execute(
                f"SELECT id, idx, " + ','.join(target_cols) +
                f" FROM task_items WHERE task_id=? AND status='done' AND ({conds})",
                (task_id,),
            ).fetchall()
            if not rows:
                # 提前算 remaining_empty 给前端展示
                empty = {}
                for p in platforms:
                    empty[p] = c.execute(
                        f"SELECT COUNT(*) FROM task_items WHERE task_id=? AND status='done' AND "
                        f"({cols[p]} IS NULL OR {cols[p]} = '')",
                        (task_id,),
                    ).fetchone()[0]
                return {'requeued': 0, 'by_platform': {p: 0 for p in platforms},
                        'remaining_empty': empty}

            # ② 只清目标列，留下其它平台的成果
            ids = [r['id'] for r in rows]
            set_clause = ', '.join(f'{c} = NULL, {c.replace("_url","_favorites")} = NULL' for c in target_cols)
            placeholders = ','.join('?' * len(ids))
            c.execute(
                f"UPDATE task_items SET {set_clause} WHERE id IN ({placeholders})",
                ids,
            )
            # ③ 这些行 status='done' 还保持 done——worker 是按列缺失补的。
            #    但 task status='completed'，下一次 _kickoff_task 不会接管。
            #    解法：把这些行单独刷成 'pending'，并把 task 拉回 'running'。
            c.execute(
                f"UPDATE task_items SET status='pending' WHERE id IN ({placeholders})",
                ids,
            )
            # ③-b 修正 task.done 计数：这批行即将从 done 变 pending，先扣掉，
            #    避免「反复重试未收录」导致 done 累加虚高(显示 >total)。
            c.execute(
                "UPDATE tasks SET done = done - ? WHERE id=? AND done >= ?",
                (len(ids), task_id, len(ids)),
            )
            c.execute(
                "UPDATE tasks SET status='running', finished_at=NULL WHERE id=?",
                (task_id,),
            )
        # ④ 把这些 row 重新扔进线程池
        for iid in ids:
            self.executor.submit(self._process_item_safe, task_id, iid)

        # 重新统计 each platform 的剩余空
        with self._connect() as c2:
            remaining = {p: c2.execute(
                f"SELECT COUNT(*) FROM task_items WHERE task_id=? AND status IN ('done','pending') "
                f"AND ({cols[p]} IS NULL OR {cols[p]} = '')",
                (task_id,),
            ).fetchone()[0] for p in platforms}
        return {
            'requeued': len(ids),
            'by_platform': by_platform,
            'remaining_empty': remaining,
            'note': f'重新入队 {len(ids)} 首，目标平台：{",".join(platforms)}',
            'mode': 'force' if force else 'safe',
        }

    # ─── 收尾：把任务真正跑干净 ───
    def _maybe_finalize(self, task_id):
        """收尾判定：当任务所有歌在三平台都已 resolved（有数据 or 确认未收录）后，
        标记 completed——此时网页上的「已完成」才是真话：每一首都对应了数据，
        未收录的歌收藏量如实留 NULL（不填 0，实事求是）。

        触发收尾需全部满足，避免把「平台还在恢复 / 还在补跑」误判成「查无此歌」：
          ① 任务处于 running；② 没有 pending 行在跑；③ 收藏量合批队列已空；
          ④ 各平台都不在熔断冷却；⑤ 连续 FINALIZE_STABLE_SWEEPS 轮补跑无新进展。
        """
        with self._db_lock, self._connect() as c:
            t = c.execute(
                "SELECT id,status,total,done,finalized FROM tasks WHERE id=?",
                (task_id,)).fetchone()
            if not t or t['status'] != 'running' or t['finalized']:
                return
            # 还有歌没「跑完」（任一平台未 resolved）：p_done=0 即该平台还没拿到数据
            # 或还没确认未收录。
            unresolved = c.execute(
                "SELECT COUNT(*) FROM task_items WHERE task_id=? "
                "AND (qq_done=0 OR kugou_done=0 OR netease_done=0)",
                (task_id,)).fetchone()[0]
        if unresolved == 0:
            # 三平台都已 resolved（有数据 or 确认未收录，均不填 0），干净收尾
            with self._db_lock, self._connect() as c:
                c.execute(
                    "UPDATE tasks SET status='completed', finished_at=?, finalized=1 WHERE id=?",
                    (time.time(), task_id))
            print(f'[batch_v2] 任务 #{task_id} 已全部跑完（含确认未收录的歌，均留 NULL 不填 0），标记完成')
            self._finalize_state.pop(task_id, None)
            return
        with self._db_lock, self._connect() as c:
            pending = c.execute(
                "SELECT COUNT(*) FROM task_items WHERE task_id=? AND status='pending'",
                (task_id,)).fetchone()[0]
        if pending > 0:
            self._finalize_state[task_id] = {'last_unresolved': None, 'stable': 0}
            return
        with self._fav_lock:
            fav_q = sum(len(v) for v in self._fav_q.values())
        if fav_q > 0:
            return  # 收藏量还在异步合批补拉，等它落库
        if any(self.gate.is_cooling(p) or self.gate.is_cooling(p + ':fav')
               for p in ('qq', 'kugou')):
            return  # 平台在冷却，可能还能恢复，先别收尾
        st = self._finalize_state.setdefault(task_id, {'last_unresolved': None, 'stable': 0})
        if st['last_unresolved'] is not None and unresolved < st['last_unresolved']:
            st['last_unresolved'] = unresolved
            st['stable'] = 0
            return  # 这轮有进展，重置稳定计数
        st['last_unresolved'] = unresolved
        st['stable'] += 1
        if st['stable'] >= FINALIZE_STABLE_SWEEPS:
            # 补跑连续 N 轮无进展且平台未限流 → 剩余 unresolved 多为「撞总上限放弃」的歌
            # （它们留 NULL，导出也是 null，界面会单独标记「未确认/可能限流」）。
            # 这里只标记完成，绝不把未搜到的歌填成 0。
            with self._db_lock, self._connect() as c:
                c.execute(
                    "UPDATE tasks SET status='completed', finished_at=?, finalized=1 WHERE id=?",
                    (time.time(), task_id))
            print(f'[batch_v2] 任务 #{task_id} 收尾：剩余 {unresolved} 首 unresolved'
                  f'（多为限流放弃，留 NULL），标记完成')
            self._finalize_state.pop(task_id, None)

    # ─── 启动恢复 ───
    def _autorun_resume(self):
        """启动时把中断的任务重新拉起。

        两类都要捞：
        1. 卡在 pending/running 的任务（阶段1 没跑完）
        2. status='completed' 但网易云没补完的任务 —— 网易云是独立的慢速第二阶段，
           2 万首规模下要跑一两个小时，中途重启/休眠很容易正好卡在这个窗口。
        """
        with self._db_lock, self._connect() as c:
            # ① 先把「上次进程被打断」而误判为 failed 的歌退回队列。
            #    这类失败跟数据无关（线程池在解释器关闭时抛的，或 SQLite 锁竞争瞬时错），
            #    不退回就永久丢。
            cur = c.execute(
                "UPDATE task_items SET status='pending', last_error='上次中断/瞬时错误，已重新排队' "
                "WHERE status IN ('failed','running') AND ("
                "  last_error LIKE '%interpreter shutdown%'"
                "  OR last_error LIKE '%cannot schedule new futures%'"
                "  OR last_error LIKE '%unable to open database%'"  # SQLite WAL 瞬时锁竞争
                "  OR last_error LIKE '%database is locked%'"       # 同上
                "  OR status='running')"
            )
            if cur.rowcount:
                print(f'[batch_v2] 启动恢复：{cur.rowcount} 首上次被中断的歌已重新排队')
            # ② 被打断的任务可能已被标成 completed，得把它拉回 running
            c.execute(
                "UPDATE tasks SET status='running', finished_at=NULL WHERE status='completed' "
                "AND id IN (SELECT DISTINCT task_id FROM task_items WHERE status='pending')"
            )
            ids = [r[0] for r in c.execute(
                "SELECT id FROM tasks WHERE status IN ('pending','running')"
            ).fetchall()]
            ne_ids = [r[0] for r in c.execute(
                "SELECT DISTINCT ti.task_id FROM task_items ti JOIN tasks t ON t.id=ti.task_id "
                "WHERE ti.status='done' AND ti.netease_done=0 AND t.status='completed'"
            ).fetchall()]
        for tid in dict.fromkeys(list(ids) + list(ne_ids)):   # 去重保序
            self._kickoff_task(tid)


    # ─── v4.25.4 跨平台缺口重试修复（一次性）───
    def _autofix_v4253(self):
        """启动一次性自查：QQ/酷狗 同属腾讯系，发行强耦合。库里「只有QQ有链接、酷狗空」
        或「只有酷狗有、QQ空」的行，弱势平台的空是误判（应为没匹配到，而非真没有）。

        把这些行的弱势平台 _done/url/match/confirms/tries 复位（保留 status='done'，
        不动已好的兄弟平台），交 sweeper 用新逻辑（歌手名归一 + 跨平台回填）重抓。
        同时把各平台「非精准匹配」的歌也复位，交 sweeper 用四维校验重抓。
        幂等：用 kv 标记位保证只跑一次。
        v4.25.4：marker 升版重跑，并把 v4.25.3 期间被「限流假完成」卡死的缺口行
        （缺失平台 done=1 但 url 空、兄弟平台有链接）一并重新入队，配合 _process_item
        的缺口行 retryable 逻辑反复重试，直到限流解除借名补回。
        """
        marker = 'autofix_v4255'
        with self._db_lock, self._connect() as c:
            c.execute("CREATE TABLE IF NOT EXISTS kv(key TEXT PRIMARY KEY, value TEXT)")
            if c.execute("SELECT 1 FROM kv WHERE key=?", (marker,)).fetchone():
                return
            tid = c.execute(
                "SELECT id FROM tasks WHERE name LIKE '%task26%' OR name LIKE '%0810%' "
                "ORDER BY id DESC LIMIT 1"
            ).fetchone()
            n = 0
            if tid:
                tid = tid['id']
                for r in c.execute(
                        "SELECT * FROM task_items WHERE task_id=? AND status='done'", (tid,)).fetchall():
                    qq_u = r['qq_url'] or ''
                    kg_u = r['kugou_url'] or ''
                    qq_m = r['qq_match'] or ''
                    kg_m = r['kugou_match'] or ''
                    ne_m = r['netease_match'] or ''
                    rq = rk = rn = False
                    # v4.25.4：只复位「单边空」缺口行（兄弟平台已有链接、本平台空）。
                    # 这些是被限流窗口「假完成」吃掉的缺口，需配合 _process_item 的缺口行
                    # retryable 逻辑反复重试借名补回。非精准复核交给 v4.25.3 已完成，不再重跑。
                    if qq_u and not kg_u:
                        rk = True
                    if kg_u and not qq_u:
                        rq = True
                    if rq or rk or rn:
                        sets = []
                        if rq:
                            sets += ['qq_done=0', 'qq_url=""', 'qq_match=""',
                                     'qq_confirms=0', 'qq_tries=0', 'qq_favorites=NULL']
                        if rk:
                            sets += ['kugou_done=0', 'kugou_url=""', 'kugou_match=""',
                                     'kugou_confirms=0', 'kugou_tries=0', 'kugou_favorites=NULL']
                        if rn:
                            sets += ['netease_done=0', 'netease_url=""', 'netease_match=""',
                                     'netease_confirms=0', 'netease_tries=0', 'netease_favorites=NULL']
                        c.execute(f"UPDATE task_items SET {','.join(sets)} WHERE id=?", (r['id'],))
                        n += 1
            c.execute("INSERT OR REPLACE INTO kv(key,value) VALUES(?,?)", (marker, '1'))
            c.commit()
        print(f'[batch_v2][v4.25.5] 缺口行复位 {n} 行（QQ/酷狗单边空 → sweeper 借名「仅歌名」重搜补回）')


# ─────────── row → API 响应 ───────────
def _row_to_task_status(row, err_rows, ne_total=0, ne_done=0, fully_resolved=None, still_unresolved=None):
    row = dict(row)
    total = row['total'] or 0
    done = row['done'] or 0
    failed = row['failed'] or 0
    processed = done + failed
    # v4.27.12 修复「进度条卡死」：进度展示改用「真实已处理首数」tasks.done
    # （每首 kugou+netease 主流程跑完即 +1，QQ 限流留空、待 sweeper 补跑也计入）。
    # 旧逻辑用 fully_resolved（要求三平台全 done）当进度分子，一旦 QQ 限流就恒为 0，
    # 于是「上传 2000 首、点搜索后进度条 0% 卡死」——实际后台一直在跑，只是数字没动。
    # 真正的「三平台全完成」度仍由 st['coverage']['fully_resolved'] 如实展示，两者不冲突。
    display_done = done
    processed_for_pct = display_done + failed
    elapsed = None
    eta = None
    speed = 0
    if row['started_at']:
        elapsed = max(0.001, (row['finished_at'] or time.time()) - row['started_at'])
        speed = processed_for_pct / elapsed * 60 if elapsed > 0 else 0  # 首/分钟
        remain = max(0, total - processed_for_pct)
        if speed > 0:
            eta = int(remain / speed * 60)  # 秒
    return {
        'id': row['id'],
        'name': row['name'],
        'status': row['status'],
        'total': total,
        'done': display_done,                       # ← 用户最该看到的「已完成」
        'failed': failed,
        'progress_pct': min(100, round((processed_for_pct / total * 100), 1)) if total else 0,
        'speed_per_min': round(speed, 2),
        'eta_sec': eta,
        'elapsed_sec': int(elapsed) if elapsed is not None else None,
        'created_at': row['created_at'],
        'started_at': row['started_at'],
        'finished_at': row['finished_at'],
        'recent_errors': [
            {'idx': e['idx'], 'song_name': e['song_name'],
             'performer': e['performer'], 'error': e['last_error']}
            for e in err_rows
        ],
        'netease_total': ne_total,
        'netease_done': ne_done,
        'netease_pct': round((ne_done / ne_total * 100), 1) if ne_total else 0,
    }


def _row_to_task_summary(row):
    total = row.get('total') or 0
    processed = (row.get('done') or 0) + (row.get('failed') or 0)
    return {
        'id': row['id'],
        'name': row['name'],
        'status': row['status'],
        'total': total,
        'done': row.get('done') or 0,
        'failed': row.get('failed') or 0,
        'progress_pct': min(100, round(processed / total * 100, 1)) if total else 0,
        'created_at': row.get('created_at'),
        'finished_at': row.get('finished_at'),
    }


# ─────────── 路由注册 ───────────
def resolve_app_module():
    """拿到「正在运行的那个 app.py 模块对象」，绝不触发二次导入。

    ⚠️ 血泪坑：app.py 以脚本方式启动时，它在 sys.modules 里的名字是 '__main__'，
    不是 'app'。此时写 `import app` Python 找不到现成的，会把整个 app.py
    **从头再执行一遍** —— 所有顶层初始化、后台线程、云同步、DB 建表全部翻倍，
    对外请求量也凭空 ×2（这是 QQ/酷狗限流被提前触发的隐形推手）。
    正确做法：优先复用 sys.modules 里已经加载好的那份。
    """
    for _name in ('__main__', 'app'):
        _m = sys.modules.get(_name)
        if _m is not None and hasattr(_m, 'search_qq'):
            return _m
    import app as _m                      # 兜底：WSGI/冻结包等确实没加载过的场景
    return _m


def register(app, start_workers=True):
    """挂在 Flask app 上。在 app.py 末尾调用（此时 app.py 已全部加载完，可安全取用）。

    start_workers：只有真正对外服务的那个进程该传 True，见 _BatchV2Engine.__init__。
    """
    _app_module = resolve_app_module()
    engine = get_engine(_app_module, start_workers=start_workers)

    # ─── 提交任务 ───
    @app.route('/api/batch_v2_submit', methods=['POST'])
    def api_batch_v2_submit():
        # 解析输入：支持文件 / JSON 两种
        songs = []
        name = (request.form.get('name') or '').strip()
        if 'file' in request.files:
            f = request.files['file']
            if not f.filename:
                return jsonify({'error': '未收到文件'}), 400
            try:
                rows = _parse_uploaded_file(f)
            except Exception as e:
                return jsonify({'error': '文件解析失败：' + str(e)}), 400
            songs = _rows_to_songs(rows)
            if not name:
                name = f.filename.rsplit('.', 1)[0][:100]
        else:
            data = request.get_json(silent=True) or {}
            raw = data.get('songs') or []
            songs = []
            for s in raw:
                if isinstance(s, dict):
                    nm = (s.get('song_name') or s.get('name') or '').strip()
                    if nm:
                        songs.append({
                            'song_name': nm,
                            'performer': (s.get('performer') or '').strip(),
                            'lyricist': (s.get('lyricist') or '').strip(),
                            'composer': (s.get('composer') or '').strip(),
                        })
            if not name:
                name = (data.get('name') or '').strip() or '批量任务'

        if not songs:
            return jsonify({'error': '请提供至少一首歌（含歌名）'}), 400
        if len(songs) > 100000:
            return jsonify({'error': '单次最多 10 万首'}), 400

        try:
            task_id = engine.submit(name, songs)
        except Exception as e:
            return jsonify({'error': '任务创建失败：' + str(e)}), 500
        return jsonify({'task_id': task_id, 'total': len(songs), 'name': name})

    # ─── 查询进度 ───
    @app.route('/api/batch_v2_status/<int:task_id>')
    def api_batch_v2_status(task_id):
        s = engine.get_status(task_id)
        if s is None:
            return jsonify({'error': '任务不存在'}), 404
        return jsonify(s)

    # ─── 任务列表 ───
    @app.route('/api/batch_v2_list')
    def api_batch_v2_list():
        return jsonify({'tasks': engine.list_tasks()})

    # ─── 取消 ───
    @app.route('/api/batch_v2_cancel/<int:task_id>', methods=['POST'])
    def api_batch_v2_cancel(task_id):
        engine.cancel(task_id)
        return jsonify({'ok': True})

    # ─── 重试 failed ───
    # 把任务的 failed 行重新入 pending 队列并立即启动。
    # 默认 only_transient=true，只补「SQLite 瞬时锁竞争」这种 99% 能成的错。
    @app.route('/api/batch_v2_retry_failed/<int:task_id>', methods=['POST'])
    def api_batch_v2_retry_failed(task_id):
        only_transient = request.form.get('all', '').lower() in ('1', 'true', 'yes')
        try:
            requeued, transient_n, all_n = engine.retry_failed(task_id, only_transient=not only_transient)
        except Exception as e:
            return jsonify({'error': 'retry_failed 失败：' + str(e)}), 500
        return jsonify({
            'ok': True,
            'task_id': task_id,
            'requeued': requeued,
            'transient_count': transient_n,
            'all_failed_count': all_n,
            'mode': 'all' if only_transient else 'transient_only',
        })

    # ─── 重试「搜不到」的行（done 但平台链接空）───
    # 跟 retry_failed 区别：那些行是 status='done'（已收口）但部分平台没搜到内容。
    # 场景：第一次跑时被瞬时限流打回了空结果，但歌其实存在；现在重置这部分的
    # 链接字段，让 worker 当成"未跑过"再跑一遍。三平台可逐个补。
    @app.route('/api/batch_v2_retry_empty/<int:task_id>', methods=['POST'])
    def api_batch_v2_retry_empty(task_id):
        platforms = (request.form.get('platforms') or 'qq,kugou,netease').lower()
        only = [p for p in ('qq', 'kugou', 'netease') if p in platforms.split(',')]
        if not only:
            return jsonify({'error': 'platforms 必须是 qq / kugou / netease 之一或组合'}), 400
        force = request.form.get('force', '').lower() in ('1', 'true', 'yes', 'f')
        try:
            stats = engine.retry_empty(task_id, only, force=force)
        except Exception as e:
            return jsonify({'error': 'retry_empty 失败：' + str(e)}), 500
        return jsonify({'ok': True, 'task_id': task_id, **stats})

    # ─── 导出 xlsx ───
    @app.route('/api/batch_v2_results/<int:task_id>')
    def api_batch_v2_results(task_id):
        """增量拉取已完成的结果行，让前端"实时追加"每首歌到表格。
        Query:
          since_idx (int, default -1): 只返回 idx > since_idx 的行
          limit     (int, default 200): 单次最多返回多少行
          include   (str, default 'done'): 'done' / 'all' (含 failed)
        返回：{rows: [...], max_idx: int, done: int, total: int, finished: bool}
        """
        try:
            since_idx = int(request.args.get('since_idx', -1))
        except (TypeError, ValueError):
            since_idx = -1
        try:
            limit = max(1, min(500, int(request.args.get('limit', 200))))
        except (TypeError, ValueError):
            limit = 200
        include = request.args.get('include', 'done')
        try:
            payload = engine.get_results_incremental(task_id, since_idx=since_idx, limit=limit, include=include)
        except KeyError:
            return jsonify({'error': '任务不存在'}), 404
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        return jsonify(payload)

    @app.route('/api/batch_v2_export/<int:task_id>.xlsx')
    def api_batch_v2_export(task_id):
        s = engine.get_status(task_id)
        if s is None:
            return jsonify({'error': '任务不存在'}), 404
        try:
            buf = _build_export_xlsx(task_id)
        except Exception as e:
            return jsonify({'error': '导出失败：' + str(e)}), 500
        from flask import send_file
        import io
        fname = f'批量歌单_{task_id}_{datetime.fromtimestamp(s["created_at"]).strftime("%Y%m%d_%H%M")}.xlsx'
        try:
            return send_file(
                io.BytesIO(buf),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=fname,
            )
        except TypeError:
            return send_file(
                io.BytesIO(buf),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                attachment_filename=fname,
            )


# ─────────── 文件解析（与 app.py batch_upload 同源，独立实现以避免耦合） ───────────
def _parse_uploaded_file(f):
    """返回二维数组 [[cell, cell, ...], ...]，跳过全空行。"""
    fname = f.filename.lower()
    if fname.endswith(('.xlsx', '.xlsm')):
        import openpyxl
        wb = openpyxl.load_workbook(f.stream, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        out = []
        for r in ws.iter_rows(values_only=True):
            row = [('' if v is None else str(v)) for v in r]
            if any(cell.strip() for cell in row):
                out.append(row)
        return out
    if fname.endswith('.docx'):
        return _parse_docx_rows(f.read())
    if fname.endswith('.xls'):
        raise ValueError('暂不支持旧版 .xls，请用 .xlsx 另存后上传')
    if fname.endswith('.doc'):
        raise ValueError('暂不支持旧版 .doc，请用 .docx 另存后上传')
    raise ValueError('仅支持 .xlsx / .xlsm / .docx 文件')


def _parse_docx_rows(content_bytes):
    """简版 docx 解析：取每个段落 / 表格单元，按 tab 拼成行。"""
    import zipfile, xml.etree.ElementTree as ET
    NS = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
    out = []
    with zipfile.ZipFile(io.BytesIO(content_bytes)) as z:
        with z.open('word/document.xml') as fp:
            tree = ET.parse(fp)
    for p in tree.iter(NS + 'p'):
        texts = [t.text or '' for t in p.iter(NS + 't')]
        line = ''.join(texts).strip()
        if line:
            out.append([line])
    return out


def _rows_to_songs(rows):
    """把 [[歌名,表演者,词,曲], ...] 标准化为 [{song_name, performer, lyricist, composer}, ...]。
    支持第一行为表头自动跳过（如果第一个单元格包含『歌』『名』『表演』『唱』『词』『曲』字样）。
    """
    if not rows:
        return []
    # 表头检测
    first = rows[0]
    header_kw = ('歌', '名', '表演', '唱', '词', '曲', 'song', 'name')
    if any(kw in (first[0] or '').lower() for kw in header_kw):
        rows = rows[1:]

    songs = []
    for row in rows:
        # 行内允许 \t / 多空格 / 中文逗号
        cells = []
        for c in row[:4]:
            if not c:
                cells.append('')
            else:
                parts = [s.strip() for s in c.replace('\t', ',').replace('，', ',').split(',')]
                cells.append(parts[0] if parts else '')
        if not cells[0]:
            continue
        songs.append({
            'song_name': cells[0],
            'performer': cells[1] if len(cells) > 1 else '',
            'lyricist': cells[2] if len(cells) > 2 else '',
            'composer': cells[3] if len(cells) > 3 else '',
        })
    return songs


CROSS_CHECK_RATIO = 50        # 某平台收藏量低于最高平台的 1/N 就判定量级异常


def _cross_check(favs):
    """跨平台量级交叉校验 —— 比解析歌手名可靠得多的「抓错版」手段。

    同一首歌在三个平台的收藏量应该是同一量级（差几倍正常，差几百倍不正常）。
    实测「晴天」QQ 5600 万、网易云 11846，差 4700 倍 —— 网易云那条一定是翻唱版
    （周杰伦全部作品早年已从网易云下架，搜到的只可能是 UGC 翻唱）。

    favs: {平台code: 收藏量}，只含拿到数的平台。
    返回：'' / '正常' / '⚠ 网易云量级异常(约1/4700)'
    """
    if len(favs) < 2:
        return ''                       # 只有一个平台有数，没法互相印证
    hi = max(favs.values())
    bad = []
    for code, v in favs.items():
        if v * CROSS_CHECK_RATIO < hi:
            bad.append(f'{TASK_PLAT_NAMES.get(code, code)}约1/{max(1, round(hi / max(v, 1)))}')
    if not bad:
        return '正常'
    return '⚠ 量级异常：' + '、'.join(bad)


def _build_export_xlsx(task_id):
    """生成 xlsx，列：序号 / 歌名 / 表演者 / 词 / 曲 / 3平台×{收藏,链接,匹配} / 状态 / 错误。

    「匹配」列用于暴露翻唱冒充：
      精准匹配(目标艺人) / 匹配(多艺人|别名:xxx) / 存疑(疑似翻唱:xxx) / 其他:xxx
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    engine = _ENGINE
    if engine is None:
        raise RuntimeError('引擎未初始化')
    with engine._db_lock, engine._connect() as c:
        rows = c.execute(
            "SELECT * FROM task_items WHERE task_id=? ORDER BY idx",
            (task_id,),
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '批量结果'
    headers = ['序号', '歌名', '表演者', '词作者', '曲作者']
    for code in TASK_PLATFORMS:
        n = TASK_PLAT_NAMES[code]
        headers += [f'{n}_收藏', f'{n}_链接', f'{n}_匹配']
    headers += ['数据校验', '处理状态', '尝试次数', '错误信息']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def g(v):
        """通用：None/空串 → 空单元格（文字字段空着该留空）。"""
        return v if v not in (None, '', 'None') else ''

    def g_fav(v):
        """收藏量专用：None → 'null'（确认未收录），0 → 0（真零收藏），其他原样。"""
        if v is None:
            return 'null'
        return v

    def rget(row, key):
        """sqlite3.Row 没有 .get()，老任务可能缺列，做安全兜底。"""
        try:
            return row[key]
        except (IndexError, KeyError):
            return None

    # 存疑行高亮：淡黄底，肉眼一眼可筛
    warn_fill = PatternFill('solid', fgColor='FFF3CD')
    warn_font = Font(color='9A6700')

    warn_cells = []          # (行号, 列号) 需要高亮的匹配单元格
    for r in rows:
        row_out = [
            g(r['idx'] + 1),
            g(r['song_name']),
            g(r['performer']),
            g(r['lyricist']),
            g(r['composer']),
        ]
        warn_cols = []
        favs = {}
        for code in TASK_PLATFORMS:
            fav = r[f'{code}_favorites']
            url = r[f'{code}_url'] or ''
            m = rget(r, f'{code}_match') or ''
            row_out += [g_fav(fav), g(url), g(m)]
            if m and ('存疑' in m or m.startswith('其他')):
                warn_cols.append(len(row_out))   # 1-based 列号 = 当前长度
            try:
                if fav is not None and int(fav) > 0:
                    favs[code] = int(fav)
            except Exception:
                pass
        verdict = _cross_check(favs)
        row_out.append(verdict)
        verdict_col = len(row_out)
        # 处理状态：终态行区分「已搜无果(未收录)」与「已完成」，避免 315 行一律显示 done 造成歧义
        raw_status = r['status']
        if raw_status == 'done':
            has_link = any((r[f'{code}_url'] or '') for code in TASK_PLATFORMS)
            status_disp = '已搜无果(未收录)' if not has_link else '已完成'
        else:
            status_disp = g(raw_status)
        row_out += [
            status_disp,
            g(r['attempts']),
            g(r['last_error']),
        ]
        ws.append(row_out)
        for col in warn_cols:
            warn_cells.append((ws.max_row, col))
        if verdict.startswith('⚠'):
            warn_cells.append((ws.max_row, verdict_col))

    for rr, cc in warn_cells:
        cell = ws.cell(row=rr, column=cc)
        cell.fill = warn_fill
        cell.font = warn_font

    # 列宽：链接列宽一点，匹配列中等
    widths = [6, 22, 18, 16, 16]
    for _ in TASK_PLATFORMS:
        widths += [12, 46, 22]
    widths += [26, 12, 10, 30]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'B2'

    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()