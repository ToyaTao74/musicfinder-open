#!/usr/bin/env python3
"""监控模块 Flask 蓝图 —— 挂载到主程序，提供「数据监控」「运营报告」两个页签的接口

路由前缀统一 /api/monitor/*，与主程序既有路由零冲突。
长任务（建档匹配、每日抓取）一律后台线程跑 + 进度轮询，绝不阻塞请求。
"""

import json
import os
import tempfile
import threading
import traceback

from flask import Blueprint, jsonify, request

from . import db
from . import importer as imp

bp = Blueprint('monitor', __name__, url_prefix='/api/monitor')

# ── 后台任务注册表（同一时刻同类型只允许一个）──────────
_jobs = {}
_jobs_lock = threading.Lock()


def _job_state(kind):
    with _jobs_lock:
        j = _jobs.get(kind)
        if not j:
            return {'running': False}
        return {'running': j['thread'].is_alive(), 'started_at': j['started_at'],
                'label': j.get('label', ''), 'error': j.get('error', '')}


def _start_job(kind, fn, label=''):
    with _jobs_lock:
        j = _jobs.get(kind)
        if j and j['thread'].is_alive():
            return False, '同类任务正在运行中'

        def wrapper():
            try:
                fn()
            except Exception as e:
                with _jobs_lock:
                    if kind in _jobs:
                        _jobs[kind]['error'] = f'{e}\n{traceback.format_exc()[-800:]}'

        t = threading.Thread(target=wrapper, daemon=True)
        _jobs[kind] = {'thread': t, 'started_at': db.now_str(), 'label': label, 'error': ''}
        t.start()
        return True, ''


def _rows(cur):
    return [dict(r) for r in cur.fetchall()]


# ═══════════════════════════════════════════════════
#  概览
# ═══════════════════════════════════════════════════

@bp.route('/stats', methods=['GET'])
def api_stats():
    db.init_db()
    s = db.stats()
    conn = db.get_conn()
    # 注意 key 必须和 _start_job 用的 kind 一致（历史上 'charts' 写错过，前端一直读不到）
    s['jobs'] = {k: _job_state(k)
                 for k in ('match', 'daily', 'chart', 'derivative', 'rematch')}
    s['recent_runs'] = _rows(conn.execute(
        'SELECT run_type, stat_date, status, total, done, failed, started_at, '
        'finished_at, duration_s FROM fetch_run ORDER BY id DESC LIMIT 8'))
    s['archive_status'] = {r['status']: r['c'] for r in conn.execute(
        'SELECT status, COUNT(*) c FROM song_archive GROUP BY status')}
    # 被暂停监控的歌数 —— 前端用于「全员暂停」告警（每日采集会跳过它们）
    s['archive_paused'] = conn.execute(
        'SELECT COUNT(*) c FROM song_archive WHERE enabled=0').fetchone()['c']
    # 匹配状态分布 —— 前端「🔍 人工审核」筛选条 + 概览徽章
    s['lock_summary'] = {
        'official': conn.execute(
            "SELECT COUNT(DISTINCT archive_id) c FROM song_platform "
            "WHERE song_id!='' AND review_status IN ('auto_locked','confirmed')"
        ).fetchone()['c'],
        'pending': conn.execute(
            "SELECT COUNT(DISTINCT archive_id) c FROM song_platform "
            "WHERE review_status='pending_review'").fetchone()['c'],
        'none': conn.execute(
            "SELECT COUNT(*) c FROM song_archive WHERE id NOT IN "
            "(SELECT archive_id FROM song_platform WHERE song_id!='' OR review_status='pending_review')"
        ).fetchone()['c'],
        'unmatched': conn.execute(
            "SELECT COUNT(*) c FROM song_archive WHERE id NOT IN "
            "(SELECT archive_id FROM song_platform WHERE song_id!='')"
        ).fetchone()['c'],
    }
    return jsonify({'ok': True, 'data': s})


# ═══════════════════════════════════════════════════
#  Excel 导入
# ═══════════════════════════════════════════════════

@bp.route('/import/preview', methods=['POST'])
def api_import_preview():
    """上传 Excel 预览：识别到哪些列、样例数据。不写库。"""
    path, err = _save_upload()
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    try:
        return jsonify({'ok': True, 'data': imp.preview(path), 'temp_path': path})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


@bp.route('/import/commit', methods=['POST'])
def api_import_commit():
    """确认导入：写入 song_archive。支持 temp_path（预览后确认）或直接传文件。"""
    body = request.get_json(silent=True) or {}
    path = body.get('temp_path') or ''
    mapping_override = body.get('mapping_override') or None
    on_conflict = body.get('on_conflict') or 'merge'   # merge=补空字段 / skip=只加新歌
    if on_conflict not in ('merge', 'skip'):
        return jsonify({'ok': False, 'error': f'on_conflict 只能是 merge/skip'}), 400
    if not path:
        path, err = _save_upload()
        if err:
            return jsonify({'ok': False, 'error': err}), 400
    if not os.path.exists(path):
        return jsonify({'ok': False, 'error': '临时文件已失效，请重新上传'}), 400
    try:
        stat = imp.import_excel(path, mapping_override=mapping_override,
                                on_conflict=on_conflict)
        return jsonify({'ok': True, 'data': stat})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


def _save_upload():
    f = request.files.get('file')
    if not f or not f.filename:
        return '', '未收到文件'
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.xlsx', '.xlsm', '.xls', '.csv', '.txt', '.tsv'):
        return '', f'不支持的文件类型：{ext}'
    fd, path = tempfile.mkstemp(suffix=ext, prefix='mf_monitor_')
    os.close(fd)
    f.save(path)
    return path, ''


# ═══════════════════════════════════════════════════
#  建档匹配
# ═══════════════════════════════════════════════════

@bp.route('/match/start', methods=['POST'])
def api_match_start():
    body = request.get_json(silent=True) or {}
    limit = body.get('limit')
    workers = int(body.get('workers') or 2)
    mode = body.get('mode', 'pending')   # pending / retry / all
    status = {'pending': ('pending', 'partial'), 'retry': ('partial',),
              'all': None}.get(mode, ('pending', 'partial'))

    from . import matcher

    def run():
        matcher.match_archive(limit=limit, only_status=status, workers=workers,
                              verbose=True)

    ok, err = _start_job('match', run, label=f'建档匹配 mode={mode}')
    return jsonify({'ok': ok, 'error': err})


@bp.route('/match/progress', methods=['GET'])
def api_match_progress():
    db.init_db()
    row = db.get_conn().execute(
        "SELECT * FROM fetch_run WHERE run_type='archive_match' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    return jsonify({'ok': True, 'data': {
        'job': _job_state('match'),
        'run': dict(row) if row else None,
    }})


# ═══════════════════════════════════════════════════
#  人工复核
# ═══════════════════════════════════════════════════

@bp.route('/review/list', methods=['GET'])
def api_review_list():
    """待复核列表。默认返回 mid/low/none —— high 已自动锁定不占用人工时间。"""
    db.init_db()
    conn = db.get_conn()
    platform = request.args.get('platform', '')
    conf = request.args.get('confidence', '')
    status = request.args.get('review_status', db.REVIEW_PENDING)
    kw = (request.args.get('q') or '').strip()
    page = max(1, int(request.args.get('page', 1)))
    size = min(200, max(1, int(request.args.get('size', 10))))

    where, args = ['a.enabled=1'], []
    if status:
        where.append('sp.review_status=?')
        args.append(status)
    if platform:
        where.append('sp.platform=?')
        args.append(platform)
    if conf:
        where.append('sp.confidence=?')
        args.append(conf)
    if kw:
        where.append('(a.song_name LIKE ? OR a.artist LIKE ?)')
        args += [f'%{kw}%', f'%{kw}%']
    w = ' AND '.join(where)

    total = conn.execute(
        f'SELECT COUNT(*) c FROM song_platform sp '
        f'JOIN song_archive a ON a.id=sp.archive_id WHERE {w}', args).fetchone()['c']

    rows = _rows(conn.execute(
        f'''SELECT sp.id sp_id, sp.archive_id, sp.platform, sp.song_id, sp.url,
                   sp.matched_name, sp.matched_artist, sp.matched_album, sp.score,
                   sp.confidence, sp.review_status, sp.candidates,
                   a.song_name, a.artist, a.lyricist, a.composer, a.album
            FROM song_platform sp JOIN song_archive a ON a.id=sp.archive_id
            WHERE {w}
            ORDER BY CASE sp.confidence WHEN 'mid' THEN 0 WHEN 'low' THEN 1
                     WHEN 'none' THEN 2 ELSE 3 END, sp.archive_id, sp.platform
            LIMIT ? OFFSET ?''', args + [size, (page - 1) * size]))
    for r in rows:
        try:
            r['candidates'] = json.loads(r['candidates']) if r['candidates'] else []
        except Exception:
            r['candidates'] = []
    return jsonify({'ok': True, 'data': {'total': total, 'page': page, 'size': size,
                                         'items': rows}})


@bp.route('/review/confirm', methods=['POST'])
def api_review_confirm():
    """确认一条匹配。

    body: {sp_id, choice_index?} 或 {archive_id, platform, song_id/url/... 手工指定}
    choice_index 指向 candidates 里的第几个候选（0 = 当前显示的那条）。
    """
    body = request.get_json(silent=True) or {}
    db.init_db()
    conn = db.get_conn()
    sp_id = body.get('sp_id')
    if not sp_id:
        return jsonify({'ok': False, 'error': '缺少 sp_id'}), 400
    row = conn.execute('SELECT * FROM song_platform WHERE id=?', (sp_id,)).fetchone()
    if not row:
        return jsonify({'ok': False, 'error': '记录不存在'}), 404

    idx = body.get('choice_index')
    if idx is not None:
        try:
            cands = json.loads(row['candidates'] or '[]')
            c = cands[int(idx)]
        except Exception:
            return jsonify({'ok': False, 'error': '候选序号无效'}), 400
        payload = {'song_id': c.get('song_id', ''), 'url': c.get('url', ''),
                   'extra': json.dumps(c.get('extra') or {}, ensure_ascii=False),
                   'matched_name': c.get('name', ''), 'matched_artist': c.get('artist', ''),
                   'matched_album': c.get('album', ''), 'score': c.get('score', 0)}
    else:
        payload = {'song_id': body.get('song_id', row['song_id']),
                   'url': body.get('url', row['url']),
                   'extra': row['extra'],
                   'matched_name': body.get('matched_name', row['matched_name']),
                   'matched_artist': body.get('matched_artist', row['matched_artist']),
                   'matched_album': body.get('matched_album', row['matched_album']),
                   'score': row['score']}

    if not payload['song_id']:
        return jsonify({'ok': False, 'error': '没有可锁定的平台 ID'}), 400

    with db.tx() as c:
        c.execute("""UPDATE song_platform SET song_id=?, url=?, extra=?, matched_name=?,
                     matched_artist=?, matched_album=?, score=?, confidence='high',
                     review_status=?, updated_at=? WHERE id=?""",
                  (payload['song_id'], payload['url'], payload['extra'],
                   payload['matched_name'], payload['matched_artist'],
                   payload['matched_album'], payload['score'],
                   db.REVIEW_CONFIRMED, db.now_str(), sp_id))
    return jsonify({'ok': True})


@bp.route('/review/reject', methods=['POST'])
def api_review_reject():
    """标记「该平台没有这首歌」—— 清空 ID，后续每日抓取自动跳过。"""
    body = request.get_json(silent=True) or {}
    sp_id = body.get('sp_id')
    if not sp_id:
        return jsonify({'ok': False, 'error': '缺少 sp_id'}), 400
    db.init_db()
    with db.tx() as c:
        c.execute("""UPDATE song_platform SET song_id='', url='', confidence='none',
                     review_status=?, updated_at=? WHERE id=?""",
                  (db.REVIEW_REJECTED, db.now_str(), sp_id))
    return jsonify({'ok': True})


@bp.route('/review/batch', methods=['POST'])
def api_review_batch():
    """批量操作复核条目。

    body:
      - {action: 'confirm_top'|'reject', sp_ids: [...]}
      - {action, select_all: True, review_filter: {q, platform, confidence, review_status}}
        按当前「机器拿不准」筛选条件（与 /review/list 同源）解析全部 sp_id，再批量处理。

    注意：select_all 默认只看 review_status=pending_review，不会误伤已 confirmed/rejected。
    """
    body = request.get_json(silent=True) or {}
    action = body.get('action')
    db.init_db()
    conn = db.get_conn()
    scope = 'page'

    if body.get('select_all'):
        rf = body.get('review_filter') or {}
        kw = (rf.get('q') or '').strip()
        platform = (rf.get('platform') or '').strip()
        conf = (rf.get('confidence') or '').strip()
        status = (rf.get('review_status') or db.REVIEW_PENDING).strip()
        where, args = ['a.enabled=1'], []
        if status:
            where.append('sp.review_status=?')
            args.append(status)
        if platform:
            where.append('sp.platform=?')
            args.append(platform)
        if conf:
            where.append('sp.confidence=?')
            args.append(conf)
        if kw:
            where.append('(a.song_name LIKE ? OR a.artist LIKE ?)')
            args += [f'%{kw}%', f'%{kw}%']
        w = ' AND '.join(where)
        ids = [r['id'] for r in _rows(conn.execute(
            f'SELECT sp.id FROM song_platform sp '
            f'JOIN song_archive a ON a.id=sp.archive_id WHERE {w}', args))]
        if not ids:
            return jsonify({'ok': False, 'error': '当前筛选条件下没有可处理的条目'}), 400
        scope = 'select_all'
    else:
        ids = body.get('sp_ids') or []
        if not ids:
            return jsonify({'ok': False, 'error': '未选中任何条目'}), 400

    n = 0
    if action == 'reject':
        with db.tx() as c:
            for i in ids:
                c.execute("""UPDATE song_platform SET song_id='', url='',
                             confidence='none', review_status=?, updated_at=?
                             WHERE id=?""", (db.REVIEW_REJECTED, db.now_str(), i))
                n += 1
    elif action == 'confirm_top':
        for i in ids:
            r = conn.execute("SELECT id, song_id FROM song_platform WHERE id=?",
                             (i,)).fetchone()
            if not r or not r['song_id']:
                continue
            with db.tx() as c:
                c.execute("""UPDATE song_platform SET confidence='high',
                             review_status=?, updated_at=? WHERE id=?""",
                          (db.REVIEW_CONFIRMED, db.now_str(), i))
            n += 1
    else:
        return jsonify({'ok': False, 'error': f'未知操作 {action}'}), 400
    return jsonify({'ok': True, 'data': {'affected': n, 'scope': scope}})


@bp.route('/review/delete', methods=['POST'])
def api_review_delete():
    """从「机器拿不准」批量删 archive：按 sp_ids 或 select_all+review_filter 解析 archive_ids（去重），
    然后复用 archive 删除链路（preview / confirm 两阶段，含 metric/platform 连带删除统计）。

    body:
      - {sp_ids?: [...], confirm?: 1}
      - {select_all: True, review_filter: {q, platform, confidence, review_status}, confirm?: 1}
    """
    body = request.get_json(silent=True) or {}
    db.init_db()
    conn = db.get_conn()

    if body.get('select_all'):
        rf = body.get('review_filter') or {}
        kw = (rf.get('q') or '').strip()
        platform = (rf.get('platform') or '').strip()
        conf = (rf.get('confidence') or '').strip()
        status = (rf.get('review_status') or db.REVIEW_PENDING).strip()
        where, args = ['a.enabled=1'], []
        if status:
            where.append('sp.review_status=?')
            args.append(status)
        if platform:
            where.append('sp.platform=?')
            args.append(platform)
        if conf:
            where.append('sp.confidence=?')
            args.append(conf)
        if kw:
            where.append('(a.song_name LIKE ? OR a.artist LIKE ?)')
            args += [f'%{kw}%', f'%{kw}%']
        w = ' AND '.join(where)
        # 同一个 archive 可能在多平台都有 sp_id，去重后归档到 archive
        ids = sorted({r['archive_id'] for r in _rows(conn.execute(
            f'SELECT sp.archive_id FROM song_platform sp '
            f'JOIN song_archive a ON a.id=sp.archive_id WHERE {w}', args))})
        scope = 'select_all'
    else:
        sp_ids = body.get('sp_ids') or []
        if not sp_ids:
            return jsonify({'ok': False, 'error': '未选中任何条目'}), 400
        qmark = ','.join('?' * len(sp_ids))
        ids = sorted({r['archive_id'] for r in _rows(conn.execute(
            f'SELECT archive_id FROM song_platform WHERE id IN ({qmark})', sp_ids))})
        scope = 'page'

    if not ids:
        return jsonify({'ok': False, 'error': '没有解析到要删除的歌曲'}), 400

    qmark = ','.join('?' * len(ids))
    songs = _rows(conn.execute(
        f'SELECT id, song_name, artist FROM song_archive WHERE id IN ({qmark})', ids))
    metric_cnt = conn.execute(
        f'SELECT COUNT(*) c FROM daily_metrics WHERE archive_id IN ({qmark})',
        ids).fetchone()['c']
    plat_cnt = conn.execute(
        f'SELECT COUNT(*) c FROM song_platform WHERE archive_id IN ({qmark})',
        ids).fetchone()['c']
    preview = {'songs': songs, 'song_count': len(songs),
               'preview_truncated': len(songs) > 30,
               'metric_rows': metric_cnt, 'platform_rows': plat_cnt}

    if not body.get('confirm'):
        return jsonify({'ok': True, 'data': {'confirmed': False, 'scope': scope, **preview}})

    with db.tx() as c:
        # foreign_keys=ON + ON DELETE CASCADE：平台锁/指标/上榜/衍生版一并清掉
        c.execute(f'DELETE FROM song_archive WHERE id IN ({qmark})', ids)
    return jsonify({'ok': True, 'data': {'confirmed': True, 'deleted': len(ids),
                                         'scope': scope, **preview}})


# ═══════════════════════════════════════════════════
#  档案浏览
# ═══════════════════════════════════════════════════

@bp.route('/archive/list', methods=['GET'])
def api_archive_list():
    db.init_db()
    conn = db.get_conn()
    kw = (request.args.get('q') or '').strip()
    lock = (request.args.get('lock') or '').strip()
    page = max(1, int(request.args.get('page', 1)))
    size = min(200, max(1, int(request.args.get('size', 10))))
    where, args = ['1=1'], []
    if kw:
        where.append('(song_name LIKE ? OR artist LIKE ? OR lyricist LIKE ? '
                     'OR composer LIKE ?)')
        args += [f'%{kw}%'] * 4
    # 人工筛选「匹配状态」：正版 / 待确认 / 平台无（前端「🔍 人工审核」用）
    lock_sub = {
        'official': "id IN (SELECT archive_id FROM song_platform WHERE song_id!='' "
                    "AND review_status IN ('auto_locked','confirmed'))",
        'pending': "id IN (SELECT archive_id FROM song_platform "
                   "WHERE review_status='pending_review')",
        'none': "id NOT IN (SELECT archive_id FROM song_platform "
                "WHERE song_id!='' OR review_status='pending_review')",
    }
    if lock in lock_sub:
        where.append(lock_sub[lock])
    w = ' AND '.join(where)
    total = conn.execute(f'SELECT COUNT(*) c FROM song_archive WHERE {w}',
                         args).fetchone()['c']
    rows = _rows(conn.execute(
        f'SELECT id, song_name, artist, lyricist, composer, album, genre, status, '
        f'enabled, note FROM song_archive WHERE {w} ORDER BY id LIMIT ? OFFSET ?',
        args + [size, (page - 1) * size]))
    ids = [r['id'] for r in rows]
    if ids:
        q = ','.join('?' * len(ids))
        plat = _rows(conn.execute(
            f'SELECT archive_id, platform, song_id, url, confidence, review_status, '
            f'matched_name, matched_artist FROM song_platform WHERE archive_id IN ({q})',
            ids))
        by = {}
        for p in plat:
            by.setdefault(p['archive_id'], {})[p['platform']] = p
        for r in rows:
            r['platforms'] = by.get(r['id'], {})
    return jsonify({'ok': True, 'data': {'total': total, 'page': page, 'size': size,
                                         'items': rows}})


@bp.route('/archive/toggle', methods=['POST'])
def api_archive_toggle():
    """暂停 / 恢复某首歌的监控。"""
    body = request.get_json(silent=True) or {}
    aid, enabled = body.get('archive_id'), body.get('enabled')
    if not aid:
        return jsonify({'ok': False, 'error': '缺少 archive_id'}), 400
    db.init_db()
    with db.tx() as c:
        c.execute('UPDATE song_archive SET enabled=?, updated_at=? WHERE id=?',
                  (1 if enabled else 0, db.now_str(), aid))
    return jsonify({'ok': True})


@bp.route('/archive/batch_lock', methods=['POST'])
def api_archive_batch_lock():
    """批量操作档案：{action: lock|pause|resume, archive_ids: [...]}。

    lock   → 把这些歌在各平台「有 ID 的锁定」全部标记为 ✅正版（高把握 + 已确认），
             相当于机器/人工确认它们就是官方原版，之后每天正常抓数。
    pause  → 批量暂停监控（enabled=0）
    resume → 批量恢复监控（enabled=1）

    v4.27.11 新增 select_all=True：用当前筛选条件（lock_filter={lock,q}）整库锁一遍，
             不需要前端先拉所有 id 再批量 POST（2277 条瓶颈）。
    """
    body = request.get_json(silent=True) or {}
    action = body.get('action')
    ids = body.get('archive_ids') or []
    if action not in ('lock', 'pause', 'resume'):
        return jsonify({'ok': False, 'error': f'未知操作 {action}'}), 400
    db.init_db()
    if body.get('all') and action in ('pause', 'resume'):
        # 「一键全部恢复/暂停」：整库操作，不用前端把 4000 个 id 传上来
        want = 1 if action == 'resume' else 0
        with db.tx() as c:
            cur = c.execute('UPDATE song_archive SET enabled=?, updated_at=? '
                            'WHERE enabled!=?', (want, db.now_str(), want))
            n = cur.rowcount
        return jsonify({'ok': True, 'data': {'affected': n, 'scope': 'all'}})
    if not ids and not body.get('select_all'):
        return jsonify({'ok': False, 'error': '未选中任何歌曲'}), 400

    # v4.27.11：按筛选条件整库批量锁定 / 暂停 / 恢复
    if body.get('select_all') and action in ('pause', 'resume'):
        filt = body.get('lock_filter') or {}
        where = ['1=1']; params = []
        q_text = (filt.get('q') or '').strip()
        if q_text:
            where.append("(song_name LIKE ? OR artist LIKE ?)")
            like = f"%{q_text}%"; params += [like, like]
        if filt.get('lock'):
            where.append(_lock_filter_sql(filt['lock']))
        sql_where = ' AND '.join(where)
        want = 1 if action == 'resume' else 0
        with db.tx() as c:
            cur = c.execute(f"UPDATE song_archive SET enabled=?, updated_at=? WHERE {sql_where} AND enabled!=?",
                            [want, db.now_str(), *params, want])
            n = cur.rowcount
        return jsonify({'ok': True, 'data': {'affected': n, 'scope': 'select_all'}})

    if action == 'lock' and body.get('select_all'):
        filt = body.get('lock_filter') or {}
        where = ['1=1']; params = []
        q_text = (filt.get('q') or '').strip()
        if q_text:
            where.append("(song_name LIKE ? OR artist LIKE ?)")
            like = f"%{q_text}%"; params += [like, like]
        lock_val = filt.get('lock') or ''
        if lock_val == 'pending':
            # 待确认：song_platform 里有 ID 且 confidence!='high'
            where.append("""id IN (SELECT archive_id FROM song_platform
                                    WHERE song_id!='' AND confidence!='high')""")
        elif lock_val == 'official':
            where.append("""id IN (SELECT archive_id FROM song_platform
                                    WHERE song_id!='' AND confidence='high')""")
        elif lock_val == 'none':
            where.append("""id IN (SELECT archive_id FROM song_platform
                                    WHERE (song_id='' OR song_id IS NULL))""")
        # 空筛选 → 全库锁（按 song_platform 有 ID 的锁）
        sql_where = ' AND '.join(where)
        with db.tx() as c:
            rows = c.execute(f"SELECT id FROM song_archive WHERE {sql_where}", params).fetchall()
            aid_list = [r[0] for r in rows]
            n = 0
            for aid in aid_list:
                r = c.execute(
                    "SELECT id FROM song_platform WHERE archive_id=? AND song_id!=''",
                    (aid,)).fetchall()
                if not r:
                    continue
                c.execute(
                    """UPDATE song_platform SET confidence='high', review_status=?,
                         updated_at=? WHERE archive_id=? AND song_id!=''""",
                    (db.REVIEW_CONFIRMED, db.now_str(), aid))
                n += 1
        return jsonify({'ok': True, 'data': {'affected': n, 'scope': 'select_all', 'considered': len(aid_list)}})

    if not ids:
        return jsonify({'ok': False, 'error': '未选中任何歌曲'}), 400
    qmark = ','.join('?' * len(ids))
    n = 0
    if action == 'lock':
        with db.tx() as c:
            for aid in ids:
                r = c.execute(
                    "SELECT id FROM song_platform WHERE archive_id=? AND song_id!=''",
                    (aid,)).fetchall()
                if not r:
                    continue
                c.execute(
                    """UPDATE song_platform SET confidence='high', review_status=?,
                         updated_at=? WHERE archive_id=? AND song_id!=''""",
                    (db.REVIEW_CONFIRMED, db.now_str(), aid))
                n += 1
    else:
        enabled = 1 if action == 'resume' else 0
        with db.tx() as c:
            c.execute(
                f"UPDATE song_archive SET enabled=?, updated_at=? WHERE id IN ({qmark})",
                [enabled, db.now_str(), *ids])
            n = len(ids)
    return jsonify({'ok': True, 'data': {'affected': n}})


def _lock_filter_sql(lock_val):
    """把 lock 筛选翻成 SQL：以 song_platform 状态为准（与前端列表保持一致）。"""
    if lock_val == 'official':
        return """id IN (SELECT archive_id FROM song_platform
                        WHERE song_id!='' AND confidence='high')"""
    if lock_val == 'pending':
        return """id IN (SELECT archive_id FROM song_platform
                        WHERE song_id!='' AND confidence!='high')"""
    if lock_val == 'none':
        return """id IN (SELECT archive_id FROM song_platform
                        WHERE (song_id='' OR song_id IS NULL))"""
    return '1=1'


# ═══════════════════════════════════════════════════
#  每日抓取
# ═══════════════════════════════════════════════════

@bp.route('/daily/start', methods=['POST'])
def api_daily_start():
    """触发一次每日指标抓取（后台线程，不阻塞）。"""
    body = request.get_json(silent=True) or {}
    platforms = body.get('platforms') or None
    only_missing = body.get('only_missing', True)

    from . import daily

    def run():
        daily.run_daily(platforms=platforms, only_missing=only_missing, verbose=True)

    ok, err = _start_job('daily', run, label='每日指标抓取')
    return jsonify({'ok': ok, 'error': err})


@bp.route('/daily/progress', methods=['GET'])
def api_daily_progress():
    db.init_db()
    row = db.get_conn().execute(
        "SELECT * FROM fetch_run WHERE run_type='daily_metrics' "
        "ORDER BY id DESC LIMIT 1").fetchone()
    return jsonify({'ok': True, 'data': {
        'job': _job_state('daily'),
        'run': dict(row) if row else None,
    }})


# ═══════════════════════════════════════════════════
#  榜单快照 + 上榜监控
# ═══════════════════════════════════════════════════

@bp.route('/chart/start', methods=['POST'])
def api_chart_start():
    """触发一次榜单抓取 + 上榜对撞（后台线程，不阻塞）。"""
    body = request.get_json(silent=True) or {}
    stat_date = body.get('stat_date') or None

    from . import chart

    def run():
        chart.run_chart(stat_date=stat_date, verbose=True)

    ok, err = _start_job('chart', run, label='榜单抓取')
    return jsonify({'ok': ok, 'error': err})


@bp.route('/chart/progress', methods=['GET'])
def api_chart_progress():
    db.init_db()
    row = db.get_conn().execute(
        "SELECT * FROM fetch_run WHERE run_type='chart' "
        "ORDER BY id DESC LIMIT 1").fetchone()
    return jsonify({'ok': True, 'data': {
        'job': _job_state('chart'),
        'run': dict(row) if row else None,
    }})


@bp.route('/chart/hits', methods=['GET'])
def api_chart_hits():
    """直接读取最近上榜命中明细（前台拉数用）。"""
    days = int(request.args.get('days', 30))
    db.init_db()
    with db.tx() as conn:
        rows = conn.execute("""
            SELECT ch.archive_id, ch.platform, ch.chart_id, ch.chart_name,
                   ch.stat_date, ch.rank, ch.prev_rank,
                   sa.song_name, sa.artist,
                   CASE WHEN sp.song_id IS NOT NULL AND sp.song_id != ''
                             AND sp.confidence IN ('high','mid','low')
                        THEN 1 ELSE 0 END AS platform_locked
            FROM chart_hit ch
            JOIN song_archive sa ON sa.id = ch.archive_id
            LEFT JOIN song_platform sp
                   ON sp.archive_id = ch.archive_id AND sp.platform = ch.platform
            WHERE ch.stat_date >= date('now', ?)
            ORDER BY ch.stat_date DESC, ch.rank ASC
            LIMIT 60
        """, (f'-{days} day',)).fetchall()
    return jsonify({'ok': True, 'items': [dict(r) for r in rows]})


# ═══════════════════════════════════════════════════
#  运营报告
# ═══════════════════════════════════════════════════

@bp.route('/report', methods=['GET'])
def api_report():
    period = request.args.get('period', 'day')
    stat_date = request.args.get('stat_date') or None
    db.init_db()
    try:
        from . import report
        data = report.get_report(period=period, stat_date=stat_date)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    return jsonify({'ok': True, 'data': data})


@bp.route('/report/artist_songs', methods=['GET'])
def api_report_artist_songs():
    """演唱者维度下钻：返回该歌手被监控的所有歌 + 各平台锁定状态 + 最新收藏量。"""
    artist = (request.args.get('artist') or '').strip()
    if not artist:
        return jsonify({'ok': False, 'error': '缺少 artist 参数'}), 400
    db.init_db()
    conn = db.get_conn()
    rows = _rows(conn.execute(
        'SELECT id, song_name, artist, lyricist, composer, album, enabled, note '
        'FROM song_archive WHERE artist=? ORDER BY id', (artist,)))
    ids = [r['id'] for r in rows]
    plat = {}
    if ids:
        qm = ','.join('?' * len(ids))
        for p in _rows(conn.execute(
                f'SELECT archive_id, platform, song_id, matched_name, matched_artist, '
                f'review_status, confidence FROM song_platform WHERE archive_id IN ({qm})',
                ids)):
            plat.setdefault(p['archive_id'], {})[p['platform']] = p
    metrics = _latest_metrics(conn, ids)
    for r in rows:
        r['platforms'] = plat.get(r['id'], {})
        r['latest'] = {p: (metrics.get((r['id'], p)) or {}).get('collection_count')
                       for p in db.MONITOR_PLATFORMS}
    return jsonify({'ok': True, 'data': {'artist': artist, 'songs': rows}})


# ═══════════════════════════════════════════════════
#  名单管理（v4.25.20）：增 / 删 / 改 / 平台级解锁重匹配 / 导出 / 历史
#
#  设计原则：
#    1. 删歌是不可逆的（历史指标一起走 CASCADE），所以接口默认先给「影响面预览」，
#       前端必须带 confirm=1 才真删 —— 防手滑清空母表。
#    2. 改歌名/歌手会让旧的平台锁定失效（锁的是别的歌的 ID），必须自动重置平台锁，
#       不能悄悄留着 —— 否则每天抓的是错的歌的数据，比没数据更糟。
#    3. 所有「重新去平台搜」的动作都走后台线程 + 单独 job 槽，不占用大批量建档的槽位。
# ═══════════════════════════════════════════════════

_LOCK_LABELS = {
    ('high', db.REVIEW_CONFIRMED): '已锁定(人工确认)',
    ('high', db.REVIEW_AUTO): '已锁定(自动)',
}


def _lock_label(p):
    """把置信度+复核状态翻译成人话，导出和前端共用同一套口径。"""
    if not p or not p.get('song_id'):
        return '未锁定' if (p or {}).get('review_status') != db.REVIEW_REJECTED else '该平台无此歌'
    key = (p.get('confidence'), p.get('review_status'))
    if key in _LOCK_LABELS:
        return _LOCK_LABELS[key]
    if p.get('review_status') == db.REVIEW_CONFIRMED:
        return '已锁定(人工确认)'
    if p.get('review_status') == db.REVIEW_REJECTED:
        return '该平台无此歌'
    return f"待复核({p.get('confidence') or 'none'})"


def _archive_brief(conn, aid):
    r = conn.execute('SELECT id, song_name, artist, lyricist, composer, album, genre, '
                     'note, status, enabled FROM song_archive WHERE id=?', (aid,)).fetchone()
    return dict(r) if r else None


def _resolve_ids(conn, body):
    """统一解析选歌方式：archive_ids 明确指定 / artist 按歌手 / batch 按导入批次。"""
    ids = [int(i) for i in (body.get('archive_ids') or []) if str(i).strip()]
    artist = (body.get('artist') or '').strip()
    batch = (body.get('batch') or '').strip()
    if artist:
        ids += [r['id'] for r in conn.execute(
            'SELECT id FROM song_archive WHERE artist=?', (artist,)).fetchall()]
    if batch:
        ids += [r['id'] for r in conn.execute(
            'SELECT id FROM song_archive WHERE batch=?', (batch,)).fetchall()]
    return sorted(set(ids))


@bp.route('/archive/delete', methods=['POST'])
def api_archive_delete():
    """从监控名单里彻底删除歌曲（含历史指标，不可恢复）。

    body: {archive_ids?: [...], artist?: '歌手名', batch?: '导入批次',
           select_all?: bool, lock_filter?: {q, lock},
           confirm?: 1}
    不带 confirm 时只返回影响面（删几首、连带删多少天历史），不动数据。
    select_all=True 时按 lock_filter={q, lock} 解析全部匹配 id（v4.27.12：
    修跨页全选删除只删 30 首 —— 前端 ids 只含当前页 → 后端真实删整库）。
    """
    body = request.get_json(silent=True) or {}
    db.init_db()
    conn = db.get_conn()

    # v4.27.12：跨页全选删除——按当前筛选条件整库解析 ids，避免弹窗只显示 30 条
    if body.get('select_all'):
        filt = body.get('lock_filter') or {}
        kw = (filt.get('q') or '').strip()
        lock = (filt.get('lock') or '').strip()
        where, args = ['1=1'], []
        if kw:
            where.append('(song_name LIKE ? OR artist LIKE ? OR lyricist LIKE ? '
                         'OR composer LIKE ?)')
            args += [f'%{kw}%'] * 4
        lock_sub = {
            'official': "id IN (SELECT archive_id FROM song_platform WHERE song_id!='' "
                        "AND review_status IN ('auto_locked','confirmed'))",
            'pending': "id IN (SELECT archive_id FROM song_platform "
                       "WHERE review_status='pending_review')",
'none': "id NOT IN (SELECT archive_id FROM song_platform "
                        "WHERE song_id!='' OR review_status='pending_review')",
        }
        if lock in lock_sub:
            where.append(lock_sub[lock])
        w = ' AND '.join(where)
        ids = [r['id'] for r in conn.execute(
            f'SELECT id FROM song_archive WHERE {w} ORDER BY id', args).fetchall()]
    else:
        ids = _resolve_ids(conn, body)

    if not ids:
        return jsonify({'ok': False, 'error': '没有选中任何歌曲'}), 400

    qmark = ','.join('?' * len(ids))
    # 预览只取前 30 条歌名（弹窗空间有限；真实影响范围看 song_count/metric_rows/platform_rows）
    songs = _rows(conn.execute(
        f'SELECT id, song_name, artist FROM song_archive WHERE id IN ({qmark}) '
        f'ORDER BY id LIMIT 30', ids))
    metric_cnt = conn.execute(
        f'SELECT COUNT(*) c FROM daily_metrics WHERE archive_id IN ({qmark})',
        ids).fetchone()['c']
    plat_cnt = conn.execute(
        f'SELECT COUNT(*) c FROM song_platform WHERE archive_id IN ({qmark})',
        ids).fetchone()['c']
    preview = {'songs': songs, 'song_count': len(ids),
               'preview_truncated': len(ids) > len(songs),
               'metric_rows': metric_cnt, 'platform_rows': plat_cnt}

    if not body.get('confirm'):
        return jsonify({'ok': True, 'data': {'confirmed': False, **preview}})

    with db.tx() as c:
        # foreign_keys=ON + ON DELETE CASCADE：平台锁/指标/上榜/衍生版一并清掉
        c.execute(f'DELETE FROM song_archive WHERE id IN ({qmark})', ids)
    return jsonify({'ok': True, 'data': {'confirmed': True, 'deleted': len(ids),
                                         **preview}})


@bp.route('/archive/add', methods=['POST'])
def api_archive_add():
    """手动往名单里加一首歌。

    body: {song_name*, artist, lyricist, composer, album, genre, note, auto_match?}
    match_key 撞车即视为同一首 —— 返回 duplicate 标记 + 已有档案，前端提示「已在名单里」。
    """
    body = request.get_json(silent=True) or {}
    name = (body.get('song_name') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': '歌名不能为空'}), 400
    artist = (body.get('artist') or '').strip()

    db.init_db()
    conn = db.get_conn()
    from .normalize import archive_key
    key = archive_key(name, artist)
    exist = conn.execute('SELECT id FROM song_archive WHERE match_key=?', (key,)).fetchone()
    if exist and not body.get('force_update'):
        return jsonify({'ok': True, 'data': {
            'duplicate': True, 'archive_id': exist['id'],
            'existing': _archive_brief(conn, exist['id'])}})

    aid = db.upsert_archive(
        song_name=name, artist=artist,
        lyricist=(body.get('lyricist') or '').strip(),
        composer=(body.get('composer') or '').strip(),
        album=(body.get('album') or '').strip(),
        match_key=key, batch=body.get('batch') or f'手动添加@{db.now_str()}',
        source_row=0, raw=None)
    genre = (body.get('genre') or '').strip()
    note = (body.get('note') or '').strip()
    if genre or note:
        with db.tx() as c:
            c.execute("UPDATE song_archive SET genre=CASE WHEN ?!='' THEN ? ELSE genre END, "
                      "note=CASE WHEN ?!='' THEN ? ELSE note END, updated_at=? WHERE id=?",
                      (genre, genre, note, note, db.now_str(), aid))

    started = False
    if body.get('auto_match'):
        started = _kick_rematch([aid], None, skip_locked=True,
                                label=f'新增歌曲建档 #{aid}')[0]
    return jsonify({'ok': True, 'data': {'duplicate': False, 'archive_id': aid,
                                         'match_started': started,
                                         'song': _archive_brief(conn, aid)}})


@bp.route('/archive/update', methods=['POST'])
def api_archive_update():
    """编辑歌曲信息。

    body: {archive_id*, song_name, artist, lyricist, composer, album, genre, note}
    歌名或歌手变了 → match_key 变 → 旧的平台锁定必然指向别的歌，自动全部重置为待匹配。
    """
    body = request.get_json(silent=True) or {}
    aid = body.get('archive_id')
    if not aid:
        return jsonify({'ok': False, 'error': '缺少 archive_id'}), 400
    db.init_db()
    conn = db.get_conn()
    old = conn.execute('SELECT * FROM song_archive WHERE id=?', (aid,)).fetchone()
    if not old:
        return jsonify({'ok': False, 'error': '歌曲不存在'}), 404

    def pick(f):
        v = body.get(f)
        return old[f] if v is None else str(v).strip()

    name, artist = pick('song_name'), pick('artist')
    if not name:
        return jsonify({'ok': False, 'error': '歌名不能为空'}), 400

    from .normalize import archive_key
    new_key = archive_key(name, artist)
    key_changed = new_key != old['match_key']
    if key_changed:
        clash = conn.execute('SELECT id, song_name, artist FROM song_archive '
                             'WHERE match_key=? AND id!=?', (new_key, aid)).fetchone()
        if clash:
            return jsonify({'ok': False, 'error': (
                f'改完会和名单里已有的「{clash["song_name"]} - {clash["artist"]}」'
                f'（#{clash["id"]}）重复，请先处理那一条')}), 409

    with db.tx() as c:
        c.execute("""UPDATE song_archive SET song_name=?, artist=?, lyricist=?, composer=?,
                     album=?, genre=?, note=?, match_key=?, updated_at=?,
                     status=CASE WHEN ? THEN 'pending' ELSE status END WHERE id=?""",
                  (name, artist, pick('lyricist'), pick('composer'), pick('album'),
                   pick('genre'), pick('note'), new_key, db.now_str(),
                   1 if key_changed else 0, aid))
        if key_changed:
            # 歌名/歌手换了，旧锁作废：清 ID 但保留历史指标（历史属于「这条档案」）
            c.execute("""UPDATE song_platform SET song_id='', url='', extra='',
                         matched_name='', matched_artist='', matched_album='', score=0,
                         confidence='none', review_status=?, candidates='', updated_at=?
                         WHERE archive_id=?""", (db.REVIEW_PENDING, db.now_str(), aid))

    started = False
    if key_changed and body.get('auto_match', True):
        started = _kick_rematch([aid], None, skip_locked=False,
                                label=f'编辑后重新建档 #{aid}')[0]
    return jsonify({'ok': True, 'data': {'key_changed': key_changed,
                                         'platforms_reset': key_changed,
                                         'match_started': started,
                                         'song': _archive_brief(conn, aid)}})


# ── 平台级操作 ────────────────────────────────────────

def _kick_rematch(archive_ids, platforms, skip_locked, label):
    """后台跑一次定向匹配。用独立 job 槽，不和「全量建档」抢位子。"""
    from . import matcher

    def run():
        matcher.match_archive(archive_ids=archive_ids, platforms=platforms,
                              only_status=None, workers=1, skip_locked=skip_locked,
                              verbose=True)

    return _start_job('rematch', run, label=label)


@bp.route('/platform/unlock', methods=['POST'])
def api_platform_unlock():
    """解锁单个平台：清掉锁定的 ID/链接，退回待匹配。历史指标保留。

    body: {archive_id*, platform*}  平台可传 all = 三平台全解锁
    """
    body = request.get_json(silent=True) or {}
    aid, plat = body.get('archive_id'), (body.get('platform') or '').strip()
    if not aid or not plat:
        return jsonify({'ok': False, 'error': '缺少 archive_id / platform'}), 400
    targets = db.MONITOR_PLATFORMS if plat == 'all' else [plat]
    bad = [p for p in targets if p not in db.MONITOR_PLATFORMS]
    if bad:
        return jsonify({'ok': False, 'error': f'不支持的平台：{bad}'}), 400
    db.init_db()
    with db.tx() as c:
        for p in targets:
            c.execute("""UPDATE song_platform SET song_id='', url='', extra='',
                         matched_name='', matched_artist='', matched_album='', score=0,
                         confidence='none', review_status=?, updated_at=?
                         WHERE archive_id=? AND platform=?""",
                      (db.REVIEW_PENDING, db.now_str(), aid, p))
        c.execute("UPDATE song_archive SET status='pending', updated_at=? WHERE id=?",
                  (db.now_str(), aid))
    return jsonify({'ok': True, 'data': {'unlocked': targets}})


@bp.route('/platform/rematch', methods=['POST'])
def api_platform_rematch():
    """重新去平台搜这首歌（可指定单平台）。后台跑，前端轮询 /platform/rematch_progress。

    body: {archive_ids?: [...], archive_id?, platform? (缺省=三平台), keep_locked?: false}
    """
    body = request.get_json(silent=True) or {}
    db.init_db()
    conn = db.get_conn()
    ids = _resolve_ids(conn, body)
    if not ids and body.get('archive_id') is not None:
        raw = body['archive_id']
        ids = [int(x) for x in (raw if isinstance(raw, (list, tuple)) else [raw]) if str(x).strip()]
    if not ids:
        return jsonify({'ok': False, 'error': '没有选中任何歌曲'}), 400
    if len(ids) > 200:
        return jsonify({'ok': False, 'error': '一次最多重匹配 200 首，'
                                             '更多请用「开始建档」全量跑'}), 400
    plat = (body.get('platform') or '').strip()
    platforms = None if plat in ('', 'all') else [plat]
    if platforms and plat not in db.MONITOR_PLATFORMS:
        return jsonify({'ok': False, 'error': f'不支持的平台：{plat}'}), 400

    label = (f'重新匹配 {len(ids)} 首'
             f'{"（" + db.PLATFORM_NAMES.get(plat, plat) + "）" if platforms else ""}')
    ok, err = _kick_rematch(ids, platforms,
                            skip_locked=bool(body.get('keep_locked')), label=label)
    return jsonify({'ok': ok, 'error': err, 'data': {'count': len(ids)}})


@bp.route('/platform/rematch_progress', methods=['GET'])
def api_platform_rematch_progress():
    return jsonify({'ok': True, 'data': {'job': _job_state('rematch')}})


# ── 历史曲线 ─────────────────────────────────────────

@bp.route('/archive/history', methods=['GET'])
def api_archive_history():
    """单首歌的历史指标曲线。

    query: archive_id*, days=90, platform?（缺省=全部平台）
    返回按平台分组的时间序列，前端直接喂图表。
    """
    aid = request.args.get('archive_id')
    if not aid:
        return jsonify({'ok': False, 'error': '缺少 archive_id'}), 400
    days = min(730, max(1, int(request.args.get('days', 90))))
    plat = (request.args.get('platform') or '').strip()
    db.init_db()
    conn = db.get_conn()
    song = _archive_brief(conn, aid)
    if not song:
        return jsonify({'ok': False, 'error': '歌曲不存在'}), 404

    where, args = ['archive_id=?', "stat_date >= date('now', ?)"], [aid, f'-{days} day']
    if plat:
        where.append('platform=?')
        args.append(plat)
    rows = _rows(conn.execute(
        f'''SELECT platform, stat_date, collection_count, listening_count,
                   comment_count, collection_capped, ok, err
            FROM daily_metrics WHERE {" AND ".join(where)}
            ORDER BY stat_date''', args))

    series = {}
    for r in rows:
        s = series.setdefault(r['platform'], {
            'platform': r['platform'], 'name': db.PLATFORM_NAMES.get(r['platform'],
                                                                     r['platform']),
            'has_listening': r['platform'] in db.PLATFORMS_WITH_LISTENING,
            'points': []})
        s['points'].append({
            'date': r['stat_date'], 'collection': r['collection_count'],
            'listening': r['listening_count'], 'comment': r['comment_count'],
            'capped': bool(r['collection_capped']), 'ok': bool(r['ok']),
            'err': r['err'] or ''})

    # 每个平台算区间涨幅（首末非空值之差），前端直接展示「7天涨了多少」
    for s in series.values():
        vals = [p for p in s['points'] if p['collection'] is not None]
        s['first'] = vals[0] if vals else None
        s['last'] = vals[-1] if vals else None
        s['delta'] = ((vals[-1]['collection'] - vals[0]['collection'])
                      if len(vals) >= 2 else None)
    plat_rows = _rows(conn.execute(
        'SELECT platform, song_id, url, matched_name, matched_artist, confidence, '
        'review_status FROM song_platform WHERE archive_id=?', (aid,)))
    for p in plat_rows:
        p['lock_label'] = _lock_label(p)
    return jsonify({'ok': True, 'data': {
        'song': song, 'days': days,
        'platforms': {p['platform']: p for p in plat_rows},
        'series': [series[p] for p in db.MONITOR_PLATFORMS if p in series]}})


# ── 导出 Excel ───────────────────────────────────────

def _latest_metrics(conn, ids):
    """{(archive_id, platform): 最新一天的指标行}"""
    if not ids:
        return {}
    qmark = ','.join('?' * len(ids))
    rows = _rows(conn.execute(f'''
        SELECT dm.* FROM daily_metrics dm
        JOIN (SELECT archive_id, platform, MAX(stat_date) md FROM daily_metrics
              WHERE archive_id IN ({qmark}) GROUP BY archive_id, platform) t
          ON t.archive_id=dm.archive_id AND t.platform=dm.platform AND t.md=dm.stat_date
    ''', ids))
    return {(r['archive_id'], r['platform']): r for r in rows}


@bp.route('/archive/export', methods=['GET'])
def api_archive_export():
    """导出监控名单为 Excel（含三平台锁定状态 + 最新一天指标）。

    query: q?（同档案搜索）, only_enabled?=1
    """
    import io
    from flask import send_file

    db.init_db()
    conn = db.get_conn()
    kw = (request.args.get('q') or '').strip()
    where, args = ['1=1'], []
    if kw:
        where.append('(song_name LIKE ? OR artist LIKE ? OR lyricist LIKE ? '
                     'OR composer LIKE ?)')
        args += [f'%{kw}%'] * 4
    if request.args.get('only_enabled') in ('1', 'true'):
        where.append('enabled=1')
    rows = _rows(conn.execute(
        f'SELECT id, song_name, artist, lyricist, composer, album, genre, status, '
        f'enabled, note, batch, created_at FROM song_archive '
        f'WHERE {" AND ".join(where)} ORDER BY id', args))
    ids = [r['id'] for r in rows]
    plat_map = {}
    if ids:
        qmark = ','.join('?' * len(ids))
        for p in _rows(conn.execute(
                f'SELECT * FROM song_platform WHERE archive_id IN ({qmark})', ids)):
            plat_map[(p['archive_id'], p['platform'])] = p
    metrics = _latest_metrics(conn, ids)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as e:
        return jsonify({'ok': False, 'error': f'缺少 openpyxl：{e}'}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = '监控名单'
    head = ['ID', '歌名', '歌手', '作词', '作曲', '专辑', '曲风', '建档状态', '是否监控']
    for p in db.MONITOR_PLATFORMS:
        n = db.PLATFORM_NAMES[p]
        head += [f'{n}-锁定状态', f'{n}-匹配到', f'{n}-收藏量', f'{n}-在听', f'{n}-评论',
                 f'{n}-数据日期', f'{n}-链接']
    head += ['备注', '导入批次', '建档时间']
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2F5597')
        c.alignment = Alignment(horizontal='center', vertical='center')

    for r in rows:
        line = [r['id'], r['song_name'], r['artist'], r['lyricist'], r['composer'],
                r['album'], r['genre'], r['status'],
                '监控中' if r['enabled'] else '已暂停']
        for p in db.MONITOR_PLATFORMS:
            sp = plat_map.get((r['id'], p)) or {}
            dm = metrics.get((r['id'], p)) or {}
            coll = dm.get('collection_count')
            if coll is not None and dm.get('collection_capped'):
                coll = f'{coll}+'
            line += [_lock_label(sp),
                     (f"{sp.get('matched_name') or ''}"
                      f"{' - ' + sp['matched_artist'] if sp.get('matched_artist') else ''}"),
                     coll if coll is not None else '',
                     dm.get('listening_count') if p in db.PLATFORMS_WITH_LISTENING else '—',
                     dm.get('comment_count') if dm.get('comment_count') is not None else '',
                     dm.get('stat_date', ''), sp.get('url', '')]
        line += [r['note'], r['batch'], r['created_at']]
        ws.append(line)

    widths = [6, 26, 20, 16, 16, 20, 10, 10, 8] + [12, 24, 10, 8, 8, 12, 40] * 3 + [20, 24, 18]
    for i, w in enumerate(widths[:len(head)], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = 'C2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"监控名单_{len(rows)}首_{db.today_str()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.'
                              'spreadsheetml.sheet')


def register(app):
    """在主程序里调用：monitor.routes.register(app)"""
    app.register_blueprint(bp)
    return bp
