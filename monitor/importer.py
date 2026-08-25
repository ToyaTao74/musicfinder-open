#!/usr/bin/env python3
"""Excel 导入 —— 把 4000 首母表读进 song_archive

列名自动识别：中文表头写法五花八门（歌名/歌曲名/曲名/作品名、歌手/演唱/演唱者/艺人…），
这里用别名表模糊匹配，识别不到时报明确错误让用户确认，绝不静默猜错列。
"""

import json
import os
import re

from . import db
from .normalize import archive_key

# 列别名表：key=标准字段，value=可能的表头写法（小写、去空格后比对）
COLUMN_ALIASES = {
    'song_name': ['歌名', '歌曲名', '歌曲名称', '曲名', '作品名', '作品名称', '歌曲',
                  'songname', 'song', 'title', 'name', '曲目', '曲目名称'],
    'artist': ['歌手', '歌手名', '歌手名称', '演唱', '演唱者', '演唱人', '艺人', '表演者',
               'artist', 'singer', 'performer', '主唱'],
    'lyricist': ['词作者', '作词', '作词人', '词', '填词', 'lyricist', 'lyric', '词作'],
    'composer': ['曲作者', '作曲', '作曲人', '曲', 'composer', 'compose', '曲作'],
    'album': ['专辑', '专辑名', '专辑名称', 'album'],
    'genre': ['曲风', '风格', '类型', 'genre', 'style'],
}

REQUIRED = ['song_name']


def _norm_header(h):
    if h is None:
        return ''
    return re.sub(r'[\s\u3000:：()（）\[\]【】/、_-]', '', str(h)).strip().lower()


def detect_columns(headers):
    """表头 -> {标准字段: 列索引}。返回 (mapping, unmatched_headers)"""
    mapping = {}
    used = set()
    normed = [_norm_header(h) for h in headers]

    # 先做完全相等匹配，再做包含匹配，避免「歌手」抢走「歌手名称」这类误配
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            a = _norm_header(alias)
            for i, h in enumerate(normed):
                if i in used or not h:
                    continue
                if h == a:
                    mapping[field] = i
                    used.add(i)
                    break
            if field in mapping:
                break
    for field, aliases in COLUMN_ALIASES.items():
        if field in mapping:
            continue
        for alias in aliases:
            a = _norm_header(alias)
            for i, h in enumerate(normed):
                if i in used or not h:
                    continue
                if a in h or h in a:
                    mapping[field] = i
                    used.add(i)
                    break
            if field in mapping:
                break

    unmatched = [headers[i] for i in range(len(headers)) if i not in used and normed[i]]
    return mapping, unmatched


def read_excel(path, sheet=None, header_row=None, max_scan=20):
    """读 Excel/CSV，返回 (headers, rows, meta)。自动找表头行（前 20 行内）。"""
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.csv', '.txt', '.tsv'):
        return _read_csv(path)

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    wb.close()

    if not all_rows:
        raise ValueError('Excel 为空')

    # 找表头行：命中标准字段最多的那一行
    best_i, best_map, best_hit = 0, {}, -1
    scan_to = min(max_scan, len(all_rows))
    rng = [header_row] if header_row is not None else range(scan_to)
    for i in rng:
        m, _ = detect_columns(all_rows[i])
        if len(m) > best_hit:
            best_i, best_map, best_hit = i, m, len(m)

    headers = [('' if c is None else str(c).strip()) for c in all_rows[best_i]]
    rows = all_rows[best_i + 1:]
    meta = {'sheet': ws.title, 'sheets': wb.sheetnames if hasattr(wb, 'sheetnames') else [],
            'header_row': best_i, 'total_rows': len(rows)}
    return headers, rows, meta


def _read_csv(path):
    import csv
    encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']
    for enc in encodings:
        try:
            with open(path, 'r', encoding=enc, newline='') as f:
                sample = f.read(4096)
                f.seek(0)
                delim = '\t' if sample.count('\t') > sample.count(',') else ','
                rows = list(csv.reader(f, delimiter=delim))
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError('CSV 编码无法识别（试过 utf-8 / gbk / gb18030）')
    if not rows:
        raise ValueError('CSV 为空')
    return rows[0], rows[1:], {'sheet': 'csv', 'sheets': [], 'header_row': 0,
                               'total_rows': len(rows) - 1}


def preview(path, limit=10, conflict_limit=50):
    """导入前预览：识别到的列 + 前 N 行样例 + 去重体检（重复了多少、和谁重复）。

    去重体检是给老板做决策用的：导入前就知道「这个表 500 行里 480 行早就在名单里了」，
    避免每次导入都以为加了 500 首新歌。
    """
    headers, rows, meta = read_excel(path)
    mapping, unmatched = detect_columns(headers)
    samples = []
    for r in rows[:limit]:
        item = {}
        for field, idx in mapping.items():
            item[field] = _cell(r, idx)
        # v4.27.14：预览时也展示「歌名-歌手」自动拆分结果，让用户导入前看到效果
        if item.get('song_name'):
            item['song_name'], item['artist'] = _maybe_split_name_artist(
                item['song_name'], item.get('artist', ''))
        samples.append(item)
    missing = [f for f in REQUIRED if f not in mapping]

    dedup = {'new': 0, 'dup_existing': 0, 'dup_in_file': 0, 'empty': 0,
             'conflicts': [], 'checked': False, 'auto_split': 0,
             'split_samples': []}
    if 'song_name' in mapping:
        try:
            dedup = _dedup_report(rows, headers, mapping, meta, conflict_limit)
        except Exception as e:
            dedup['error'] = f'去重体检失败（不影响导入）：{e}'

    return {
        'file': os.path.basename(path),
        'headers': headers,
        'mapping': {f: headers[i] for f, i in mapping.items()},
        'unmatched_headers': unmatched,
        'missing_required': missing,
        'total_rows': meta['total_rows'],
        'header_row': meta['header_row'],
        'sheet': meta['sheet'],
        'samples': samples,
        'dedup': dedup,
    }


def _dedup_report(rows, headers, mapping, meta, conflict_limit=50):
    """逐行算 match_key，和「文件内其他行」「库里已有档案」双向查重。

    v4.27.14：先对每行应用「歌名-歌手」自动拆分，再算 match_key——
    否则「嚣张-en」会算成「song_name=嚣张-en, artist=''」的 key，与库里
    已有的「song_name=嚣张, artist=en」判成两首歌，去重体检完全错。
    """
    db.init_db()
    conn = db.get_conn()
    existing = {r['match_key']: r for r in conn.execute(
        'SELECT id, match_key, song_name, artist FROM song_archive')}

    out = {'new': 0, 'dup_existing': 0, 'dup_in_file': 0, 'empty': 0,
           'conflicts': [], 'checked': True,
           'auto_split': 0, 'split_samples': []}
    seen = {}
    for i, row in enumerate(rows):
        raw_name = _cell(row, mapping.get('song_name'))
        if not raw_name:
            out['empty'] += 1
            continue
        raw_artist = _cell(row, mapping.get('artist'))
        # v4.27.14：拆分「歌名-歌手」复合格式
        name, artist = _maybe_split_name_artist(raw_name, raw_artist)
        if not raw_artist and artist and '-' in raw_name and raw_name != name:
            out['auto_split'] += 1
            if len(out['split_samples']) < 5:
                out['split_samples'].append({
                    'raw': raw_name, 'song_name': name, 'artist': artist})
        key = archive_key(name, artist)
        excel_row = meta['header_row'] + 2 + i
        if key in seen:
            out['dup_in_file'] += 1
            if len(out['conflicts']) < conflict_limit:
                out['conflicts'].append({
                    'kind': 'file', 'excel_row': excel_row, 'song_name': name,
                    'artist': artist, 'first_row': seen[key]})
            continue
        seen[key] = excel_row
        hit = existing.get(key)
        if hit:
            out['dup_existing'] += 1
            if len(out['conflicts']) < conflict_limit:
                out['conflicts'].append({
                    'kind': 'db', 'excel_row': excel_row, 'song_name': name,
                    'artist': artist, 'existing_id': hit['id'],
                    'existing_name': hit['song_name'],
                    'existing_artist': hit['artist']})
        else:
            out['new'] += 1
    out['conflict_total'] = out['dup_existing'] + out['dup_in_file']
    return out


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ''
    v = row[idx]
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _maybe_split_name_artist(song_name, artist=''):
    """「歌名-歌手」复合格式自动拆分。

    v4.27.14 新增：用户上传的 Excel 经常把歌名和歌手挤在同一列（如「嚣张-en」、
    「下山-要不要买菜」），歌手列空着或根本不存在。2280 行的批量表里一半以上
    歌手都丢。先按第一个 '-' 拆，防御性校验避免误拆。

    不拆的边界：
      - artist 已有值（用户明示，优先尊重）
      - song_name 不含 '-' 或拆后任一段为空
      - 拆出的 artist 是纯数字（避免「编号-歌手」误拆）
      - 拆出的 artist > 30 字（基本是歌名延续，不是歌手）
      - 拆出的 song_name > 50 字（很可能是别的复合结构）
    """
    if not song_name or not isinstance(song_name, str):
        return song_name, artist
    if artist:
        return song_name, artist
    if '-' not in song_name:
        return song_name, artist
    parts = song_name.split('-', 1)
    if len(parts) != 2:
        return song_name, artist
    name = parts[0].strip()
    art = parts[1].strip()
    if not name or not art:
        return song_name, artist
    if art.isdigit():
        return song_name, artist
    if len(art) > 30:
        return song_name, artist
    if len(name) > 50:
        return song_name, artist
    return name, art


def import_excel(path, batch=None, mapping_override=None, dry_run=False,
                 on_conflict='merge'):
    """导入到 song_archive。同 match_key 视为同一首（幂等，可重复导入补字段）。

    on_conflict:
      merge —— 默认。已在名单里的歌只补空字段（词/曲/专辑），不覆盖已有内容。
      skip  —— 已在名单里的歌完全不动（连空字段也不补），只加新歌。
                适合「只想知道这批里有几首新歌」的场景，不改动已建档数据。
    """
    db.init_db()
    headers, rows, meta = read_excel(path)
    mapping, _ = detect_columns(headers)
    if mapping_override:
        # {'song_name': '歌曲名称'} 形式，按表头名覆盖
        for field, header_name in mapping_override.items():
            if header_name in headers:
                mapping[field] = headers.index(header_name)

    missing = [f for f in REQUIRED if f not in mapping]
    if missing:
        raise ValueError(f'必需列未识别到：{missing}；实际表头={headers}')

    batch = batch or f'{os.path.basename(path)}@{db.now_str()}'
    stat = {'total': 0, 'inserted': 0, 'updated': 0, 'skipped_empty': 0,
            'dup_in_file': 0, 'skipped_existing': 0, 'on_conflict': on_conflict,
            'batch': batch, 'auto_split': 0}
    seen_keys = set()
    conn0 = db.get_conn()
    before = conn0.execute('SELECT COUNT(*) c FROM song_archive').fetchone()['c']
    existing_keys = set()
    if on_conflict == 'skip':
        existing_keys = {r['match_key'] for r in
                         conn0.execute('SELECT match_key FROM song_archive')}

    for i, row in enumerate(rows):
        raw_name = _cell(row, mapping.get('song_name'))
        if not raw_name:
            stat['skipped_empty'] += 1
            continue
        raw_artist = _cell(row, mapping.get('artist'))
        # v4.27.14：「歌名-歌手」复合格式自动拆分（Excel 经常把两列挤一起）
        name, artist = _maybe_split_name_artist(raw_name, raw_artist)
        # 只有原始 artist 为空且确实发生拆分才算「自动拆分」——用户已明示的不算
        if not raw_artist and artist and '-' in raw_name and raw_name != name:
            stat['auto_split'] += 1
        key = archive_key(name, artist)
        if key in seen_keys:
            stat['dup_in_file'] += 1
            continue
        seen_keys.add(key)
        if on_conflict == 'skip' and key in existing_keys:
            stat['skipped_existing'] += 1
            continue
        stat['total'] += 1
        if dry_run:
            continue
        raw = {headers[j]: _cell(row, j) for j in range(min(len(headers), len(row)))}
        db.upsert_archive(
            song_name=name,
            artist=artist,
            lyricist=_cell(row, mapping.get('lyricist')),
            composer=_cell(row, mapping.get('composer')),
            album=_cell(row, mapping.get('album')),
            match_key=key,
            batch=batch,
            source_row=meta['header_row'] + 2 + i,
            raw=raw,
        )
        # 曲风若 Excel 自带则先落库（后续抓平台标签会覆盖空值）
        g = _cell(row, mapping.get('genre'))
        if g:
            with db.tx() as conn:
                conn.execute("UPDATE song_archive SET genre=?, genre_source='excel' "
                             "WHERE match_key=? AND genre=''", (g, key))

    after = db.get_conn().execute('SELECT COUNT(*) c FROM song_archive').fetchone()['c']
    stat['inserted'] = after - before
    stat['updated'] = stat['total'] - stat['inserted'] if not dry_run else 0
    stat['archive_total'] = after
    return stat


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print('usage: python -m monitor.importer <excel_path> [--preview|--import]')
        sys.exit(1)
    p = os.path.expanduser(sys.argv[1])
    mode = sys.argv[2] if len(sys.argv) > 2 else '--preview'
    if mode == '--import':
        print(json.dumps(import_excel(p), ensure_ascii=False, indent=2))
    else:
        print(json.dumps(preview(p), ensure_ascii=False, indent=2))
