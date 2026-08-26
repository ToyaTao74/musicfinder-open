#!/usr/bin/env python3
"""
多平台音乐搜索聚合 API
支持：QQ音乐、酷狗音乐、酷我音乐、网易云音乐、汽水音乐
字段：歌名、歌手、专辑、词作者、曲作者、收藏、在听、评论、唱片公司、链接

# 版本号 (v4.17.2) — 单一权威来源，所有前后端展示从这里取
"""

# ════════════════════════════════════════════════════════════════════════════
# 版本号 — 单一权威来源，所有前后端展示从这里取
# ════════════════════════════════════════════════════════════════════════════
APP_VERSION        = '4.28.5'
_BUILD_STAMP        = '20260824.05'  # v4.28.0：匹配器根因修复（歌词演唱者解析 _parse_lyric_performer + 脏数据bug修复 + 批量 _enrich_result 兜底）。 // v4.27.34：搜索真实进度。① 新增内存进度注册表 SEARCH_PROGRESS + 打点函数（_sp_start/_sp_platform_done/_sp_stage/_sp_finish），search_all 每个「平台×关键词」任务完成即累加条数（失败也计数，分母不悬空），_search_core 在补全/聚合阶段切 stage。② 新增 GET /api/search_progress?sid=，返回 stage/total/各平台条数/任务完成数/耗时。③ 前端生成 search_id 随 POST 发出，复用原 1 秒定时器轮询进度，横幅副标题实时显示「已抓到 N 条（QQ x · 酷狗 y）· 正在抓取剩余平台/补全详情/聚合」，取代原来只有「已等待 N 秒」的黑盒。④ 修既有假死 bug：软超时(150s)后 fetch 返回时旧代码 `if (timedOut) return` 吞掉结果，横幅一直转、搜索按钮永久 disabled；现在超时只弹 toast，结果照常渲染、UI 正常收尾。 // 上版 v4.27.33：提高每平台搜索上限并让大数量真正有用。① fetch_limit 去掉打折/地板，用户选 100/500 如实抓取（输入上界由 api_search min(limit,1000) 兜底）。② 详情补全不再硬编码 results[:30]，改为 results[:SEARCH_ENRICH_CAP=100]：选 100/500 时补齐前 100 条的词曲/发行方/收藏量，长尾保留搜索接口基础字段；补全耗时框死在 100 条内。③ 单平台 future 超时 70s→120s（500 大数量最慢单平台任务逼近 90s，放宽避免截断丢结果）；前端软超时 120s→150s + 文案改为「每平台大数量搜索并补全详情中」。
APP_VERSION_NAME   = 'v4.28.0 匹配器根因修复（歌词演唱者认回+脏数据修复+批量enrich兜底）'
APP_VERSION_DATE   = '2026-08-24'
# _APP_START_TS 在 main() 第一行设置（避免在此 global 声明失败）

from flask import Flask, g, render_template, jsonify, request, make_response, Response, stream_with_context, session
import requests
import concurrent.futures
import random
import string
import urllib3
import urllib.parse
import json
import ast
import re
import os
import html as html_mod
import base64
import time
import subprocess
import shutil
import hashlib
import sys
import webbrowser
import threading
import argparse
import logging
import uuid
import tempfile
import binascii
import secrets
import config  # 云端配置默认值兜底（CLOUDBASE_URL / CLOUDBASE_TOKEN）

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
logger = logging.getLogger('musicfinder')

# 打包感知：PyInstaller 冻结后资源位于 sys._MEIPASS，开发/本地运行时在同目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BUNDLE_DIR = getattr(sys, '_MEIPASS', BASE_DIR)

# 打包后 Playwright 浏览器路径（分平台）：
#   Windows：Chromium 已烤进安装包（bundle 内 playwright_browsers），用户零下载、零等待。
#   macOS  ：PyInstaller 给嵌套 Chromium.app 做 ad-hoc 重签会失败，故不烤包，
#            改为首次一键登录时按需下载到用户级缓存（PLAYWRIGHT_BROWSERS_PATH='0'）。
#            且登录优先复用系统 Chrome（channel='chrome'），多数 Mac 装了 Chrome 则根本不下载。
if getattr(sys, '_MEIPASS', None):
    if sys.platform.startswith('win'):
        os.environ.setdefault('PLAYWRIGHT_BROWSERS_PATH', os.path.join(BUNDLE_DIR, 'playwright_browsers'))
    else:
        os.environ.setdefault('PLAYWRIGHT_BROWSERS_PATH', '0')

app = Flask(
    __name__,
    template_folder=os.path.join(BUNDLE_DIR, 'templates'),
    static_folder=os.path.join(BUNDLE_DIR, 'static'),
    static_url_path='/static',
)

COMMON_UA = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'

# ═══════════════════════════════════════════════════
#  Cookie 管理
# ═══════════════════════════════════════════════════

# Cookie 存储：优先用户可写目录（打包后 .app 内部只读），打包内 cookies.json 作为只读默认
COOKIE_DIR = os.path.expanduser('~/.musicfinder')
COOKIE_FILE = os.path.join(COOKIE_DIR, 'cookies.json')
BUNDLED_COOKIE_FILE = os.path.join(BUNDLE_DIR, 'cookies.json')

# 网易云收藏量(红心数) 走 eapi 加密接口，Python 标准库无 AES，改用 Node 脚本
NODE_BIN = '/Users/toya/.workbuddy/binaries/node/versions/22.22.2/bin/node'
if not os.path.exists(NODE_BIN):
    NODE_BIN = shutil.which('node') or 'node'
NETEASE_EAPI_JS = os.path.join(BASE_DIR, 'netease_eapi.js')

PLATFORM_NAMES = {
    'qq': 'QQ音乐', 'kugou': '酷狗音乐', 'kuwo': '酷我音乐',
    'netease': '网易云音乐', 'qishui': '汽水音乐',
}
PLATFORM_ORDER = ['qq', 'kugou', 'kuwo', 'netease', 'qishui']


def _effective_cookies_file():
    """返回搜索用 Cookie 文件路径，按当前登录用户隔离；未登录走旧路径。"""
    ud = _user_dir()
    if ud:
        return os.path.join(ud, 'cookies.json')
    return COOKIE_FILE


def load_cookies():
    """从文件加载 Cookie（按当前用户隔离；优先用户目录，回退打包内默认）"""
    cf = _effective_cookies_file()
    if os.path.exists(cf):
        try:
            with open(cf, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    if not _cur_user() and os.path.exists(BUNDLED_COOKIE_FILE):
        try:
            with open(BUNDLED_COOKIE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}


def save_cookies(cookies):
    """保存 Cookie 到用户可写目录（按当前用户隔离）"""
    try:
        cf = _effective_cookies_file()
        d = os.path.dirname(cf)
        os.makedirs(d, exist_ok=True)
        with open(cf, 'w', encoding='utf-8') as f:
            json.dump(cookies, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"[Cookie] Save error: {e}")
        return False


def get_cookie_string(platform):
    """获取指定平台的 Cookie 字符串。
    优先级：1. 服务器本地 cookies.json（由「浏览器登录」写入，最权威）
            2. 请求携带的 per-device cookies（前端 localStorage，作兜底/多设备隔离）
    说明：早期实现优先用前端 localStorage，导致用户在设置页换账号登录（写入文件）
    后，前端仍发送旧账号 cookie 覆盖新账号，表现为「换账号仍报风控」。现改为文件优先，
    确保「浏览器登录」写入的新账号立即生效。"""
    # 1. 优先用服务器本地 cookies.json（浏览器登录写入的权威值）
    try:
        cookies = load_cookies()
        file_ck = cookies.get(platform, '')
        if file_ck:
            return file_ck
    except RuntimeError:
        pass
    # 2. 回退到请求携带的 per-device cookies（必须在 Flask 请求上下文内）
    try:
        if hasattr(g, 'request_cookies') and g.request_cookies.get(platform):
            return g.request_cookies.get(platform, '')
    except RuntimeError:
        pass
    return ''


@app.before_request
def load_request_cookies():
    """每个请求进来时，先从 JSON body 或 Header 读取 per-device cookies，注入 g 对象。"""
    g.request_cookies = {}
    # POST/PUT/PATCH 请求优先从 JSON body 读取
    if request.method in ('POST', 'PUT', 'PATCH'):
        try:
            data = request.get_json(silent=True) or {}
            if isinstance(data, dict) and 'cookies' in data and isinstance(data['cookies'], dict):
                g.request_cookies = data['cookies']
                return
        except Exception:
            pass
    # GET 或其他请求从 Header 读取（也作为 body 缺失的兜底）
    header = request.headers.get('X-Per-Device-Cookies')
    if header:
        try:
            cookies = json.loads(header)
            if isinstance(cookies, dict):
                g.request_cookies = cookies
        except Exception:
            pass
    # 解析当前登录用户（账号体系 v4.7）：未登录则 current_user=None（退回单用户旧行为）
    try:
        g.current_user = session.get('username') or None
    except RuntimeError:
        g.current_user = None


# ═══════════════════════════════════════════════════
#  多用户 / 账号体系 (v4.7)
#  设计：未登录时 current_user=None，所有数据读写走「旧路径」（向后兼容，
#  现有单用户用法不受任何影响）。登录后按用户名分目录隔离，互不串。
#  凭证（搜索 Cookie / 红心 Cookie / 识曲 Key）也按用户隔离（每人各自登录音乐账号）。
# ═══════════════════════════════════════════════════

# 应用密钥：session 签名用。持久化到 ~/.musicfinder/.secret，重启后登录不失效。
_SECRET_FILE = os.path.join(COOKIE_DIR, '.secret')
try:
    if not os.path.exists(_SECRET_FILE):
        with open(_SECRET_FILE, 'w') as _sf:
            _sf.write(secrets.token_hex(32))
    with open(_SECRET_FILE, 'r') as _sf:
        app.secret_key = (_sf.read().strip() or secrets.token_hex(32))
except Exception:
    app.secret_key = secrets.token_hex(32)

# 账号本地存储（云端未配置时的兜底；云端配置后账号也镜像到云端以支持跨设备登录）
_USERS_AUTH_FILE = os.path.join(COOKIE_DIR, 'users_auth.json')
_users_lock = threading.Lock()


# ═══════════════════════════════════════════════════════════════
#  艺人变体别名表（v4.25.x 起）
#  同一艺人在不同平台可能用不同账号发布作品，如「钦觉 / 钦觉呀 / 钦觉(微唱)」
#  皆为同一人。picker 在严谨归一（CJK 尾缀不放行→v4.25.16 教训）的前提下，
#  允许用户通过白名单显式追加「确认是同一人」的变体，认回正主。
#  持久化在 ~/.musicfinder/performer_aliases.json，可通过 admin 端点维护。
# ═══════════════════════════════════════════════════════════════
_PERFORMER_ALIASES_FILE = os.path.join(COOKIE_DIR, 'performer_aliases.json')
_performer_aliases_lock = threading.Lock()
_performer_aliases_loaded = False


def _cur_user():
    """当前请求的用户名；未在 Flask 请求上下文或已登出返回 None。"""
    try:
        return getattr(g, 'current_user', None)
    except RuntimeError:
        return None


def _user_dir(username=None):
    """返回某用户的私有数据目录；未指定则取当前登录用户。无用户返回 None（旧路径）。"""
    u = username or _cur_user()
    if not u:
        return None
    d = os.path.join(COOKIE_DIR, 'users', u)
    os.makedirs(d, exist_ok=True)
    return d


def _hash_password(pw):
    salt = os.urandom(16)
    dk = hashlib.pbkdf2_hmac('sha256', pw.encode('utf-8'), salt, 100000)
    return 'pbkdf2$' + binascii.hexlify(salt).decode() + '$' + binascii.hexlify(dk).decode()


def _verify_password(pw, stored):
    try:
        _, salt_hex, hash_hex = stored.split('$')
        salt = binascii.unhexlify(salt_hex)
        dk = binascii.unhexlify(hash_hex)
        return hashlib.pbkdf2_hmac('sha256', pw.encode('utf-8'), salt, 100000) == dk
    except Exception:
        return False


def _load_users_auth():
    try:
        with open(_USERS_AUTH_FILE, 'r', encoding='utf-8') as f:
            d = json.load(f)
        return d.get('users', {}) if isinstance(d, dict) else {}
    except (OSError, ValueError):
        return {}


def _save_users_auth(users):
    with _users_lock:
        with open(_USERS_AUTH_FILE, 'w', encoding='utf-8') as f:
            json.dump({'users': users}, f, ensure_ascii=False, indent=2)
    # v4.20: 本地保存后异步推云端
    try:
        _save_users_auth_cloud(users)
    except Exception:
        logger.debug('云端推送跳过（可能未配置）')


# 唯一管理员账号：只有它拥有禁用/删除等其他账号的权限；其他账号均为纯使用者。
ADMIN_USERNAME = '杨颖大王'


def _is_admin(username):
    """判断给定用户名是否为管理员账号。"""
    if not username:
        return False
    users = _load_users_auth()
    return bool(users.get(username, {}).get('is_admin'))


def _touch_last_login(username):
    """记录账号最后登录时间。"""
    if not username:
        return
    users = _load_users_auth()
    u = users.get(username)
    if not u:
        return
    u['last_login'] = time.strftime('%Y-%m-%d %H:%M:%S')
    _save_users_auth(users)


def _ensure_admin_exists():
    """锁定管理员为 ADMIN_USERNAME（杨颖大王）单一账号：
    - 若清单中存在该账号，强制 is_admin=True，并清掉其它账号可能残留的 is_admin（防升级抢权）。
    - 若清单中不存在该账号，绝不自动把别的账号提为管理员（避免下一个注册的人抢到权限）。
    """
    users = _load_users_auth()
    if not users:
        return
    changed = False
    if ADMIN_USERNAME in users:
        if not users[ADMIN_USERNAME].get('is_admin'):
            users[ADMIN_USERNAME]['is_admin'] = True
            changed = True
    for nm, u in users.items():
        if nm != ADMIN_USERNAME and u.get('is_admin'):
            u['is_admin'] = False
            changed = True
    if changed:
        _save_users_auth(users)


def _require_admin():
    """管理员权限守卫：返回 None 表示通过，否则返回 403 响应。"""
    u = _cur_user()
    if not u or not _is_admin(u):
        return jsonify({'error': '需要管理员权限'}), 403
    return None


# ── v4.25.7 细粒度标记权限（跨账号可见 + 可管理） ──
# 模型：每个账号带 perms = {view_all_marks, manage_marks}
#   view_all_marks：能否看「共享」视图里其他账号的标记（默认 True，全员可读）
#   manage_marks：能否在共享视图里修改/删除他人打的标（默认 False，admin 点名才开）
#   admin（杨颖大王）恒拥有全部权限，不走 perms 位。
def _def_perms(username=None):
    """默认权限位。管理员双 True；其余默认可读、不可管理。"""
    if username and _is_admin(username):
        return {'view_all_marks': True, 'manage_marks': True}
    return {'view_all_marks': True, 'manage_marks': False}


def _can_user(username, perm):
    """细粒度权限判断：admin 恒 True；其余按 perms 位（带安全默认值）。"""
    if not username:
        return False
    if _is_admin(username):
        return True
    users = _load_users_auth()
    u = users.get(username)
    if not u:
        return False
    perms = u.get('perms')
    if not isinstance(perms, dict):
        perms = _def_perms(username)
    if perm == 'view_all_marks':
        return bool(perms.get('view_all_marks', True))
    if perm == 'manage_marks':
        return bool(perms.get('manage_marks', False))
    return False


def _ensure_perms_all():
    """启动时幂等迁移：给每个老账号补 perms 字段（view_all_marks 默认 True / manage_marks 默认 False）。
    admin 自动双 True。仅在有变化时落盘。"""
    users = _load_users_auth()
    if not users:
        return
    changed = False
    for name, u in users.items():
        perms = u.get('perms')
        if not isinstance(perms, dict):
            u['perms'] = _def_perms(name)
            changed = True
        else:
            if 'view_all_marks' not in perms:
                perms['view_all_marks'] = True
                changed = True
            if 'manage_marks' not in perms:
                perms['manage_marks'] = False
                changed = True
    if changed:
        _save_users_auth(users)


def _account_exists(username):
    return bool(username) and username in _load_users_auth()


def _register_user(username, password):
    username = (username or '').strip()
    if not username or not password:
        return False, '用户名和密码都不能为空'
    if len(password) < 4:
        return False, '密码至少 4 位'
    users = _load_users_auth()
    was_empty = not users
    if username in users:
        return False, '该用户名已存在'
    users[username] = {
        'pw': _hash_password(password),
        'created_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        # 只有 ADMIN_USERNAME（杨颖大王）注册时才是管理员；其他账号一律纯使用者
        'is_admin': username == ADMIN_USERNAME,
    }
    _save_users_auth(users)
    # 第一个账号：把本机原有单用户数据（标记/歌单/红心/登录态）复制进该账号，避免丢失
    if was_empty:
        _migrate_legacy_to_user(username)
    return True, 'ok'


def _migrate_legacy_to_user(username):
    """把本机原有单用户数据文件复制到首个账号目录，保证升级到多用户不丢历史。

    旧路径常量在模块后面才定义，故在此惰性求值（调用时模块已加载完）。
    """
    ud = _user_dir(username)
    if not ud:
        return
    legacy = [
        _MARKS_FILE, _PLAYLIST_FILE, _PLAYLIST_RESULTS_FILE,
        _PINS_FILE, _DELETED_FILE, _HEARTS_FILE, _HEART_COOKIE_FILE, COOKIE_FILE,
    ]
    for lf in legacy:
        if lf and os.path.exists(lf):
            dest = os.path.join(ud, os.path.basename(lf))
            if not os.path.exists(dest):
                try:
                    shutil.copy2(lf, dest)
                except Exception:
                    logger.exception('迁移旧数据失败: %s', lf)


def _login_user(username, password):
    username = (username or '').strip()
    users = _load_users_auth()
    u = users.get(username)
    # 1) 本地命中且密码匹配
    if u and _verify_password(password, u.get('pw', '')):
        if u.get('disabled'):
            try: _audit_log(username, 'login_denied', 'disabled', request.remote_addr if request else '')
            except: pass
            return False, '该账号已被禁用，请联系管理员'
        try: _audit_log(username, 'login', 'ok', request.remote_addr if request else '')
        except: pass
        return True, 'ok'
    # 2) 本地未命中 / 本地密码不符 → 回退云端（支持他处改密后任意设备登录、新设备登录）
    cloud = _cloud_load_users()
    if cloud and username in cloud:
        cu = cloud[username]
        if _verify_password(password, cu.get('pw', '')):
            if cu.get('disabled'):
                try: _audit_log(username, 'login_denied', 'disabled', request.remote_addr if request else '')
                except: pass
                return False, '该账号已被禁用，请联系管理员'
            # 回写本地缓存：便于离线登录与后续管理
            users[username] = cu
            _save_users_auth(users)
            try: _audit_log(username, 'login', 'ok(cloud)', request.remote_addr if request else '')
            except: pass
            return True, 'ok'
    return False, '用户名或密码错误'


# ════════ v4.20 云端用户中心（里程碑1：上云读写 + 本地降级 + 审计日志） ════════
_cloud_auth_executor = None  # 惰性初始化（避免模块级 threading 限制）
_cloud_auth_synced = False

def _get_cloud_auth_executor():
    global _cloud_auth_executor
    if _cloud_auth_executor is None:
        _cloud_auth_executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
    return _cloud_auth_executor

def _cloud_auth_cfg():
    """云端用户中心配置（与标记同步共用 url/token）。

    解析优先级（实现「新设备零配置自动上云」）：
      1. ~/.musicfinder/config.json 的 cloudbase.url / cloudbase.token（最优先）
      2. 环境变量 MUSICFINDER_CLOUDBASE_URL / MUSICFINDER_CLOUDBASE_TOKEN
      3. config.py 内置默认值（CLOUDBASE_URL / CLOUDBASE_TOKEN）
    仅当用户显式声明了非 cloudbase 后端（如 local）时才回落纯本地。
    """
    try:
        cfg = _marks_backend_cfg()
    except Exception:
        return None
    backend = (cfg.get('backend') or '').strip().lower()
    # 显式选择了其他后端（非 cloudbase）→ 尊重用户，纯本地
    if backend and backend != 'cloudbase':
        return None
    cb = cfg.get('cloudbase') or {}
    url = (cb.get('url') or '').strip().rstrip('/') \
        or (getattr(config, 'CLOUDBASE_URL', '') or '').strip().rstrip('/') \
        or (os.environ.get('MUSICFINDER_CLOUDBASE_URL') or '').strip().rstrip('/')
    token = (cb.get('token') or '').strip() \
        or (getattr(config, 'CLOUDBASE_TOKEN', '') or '').strip() \
        or (os.environ.get('MUSICFINDER_CLOUDBASE_TOKEN') or '').strip()
    # 未声明后端 + 仍无任何云端凭据 → 纯本地（不强行上云）
    if not backend and (not url or not token):
        return None
    if not url or not token:
        return None
    return {'url': url, 'token': token}

# 业务集合 → mark_key 前缀映射（mark-sync 云函数只支持 song_marks 集合）
_AUTH_BIZ_MAP = {
    'mf_users':        'user',
    'mf_audit':        'audit',
    'mf_invite_codes': 'invite',
}

def _cloud_auth_call(cb, action, **kwargs):
    """云函数调用（用户中心/审计/邀请码）。

    注意：mark-sync 云函数硬编码只支持 song_marks 集合。
    业务数据通过 mark_key 字段前缀区分（song_marks 写入用 mark_key 字段，不是 _id）：
      - mf_users         → mark_key 'user:<username>'
      - mf_audit         → mark_key 'audit:<ts>_<username>_<action>'
      - mf_invite_codes  → mark_key 'invite:<code>'

    调用方仍可传 'class': 'mf_xxx'，本函数自动转换前缀并落到 song_marks 集合。
    """
    body = {'action': action, 'token': cb['token'], 'collection': 'song_marks'}

    biz = None
    if 'class' in kwargs:
        cls_name = kwargs.pop('class')
        biz = _AUTH_BIZ_MAP.get(cls_name)

    if action == 'batch_upsert' and biz and 'payloads' in kwargs:
        prefix = biz + ':'
        new_payloads = []
        for p in kwargs['payloads']:
            old_mk = p.get('mark_key') or p.get('_id') or ''
            if not old_mk.startswith(prefix):
                new_mk = prefix + old_mk
            else:
                new_mk = old_mk
            # song_marks 标准格式：顶层 mark_key + 嵌套 data
            inner_data = p.get('data') or {}
            # 去掉 p 里的顶层 mark_key/_id，只留 data
            payload_data = {k: v for k, v in p.items()
                            if k not in ('mark_key', '_id', 'data')}
            payload_data.update(inner_data)
            new_payloads.append({
                'mark_key': new_mk,
                'datatype': biz,
                'data': payload_data,
            })
        kwargs['payloads'] = new_payloads
        kwargs.pop('_id', None)

    if action == 'delete' and biz:
        mk = kwargs.get('mark_key') or ''
        prefix = biz + ':'
        if not mk.startswith(prefix):
            kwargs['mark_key'] = prefix + mk
        # delete 也走 mark_key 匹配而非 _id

    body.update(kwargs)
    r = requests.post(cb['url'], json=body, timeout=6,
                      proxies={'http': None, 'https': None})
    r.raise_for_status()
    res = r.json()
    if res.get('code') not in (0, None):
        raise RuntimeError('用户中心云函数错误: %s' % (
            res.get('msg') or res.get('error') or '未知'))
    data = res.get('data')

    if action == 'get_all' and biz and isinstance(data, list):
        # song_marks 标准格式：datatype 在 data 嵌套层
        return [it for it in data
                if ((it.get('data') or {}).get('datatype') or '') == biz]

    return data

def _cloud_load_users():
    """从云端拉取全部用户。返回 {username: userdict} 或 None（未配置/失败）。

    关键：用 data.username 作为 key（不是 _id / mark_key）。
    云端记录主键 _id 可能是 objectId；mark_key 又带 'user:' 前缀 —— 直接当 username
    会构造出 'user:杨颖大王' 这种伪账号，必须剥前缀。
    """
    cb = _cloud_auth_cfg()
    if not cb:
        return None
    try:
        items = _cloud_auth_call(cb, 'get_all', **{'class': 'mf_users'}) or []
        out = {}
        for item in items:
            # 云端结构 { _id, data: { mark_key, datatype, data: { 业务字段 } } } ---- 业务字段在 .data.data
            wrapper = item.get('data') or {}
            inner = wrapper.get('data') or {}
            # 防御：必须是用户业务（_cloud_auth_call 已按 datatype 过滤一遍，这里再核一次）
            if (wrapper.get('datatype') or '') != 'user':
                continue
            # 用户名只信任 inner.username，不再用 _id / mark_key —— 防止「user:xxx」假账号
            uname = (inner.get('username') or '').strip()
            if not uname or not isinstance(inner, dict):
                continue
            # 跳过任何看起来像前缀污染的（极端防御：data.username 里再带 'user:' 也算脏）
            if uname.startswith(('user:', 'audit:', 'invite:')):
                logger.warning('云端用户条目 username 含业务前缀，已跳过: %r', uname)
                continue
            # 写入本地前清掉外层 username 字段，保持顶层 schema 单一权威（与 _register_user 一致）
            clean = {k: v for k, v in inner.items() if k != 'username'}
            clean.setdefault('pw', '')  # 必填兜底
            out[uname] = clean
        return out
    except Exception:
        logger.exception('云端用户中心读取失败')
        return None

def _cloud_push_user(username, data):
    """把单个用户 upsert 到云端 mf_users。返回是否成功。"""
    cb = _cloud_auth_cfg()
    if not cb:
        return False
    try:
        payload = dict(data)
        payload['username'] = username
        _cloud_auth_call(cb, 'batch_upsert', **{
            'class': 'mf_users', 'payloads': [{'_id': username, 'data': payload}]})
        return True
    except Exception:
        logger.exception('云端用户推送失败: %s', username)
        return False

def _sync_users_from_cloud():
    """启动时从云端同步用户到本地（合并补缺，不覆盖本地已有）。"""
    global _cloud_auth_synced
    if _cloud_auth_synced:
        return
    _cloud_auth_synced = True
    cloud = _cloud_load_users()
    if cloud is None:
        logger.info('云端用户中心未配置或读取失败，维持纯本地模式')
        return
    local = _load_users_auth()
    changed = False
    for uname, data in cloud.items():
        if uname not in local:
            local[uname] = data
            changed = True
    if changed:
        _save_users_auth(local)
    logger.info('v4.20 云端用户中心同步完成: 本地 %d 用户, 云端 %d 用户',
                len(local), len(cloud))


def _cloud_delete_user(username):
    """从云端 mf_users 删除账号（与本地删除对称，保证跨设备移除生效）。"""
    cb = _cloud_auth_cfg()
    if not cb:
        return False
    try:
        _cloud_auth_call(cb, 'delete', **{'class': 'mf_users', 'mark_key': username})
        return True
    except Exception:
        logger.exception('云端删除用户失败: %s', username)
        return False


def _fetch_user_from_cloud_if_missing(users, name):
    """若本地 users 不含 name，尝试从云端拉取并回写本地缓存；返回 name 是否已就位。

    用于管理接口（改密/禁用/权限）：让「仅存在于云端」的账号也能被任意设备操作，
    操作后 _save_users_auth 会把变更推回云端，形成闭环。
    """
    if name in users:
        return True
    cloud = _cloud_load_users()
    if cloud and name in cloud:
        users[name] = cloud[name]
        _save_users_auth(users)
        return True
    return False


def _save_users_auth_cloud(users):
    """v4.20: 本地保存后异步推每个用户到云端。"""
    ex = _get_cloud_auth_executor()
    for uname, data in users.items():
        fut = ex.submit(_cloud_push_user, uname, dict(data))
        fut.add_done_callback(lambda f, u=uname: f.exception() and
                              logger.error('云端用户推送失败 [%s]: %s', u, f.exception()))


def _push_local_users_to_cloud():
    """v4.27.3: 启动时把本地用户推到云端（补缺，不覆盖云端已有）。

    修复历史问题：_sync_users_from_cloud 只做 cloud→local 单向同步，
    若用户在云同步启用前已创建，本地有但云端没有 → 新设备登录失败。
    此函数在启动时反向同步 local→cloud，确保跨设备登录可用。
    """
    cb = _cloud_auth_cfg()
    if not cb:
        return
    try:
        cloud = _cloud_load_users() or {}
    except Exception:
        logger.exception('启动推送本地用户到云端：读取云端用户失败')
        return
    local = _load_users_auth()
    pushed = 0
    for uname, data in local.items():
        if uname not in cloud:
            try:
                _cloud_push_user(uname, dict(data))
                pushed += 1
            except Exception:
                logger.exception('启动推送本地用户到云端失败: %s', uname)
    if pushed:
        logger.info('v4.27.3 启动补推 %d 个本地用户到云端', pushed)

def _audit_log(username, action, result='ok', ip=''):
    """写审计日志到云端 mf_audit 集合。异步不阻塞。"""
    cb = _cloud_auth_cfg()
    if not cb:
        return
    def _do():
        try:
            _cloud_auth_call(cb, 'batch_upsert', **{
                'class': 'mf_audit',
                'payloads': [{'_id': '%s_%d' % (username or 'x', int(time.time()*1000)),
                              'data': {'username': username or '', 'action': action,
                                       'result': result, 'ip': ip, 'ts': time.time()}}]})
        except Exception:
            logger.exception('审计日志写入失败')
    _get_cloud_auth_executor().submit(_do)

# v4.20 邀请码制（里程碑2预埋）
def _cloud_create_invite(invited_by, note=''):
    """admin 发邀请码 → 云端 mf_invite_codes 集合，用完即废。"""
    cb = _cloud_auth_cfg()
    if not cb:
        return None
    import secrets as _s
    code = _s.token_hex(4).upper()  # 8 位邀请码
    try:
        _cloud_auth_call(cb, 'batch_upsert', **{
            'class': 'mf_invite_codes',
            'payloads': [{'_id': code, 'data': {'code': code, 'invited_by': invited_by,
                         'note': note, 'used': False, 'created_at': time.time()}}]})
        return code
    except Exception:
        logger.exception('邀请码创建失败')
        return None

def _cloud_use_invite(code, username):
    """注册时验证并消费邀请码。返回 True=有效/False=无效或已用。"""
    cb = _cloud_auth_cfg()
    if not cb:
        return True  # 云端未配置时不拦截
    try:
        items = _cloud_auth_call(cb, 'get_all', **{'class': 'mf_invite_codes'}) or []
        for item in items:
            # 云端结构 { _id, data: { mark_key, datatype, data: { 业务字段 } } }
            d = ((item.get('data') or {}).get('data')) or {}
            if d.get('code') == code and not d.get('used'):
                # 标记已用
                _cloud_auth_call(cb, 'batch_upsert', **{
                    'class': 'mf_invite_codes',
                    'payloads': [{'_id': item.get('_id') or code,
                                  'data': {**d, 'used': True, 'used_by': username,
                                           'used_at': time.time()}}]})
                return True
        return False
    except Exception:
        logger.exception('邀请码验证失败')
        return True  # 出错不拦截（降级）


# ════════ 版本号 (v4.17.1) — 单一权威来源，前端从这里取 ════════
@app.route('/api/version', methods=['GET'])
def api_version():
    import time as _t
    global _APP_START_TS
    if _APP_START_TS is None:
        _APP_START_TS = _t.time()
    return jsonify({
        'version': APP_VERSION,
        'name': APP_VERSION_NAME,
        'date': APP_VERSION_DATE,
        'build': _BUILD_STAMP,
        'uptime_sec': int(_t.time() - _APP_START_TS),
    })


@app.route('/api/auth/me', methods=['GET'])
def api_auth_me():
    u = _cur_user()
    return jsonify({'logged_in': bool(u), 'username': u, 'is_admin': _is_admin(u),
                   'manage_marks': _can_user(u, 'manage_marks'),
                   'view_all_marks': _can_user(u, 'view_all_marks')})


@app.route('/api/auth/register', methods=['POST'])
def api_auth_register():
    data = request.get_json(silent=True) or {}
    username = (data.get('username', '') or '').strip()
    password = data.get('password', '')
    invite_code = (data.get('invite_code', '') or '').strip()
    # v4.20 邀请码制：杨颖大王本人免码，其他人必须持有效邀请码
    if username != ADMIN_USERNAME:
        if not invite_code:
            return jsonify({'error': '需要邀请码才能注册，请联系管理员获取'}), 403
        if not _cloud_use_invite(invite_code, username):
            return jsonify({'error': '邀请码无效或已被使用'}), 403
    ok, msg = _register_user(username, password)
    if not ok:
        try: _audit_log(username, 'register_fail', 'denied', request.remote_addr if request else '')
        except: pass
        return jsonify({'error': msg}), 400
    session['username'] = username
    g.current_user = session['username']
    # v4.20 异步化：三个云同步函数每个 timeout=15s，串行执行最慢可拖到 45s。
    # 全部挪到后台线程，立刻返 200；后台异常 logger.exception 不影响登录。
    try:
        threading.Thread(target=_migrate_legacy_cloud_marks, args=(session['username'],),
                         daemon=True).start()
        threading.Thread(target=_pull_cloud_marks_once, daemon=True).start()
        threading.Thread(target=_pull_cloud_blobs_once, daemon=True).start()
    except Exception:
        logger.exception('提交后台同步任务失败')
    try: _audit_log(username, 'register', 'ok', request.remote_addr if request else '')
    except: pass
    return jsonify({'ok': True, 'username': session['username']})


@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    data = request.get_json(silent=True) or {}
    ok, msg = _login_user(data.get('username', ''), data.get('password', ''))
    if not ok:
        return jsonify({'error': msg}), 401
    session['username'] = data.get('username', '').strip()
    g.current_user = session['username']  # 本次请求内立即生效
    # v4.20 异步化：legacy 迁移 + 标记云拉 + blob 云拉都丢后台，绝不阻塞登录响应
    try:
        threading.Thread(target=_migrate_legacy_cloud_marks, args=(session['username'],),
                         daemon=True).start()  # legacy 标记归属到本账号
        threading.Thread(target=_pull_cloud_marks_once, daemon=True).start()  # 拉取本账号云端标记
        threading.Thread(target=_pull_cloud_blobs_once, daemon=True).start()  # 把云端个人数据拉回本地（换设备即用）
    except Exception:
        logger.exception('提交后台同步任务失败')
    try:
        _touch_last_login(session['username'])
    except Exception:
        logger.exception('更新最后登录时间失败')
    return jsonify({'ok': True, 'username': session['username']})


@app.route('/api/auth/logout', methods=['POST'])
def api_auth_logout():
    session.pop('username', None)
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════
#  管理员后台：账号管理（仅管理员可访问）
# ═══════════════════════════════════════════════════

def _user_stats(username):
    """统计某账号的歌单/标记/固定/红心数量（try 保护，文件缺失不影响）。"""
    ud = _user_dir(username) or ''
    stats = {'playlist': 0, 'marks': 0, 'pins': 0, 'hearts': 0, 'has_dir': False}
    if not ud or not os.path.isdir(ud):
        return stats
    stats['has_dir'] = True
    try:
        d = json.load(open(os.path.join(ud, 'my_playlist.json'), encoding='utf-8'))
        stats['playlist'] = len(d.get('songs', [])) if isinstance(d, dict) else 0
    except Exception:
        pass
    try:
        d = json.load(open(os.path.join(ud, 'song_marks.json'), encoding='utf-8'))
        if isinstance(d, dict):
            m = d.get('marks')
            stats['marks'] = len(m) if m is not None else 0
        elif isinstance(d, list):
            stats['marks'] = len(d)
    except Exception:
        pass
    try:
        d = json.load(open(os.path.join(ud, 'my_playlist_pins.json'), encoding='utf-8'))
        stats['pins'] = len(d) if isinstance(d, list) else 0
    except Exception:
        pass
    try:
        d = json.load(open(os.path.join(ud, 'platform_hearts.json'), encoding='utf-8'))
        if isinstance(d, dict):
            stats['hearts'] = sum(len(v) for v in d.values() if isinstance(v, list))
    except Exception:
        pass
    return stats


@app.route('/api/admin/users', methods=['GET'])
def api_admin_users():
    err = _require_admin()
    if err:
        return err
    users = _load_users_auth()
    # v4.27: 合并云端账号，使管理员在任意设备上都能看到并管理全部账号（不依赖重启快照）
    cloud = _cloud_load_users()
    merged = dict(users)
    if cloud:
        for un, ud in cloud.items():
            if un not in merged:
                merged[un] = ud
            else:
                # 云端为权威：is_admin / disabled 以云端为准
                merged[un]['is_admin'] = ud.get('is_admin', merged[un].get('is_admin'))
                merged[un]['disabled'] = ud.get('disabled', merged[un].get('disabled'))
    out = []
    for name, u in merged.items():
        out.append({
            'username': name,
            'created_at': u.get('created_at', ''),
            'last_login': u.get('last_login', ''),
            'is_admin': bool(u.get('is_admin')),
            'disabled': bool(u.get('disabled')),
            'perms': u.get('perms') or _def_perms(name),
            'stats': _user_stats(name),
            'source': 'cloud' if (cloud and name in cloud and name not in users) else 'local',
        })
    out.sort(key=lambda x: (not x['is_admin'], x['created_at']))
    return jsonify({'ok': True, 'users': out, 'current': _cur_user()})


@app.route('/api/admin/sync_local_users', methods=['POST'])
def api_admin_sync_local_users():
    """把本机 users_auth.json 里的全部账号上传到云端（含密码哈希），供其他设备拉取登录。

    用于：老设备/离线注册的账号从未上云，在新版本「零配置自动上云」生效前，
    点一下即可把存量账号补齐到云端。已存在的账号会被覆盖更新。
    """
    err = _require_admin()
    if err:
        return err
    cb = _cloud_auth_cfg()
    if not cb:
        return jsonify({'error': '云端未配置（内置默认值应自动生效，请检查网络或 config.json）'}), 500
    users = _load_users_auth()
    synced, failed = [], []
    for name, data in users.items():
        try:
            if _cloud_push_user(name, dict(data)):
                synced.append(name)
            else:
                failed.append(name)
        except Exception:
            logger.exception('同步用户失败: %s', name)
            failed.append(name)
    return jsonify({'ok': True, 'synced': synced, 'failed': failed, 'total': len(users)})


@app.route('/api/admin/user/delete', methods=['POST'])
def api_admin_user_delete():
    err = _require_admin()
    if err:
        return err
    data = request.get_json(silent=True) or {}
    name = (data.get('username') or '').strip()
    if not name:
        return jsonify({'error': '缺少用户名'}), 400
    if name == ADMIN_USERNAME:
        return jsonify({'error': '管理员账号不可删除'}), 400
    if name == _cur_user():
        return jsonify({'error': '不能删除当前登录的账号'}), 400
    users = _load_users_auth()
    if name not in users:
        return jsonify({'error': '账号不存在'}), 404
    if users[name].get('is_admin') and sum(1 for u in users.values() if u.get('is_admin')) <= 1:
        return jsonify({'error': '不能删除最后一个管理员账号'}), 400
    users.pop(name)
    _save_users_auth(users)
    # v4.27: 同步删除云端记录，防止其他设备重启同步后又冒出来
    try:
        _cloud_delete_user(name)
    except Exception:
        logger.debug('云端删除用户失败（可能未配置或网络问题）: %s', name)
    ud = _user_dir(name)
    if ud and os.path.isdir(ud):
        try:
            shutil.rmtree(ud)
        except Exception:
            logger.exception('删除用户目录失败: %s', ud)
    return jsonify({'ok': True})


@app.route('/api/admin/user/reset_password', methods=['POST'])
def api_admin_user_reset_password():
    err = _require_admin()
    if err:
        return err
    data = request.get_json(silent=True) or {}
    name = (data.get('username') or '').strip()
    new_pw = data.get('password', '')
    if not name or len(new_pw) < 4:
        return jsonify({'error': '用户名或新密码无效（密码至少 4 位）'}), 400
    users = _load_users_auth()
    if name not in users:
        return jsonify({'error': '账号不存在'}), 404
    users[name]['pw'] = _hash_password(new_pw)
    _save_users_auth(users)
    return jsonify({'ok': True})


@app.route('/api/admin/user/toggle_disable', methods=['POST'])
def api_admin_user_toggle_disable():
    err = _require_admin()
    if err:
        return err
    data = request.get_json(silent=True) or {}
    name = (data.get('username') or '').strip()
    if not name:
        return jsonify({'error': '缺少用户名'}), 400
    if name == ADMIN_USERNAME:
        return jsonify({'error': '管理员账号不可禁用'}), 400
    if name == _cur_user():
        return jsonify({'error': '不能禁用当前登录的账号'}), 400
    users = _load_users_auth()
    if name not in users:
        return jsonify({'error': '账号不存在'}), 404
    if users[name].get('is_admin') and sum(1 for u in users.values() if u.get('is_admin')) <= 1:
        return jsonify({'error': '不能禁用最后一个管理员账号'}), 400
    users[name]['disabled'] = not users[name].get('disabled', False)
    _save_users_auth(users)
    return jsonify({'ok': True, 'disabled': users[name]['disabled']})


@app.route('/api/admin/user/set_perms', methods=['POST'])
def api_admin_user_set_perms():
    """admin 给指定账号设置细粒度标记权限：view_all_marks / manage_marks。
    - view_all_marks：能否看共享视图里其他账号的标记（默认 True）
    - manage_marks：能否在共享视图里修改/删除他人打的标（默认 False，点名才开）
    """
    err = _require_admin()
    if err:
        return err
    data = request.get_json(silent=True) or {}
    name = (data.get('username') or '').strip()
    if not name:
        return jsonify({'error': '缺少用户名'}), 400
    perms_in = data.get('perms') or {}
    users = _load_users_auth()
    if name not in users:
        return jsonify({'error': '账号不存在'}), 404
    if _is_admin(name):
        return jsonify({'error': '管理员账号拥有全部权限，无需单独设置'}), 400
    u = users[name]
    perms = u.get('perms')
    if not isinstance(perms, dict):
        perms = _def_perms(name)
    perms['view_all_marks'] = bool(perms_in.get('view_all_marks', perms.get('view_all_marks', True)))
    perms['manage_marks'] = bool(perms_in.get('manage_marks', perms.get('manage_marks', False)))
    u['perms'] = perms
    _save_users_auth(users)
    return jsonify({'ok': True, 'username': name, 'perms': perms})


# ════════ v4.20 邀请码管理 + 审计日志接口（里程碑2+3） ════════
@app.route('/api/admin/invite', methods=['POST'])
def api_admin_invite():
    err = _require_admin()
    if err:
        return err
    data = request.get_json(silent=True) or {}
    count = min(int(data.get('count', 1) or 1), 20)
    note = (data.get('note', '') or '').strip()[:200]
    codes = []
    for _ in range(count):
        code = _cloud_create_invite(_cur_user(), note)
        if code:
            codes.append(code)
    if not codes:
        return jsonify({'error': '邀请码生成失败，请检查云端配置'}), 500
    return jsonify({'ok': True, 'codes': codes})


# v4.26 修复：云端用户中心链路已验证正常（mark-sync 仅认 collection 字段，
# 客户端 _cloud_auth_call 已正确转换为 mark_key 前缀 + datatype）。
# 邀请码直接读云端 mf_invite_codes，不再依赖本地兜底文件。
@app.route('/api/admin/invites', methods=['GET'])
def api_admin_invites():
    err = _require_admin()
    if err:
        return err
    cb = _cloud_auth_cfg()
    if not cb:
        return jsonify({'invites': [], 'note': '云端未配置'})
    try:
        items = _cloud_auth_call(cb, 'get_all', **{'class': 'mf_invite_codes'}) or []
        out = []
        for item in items:
            # 云端结构 { _id, data: { mark_key, datatype, data: { 业务字段 } } }
            d = ((item.get('data') or {}).get('data')) or {}
            out.append({
                'code': d.get('code', ''), 'invited_by': d.get('invited_by', ''),
                'note': d.get('note', ''), 'used': d.get('used', False),
                'used_by': d.get('used_by', ''), 'used_at': d.get('used_at'),
                'created_at': d.get('created_at'), 'source': 'cloud',
            })
        out.sort(key=lambda x: x.get('created_at') or 0, reverse=True)
        return jsonify({'invites': out})
    except Exception:
        logger.exception('获取邀请码列表失败')
        return jsonify({'error': '获取失败'}), 500


# ═══════════════════════════════════════════════════════════════
#  艺人变体白名单 admin 端点（v4.25.x 起）
#  picker 把「同艺人多账号发歌」（钦觉 / 钦觉呀 / 钦觉(微唱)）显式认回正主，
#  避免归一核误判不同人。维护入口只暴露给 admin，避免随意添加。
# ═══════════════════════════════════════════════════════════════
@app.route('/api/admin/performer_aliases', methods=['GET'])
def api_admin_performer_aliases_get():
    err = _require_admin()
    if err:
        return err
    _load_performer_aliases()
    _load_alias_suggestions()
    return jsonify({'aliases': _all_performer_aliases(),
                    'suggestions': list(_performer_alias_sugs.values())})


@app.route('/api/admin/performer_aliases', methods=['POST'])
def api_admin_performer_aliases_add():
    err = _require_admin()
    if err:
        return err
    _load_performer_aliases()
    data = request.get_json(silent=True) or {}
    base = (data.get('base') or '').strip()
    alias = (data.get('alias') or '').strip()
    if not base or not alias or base == alias:
        return jsonify({'error': '基名与变体均必填且不能相同'}), 400
    with _performer_aliases_lock:
        cur = _PERFORMER_ALIASES.setdefault(base, set())
        cur.add(alias)
        _save_performer_aliases()
    # 批准后从待确认队列移除（若来自建议）
    _load_alias_suggestions()
    with _performer_alias_sug_lock:
        _performer_alias_sugs.pop(f"{base}||{alias}", None)
        _save_alias_suggestions()
    return jsonify({'ok': True, 'aliases': _all_performer_aliases(),
                    'suggestions': list(_performer_alias_sugs.values())})


@app.route('/api/admin/performer_aliases', methods=['DELETE'])
def api_admin_performer_aliases_del():
    err = _require_admin()
    if err:
        return err
    _load_performer_aliases()
    data = request.get_json(silent=True) or {}
    base = (data.get('base') or '').strip()
    alias = (data.get('alias') or '').strip()
    if not base or not alias:
        return jsonify({'error': '基名与变体均必填'}), 400
    with _performer_aliases_lock:
        if base in _PERFORMER_ALIASES:
            _PERFORMER_ALIASES[base].discard(alias)
            if not _PERFORMER_ALIASES[base]:
                del _PERFORMER_ALIASES[base]
        _save_performer_aliases()
    return jsonify({'ok': True, 'aliases': _all_performer_aliases(),
                    'suggestions': list(_performer_alias_sugs.values())})


@app.route('/api/admin/performer_alias_suggestions/dismiss', methods=['POST'])
def api_admin_performer_alias_suggestions_dismiss():
    """忽略一条待确认建议（不并入白名单，仅从队列移除）。"""
    err = _require_admin()
    if err:
        return err
    _load_alias_suggestions()
    data = request.get_json(silent=True) or {}
    base = (data.get('base') or '').strip()
    alias = (data.get('alias') or '').strip()
    if not base or not alias:
        return jsonify({'error': '基名与变体均必填'}), 400
    with _performer_alias_sug_lock:
        _performer_alias_sugs.pop(f"{base}||{alias}", None)
        _save_alias_suggestions()
    return jsonify({'ok': True, 'suggestions': list(_performer_alias_sugs.values())})


@app.route('/api/admin/audit', methods=['GET'])
def api_admin_audit():
    err = _require_admin()
    if err:
        return err
    cb = _cloud_auth_cfg()
    if not cb:
        return jsonify({'logs': [], 'note': '云端未配置'})
    try:
        items = _cloud_auth_call(cb, 'get_all', **{'class': 'mf_audit'}) or []
        out = []
        for item in items:
            # 云端结构 { _id, data: { mark_key, datatype, data: { 业务字段 } } }
            d = ((item.get('data') or {}).get('data')) or {}
            out.append({
                'username': d.get('username', ''), 'action': d.get('action', ''),
                'result': d.get('result', ''), 'ip': d.get('ip', ''),
                'ts': d.get('ts'),
            })
        out.sort(key=lambda x: x.get('ts') or 0, reverse=True)
        filter_result = (request.args.get('result') or '').strip()
        if filter_result:
            out = [x for x in out if x.get('result') == filter_result]
        limit = min(int(request.args.get('limit', 500) or 500), 2000)
        out = out[:limit]
        return jsonify({'logs': out, 'total': len(out)})
    except Exception:
        logger.exception('获取审计日志失败')
        return jsonify({'error': '获取失败'}), 500


@app.route('/admin')
def admin_page():
    err = _require_admin()
    if err:
        return err  # 非管理员返回 403
    return render_template('admin.html')


# ═══════════════════════════════════════════════════
#  网易云红心数(eapi) —— 纯 Python 实现（甩掉 Node 依赖）
# ═══════════════════════════════════════════════════

def _netease_eapi_red_count(ids):
    """纯 Python 复刻 netease_eapi.js：AES-128-ECB 加密请求网易云红心数接口。
    返回 {songId(str): {count, countDesc}}。失败时回退 Node 脚本。"""
    out = {}
    try:
        from Crypto.Cipher import AES
        from Crypto.Util.Padding import pad
        import hashlib as _hl

        def eapi(uri, data):
            json_data = json.dumps(data)
            digest = _hl.md5(f'nobody{uri}use{json_data}md5forencrypt'.encode('utf-8')).hexdigest()
            payload = f'{uri}-36cd479b6b5-{json_data}-36cd479b6b5-{digest}'
            cipher = AES.new(b'e82ckenh8dichen8', AES.MODE_ECB)
            ct = cipher.encrypt(pad(payload.encode('utf-8'), 16))
            return ct.hex().upper()

        for sid in ids:
            uri = '/api/song/red/count'
            header = {
                'osver': '16.2', 'deviceId': '', 'os': 'iPhone OS', 'appver': '9.0.90',
                'versioncode': '140', 'mobilename': '', 'buildver': str(int(time.time())),
                'resolution': '1170x2532', '__csrf': '', 'channel': 'distribution',
                'requestId': f'{int(time.time() * 1000)}_{str(random.randint(0, 9999)).zfill(4)}',
            }
            body = {'songId': int(sid), 'header': header}
            params = eapi(uri, body)
            cookie = '; '.join(
                f'{urllib.parse.quote(str(k))}={urllib.parse.quote(str(v))}' for k, v in header.items()
            )
            try:
                resp = requests.post(
                    'https://interface.music.163.com/eapi/song/red/count',
                    data={'params': params},
                    headers={
                        'Content-Type': 'application/x-www-form-urlencoded',
                        'User-Agent': 'NeteaseMusic 9.0.90/5038 (iPhone; iOS 16.2; zh_CN)',
                        'Cookie': cookie,
                    },
                    timeout=8,
                )
                d = resp.json()
                out[str(sid)] = {
                    'count': d.get('data', {}).get('count'),
                    'countDesc': d.get('data', {}).get('countDesc'),
                }
            except Exception as e:
                print(f'[NetEase] eapi single error {sid}: {e}')
                out[str(sid)] = None
        return out
    except Exception as e:
        print(f'[NetEase] python eapi error: {e}，回退 Node')
        return _netease_eapi_red_count_node(ids)


def _netease_eapi_red_count_node(ids):
    """回退方案：用 Node 脚本取红心数（需要 Node 环境）"""
    out = {}
    if not (os.path.exists(NETEASE_EAPI_JS) and NODE_BIN):
        return out
    try:
        proc = subprocess.run(
            [NODE_BIN, NETEASE_EAPI_JS, json.dumps([str(i) for i in ids])],
            capture_output=True, text=True, timeout=40,
        )
        if proc.returncode == 0 and proc.stdout.strip():
            out = json.loads(proc.stdout.strip())
        else:
            print(f"[NetEase] eapi stderr: {proc.stderr[:200]}")
    except Exception as e:
        print(f"[NetEase] eapi error: {e}")
    return out


# ═══════════════════════════════════════════════════
#  浏览器登录（弹出真实浏览器窗口，登录后自动保存 Cookie）
# ═══════════════════════════════════════════════════

LOGIN_HOME = {
    # QQ 音乐：直接打开登录页，避免首页需要用户再点登录
    'qq': 'https://y.qq.com/n/ryqq/login',
    'kugou': 'https://www.kugou.com/',
    'kuwo': 'https://www.kuwo.cn/',
    'netease': 'https://music.163.com/',
    # 抖音登录接口直接访问会返回 JSON 错误（非法应用），只能先打开首页，让用户点右上角「登录」
    'qishui': 'https://www.douyin.com/',
}
# 各平台“已登录”判定 cookie（存在其一即视为登录态，绝对判定）
LOGIN_AUTH_COOKIES = {
    'qq': ['uin', 'p_uin', 'skey'],
    'kugou': ['KuGoo', 'USERID', 'token'],
    # 酷我登录后的 Cookie 名不固定（uid/name/token/kwo/userid/websid 等都可能），
    # 这里列常用候选，配合 run_browser_login 的「非统计类 Cookie 兜底」判定，避免漏识别
    'kuwo': ['uid', 'userid', 'name', 'token', 'kwo', 'kwo_token',
             'music_u', 'websid', 'uname', 'uname3', 'sid', 'passport', 'auth'],
    'netease': ['MUSIC_U', 'ntes_utid'],
    # 注意：抖音打开首页即自动下发 sessionid_ss / sid_tt / sid_guard / passport_csrf_token 等“游客 cookie”，
    # 这些不能用来判定登录。只有真正登录后才会下发 sid_ucp_v1 / passport_auth_status / uid_tt / user_id。
    # 注意：游客态也可能下发 sessionid（无 _ss 后缀），所以这里不再用 sessionid 判定，避免误判。
    'qishui': ['sid_ucp_v1', 'passport_auth_status', 'uid_tt', 'user_id'],
}

# 通用“会话类”cookie 关键字已弃用：抖音等平台会在未登录时就下发游客 Cookie（如
# passport_csrf_token / passport_auth_mix_state / sessionid_ss / sid_tt 等），
# 按关键字匹配极易误判。改为只相信各平台 LOGIN_AUTH_COOKIES 中的精确 Cookie 名。
AUTH_COOKIE_KEYWORDS = []


LOGIN_STATUS = {}  # platform -> {running, done, success, message, cookie}

# 红心管理页专用的浏览器登录状态（与 Cookie 设置页独立，互不覆盖）
HEART_LOGIN_STATUS = {}  # platform -> {running, done, success, message, cookie, count}

# 实时登录会话：浏览器登录成功后，不关闭浏览器，而是保留（browser/context/page），
# 后续搜索复用同一「真实、已登录、指纹一致」的会话，绕过抖音把「注入 Cookie 的无头
# 浏览器」判为未登录（登录墙）的反爬限制。key=platform。
LIVE_SESSIONS = {}
LIVE_SESSIONS_LOCK = threading.Lock()
# 串行化对同一个实时浏览器会话的搜索，避免多线程同时操作同一 page 引发竞态
QISHUI_LIVE_LOCK = threading.Lock()


def _close_live_session(platform):
    """关闭并清理某个平台的实时登录会话。调用方需先持有 LIVE_SESSIONS_LOCK。"""
    sess = LIVE_SESSIONS.pop(platform, None)
    if not sess:
        return
    try:
        if sess.get('browser') is not None:
            sess['browser'].close()
    except Exception:
        pass
    try:
        if sess.get('p') is not None:
            sess['p'].stop()
    except Exception:
        pass


def _activate_chrome():
    """把刚启动的 Chrome 最后一个窗口置顶到最前，避免被压在当前窗口后面看不到。"""
    try:
        import subprocess
        script = '''tell application "Google Chrome"
    set winList to windows
    if (count winList) > 0 then
        set targetWin to item (count winList) of winList
        set index of targetWin to 1
    end if
    activate
end tell'''
        subprocess.run(["osascript", "-e", script], capture_output=True, timeout=5)
    except Exception:
        pass


def _login_validate(platform, cookie_str):
    """用「必须登录才返回数据」的真实接口验证 Cookie 是否有效（方法无关，不挑 Cookie 名）。

    返回 True 表示登录态有效。任何异常 / 不确定 / 未登录都返回 False（绝不误存）。
    这是为解决「微信 / 手机号登录的 Cookie 名与预设名单不一致导致自动判定失败」而加的
    兜底判定：只要 Cookie 真能拉到登录态数据，就认为登录成功。
    """
    try:
        import requests
        if platform == 'qq':
            url = 'https://c.y.qq.com/rsc/fcgi-bin/fcg_get_profile_homepage.fcg?format=json'
            headers = {
                'Cookie': cookie_str,
                'User-Agent': COMMON_UA,
                'Referer': 'https://y.qq.com/',
            }
            r = requests.get(url, headers=headers, timeout=6)
            try:
                d = r.json()
            except Exception:
                return False
            # 登录态：code=0 且带回昵称；匿名时通常 code!=0 或 data 为空
            if d.get('code') == 0 and d.get('data', {}).get('nickname'):
                return True
            return False
        # 其他平台暂不内置接口验证，交给「用户按钮 / 预设 Cookie 名 / meaningful 兜底」
        return False
    except Exception:
        return False


def ensure_playwright_chromium(timeout=120, on_status=None):
    """确保 Playwright Chromium 已安装（运行 exe 首次一键登录时按需下载）。

    返回 (ok, msg)。PLAYWRIGHT_BROWSERS_PATH 已由上方 bootstrap 设好：
      - Windows：指向 bundle 内的 playwright_browsers（已烤进安装包），通常直接就绪、零下载。
      - macOS  ：指向用户级缓存（'0'），未装则此处按需下载（约 180MB，仅一次）。
    探测顺序：先试系统 Chrome（channel='chrome'），再试已装 Chromium，都没有才下载。
    on_status(msg)：下载/准备过程中回传状态，供前端实时展示（避免慢网看起来像卡死）。
    """
    def _say(m):
        if on_status:
            try:
                on_status(m)
            except Exception:
                pass
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return False, '本机没装 Playwright Python 包'
    try:
        with sync_playwright().start() as p:
            # 优先系统 Chrome（多数 Mac/Win 已装，直接复用，无需下载）
            try:
                b = p.chromium.launch(headless=True, channel='chrome')
                b.close()
                return True, '已就绪（系统 Chrome）'
            except Exception:
                pass
            # 其次试已烤进包 / 已下载的 Chromium
            try:
                b = p.chromium.launch(headless=True)
                b.close()
                return True, '已就绪'
            except Exception:
                pass
    except Exception:
        pass
    # 浏览器缺失 → 调 playwright CLI 安装（仅 macOS 无 Chrome 时需要；Windows 烤包基本不会走到这）
    def _have_cli():
        from playwright._impl._driver import compute_driver_executable
        try:
            return compute_driver_executable()
        except Exception:
            return None
    cli = _have_cli()
    if not cli:
        return False, '未找到 playwright 安装命令'
    _say('正在准备浏览器：本机未检测到可用浏览器，开始下载 Chromium（约 180MB，仅首次需要，网速慢请稍候）…')
    try:
        import subprocess
        rc = subprocess.run(
            [cli, 'install', 'chromium'],
            capture_output=True, text=True, timeout=timeout,
        )
        if rc.returncode == 0:
            return True, f'已自动下载（{timeout}s 限时）'
        return False, (rc.stderr or rc.stdout or '未知错误').strip().splitlines()[-1][:200]
    except subprocess.TimeoutExpired:
        return False, f'下载超时（>={timeout}s），请检查网络后重试'
    except Exception as e:
        return False, str(e)[:200]


def run_browser_login(platform, timeout=360, target='cookies'):
    """启动系统 Chrome 打开登录页，等待用户登录，自动保存 Cookie。

    target='cookies'：保存到搜索用的 cookies.json，并保留实时浏览器会话供搜索复用。
    target='hearts'：保存到 hearts_cookies.json，并自动抓取「我喜欢」歌单；完成后关闭浏览器。
    返回 dict: {success, message, cookie, count}（count 仅在 hearts 目标且成功时填充）
    """
    if platform not in PLATFORM_NAMES:
        return {'success': False, 'message': '无效平台', 'cookie': '', 'count': 0}
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return {'success': False, 'message': '未安装 Playwright，无法弹出浏览器', 'cookie': '', 'count': 0}

    result = {'success': False, 'message': '登录超时，请重试', 'cookie': '', 'count': 0}
    p = None
    browser = None
    try:
        # 打包环境：Windows 浏览器已烤进包（零下载）；macOS 若本机无 Chrome 则按需下载到
        # 用户级缓存。下载过程中通过 on_status 实时回传状态，慢网也不像卡死。
        if getattr(sys, '_MEIPASS', None):
            _status_reg = LOGIN_STATUS if target == 'cookies' else HEART_LOGIN_STATUS
            def _on_prep(m):
                st = _status_reg.get(platform)
                if isinstance(st, dict):
                    st['message'] = m
            ok, msg = ensure_playwright_chromium(timeout=120, on_status=_on_prep)
            if not ok:
                return {'success': False, 'message': f'浏览器未就绪：{msg}。可改用「打开登录页」手动复制 Cookie', 'cookie': '', 'count': 0}
        p = sync_playwright().start()
        try:
            browser = p.chromium.launch(headless=False, channel='chrome')
        except Exception:
            browser = p.chromium.launch(headless=False)
        _activate_chrome()  # 关键：置顶窗口，确保用户看得到
        context = browser.new_context(user_agent=COMMON_UA)
        page = context.new_page()

        # 用 init script 在「浏览器上下文」层面注入 banner 脚本——
        # 这样不管 QQ/酷狗登录后页面跳几次（QQ 跳到 /notfound、手机号登录跳到确认页等），
        # 都会重新执行，红条和「✅ 我已登录，保存 Cookie」按钮都会重新出现。
        # 关键改进（v4.25.10）：不再只靠「预设 Cookie 名」判定登录（微信 / 手机号登录写下的 Cookie 名
        # 与预设名单不一致会漏判），改为三层判定：
        #   (1) 用户主动点按钮（方法无关，最可靠）
        #   (2) 真实接口验证（不挑 Cookie 名，登录方式无关）
        #   (3) 预设 Cookie 名兜底（密码登录等，向后兼容）
        try:
            context.add_init_script(r"""
                (function() {
                    var __mf_banner_collapsed = false;
                    function showCorner() {
                        if (document.getElementById('__mf_corner_btn')) return;
                        var c = document.createElement('button');
                        c.id = '__mf_corner_btn';
                        c.textContent = '✅ 已登录？点这里保存 Cookie';
                        c.title = '点我重新保存 Cookie';
                        c.style.cssText = 'position:fixed;bottom:16px;right:16px;z-index:99999;background:#16a34a;color:#fff;border:none;border-radius:8px;padding:8px 14px;font-size:13px;font-weight:bold;font-family:sans-serif;cursor:pointer;box-shadow:0 2px 8px rgba(0,0,0,0.25)';
                        c.addEventListener('click', function() {
                            window.__mf_login_done = true;
                            var c2 = document.getElementById('__mf_corner_btn');
                            if (c2) c2.remove();
                            injectBanner();
                        });
                        (document.body || document.documentElement).appendChild(c);
                    }
                    function injectBanner() {
                        if (document.getElementById('__mf_login_banner')) return;
                        var b = document.createElement('div');
                        b.id = '__mf_login_banner';
                        if (window.__mf_login_done) {
                            b.style.cssText = 'position:fixed;top:0;left:0;right:0;z-index:1000;background:#16a34a;color:#fff;font-size:14px;padding:8px 12px;text-align:center;font-family:sans-serif;box-shadow:0 2px 8px rgba(0,0,0,0.15)';
                            b.textContent = '✅ Cookie 已自动保存，可以关闭此窗口';
                            (document.body || document.documentElement).appendChild(b);
                            setTimeout(function() { b.style.transition='opacity .4s'; b.style.opacity='0'; setTimeout(function(){ if(b.parentNode) b.parentNode.removeChild(b); }, 500); }, 3000);
                            return;
                        }
                        // 红条层级 z-index=1000，QQ/酷狗登录弹窗通常 ≥2000，会自动盖住红条；
                        // 同时 6 秒后自动收起为右下角小按钮，不再挡用户视线。
                        b.style.cssText = 'position:fixed;top:0;left:0;right:0;z-index:1000;background:#ff5e5e;color:#fff;font-size:14px;padding:8px 12px;display:flex;gap:12px;justify-content:center;align-items:center;font-family:sans-serif;box-shadow:0 2px 8px rgba(0,0,0,0.15)';
                        var span = document.createElement('span');
                        span.textContent = '请登录（微信 / 手机号 / 密码 均可）。看到头像即代表已登录，点右侧按钮即可保存';
                        b.appendChild(span);
                        var btn = document.createElement('button');
                        btn.id = '__mf_save_cookie';
                        btn.textContent = '✅ 我已登录，保存 Cookie';
                        btn.style.cssText = 'cursor:pointer;background:#fff;color:#ff5e5e;border:none;border-radius:6px;padding:6px 14px;font-weight:bold;font-size:13px;font-family:inherit';
                        btn.addEventListener('click', function() {
                            window.__mf_login_done = true;
                            var old = document.getElementById('__mf_login_banner');
                            if (old) old.remove();
                            injectBanner();
                        });
                        b.appendChild(btn);
                        (document.body || document.documentElement).appendChild(b);
                        // 6 秒不操作就淡出收为右下角小按钮，不挡登录弹窗
                        setTimeout(function() {
                            if (window.__mf_login_done) return;
                            if (document.getElementById('__mf_login_banner') !== b) return;
                            b.style.transition = 'opacity .5s';
                            b.style.opacity = '0';
                            setTimeout(function() {
                                if (b.parentNode) b.parentNode.removeChild(b);
                                if (document.getElementById('__mf_corner_btn')) return;
                                showCorner();
                            }, 600);
                        }, 6000);
                    }
                    if (document.readyState === 'loading') {
                        document.addEventListener('DOMContentLoaded', injectBanner);
                    } else {
                        injectBanner();
                    }
                })();
            """)
        except Exception:
            pass

        page.goto(LOGIN_HOME[platform], wait_until='domcontentloaded', timeout=30000)

        logged_in = False
        steps = timeout // 2
        known = LOGIN_AUTH_COOKIES.get(platform, [])
        strong = STRONG_AUTH_COOKIES.get(platform, [])
        for _ in range(steps):
            try:
                cookies = context.cookies()
                names = {c['name'] for c in cookies}
                # 方法 1：用户主动点「我已登录」按钮（最可靠，微信 / 手机号 / 密码均可）
                try:
                    if page.evaluate('window.__mf_login_done === true'):
                        logged_in = True
                        break
                except Exception:
                    pass
                # 方法 2：真实接口验证（不挑 Cookie 名，登录方式无关）
                cookie_str = '; '.join(f"{c['name']}={c['value']}" for c in cookies)
                if cookie_str and _login_validate(platform, cookie_str):
                    logged_in = True
                    break
                # 方法 3：预设登录 Cookie 名（密码登录等，向后兼容）
                hit = any(k in names for k in known)
                if hit:
                    logged_in = True
                    # 若平台有强鉴权 Cookie（如网易云 MUSIC_U），等其落盘后再存，
                    # 否则存到的可能是游客态 Cookie，后续仍拿不到播放/收藏量
                    if strong and not any(s in names for s in strong):
                        time.sleep(2)
                        continue
                    break
            except Exception:
                pass
            time.sleep(2)

        # 兜底判定：若上下文里出现了「非统计类」的业务 Cookie（如 uid/token/kwo 等），
        # 即使不在我们预设的精确名单里，也视为登录成功（解决酷我等平台 Cookie 名不确定问题）。
        # 但 v4.25.9 收紧：纯游客态也会下 ts_uid 等追踪 Cookie，盲目存会误存匿名 Cookie 还报成功，
        # 因此要求「非统计类 Cookie」里至少出现一个真正的会话型 Cookie 名，或用户已点按钮。
        ANALYTICS = {'hm_lvt', 'hm_lpvt', 'hmaccount', 'hm_iuvt',
                     '_ga', '_gid', '_gat', 'hacid', 'hmsr', '_hmid'}
        # 会话型 Cookie 关键字（用于区分「真登录」与「游客追踪」）
        SESSION_HINT = {'token', 'skey', 'uin', 'sid', 'kugou', 'user_id',
                        'userid', 'passport', 'auth', 'mid', 'oauth', 'login'}
        cookies = context.cookies()
        meaningful = [c for c in cookies
                      if c['name'].lower() not in ANALYTICS
                      and c['name'].lower() != 'kw_token']
        meaningful_has_session = any(
            any(h in c['name'].lower() for h in SESSION_HINT) for c in meaningful
        )
        if logged_in or (meaningful and meaningful_has_session):
            # 额外等待几秒，让剩余会话 Cookie 完全写入
            time.sleep(3)
            cookies = context.cookies()
            cookie_str = '; '.join(f"{c['name']}={c['value']}" for c in cookies)

            # 切到绿色「已保存」banner（让用户看到反馈，即便页面再跳也保持）
            try:
                page.evaluate("window.__mf_login_done = true; "
                              "var b=document.getElementById('__mf_login_banner'); "
                              "if(b){b.remove();} "
                              "var nb=document.createElement('div'); "
                              "nb.id='__mf_login_banner'; "
                              "nb.style.cssText='position:fixed;top:0;left:0;right:0;z-index:99999;background:#16a34a;color:#fff;font-size:14px;padding:8px 12px;text-align:center;font-family:sans-serif;box-shadow:0 2px 8px rgba(0,0,0,0.15)'; "
                              "nb.textContent='✅ Cookie 已自动保存，可以关闭此窗口'; "
                              "if(document.body)document.body.appendChild(nb);")
            except Exception:
                pass

            if target == 'hearts':
                # 保存到独立的红心 Cookie 文件
                with _heart_cookie_lock:
                    hcookies = _load_heart_cookies()
                    hcookies[platform] = {
                        'cookie': cookie_str,
                        'updated_at': time.strftime('%Y-%m-%d %H:%M:%S'),
                    }
                    _save_heart_cookies(hcookies)
                # 自动抓取「我喜欢」歌单
                fav_count = 0
                try:
                    favs = fetch_fav_songs(platform, cookie_str)
                    if favs:
                        with _hearts_lock:
                            data_all = _load_hearts()
                            p_data = _hearts_platform(data_all, platform)
                            p_data['songs'] = _favs_to_dict_list(favs)
                            p_data['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
                            p_data['source'] = 'cookie'
                            data_all[platform] = p_data
                            _save_hearts(data_all)
                        fav_count = len(favs)
                except Exception as e:
                    logger.exception('浏览器登录后抓取 %s 红心失败', platform)
                    # Cookie 已保存，但抓取失败——返回成功登录 + 友好提示
                    result = {
                        'success': True,
                        'message': f"{PLATFORM_NAMES[platform]} 登录成功，Cookie 已保存，但自动抓取歌单失败：{e}",
                        'cookie': cookie_str,
                        'count': 0,
                    }
                else:
                    result = {
                        'success': True,
                        'message': f"{PLATFORM_NAMES[platform]} 登录成功，已识别 {fav_count} 首红心歌曲",
                        'cookie': cookie_str,
                        'count': fav_count,
                    }
            else:
                # 默认：保存到搜索 Cookie
                store = load_cookies()
                store[platform] = cookie_str
                save_cookies(store)
                result = {'success': True,
                          'message': f"{PLATFORM_NAMES[platform]} 登录成功，Cookie 已自动保存（登录窗口保留中，可直接搜索）",
                          'cookie': cookie_str,
                          'count': 0}
                # 保留实时会话：不关闭浏览器，供汽水音乐搜索复用（真实已登录 + 指纹一致，
                # 绕过抖音把「注入 Cookie 的无头浏览器」判为未登录而弹出登录墙的反爬限制）。
                try:
                    with LIVE_SESSIONS_LOCK:
                        _close_live_session(platform)  # 先清掉旧会话，避免泄漏
                        LIVE_SESSIONS[platform] = {
                            'p': p, 'browser': browser, 'context': context, 'page': page,
                        }
                    p = None  # 已移交，避免下方 finally 误关
                    browser = None
                except Exception as e:
                    print(f'[Login] 保留实时会话失败: {e}')
        if browser is not None:
            try:
                browser.close()
            except Exception:
                pass
        if p is not None:
            try:
                p.stop()
            except Exception:
                pass
    except Exception as e:
        result = {'success': False,
                  'message': f"无法打开浏览器：{e}（可改用「打开登录页」手动获取）",
                  'cookie': '', 'count': 0}
        if browser is not None:
            try:
                browser.close()
            except Exception:
                pass
        if p is not None:
            try:
                p.stop()
            except Exception:
                pass
    return result


def start_browser_login(platform):
    """后台线程启动浏览器登录，结果写入 LOGIN_STATUS 供前端轮询。"""
    if LOGIN_STATUS.get(platform, {}).get('running'):
        return  # 已在登录中，避免重复拉起
    def _worker():
        LOGIN_STATUS[platform] = {
            'running': True, 'done': False, 'success': False,
            'message': f'正在打开 {PLATFORM_NAMES.get(platform, platform)} 登录页…',
            'cookie': '',
        }
        try:
            res = run_browser_login(platform, target='cookies')
        except Exception as e:
            res = {'success': False, 'message': f'登录异常：{e}', 'cookie': '', 'count': 0}
        LOGIN_STATUS[platform] = {
            'running': False, 'done': True,
            'success': res.get('success', False),
            'message': res.get('message', '未知结果'),
            'cookie': res.get('cookie', ''),
        }
    threading.Thread(target=_worker, daemon=True).start()


def start_heart_browser_login(platform):
    """后台线程启动浏览器登录（红心管理页专用），结果写入 HEART_LOGIN_STATUS 供前端轮询。"""
    if HEART_LOGIN_STATUS.get(platform, {}).get('running'):
        return
    def _worker():
        HEART_LOGIN_STATUS[platform] = {
            'running': True, 'done': False, 'success': False,
            'message': f'正在打开 {PLATFORM_NAMES.get(platform, platform)} 登录页…',
            'cookie': '', 'count': 0,
        }
        try:
            res = run_browser_login(platform, target='hearts')
        except Exception as e:
            res = {'success': False, 'message': f'登录异常：{e}', 'cookie': '', 'count': 0}
        HEART_LOGIN_STATUS[platform] = {
            'running': False, 'done': True,
            'success': res.get('success', False),
            'message': res.get('message', '未知结果'),
            'cookie': res.get('cookie', ''),
            'count': res.get('count', 0),
        }
    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════
#  通用工具函数
# ═══════════════════════════════════════════════════

def _decode_escapes(s):
    if not s:
        return s
    return (s.replace('\\u0026', '&')
             .replace('\\u0027', "'")
             .replace('\\u0022', '"'))


def _parse_lyricist_composer(lyric_text):
    """从歌词文本中解析词曲作者。

    覆盖的常见格式（逐行解析，不依赖结尾换行，最后一行无换行也能匹配）：
      [00:00.000] 作词 : 施人诚        (网易云/QQ 中文)
      [00:03.86] 词：林夕              (QQ)
      [00:01.000] 作曲 : 玉城千春
      作词：xxx  作曲：xxx
      [作词：xxx] [作曲：xxx]
      [00:00.06] 词Lyricist：侯波      (QQ AI 生成歌词，中英混排)
      [00:00.12] 曲Composer：玄昌俊
      [00:00.44]Lyrics by：Tommy Brown (QQ 英文歌，全角冒号)
      [00:00.88]Composed by：Tommy Brown (QQ 英文歌)
      [00:01.32]Arranged by：TEDDY     (QQ 编曲，附带给)
      Lyricist : Park Jinyoung          (网易云英文版)
      Composer : Teddy                  (网易云英文版)
    """
    if not lyric_text:
        return None, None

    lyricist = None
    composer = None

    # 标签候选：长标签（含英文后缀/前缀）放前面优先匹配
    # 顺序很关键：含 by 的要在不含 by 的前面，避免 "Lyricist" 被单独的 "词" 截断
    lyric_keys = [
        r'Lyrics?\s*by',           # Lyrics by / Lyric by
        r'Lyricist',               # 词 Lyricist / Lyricist: (全英)
        r'作词',
        r'词',
    ]
    comp_keys = [
        r'Composed?\s*by',         # Composed by / Compose by
        r'Composer',               # 曲 Composer / Composer: (全英)
        r'作曲',
        r'曲',
    ]

    def _find(keys, content):
        for k in keys:
            m = re.search(r'(?:' + k + r')\s*[:：]\s*(.+)$', content)
            if m:
                # 同行可能跟了另一标签（如「词：林夕 曲：陈辉阳」），截断到下一个标签前
                # 扩展截断列表，覆盖中英所有变体
                # v4.28.x 修正 + 扩充：
                #   * 旧词表把「词：/曲：/词:/曲:」写成带冒号 token，又在外层追加
                #     \s*[:：]，等于要求「曲：」后再跟一个冒号 → 永远匹配不上，
                #     导致「词：重乐 曲：方杨…」的 曲： 切断失效，词作者被污染成
                #     「重乐 曲：方杨…」脏值（②d 作者铁证比对随之失败，改名歌误判未收录）。
                #     改为裸关键词 词/曲 + 统一后缀 \s*[:：]。
                #   * 补 编曲/混音/录音/缩混/母带/制作人/监制/出品/发行/统筹/和声/伴唱 等
                #     制作角色，避免「曲：方杨@小分队 编曲：何佳 和声：任芯冉/化十…」整行被吞。
                val = re.split(
                    r'\s*(?:'
                    r'作词|作曲|词\s*Lyricist|曲\s*Composer|词|曲|'
                    r'编曲|混音|录音|缩混|母带|制作人|监制|出品|发行|统筹|和声|伴唱|'
                    r'Lyrics?\s*by|Composed?\s*by|Lyricist|Composer|'
                    r'Arranged?\s*by|Arranger|'
                    r'Piano|Guitar|Bass|Drums|Strings|Violin|Cello'
                    r')\s*[:：]',
                    m.group(1)
                )[0].strip()
                # 去掉残留的英文后缀
                val = re.sub(r'\s*[Ll]yricist$', '', val).strip()
                val = re.sub(r'\s*[Cc]omposer$', '', val).strip()
                return val
        return None

    for line in lyric_text.splitlines():
        # 去掉时间轴前缀 [mm:ss.xx]
        content = re.sub(r'^\[\d+:\d+(?:\.\d+)?\]', '', line).strip()
        if not content:
            continue
        if lyricist is None:
            v = _find(lyric_keys, content)
            if v and len(v) < 200:
                lyricist = v
        if composer is None:
            v = _find(comp_keys, content)
            if v and len(v) < 200:
                composer = v
        if lyricist and composer:
            break

    return lyricist, composer


def _parse_lyric_performer(lyric_text):
    """从歌词文本中解析【演唱者/歌手】名（用于认回改名、多账号、原唱版本）。

    与 _parse_lyricist_composer（只挖词/曲作者）互补：歌词里常直接白描歌手，
    但旧匹配器从未利用，导致「歌手改名但歌词仍署原名」的歌被误判未收录。典型来源：
      - LRC 头部「歌名 - 歌手」行（无时间戳或首行白描）：
          醉花间 - 任芯冉
      - 制作信息行内嵌的演唱/原唱/和声等标签：
          演唱：任芯冉   原唱：xxx   主唱：xxx   和声：任芯冉/化十

    返回候选名列表（去重保序）。取不到返回 []。仅作「字符串包含」级认回，
    配合 picker 的歌名强匹配(ns>=85)使用，避免被随机歌词文本误伤。
    """
    if not lyric_text:
        return []
    _ts = re.compile(r'^\[\d{1,2}:\d{2}(?:\.\d{1,3})?\]')
    # 头部「歌名 - 歌手」：取「-」后段为歌手候选
    _header = re.compile(r'^\s*(.{1,40}?)\s*[-－—~]\s*(.{1,40}?)\s*$')
    # 演唱/原唱/主唱/和声/歌手/人声 : 名字（不含后续标点/换行）
    _credit = re.compile(
        r'(?:演唱|原唱|主唱|和声|歌手|人声|vocal)\s*[:：]\s*([^，,。\.\n【】\[\]]{1,40})'
    )
    cands = []
    for raw in lyric_text.splitlines():
        content = _ts.sub('', raw).strip()
        if not content:
            continue
        # 头部行：歌名 - 歌手
        m = _header.match(content)
        if m:
            singer_part = m.group(2).strip()
            # 后段若只是制作信息（含 词/曲/编 等标签）不算歌手
            if singer_part and not re.search(r'(词|曲|编|混|录|制|监|伴|和|缩|母|统|出|发)', singer_part):
                cands.append(singer_part)
        # 制作信息行：演唱/原唱/和声：名字（可能多人 任芯冉/化十）
        for cm in _credit.finditer(content):
            val = cm.group(1).strip()
            for part in re.split(r'[/&,，、]', val):
                part = part.strip()
                if part and len(part) < 30:
                    cands.append(part)
    # 去重保序
    seen = set()
    out = []
    for c in cands:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out


# 详情字段抓取状态：ok=有数据；missing=平台确实未提供；error=我们抓取失败（程序问题）
def _mark_detail(r, field, status):
    """记录某结果某字段的抓取状态。

    核心规则：一旦已是 'ok'（已有有效数据，可能来自搜索结果），绝不被 'missing'/'error' 覆盖，
    避免详情接口偶发失败把已有数据误标红。其余情况（当前为 missing/error/none）按真实抓取结果更新，
    因此 enrichment 成功拿到数据时可从 missing 升级为 ok，抓取失败则从 missing 升级为 error。
    """
    st = r.setdefault('_detail_status', {})
    if st.get(field) == 'ok':
        return
    st[field] = status


import time as _time


def _req_json(method, url, tries=2, **kw):
    """带一次重试的 JSON 请求，降低偶发超时/风控造成的误报红。"""
    last = None
    timeout = kw.pop('timeout', 10)
    for _ in range(max(1, tries)):
        try:
            return requests.request(method, url, timeout=timeout, **kw)
        except Exception as e:
            last = e
            _time.sleep(0.4)
    raise last


# 需要在搜索结果上初始化的详情字段
DETAIL_FIELDS = ['record_label', 'collection_count', 'comment_count',
                 'listening_count', 'lyricist', 'composer']


def _parse_lyric_copyright(lyric_text):
    """从歌词文本解析唱片公司/厂牌信息（酷狗歌词常内嵌版权行）。

    酷狗歌词常在首尾写出，例如：
        OP：Hikoon Music
        出品方：乐过山丘Melody Mound
        发行公司：壹歌传媒
        SP：XXX Entertainment
    优先级：发行公司/发行方 > 录音版权(录音作品/录音制品版权，即录音制作者=厂牌) >
            出品公司/出品方 > 制作公司 > OP/SP > 出版/录制。
    注意：刻意不采集「词/曲版权管理方」「联合出品方」(多为发行代理/出品监制，非厂牌)。
    返回公司名字符串或 None。
    """
    if not lyric_text:
        return None
    patterns = [
        r'发行公司\s*[:：]\s*([^\n\r（）()]+)',
        r'发行方\s*[:：]\s*([^\n\r（）()]+)',
        # 录音作品/MV/制品版权 = 录音制作者，通常即唱片厂牌（如「录音作品及MV版权：EAS MUSIC LTD」）
        r'录音作品及?MV?版权\s*[:：]\s*([^\n\r（）()]+)',
        r'录音制品版权\s*[:：]\s*([^\n\r（）()]+)',
        r'录音版权\s*[:：]\s*([^\n\r（）()]+)',
        r'出品公司\s*[:：]\s*([^\n\r（）()]+)',
        # 负向断言：排除「联合出品方」(多为出品监制/联合投资方，非厂牌)
        r'(?<!联合)出品方\s*[:：]\s*([^\n\r（）()]+)',
        r'制作公司\s*[:：]\s*([^\n\r（）()]+)',
        r'(?<![A-Za-z])OP\s*[:：]\s*([^\n\r（）()]+)',
        r'(?<![A-Za-z])SP\s*[:：]\s*([^\n\r（）()]+)',
        r'出版\s*[:：]\s*([^\n\r（）()]+)',
        r'录制\s*[:：]\s*([^\n\r（）()]+)',
    ]
    for pat in patterns:
        m = re.search(pat, lyric_text, re.I)
        if m:
            val = m.group(1).strip().strip('。；;,.，、 ')
            if val and 1 < len(val) < 60:
                return val
    return None


def _safe_int(val):
    """安全转换为整数"""
    try:
        if val and int(val) > 0:
            return int(val)
    except:
        pass
    return None


def _safe_int_keep_zero(val):
    """安全转整数但【保留 0】（0 是合法精确值，不可当缺失）。
    _safe_int 内部 `if val and int(val)>0` 会把 0 收藏量吞成 None → 冷门歌被误标「未查到」。
    本函数仅当 val 为 None/非数字时返回 None，val=0/'0' 正常返回 0。
    v4.25.13 为修 _safe_int 吞 0 隐患（网易云/汽水同 QQ/酷狗）而加。"""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


# ═══════════════════════════════════════════════════
#  QQ音乐
# ═══════════════════════════════════════════════════

def _ts_to_date(ts, is_ms=False):
    """Unix 时间戳(秒/毫秒) → 'YYYY-MM-DD'（按北京时间解释，避免跨日偏差）；解析失败返回 None。"""
    try:
        if not ts:
            return None
        ts = float(ts)
        if is_ms:
            ts = ts / 1000.0
        from datetime import datetime, timezone, timedelta
        dt = datetime(1970, 1, 1, tzinfo=timezone.utc) + timedelta(seconds=ts + 8 * 3600)
        return dt.strftime('%Y-%m-%d')
    except Exception:
        return None


def _qq_availability(song):
    """判断 QQ 音乐单曲的可用状态。

    实测字段规律（2026-08-04，client_search_cp 搜索结果）：
      正常可播：alertid=2/41/42, pay.payplay=1, msgid=15
      灰色下架：alertid=11,      pay.payplay=0, msgid=0   （如「劫后重逢(克保和政芬)/罗大佑」）
    alertid 是点击时的提示框 ID，11 = 「因版权保护暂无法播放」，是最可靠的下架信号。
    payplay=0 单独不代表下架（免费歌也是 0），必须结合 alertid/msgid 判断。
    """
    def _i(v):
        # 注意：不能用 _safe_int，它把 0 也返回 None，会让 payplay==0 的判定失效
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    try:
        pay = song.get('pay') or {}
        alertid = _i(song.get('alertid'))
        msgid = _i(song.get('msgid'))
        payplay = _i(pay.get('payplay'))
        if alertid == 11:
            return '已下架'
        if payplay == 0 and msgid == 0:
            return '已下架'
        # payplay=1 只表示 VIP 曲目，QQ 上绝大多数歌都是，属正常上架
        if payplay == 1:
            return '在架(VIP)'
        return '在架'
    except Exception:
        return ''


def _search_qq_smartbox(keyword, max_results=10):
    """QQ 搜索的兜底通道：智能联想接口 smartbox_new.fcg。

    为什么需要它（2026-08-07 实测）：
      正式搜索 client_search_cp 被 IP 级风控后会返回
      {"code":0, "message":"query error", "subcode":-10003, song.list:[]}，
      新版 musicu.fcg DoSearchForQQMusicDesktop 也同步变空（sum=0），
      而且这一封持续 3 小时以上——2 万首的批量跑必然撞上。
      联想接口走的是另一个限流桶，正式搜索全灭时它仍能出数据，
      返回的 mid（拼链接）+ id（查收藏量）+ 歌名歌手（做匹配）刚好够批量场景用。
    局限：只返回 Top ~10 条且没有专辑/发行日期，所以只当降级用，不当主力。
    """
    try:
        resp = requests.get(
            'https://c.y.qq.com/splcloud/fcgi-bin/smartbox_new.fcg',
            params={'is_xml': 0, 'format': 'json', 'key': keyword, 'utf8': 1, 'g_tk': 5381},
            headers={'User-Agent': COMMON_UA, 'Referer': 'https://y.qq.com/'},
            timeout=10,
        )
        items = (resp.json().get('data') or {}).get('song', {}).get('itemlist', []) or []
    except Exception as e:
        print(f'[QQ] smartbox error: {e}')
        return []
    out = []
    for s in items[:max_results]:
        mid = s.get('mid', '')
        if not mid:
            continue
        out.append({
            'song_name': s.get('name', ''),
            'performer': s.get('singer', ''),
            'platform': 'QQ音乐',
            'platform_code': 'qq',
            'availability': '在架',
            'song_url': f'https://y.qq.com/n/ryqq/songDetail/{mid}',
            'album': '',
            'release_date': None,
            'lyricist': None,
            'composer': None,
            'collection_count': None,
            'listening_count': None,
            'comment_count': None,
            'share_count': None,
            'record_label': None,
            '_songmid': mid,
            '_songid': _safe_int(s.get('id')) or 0,
            '_albummid': '',
            '_from_smartbox': True,
        })
    return out


def search_qq(keyword, max_results=30):
    """QQ音乐搜索（支持多页抓取）。

    正式搜索被风控打空时自动降级到 _search_qq_smartbox（见该函数注释）。
    """
    results = []
    page_size = 30
    max_pages = (max_results + page_size - 1) // page_size
    cookie_str = get_cookie_string('qq')
    try:
        for page in range(1, max_pages + 1):
            if len(results) >= max_results:
                break
            url = 'https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
            params = {
                'w': keyword, 'format': 'json', 'p': page, 'n': page_size,
                'cr': 1, 'g_tk': 5381, 't': 0, 'aggr': 1,
                'catZhida': 1, 'lossless': 0, 'flag_qc': 0,
            }
            headers = {
                'User-Agent': COMMON_UA,
                'Referer': 'https://y.qq.com/',
            }
            if cookie_str:
                headers['Cookie'] = cookie_str
            resp = requests.get(url, params=params, headers=headers, timeout=10)
            data = resp.json()
            song_list = data.get('data', {}).get('song', {}).get('list', [])
            if not song_list:
                break
            for song in song_list:
                if len(results) >= max_results:
                    break
                singers = '/'.join(s.get('name', '') for s in song.get('singer', []))
                songmid = song.get('songmid', '')
                results.append({
                    'song_name': song.get('songname', ''),
                    'performer': singers,
                    'platform': 'QQ音乐',
                    'platform_code': 'qq',
                    'availability': _qq_availability(song),
                    'song_url': f'https://y.qq.com/n/ryqq/songDetail/{songmid}',
                    'album': song.get('albumname', ''),
                    'release_date': _ts_to_date(song.get('pubtime')),  # 搜索结果自带 pubtime（秒级时间戳）
                    'lyricist': None,
                    'composer': None,
                    'collection_count': None,
                    'listening_count': None,
                    'comment_count': None,
                    'share_count': None,
                    'record_label': None,
                    '_songmid': songmid,
                    '_songid': song.get('songid', 0),
                    '_albummid': song.get('albummid', ''),
                    # v4.25.12：时长维度（秒）。QQ 搜索返回自带 interval，归一成 duration
                    # 供匹配器「片段救援」使用（官方只发 30s 试听片段、完整版在民间重传的场景）。
                    'duration': _safe_int(song.get('interval')),
                })

    except Exception as e:
        print(f"[QQ] Error: {e}")

    if not results:
        # 正式搜索被风控（query error / subcode -10003）→ 走联想接口兜底
        results = _search_qq_smartbox(keyword, min(max_results, 10))
    # v4.27.20：监控建档要词曲作者匹配，必须抓详情（含 _fetch_qq_details 解析歌词里的
    # 作词/作曲）。之前 search_qq 末尾直接 return 漏了这步，导致 QQ 候选词曲恒为 None，
    # 词曲一致性过滤对 QQ 完全失效。netease/kugou 已有等价调用。
    if results:
        try:
            _fetch_qq_details(results[:SEARCH_ENRICH_CAP], cookie_str)
        except Exception as e:
            print(f'[QQ] details 抓取失败(词曲等): {e}')
    return results


def _parse_qq_show_count(s):
    """解析 QQ 音乐 m_show 计数，如 '5600w+' -> 56000000, '10w+' -> 100000, '1.2k+' -> 1200。"""
    if not s:
        return None
    s = s.strip().replace(',', '').replace('，', '')
    try:
        if s.endswith('w+'):
            return int(float(s[:-2]) * 10000)
        if s.endswith('万+'):
            return int(float(s[:-2]) * 10000)
        if s.endswith('k+'):
            return int(float(s[:-2]) * 1000)
        if s.endswith('+'):
            return int(float(s[:-1]))
        if s.lower().endswith('w'):
            return int(float(s[:-1]) * 10000)
        if s.endswith('万'):
            return int(float(s[:-1]) * 10000)
        if s.lower().endswith('k'):
            return int(float(s[:-1]) * 1000)
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _parse_qq_listen_text(text):
    """解析 QQ 音乐在听文本，如 '1.2万人在听' -> 12000；返回 (count, raw_text)。"""
    if not text:
        return None, None
    m = re.search(r'([\d.]+)\s*(万|w|W|k|K)?\s*人?在听', str(text))
    if not m:
        return None, str(text).strip()
    try:
        n = float(m.group(1))
    except (ValueError, TypeError):
        return None, str(text).strip()
    unit = m.group(2) or ''
    if unit == '万' or unit.lower() == 'w':
        return int(round(n * 10000)), str(text).strip()
    if unit.lower() == 'k':
        return int(round(n * 1000)), str(text).strip()
    return int(round(n)), str(text).strip()


def _extract_qq_song_company(data):
    """从 QQ get_song_detail_yqq 返回的 data.info 中提取唱片公司。

    结构示例：data.info.company.content = [{"value": "Super Mark", ...}]
    """
    if not isinstance(data, dict):
        return None
    info = data.get('info') or {}
    if not isinstance(info, dict):
        return None
    company = info.get('company') or {}
    if isinstance(company, dict):
        for item in company.get('content', []):
            val = item.get('value')
            if isinstance(val, str) and val.strip():
                return val.strip()
    # 有些接口把 company 直接放字符串
    if isinstance(company, str) and company.strip():
        return company.strip()
    return None


def _fetch_qq_details(results, cookie_str='', only_fav=False, skip_comments=False):
    """获取QQ音乐详情：词曲作者、收藏量、评论数、在听人数。

    only_fav=True：只跑「收藏量」这一段（1 个批量请求就够），跳过评论数 / 在听人数 /
    歌词词曲作者。大批量任务（2 万首）只要收藏量+链接，跑全量详情等于每首多发几十个
    请求，既慢又必被风控。

    关键发现（来自公开 PHP API 源码 / music-link-filler-cloud-hybrid）：
    - 收藏量：musicu.fcg + music.musicasset.SongFavRead.GetSongFansNumberById
      传 songId 数组，返回 data.m_numbers / data.m_show（无需签名）
    - 评论数：c.y.qq.com/base/fcgi-bin/fcg_global_comment_h5.fcg?cmd=4&biztype=1
      传 topid（歌曲ID，可逗号分隔），返回 batch_commenttotal（无需签名）
    - 在听人数：musicu.fcg + music.sharing.PlayPageSvr.GetSongTag
      需要先通过 music.pf_song_detail_svr.get_song_detail_yqq 取 songID，再调 GetSongTag 取 songTag。
      只有部分歌曲返回 "X人在听"，热门歌曲常返回榜单标签（如 "热歌榜No.37"）。
    """
    if not results:
        return

    headers = {
        'User-Agent': COMMON_UA,
        'Referer': 'https://y.qq.com/',
    }
    if cookie_str:
        headers['Cookie'] = cookie_str

    # 批量获取收藏量（分批，每批最多 50 个 songId）。
    # QQ 音乐 GetSongFansNumberById 接口一次传超过 50 个会返回 code 80071 且空数据。
    # ⚠️ 同一 songId 可能对应多行（2 万首曲库里同名同版本重复常见），用 list 收集，
    # 避免后行覆盖前行导致静默漏数。
    songid_map = {}
    for r in results:
        sid = r.get('_songid')
        if sid:
            songid_map.setdefault(str(sid), []).append(r)
    BATCH_SIZE = 50
    if songid_map:
        sids = list(songid_map.keys())
        for i in range(0, len(sids), BATCH_SIZE):
            batch_sids = sids[i:i + BATCH_SIZE]
            try:
                body = {
                    'result': {
                        'module': 'music.musicasset.SongFavRead',
                        'method': 'GetSongFansNumberById',
                        'param': {'v_songId': [int(s) for s in batch_sids]},
                    }
                }
                resp = _req_json('POST', 'https://u.y.qq.com/cgi-bin/musicu.fcg',
                                params={'format': 'json'}, json=body,
                                headers={**headers, 'Content-Type': 'application/json'},
                                timeout=10, tries=2)
                data = resp.json().get('result', {}).get('data', {})
                numbers = data.get('m_numbers', {})
                shows = data.get('m_show', {})
                for sid in batch_sids:
                    for r in songid_map[sid]:
                        # ⚠️ 必须用显式 None 判断，不能写 `numbers.get(sid) or numbers.get(int(sid))`！
                        # 0 收藏量是合法精确值，写成 `or` 会把 0（falsy）短路成 None →
                        # 冷门歌被错标成 'missing' 而非记 0（v4.25.11 补修，同酷狗 0 收藏量 bug）
                        num_val = numbers.get(sid)
                        if num_val is None:
                            try:
                                num_val = numbers.get(int(sid))
                            except (ValueError, TypeError):
                                num_val = None
                        show_val = shows.get(sid)
                        if show_val is None:
                            try:
                                show_val = shows.get(int(sid))
                            except (ValueError, TypeError):
                                show_val = None
                        # 优先用精确数字 m_numbers（含 0）
                        val = None
                        if num_val is not None:
                            try:
                                val = int(num_val)
                            except (ValueError, TypeError):
                                val = None
                        real = _parse_qq_show_count(show_val)
                        # QQ 收藏量取值规则（v4.28.x 修正，原逻辑见下「旧坑」）：
                        #   QQ 对小收藏量歌曲会返回 m_numbers=999 哨兵，真实值只藏在 m_show（如 "5k+"=5000）。
                        #   原逻辑「m_numbers<10000 直接当精确值保留」会把 999 当真值写入，
                        #   导致批量任务里大量冷门歌 QQ 收藏量被错写成 999（见 需入库表 4040 首问题）。
                        #   修正取值优先级：
                        #     1) m_show 可解析出真实值 → 一律优先用 m_show（同时覆盖 万级封顶 与 <10000 哨兵）；
                        #     2) m_show 缺失且 m_numbers 非哨兵(≠999) → 才用 m_numbers 精确值（含 0）；
                        #     3) m_numbers=999 且 m_show 缺失 → 视为未提供(missing)，绝不写假值。
                        if real is not None:
                            val = real
                        elif val is not None and val != 999:
                            pass  # m_show 缺失，m_numbers 为可信精确值（含 0），保持
                        else:
                            val = None  # 999 哨兵且无 m_show → 未知，留给 missing 分支处理
                        if val is not None:
                            r['collection_count'] = val
                            _mark_detail(r, 'collection_count', 'ok')
                        else:
                            # 接口成功返回但无该歌曲收藏数 → 平台确实未提供
                            _mark_detail(r, 'collection_count', 'missing')
            except Exception as e:
                print(f"[QQ] fav fetch error batch {i//BATCH_SIZE}: {e}")
                for sid in batch_sids:
                    for r in songid_map[sid]:
                        _mark_detail(r, 'collection_count', 'error')

    if only_fav:
        return

    if skip_comments:
        return

    # 批量获取评论数（分批，每批最多 50 个 topid，避免 URL 过长/接口拒绝）
    if songid_map:
        sids = list(songid_map.keys())
        for i in range(0, len(sids), BATCH_SIZE):
            batch_sids = sids[i:i + BATCH_SIZE]
            try:
                ids = ','.join(batch_sids)
                url = 'https://c.y.qq.com/base/fcgi-bin/fcg_global_comment_h5.fcg'
                params = {
                    'inCharset': 'utf8', 'outCharset': 'utf-8', 'format': 'json',
                    'topid': ids, 'cmd': 4, 'biztype': 1,
                }
                resp = _req_json('GET', url, params=params, headers=headers, timeout=10, tries=2)
                got = set()
                for item in resp.json().get('batch_commenttotal', []):
                    sid = str(item.get('topid'))
                    total = item.get('commenttotal')
                    if sid in songid_map:
                        got.add(sid)
                        if total:
                            for r in songid_map[sid]:
                                r['comment_count'] = _safe_int(total)
                                _mark_detail(r, 'comment_count', 'ok')
                        else:
                            # 接口返回了该歌曲条目但评论数为 0/空 → 平台无评论
                            for r in songid_map[sid]:
                                _mark_detail(r, 'comment_count', 'missing')
                # 批量里未出现在返回中的歌曲 → 平台未提供评论数
                for sid in batch_sids:
                    if sid not in got:
                        for r in songid_map[sid]:
                            _mark_detail(r, 'comment_count', 'missing')
            except Exception as e:
                print(f"[QQ] comment fetch error batch {i//BATCH_SIZE}: {e}")
                for sid in batch_sids:
                    for r in songid_map[sid]:
                        _mark_detail(r, 'comment_count', 'error')

    # 并行获取歌词（词曲作者）
    def fetch_lyric(r):
        songmid = r.get('_songmid')
        if not songmid:
            # 搜索结果未返回 songmid，无法取歌词 → 平台未提供
            _mark_detail(r, 'lyricist', 'missing')
            _mark_detail(r, 'composer', 'missing')
            return
        try:
            lyric_url = 'https://c.y.qq.com/lyric/fcgi-bin/fcg_query_lyric_new.fcg'
            params = {
                'songmid': songmid, 'g_tk': 5381, 'format': 'json',
                'inCharset': 'utf8', 'outCharset': 'utf-8',
            }
            lyric_headers = {
                'User-Agent': COMMON_UA,
                'Referer': 'https://y.qq.com/',
            }
            if cookie_str:
                lyric_headers['Cookie'] = cookie_str
            resp = _req_json('GET', lyric_url, params=params, headers=lyric_headers, timeout=6, tries=2)
            ldata = resp.json()
            lyric_b64 = ldata.get('lyric', '')
            if lyric_b64:
                lyric_text = base64.b64decode(lyric_b64).decode('utf-8', errors='ignore')
                lyricist, composer = _parse_lyricist_composer(lyric_text)
                # v4.28.x 根因修复：顺手解析歌词里的「歌名 - 歌手」头部 / 演唱/和声 标签，
                # 还原真实演唱者，供 picker ②g 认回改名/多账号艺人。
                _lps = _parse_lyric_performer(lyric_text)
                if _lps:
                    r['_lyric_performers'] = _lps
                if lyricist:
                    r['lyricist'] = lyricist
                    _mark_detail(r, 'lyricist', 'ok')
                else:
                    _mark_detail(r, 'lyricist', 'missing')
                if composer:
                    r['composer'] = composer
                    _mark_detail(r, 'composer', 'ok')
                else:
                    _mark_detail(r, 'composer', 'missing')
            else:
                _mark_detail(r, 'lyricist', 'missing')
                _mark_detail(r, 'composer', 'missing')
        except Exception as e:
            print(f"[QQ] lyric fetch error {songmid}: {e}")
            _mark_detail(r, 'lyricist', 'error')
            _mark_detail(r, 'composer', 'error')

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_lyric, r): r for r in results}
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except Exception:
                pass

    # 在听人数（QQ 音乐 App/H5 接口，部分歌曲返回）
    listen_headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 Mobile/15E148',
        'Referer': 'https://i2.y.qq.com/n3/other/pages/playsong/index.html',
        'Content-Type': 'application/json',
    }
    if cookie_str:
        listen_headers['Cookie'] = cookie_str

    def fetch_listen(r):
        songmid = r.get('_songmid')
        if not songmid:
            _mark_detail(r, 'listening_count', 'missing')
            return
        try:
            # 1) 取 songID
            detail_payload = {
                'songinfo': {
                    'method': 'get_song_detail_yqq',
                    'module': 'music.pf_song_detail_svr',
                    'param': {'song_mid': songmid},
                }
            }
            resp = _req_json('POST', 'https://u.y.qq.com/cgi-bin/musicu.fcg?format=json',
                             json=detail_payload, headers=listen_headers, timeout=10, tries=2)
            resp_data = resp.json().get('songinfo', {}).get('data', {})
            track = resp_data.get('track_info', {})

            # 顺手取歌曲详情页 info.company，作为唱片公司兜底（不少单曲/albummid 为空时这里仍有公司）
            song_company = _extract_qq_song_company(resp_data)
            if song_company and not r.get('record_label'):
                r['record_label'] = song_company
                _mark_detail(r, 'record_label', 'ok')

            song_id = track.get('id')
            if not song_id:
                _mark_detail(r, 'listening_count', 'missing')
                return
            # 2) 取在听标签
            listen_payload = {
                'comm': {'ct': 23, 'cv': 1},
                'req_listen': {
                    'module': 'music.sharing.PlayPageSvr',
                    'method': 'GetSongTag',
                    'param': {'songID': song_id},
                }
            }
            resp2 = _req_json('POST', 'https://u.y.qq.com/cgi-bin/musicu.fcg?format=json',
                              json=listen_payload, headers=listen_headers, timeout=10, tries=2)
            tag = resp2.json().get('req_listen', {}).get('data', {}).get('songTag', '')
            count, raw = _parse_qq_listen_text(tag)
            if count:
                # 仅当解析出真实在听人数时才记录。QQ 对多数歌曲不暴露真实在听
                # （GetSongTag 仅返回榜单标签），此时严格留空，绝不用收藏数兜底，避免造假。
                r['listening_count'] = count
                _mark_detail(r, 'listening_count', 'ok')
            else:
                # QQ 未暴露真实在听（仅榜单标签）→ 平台无此数据
                _mark_detail(r, 'listening_count', 'missing')
        except Exception as e:
            print(f"[QQ] listen fetch error {songmid}: {e}")
            _mark_detail(r, 'listening_count', 'error')

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_listen, r): r for r in results}
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except Exception:
                pass


def _fetch_qq_album_labels(results, cookie_str=''):
    """通过专辑详情接口补全唱片公司（record_label）和专辑名（album）。

    关键发现：QQ 音乐的专辑详情接口 fcg_v8_album_info_cp.fcg 不需要 sign 签名，
    直接返回 company 字段（唱片公司名）。按 albummid 去重批量抓取，避免重复请求。

    修复点：
      - 2026-07-30 带上传入的 QQ cookie，解决登录态下仍无 record_label 的问题；
      - 兼容 company/company_new/singer_company/label/publisher 等多个可能字段；
      - 顺手补回缺失的 album 字段（搜索结果可能为空，但专辑详情有）。
    """
    if not results:
        return
    by_album = {}
    for r in results:
        am = r.get('_albummid')
        if am:
            by_album.setdefault(am, []).append(r)
    if not by_album:
        return
    headers = {
        'User-Agent': COMMON_UA,
        'Referer': 'https://y.qq.com/',
    }
    if cookie_str:
        headers['Cookie'] = cookie_str
    def _fetch_one_album(albummid_group, attempt=1):
        albummid, group = albummid_group
        try:
            url = 'https://c.y.qq.com/v8/fcg-bin/fcg_v8_album_info_cp.fcg'
            params = {
                'albummid': albummid, 'g_tk': 5381, 'format': 'json',
                'inCharset': 'utf8', 'outCharset': 'utf-8',
                'platform': 'yqq.json',
            }
            resp = _req_json('GET', url, params=params, headers=headers, timeout=8, tries=2)
            data = resp.json().get('data', {})

            # 顺手补回专辑名（搜索结果可能为空，专辑详情更全）
            album_name = data.get('name') or data.get('albumname') or ''
            if album_name:
                for r in group:
                    if not r.get('album'):
                        r['album'] = album_name

            # 唱片公司：尝试多个可能字段
            company = ''
            for key in ('company', 'company_new', 'singer_company', 'label', 'publisher', 'record_label'):
                val = data.get(key)
                if isinstance(val, str) and val.strip():
                    company = val.strip()
                    break
            if company:
                for r in group:
                    r['record_label'] = company
                    _mark_detail(r, 'record_label', 'ok')
            else:
                for r in group:
                    _mark_detail(r, 'record_label', 'missing')
        except Exception as e:
            print(f"[QQ album] {albummid} Error: {e}")
            for r in group:
                _mark_detail(r, 'record_label', 'error')

    # 并发 5 个：前 100 条通常对应 <=100 个不同专辑，5 并发既能控制总耗时，
    # 又避免对 QQ 服务器造成过大压力导致部分请求被风控/超时。
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        list(executor.map(_fetch_one_album, by_album.items()))


# ═══════════════════════════════════════════════════
#  酷狗音乐
# ═══════════════════════════════════════════════════

# ── 持久化缓存层（"给一次、永不二次操作"的核心）────────
# 酷狗签名网关依赖会过期的账号凭证（userid+token），匿名请求一律 20006。
# 因此把"抓到的数据"与"易过期凭证"解耦：任何一次成功抓取都按歌曲 hash 落盘，
# 之后即便凭证过期、网关全挂，已缓存的歌依旧返回完整唱片公司/发行时间等字段。
import sqlite3 as _sqlite3
# 缓存必须用用户目录(~/.musicfinder)，而非 os.path.abspath(__file__)：
# PyInstaller 冻结后 __file__ 指向 _MEIPASS 临时只读目录，缓存会失效（读不到、写不进）
_KUGOGU_CACHE_DB = os.path.join(COOKIE_DIR, 'kugou_cache.db')
_KUGOGU_CACHE_FIELDS = ['song_name', 'performer', 'album', 'release_date', 'lyricist',
                        'composer', 'collection_count', 'listening_count', 'comment_count',
                        'record_label', 'share_count', '_album_id', '_mixsongid']


def _kugou_cache_conn():
    _db_dir = os.path.dirname(_KUGOGU_CACHE_DB)
    os.makedirs(_db_dir, exist_ok=True)
    conn = _sqlite3.connect(_KUGOGU_CACHE_DB)
    conn.execute('''CREATE TABLE IF NOT EXISTS kugou_song (
        hash TEXT PRIMARY KEY, data TEXT, updated_at REAL)''')
    return conn


def _kugou_cache_get(hash_val):
    if not hash_val:
        return None
    try:
        conn = _kugou_cache_conn()
        row = conn.execute('SELECT data FROM kugou_song WHERE hash=?', (hash_val,)).fetchone()
        conn.close()
        return json.loads(row[0]) if row else None
    except Exception:
        return None


def _kugou_cache_put(r):
    hash_val = r.get('_hash')
    if not hash_val:
        return
    try:
        payload = {k: r.get(k) for k in _KUGOGU_CACHE_FIELDS}
        conn = _kugou_cache_conn()
        conn.execute('INSERT INTO kugou_song(hash, data, updated_at) VALUES(?,?,?) '
                     'ON CONFLICT(hash) DO UPDATE SET data=excluded.data, updated_at=excluded.updated_at',
                     (hash_val, json.dumps(payload, ensure_ascii=False), time.time()))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[Kugou cache] put error: {e}")


def _search_kugou_songsearch(keyword, max_results=20):
    """酷狗兜底搜索通道：songsearch.kugou.com/song_search_v2。

    背景：主通道 mobilecdn /api/v3/search/song 被 IP 软限流时，会返回
    `status=1, errcode=0` 但 `data.info` 为空数组 —— 接口「通」但结果被清空，
    从外面看跟「查无此歌」一模一样，极具欺骗性。
    songsearch 是**独立限流桶**，主通道被清空时它照样出数据，
    字段含 FileHash / MixSongID / Audioid，做「链接 + 收藏量」完全够用。
    """
    out = []
    try:
        resp = requests.get(
            'https://songsearch.kugou.com/song_search_v2',
            params={'keyword': keyword, 'page': 1,
                    'pagesize': max(1, min(int(max_results or 20), 30))},
            headers={'User-Agent': COMMON_UA, 'Referer': 'https://www.kugou.com/'},
            timeout=10, verify=False,
        )
        lists = ((resp.json() or {}).get('data') or {}).get('lists') or []
    except Exception:
        return out

    def _clean(s):
        # songsearch 会在命中词上包 <em>高亮</em>，必须剥掉否则污染歌名比对
        return re.sub(r'</?em>', '', str(s or '')).strip()

    for it in lists:
        if len(out) >= max_results:
            break
        h = it.get('FileHash') or it.get('HQFileHash') or it.get('SQFileHash') or ''
        if not h:
            continue
        out.append({
            'song_name': _clean(it.get('SongName')),
            'performer': _clean(it.get('SingerName')),
            'platform': '酷狗音乐',
            'platform_code': 'kugou',
            'song_url': f'https://www.kugou.com/song/#hash={h}',
            'album': _clean(it.get('AlbumName')),
            'release_date': None, 'lyricist': None, 'composer': None,
            'collection_count': None, 'listening_count': None,
            'heat_count': None, 'comment_count': None, 'share_count': None,
            'record_label': None,
            '_hash': h,
            '_album_id': it.get('AlbumID') or '',
            # 收藏量接口认 album_audio_id ≙ MixSongID；Audioid 是 audio_id，
            # 传错会返回一个极小的数（看起来像正常值，实则完全不对），必须按这个顺序取
            '_mixsongid': it.get('MixSongID') or it.get('ID') or it.get('Audioid') or '',
            '_from_songsearch': True,
        })
    return out


def search_kugou(keyword, max_results=30, light=False):
    """酷狗音乐搜索（支持多页抓取）

    light=True：只返回搜索列表（歌名/歌手/专辑/hash/album_id/mixsongid），
    跳过 6 个详情接口。用于「监控建档」批量匹配 4000 首——建档只需要锁定
    平台 ID，不需要收藏/评论/版权，跳过后单首耗时从 ~6s 降到 ~0.5s。
    """
    results = []
    page_size = 20
    max_pages = (max_results + page_size - 1) // page_size
    cookie_str = get_cookie_string('kugou')
    try:
        for page in range(1, max_pages + 1):
            if len(results) >= max_results:
                break
            url = 'http://mobilecdn.kugou.com/api/v3/search/song'
            params = {
                'keyword': keyword, 'page': page, 'pagesize': page_size,
                'showtype': 10,
            }
            headers = {'User-Agent': COMMON_UA, 'Referer': 'http://www.kugou.com/'}
            if cookie_str:
                headers['Cookie'] = cookie_str
            resp = requests.get(url, params=params, headers=headers, timeout=10, verify=False)
            data = resp.json()
            info_list = data.get('data', {}).get('info', [])
            if not info_list:
                break
            for item in info_list:
                if len(results) >= max_results:
                    break
                results.append({
                    'song_name': item.get('songname', ''),
                    'performer': item.get('singername', ''),
                    'platform': '酷狗音乐',
                    'platform_code': 'kugou',
                    'song_url': f'https://www.kugou.com/song/#hash={item.get("hash", "")}',
                    'album': item.get('album_name', ''),
                    'release_date': None,
                    'lyricist': None,
                    'composer': None,
                    'collection_count': None,
                    'listening_count': None,
                    'heat_count': None,
                    'comment_count': None,
                    'share_count': None,
                    'record_label': None,
                    '_hash': item.get('hash', ''),
                    '_album_id': item.get('album_id', ''),
                    '_mixsongid': item.get('album_audio_id', ''),
                    # v4.25.12：时长维度（秒）。酷狗搜索返回自带 duration。
                    'duration': _safe_int(item.get('duration')),
                })
                # 酷狗搜索结果自带 ownercount（被收藏/拥有的累计用户数）→ 作为「热度」列
                # 实时「在听」人数走 listening_user_rank 接口（见 _fetch_kugou_listening，免登录）
                oc = _safe_int(item.get('ownercount'))
                if oc:
                    results[-1]['heat_count'] = oc

        # 主通道空结果 → 走独立限流桶的兜底通道（被软限流时的救命稻草）
        if not results:
            results = _search_kugou_songsearch(keyword, min(max_results, 20))
            if results:
                print(f'[酷狗] 主通道空，songsearch 兜底命中 {len(results)} 条 · {keyword}')

        if light:
            return results

        _fetch_kugou_details(results[:SEARCH_ENRICH_CAP], cookie_str)
        _fetch_kugou_collection_counts(results[:SEARCH_ENRICH_CAP])
        _fetch_kugou_listening(results[:SEARCH_ENRICH_CAP])
        _fetch_kugou_extra(results[:SEARCH_ENRICH_CAP])
        _fetch_kugou_release_date(results[:SEARCH_ENRICH_CAP])
        _fetch_kugou_copyright(results[:SEARCH_ENRICH_CAP])
        # ── 缓存回填 + 落盘 ──
        # 1) 用缓存补全本次网络未拿到的字段（凭证过期时仍返回历史完整数据）
        for r in results:
            cached = _kugou_cache_get(r.get('_hash'))
            if cached:
                for f in _KUGOGU_CACHE_FIELDS:
                    if (r.get(f) in (None, '')) and cached.get(f) not in (None, ''):
                        r[f] = cached[f]
        # 2) 把本次成功抓取的结果写入缓存（供日后免凭证复用）
        for r in results:
            _kugou_cache_put(r)
        return results
    except Exception as e:
        print(f"[Kugou] Error: {e}")
        return results


def _fetch_kugou_details(results, cookie_str=''):
    """获取酷狗详情：词曲作者（通过歌词API）"""

    def fetch_one(r):
        hash_val = r.get('_hash')
        if not hash_val:
            return
        headers = {'User-Agent': COMMON_UA, 'Referer': 'https://www.kugou.com/'}
        if cookie_str:
            headers['Cookie'] = cookie_str

        # 1. 通过歌词搜索API获取词曲作者
        try:
            keyword = r.get('song_name', '')
            search_url = f'http://lyrics.kugou.com/search?ver=1&man=yes&client=pc&keyword={keyword}&hash={hash_val}'
            resp = requests.get(search_url, headers=headers, timeout=6)
            ldata = resp.json()
            candidates = ldata.get('candidates', [])
            if candidates:
                c = candidates[0]
                lyric_id = c.get('id')
                accesskey = c.get('accesskey')
                if lyric_id and accesskey:
                    dl_url = f'http://lyrics.kugou.com/download?ver=1&client=pc&id={lyric_id}&accesskey={accesskey}&fmt=lrc&charset=utf8'
                    resp2 = requests.get(dl_url, headers=headers, timeout=6)
                    dl_data = resp2.json()
                    content_b64 = dl_data.get('content', '')
                    if content_b64:
                        lyric_text = base64.b64decode(content_b64).decode('utf-8', errors='ignore')
                        lyricist, composer = _parse_lyricist_composer(lyric_text)
                        # v4.28.x 根因修复：解析歌词演唱者，供 picker ②g 认回改名艺人
                        _lps = _parse_lyric_performer(lyric_text)
                        if _lps:
                            r['_lyric_performers'] = _lps
                        if lyricist:
                            r['lyricist'] = lyricist
                        if composer:
                            r['composer'] = composer
                        # 来源 A：歌词内嵌的 OP/SP/出品方/发行公司 → 唱片公司
                        rc = _parse_lyric_copyright(lyric_text)
                        if rc and not r.get('record_label'):
                            r['record_label'] = rc
        except:
            pass

        # 2. 评论数（mcomment.kugou.com 接口，免登录）
        try:
            c_url = 'https://mcomment.kugou.com/index.php'
            c_params = {
                'r': 'commentsv2/getCommentWithLike',
                'code': 'fc4be23b4e972707f36b8a828a93ba8a',
                'extdata': hash_val,
                'p': '1', 'pagesize': '1', 'kugouid': '0',
                'clientver': '1000', 'appid': '1005',
            }
            c_resp = requests.get(c_url, params=c_params, headers=headers, timeout=6)
            c_data = c_resp.json()
            cnt = c_data.get('combine_count') or c_data.get('count')
            if cnt:
                r['comment_count'] = _safe_int(cnt)
        except:
            pass

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_one, r): r for r in results}
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except:
                pass


# 酷狗收藏数（免登录，Android 签名）：gateway.kugou.com/count/v1/audio/mget_collect
KUGOU_ANDROID_SECRET = 'OIlwieks28dk2k092lksi2UIkp'
KUGOU_APPID = 1005
KUGOU_CLIENTVER = 12329


def _kugou_android_signature(params, data=''):
    """酷狗 Android 签名：md5(SECRET + 排序 key=value + SECRET)"""
    s = ''.join(
        f'{k}={json.dumps(params[k]) if isinstance(params[k], (dict, list)) else params[k]}'
        for k in sorted(params)
    )
    return hashlib.md5(f'{KUGOU_ANDROID_SECRET}{s}{data}{KUGOU_ANDROID_SECRET}'.encode()).hexdigest()


_KG_DEV_CACHE = {'ts': 0.0, 'val': None}


def _kugou_device_ids():
    """取酷狗的真实设备指纹 (dfid, mid, cookie)，拿不到就退回匿名。

    ⚠️ 这是 2026-08-10 排查出的关键点：
    酷狗风控升级后，匿名 dfid='-' 打 count/v1/audio/mget_collect 会被判定
    「本次操作需进行验证」(errcode 20028)，整个收藏量通道全线阵亡，随机伪造
    dfid 同样无效（风控看的是「这个指纹有没有真实浏览历史」）。
    但用浏览器 Cookie 里那份真实的 dfid + kg_mid 就能正常返回。
    所以收藏量能不能查，取决于酷狗 Cookie 是否有效 —— Cookie 过期时
    表现为收藏量整列为空，这时去「Cookie 管理」重新粘一份即可。
    """
    now = time.time()
    if _KG_DEV_CACHE['val'] and now - _KG_DEV_CACHE['ts'] < 300:
        return _KG_DEV_CACHE['val']
    dfid, mid, ck = '-', hashlib.md5(b'-').hexdigest(), ''
    try:
        ck = get_cookie_string('kugou') or ''
        if ck:
            jar = dict(re.findall(r'([^=;\s]+)=([^;]*)', ck))
            d = (jar.get('dfid') or jar.get('kg_dfid') or '').strip()
            m = (jar.get('kg_mid') or jar.get('mid') or '').strip()
            if d and d != '-':
                dfid = d
                mid = m or hashlib.md5(d.encode()).hexdigest()
    except Exception:
        pass
    _KG_DEV_CACHE.update(ts=now, val=(dfid, mid, ck))
    return dfid, mid, ck


def _fetch_kugou_collection_counts(results):
    """批量获取酷狗收藏量（favorite count）。

    关键发现（来自开源 KuGouMusicApi 项目 + 实测验证）：
    - 端点：GET https://gateway.kugou.com/count/v1/audio/mget_collect
    - 需 Android 签名（encryptType: android），但免登录（userid=0, dfid='-'）
    - 入参 mixsongids 用的是 album_audio_id（不是 audio_id / hash！用错 ID 会返回极小值）
    - 返回 data.list[].count 即收藏数（如晴天/周杰伦 = 46397374）
    """
    if not results:
        return
    # ⚠️ 一个 ID 可能对应多行（2 万首的曲库表里同名同版本重复很常见）。
    # 早期这里写成 id_map[id] = r，后来的行会把前面的覆盖掉，
    # 结果整批只有最后那行拿到收藏量，其余静默留空 —— 排查起来极难。
    id_map = {}
    for r in results:
        mid_val = r.get('_mixsongid')
        if mid_val:
            id_map.setdefault(str(mid_val), []).append(r)
    if not id_map:
        return
    try:
        dfid, mid, cookie = _kugou_device_ids()
        uuid = hashlib.md5(f'{dfid}{mid}'.encode()).hexdigest()
        params = {
            'dfid': dfid,
            'mid': mid,
            'uuid': uuid,
            'appid': KUGOU_APPID,
            'clientver': KUGOU_CLIENTVER,
            'userid': 0,
            'clienttime': int(time.time()),
            'mixsongids': ','.join(id_map.keys()),
        }
        params['signature'] = _kugou_android_signature(params)
        _hdr = {'User-Agent': COMMON_UA, 'Referer': 'https://www.kugou.com/'}
        if cookie:
            _hdr['Cookie'] = cookie
        resp = requests.get(
            'https://gateway.kugou.com/count/v1/audio/mget_collect',
            params=params,
            headers=_hdr,
            timeout=12,
        )
        data = resp.json().get('data', {}).get('list', [])
        for item in data:
            key = str(item.get('mixsongid'))
            if key in id_map:
                # ⚠️ 不能用 `_safe_int(item.get('count'))`！`_safe_int` 内部 `if val and int(val)>0`
                # 会把 count=0（真 0 收藏量，冷门歌合法值）吞成 None → 整列被误判限流熔断。
                # 这里必须保留 0：显式 int 转换，仅非数字才 None。v4.25.11 续修。
                cv = item.get('count')
                cnt = None
                if cv is not None:
                    try:
                        cnt = int(cv)
                    except (ValueError, TypeError):
                        cnt = None
                # ⚠️ 必须用 `is not None`，不能写 `if cnt`！count=0 是合法结果，
                # 写成 `if cnt` 会把 0 当 falsy 跳过 → collection_count 留 None →
                # 误解「接口被限流」→ 酷狗通道永久熔断。
                if cnt is not None:
                    for r in id_map[key]:
                        r['collection_count'] = cnt
    except Exception as e:
        print(f"[Kugou] collection fetch error: {e}")


# 酷狗在听人数（免登录，V5/URL 签名）：listening/v1/play_page_left/listening_user_rank
# 2026.04 新版签名 salt：y9tjae~n)k)vn[8（来源 992537199/KuGou 开源项目，appid=1000 同款）
KUGOU_LISTEN_SALT = 'y9tjae~n)k)vn[8'
KUGOU_LISTEN_APPID = 1000
KUGOU_LISTEN_CLIENTVER = 20729


def _kugou_listen_signature(params):
    """V5/URL 签名：md5(salt + 排序 key=value + salt)，key=value 之间无 & 分隔"""
    s = ''.join(f'{k}={params[k]}' for k in sorted(params))
    return hashlib.md5(f'{KUGOU_LISTEN_SALT}{s}{KUGOU_LISTEN_SALT}'.encode()).hexdigest()


def _fetch_kugou_listening(results):
    """批量获取酷狗在听人数（免登录，游客参数即可调通）。

    端点：GET https://gateway.kugou.com/listening/v1/play_page_left/listening_user_rank
    入参 mixsongid 用 album_audio_id；返回 data.user_total 即在听人数（实时）。
    签名用 V5/URL 算法（salt='y9tjae~n)k)vn[8'），游客参数（userid=0, 无 token）即可调通。
    来源：用户用 Charles 抓酷狗 Mac 客户端播放页得到真实接口，再反推签名算法落地。
    """
    if not results:
        return
    targets = [r for r in results if r.get('_mixsongid')]
    if not targets:
        return
    dfid = '-'
    mid = hashlib.md5(dfid.encode()).hexdigest()

    def fetch_one(r):
        params = {
            'appid': KUGOU_LISTEN_APPID,
            'clienttime': int(time.time()),
            'clientver': KUGOU_LISTEN_CLIENTVER,
            'dfid': dfid,
            'mid': mid,
            'mixsongid': str(r['_mixsongid']),
            'userid': 0,
        }
        params['signature'] = _kugou_listen_signature(params)
        try:
            resp = requests.get(
                'https://gateway.kugou.com/listening/v1/play_page_left/listening_user_rank',
                params=params,
                headers={'User-Agent': COMMON_UA, 'Referer': 'https://www.kugou.com/'},
                timeout=10,
            )
            data = resp.json().get('data', {})
            cnt = _safe_int(data.get('user_total'))
            if cnt:
                r['listening_count'] = cnt
        except Exception as e:
            print(f"[Kugou] listening fetch error: {e}")

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(fetch_one, r) for r in targets]
        for f in concurrent.futures.as_completed(futures):
            try:
                f.result()
            except:
                pass


# 酷狗「转发/分享数」与「厂牌」——免费 Web 接口均无，需酷狗客户端本体数据。
# 渠道：用 Charles 抓酷狗 Mac 客户端播放页/歌曲详情响应（含分享数 + 厂牌/发行公司）。
# 抓到后把请求 URL 模板（用 {mixsongid} 占位）填到 KUGOU_EXTRA_ENDPOINT，并把样本
# JSON 反馈，下面的通用解析器会自动定位字段（不依赖具体字段名），无需反复改代码。
KUGOU_EXTRA_ENDPOINT = ''  # 例: 'https://gateway.kugou.com/xxx/play_page?mixsongid={mixsongid}&appid=...'

# 酷狗本体「发行公司」——来源：Charles 抓到的客户端 song_related_v2 接口（响应含 copyright 字段）。
# ⚠️ 2026-07-28 证伪：song_related_v2 组合请求的 copyright 是权利元数据、非厂牌，
# 且 gateway 签名用混淆 salt 无法离线复算 → 该接口不作为厂牌来源。酷狗厂牌请看
# _fetch_kugou_details（歌词 OP/SP 解析）。以下常量仅作历史存档，勿再启用。
KUGOU_COMBO_DFID = '2B43IL1lR1yp3IetGp1twQ2h'  # 抓包 URL 里的 dfid（与音频接口不同）
KUGOU_COMBO_BODY = ''  # 已证伪：combo 不含厂牌，保持空 = 该路径禁用


def _kugou_extra_find(obj, keys, limit=8000):
    """递归在任意 JSON 中找第一处命中出现 keys 之一的标量值。"""
    stack = [obj]
    n = 0
    while stack and n < limit:
        node = stack.pop()
        n += 1
        if isinstance(node, dict):
            for k, v in node.items():
                if isinstance(v, (str, int, float)) and any(kk.lower() in str(k).lower() for kk in keys):
                    return v
                if isinstance(v, (dict, list)):
                    stack.append(v)
        elif isinstance(node, list):
            for v in node:
                if isinstance(v, (dict, list)):
                    stack.append(v)
    return None


def _fetch_kugou_extra(results):
    """补酷狗「转发/分享数」与「厂牌」。依赖 Charles 抓到的客户端播放页接口。

    未配置 KUGOU_EXTRA_ENDPOINT 时直接返回（不影响现有已填充列）。
    """
    if not KUGOU_EXTRA_ENDPOINT or not results:
        return
    targets = [r for r in results if r.get('_mixsongid')]
    if not targets:
        return
    SHARE_KEYS = ('share_count', 'sharecount', 'share_num', 'sharenum', 'count_shared',
                  'shared_count', 'forward_count', 'share_total', '转发')
    LABEL_KEYS = ('label', 'company', 'publish_company', 'record_label', 'copyright',
                  'album_company', '厂牌', '发行公司', 'singer_company')
    for r in targets:
        try:
            url = KUGOU_EXTRA_ENDPOINT.replace('{mixsongid}', str(r['_mixsongid']))
            resp = requests.get(url, headers={'User-Agent': COMMON_UA,
                                              'Referer': 'https://www.kugou.com/'},
                                 timeout=10, verify=False)
            data = resp.json()
            sh = _kugou_extra_find(data, SHARE_KEYS)
            if sh is not None:
                r['share_count'] = _safe_int(sh)
            lb = _kugou_extra_find(data, LABEL_KEYS)
            if lb:
                # 专辑页厂牌（来源 B）优先于歌词 OP/SP（来源 A）
                r['record_label'] = str(lb).strip()
        except Exception as e:
            print(f"[Kugou] extra fetch error: {e}")


def _fetch_kugou_copyright(results):
    """【已证伪/废弃】原设想用 Charles 抓到的 song_related_v2 组合请求取酷狗本体厂牌。

    ⚠️ 2026-07-28 实测结论：该 combo 响应里的 `copyright` 字段**不是发行公司**，而是
    权利/授权元数据（rp_type_flac / privilege / price 等）；整套响应 378 个字段中
    无任何 厂牌/发行公司/版权方 类字段（cpy_map 等只是数字代码）。且 gateway 签名用
    混淆 salt 无法离线复算、kmr token 已过期，无法对任意歌曲复用。
    → 酷狗发行公司的**真实且已落地**来源是 `_fetch_kugou_details` 里的歌词解析
      （`_parse_lyric_copyright` 抓「发行公司：/OP：/SP：」，实测「分飞」→ 滚石国际音乐股份有限公司）。
    本函数保留为 no-op：KUGOU_COMBO_BODY 默认空，直接返回，不影响现有流程。
    """
    if not KUGOU_COMBO_BODY or not results:
        return
    targets = [r for r in results if r.get('_hash')]
    if not targets:
        return
    for r in targets:
        try:
            ct = int(time.time())
            q = {'appid': '1000', 'clienttime': str(ct), 'clientver': '20729',
                 'dfid': KUGOU_COMBO_DFID, 'gzip_encoding': '0', 'mid': KUGOU_AUDIO_MID}
            body = KUGOU_COMBO_BODY.replace('{hash}', r['_hash'])
            q['signature'] = _kugou_audio_signature(q, body)
            url = 'https://gateway.kugou.com/pxy/v1/combo/song_related_v2?' + '&'.join(
                f'{k}={q[k]}' for k in ['appid', 'clienttime', 'clientver', 'dfid', 'gzip_encoding', 'mid', 'signature'])
            h = {'User-Agent': 'Mozilla/5.0', 'mid': KUGOU_AUDIO_MID, 'KG-TID': '238',
                 'clientver': '20729', 'KG-RC': '1', 'Content-Type': 'application/json'}
            h['KG-CLIENTTIMEMS'] = str(ct * 1000)
            resp = requests.post(url, headers=h, data=body.encode(), timeout=10,
                                 verify=False, proxies={'http': None, 'https': None})
            data = resp.json()
            # 优先取 copyright（用户已确认该字段存在），退而求其次 label/company 等
            lb = _kugou_extra_find(data, ['copyright', 'company', 'label', '唱片', '厂牌', '发行'])
            if lb:
                # combo 接口返回的是「相关歌曲」列表，copyright 可能在相关条目上；
                # 仅当当前歌曲本身无厂牌时才用（不覆盖歌词 OP/SP 已得的更可靠的本体厂牌）
                if not r.get('record_label'):
                    r['record_label'] = str(lb).strip()
        except Exception as e:
            print(f"[Kugou] copyright fetch error: {e}")


# ── 酷狗「发行时间」补全：kmr/v2/audio（Charles 抓包得到）─────────────────
# 端点 POST https://gateway.kugou.com/kmr/v2/audio
# 请求体指定 hash + entity_id(=album_audio_id)，返回 base.publish_date（歌曲发行日期）。
# 签名含请求体原文：md5(SALT + 排序 k=v + 请求体 + SALT)，salt 同 listening 端点（已反推）。
# 会话参数(dfid/mid/token/userid)来自用户 Charles 抓包；过期后需用 Charles 重新抓一份刷新。
KUGOU_AUDIO_SALT = 'y9tjae~n)k)vn[8'
KUGOU_AUDIO_DFID = '3LjMar1lR21v38RMb90g1kOf'
KUGOU_AUDIO_MID = '164d201bd964acb690bf304d674b672be900164e'
KUGOU_AUDIO_TOKEN = 'b274e2fde6f732ce0eac5c26d5560a4ca7aaf36f245c297eaab1f917d33bddff'
KUGOU_AUDIO_USERID = '2021478038'
KUGOU_AUDIO_HEADERS = {
    'Host': 'gateway.kugou.com', 'KG-RFB': '0',
    'User-Agent': 'IPhone-20729-KGTweLisRecHome#183534257/723988397/625045823/673508761-GetSingerBySong',
    'mid': KUGOU_AUDIO_MID, 'userid': KUGOU_AUDIO_USERID, 'KG-TID': '238',
    'UNI-UserAgent': 'iOS18.7-Phone20729-1009-0-WiFi', 'KG-FAKE': KUGOU_AUDIO_USERID,
    'KG-DEVID': 'dc7d829dd59d985d6822145dab44f1eafdb25136d3132af53dcc0c1db41b4d623dbb5e3b8798fed84977ca52527994d91058f31cfb80cfe1b01f70756d93599d0f67d695d936510eb6ded797faa6a20bd19f3f85f6de213665f513a560e2ebd4819f5b05b1d878c959a2c4ff5c5fba3a005ecea9744ea3e63b023fc0c35bf96b',
    'KG-RC': '1', 'KG-FAKE-SUBTYPE': '0', 'token': KUGOU_AUDIO_TOKEN,
    'x-router': 'openapi.kugou.com', 'KG-FAKE-TYPE': '0,0',
    'Accept-Language': 'zh-Hans-CN;q=1, en-CN;q=0.9, ko-CN;q=0.8, zh-Hant-CN;q=0.7',
    'KG-RF': '771D57926336C017164D201BD964ACB6', 'dfid': KUGOU_AUDIO_DFID,
    'Content-Type': 'application/json', 'Accept': '*/*',
}


def _kugou_audio_signature(params, body):
    """md5(SALT + 排序 k=v + 请求体原文 + SALT)"""
    kv = ''.join(f'{k}={params[k]}' for k in sorted(params))
    return hashlib.md5(f'{KUGOU_AUDIO_SALT}{kv}{body}{KUGOU_AUDIO_SALT}'.encode()).hexdigest()


def _fetch_kugou_release_date(results):
    """用 kmr/v2/audio 补全酷狗「发行时间」(release_date)。"""
    targets = [r for r in results if r.get('_hash') and r.get('_mixsongid')]
    if not targets:
        return
    for r in targets:
        try:
            ct = int(time.time())
            q = {'appid': '1000', 'clienttime': str(ct), 'clientver': '20729',
                 'dfid': KUGOU_AUDIO_DFID, 'mid': KUGOU_AUDIO_MID,
                 'token': KUGOU_AUDIO_TOKEN, 'userid': KUGOU_AUDIO_USERID}
            body = '{"data":[{"hash":"%s","entity_id":"%s"}],"fields":"authors.base,authors.ip"}' % (
                r['_hash'], r['_mixsongid'])
            q['signature'] = _kugou_audio_signature(q, body)
            url = 'https://gateway.kugou.com/kmr/v2/audio?' + '&'.join(
                f'{k}={q[k]}' for k in ['appid', 'clienttime', 'clientver', 'dfid', 'mid', 'token', 'userid', 'signature'])
            h = dict(KUGOU_AUDIO_HEADERS)
            h['KG-CLIENTTIMEMS'] = str(ct * 1000)
            resp = requests.post(url, headers=h, data=body.encode(), timeout=10,
                                 verify=False, proxies={'http': None, 'https': None})
            j = resp.json()
            if j.get('error_code'):
                continue
            d = (j.get('data') or [{}])[0]
            pd = (d.get('base') or {}).get('publish_date')
            if pd:
                r['release_date'] = pd
        except Exception as e:
            print(f"[Kugou] release_date fetch error: {e}")


# ═══════════════════════════════════════════════════
#  酷我音乐
# ═══════════════════════════════════════════════════

def _kuwo_token():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=16))


def search_kuwo(keyword, max_results=30):
    """酷我音乐搜索（支持多页抓取）"""
    results = []
    cookie_str = get_cookie_string('kuwo')

    # 方案一：新版网页搜索 API（结果更准确，无需 CSRF）
    try:
        page_size = 20
        max_pages = (max_results + page_size - 1) // page_size
        for page in range(max_pages):
            if len(results) >= max_results:
                break
            encoded_kw = urllib.parse.quote(keyword)
            url = f'https://www.kuwo.cn/search/searchMusicBykeyWord?vipver=1&client=kt&ft=music&cluster=0&strategy=2012&encoding=utf8&rformat=json&source=kwplayer_pc_5.1&from=pc&pn={page}&rn={page_size}&all={encoded_kw}'
            headers = {
                'User-Agent': COMMON_UA,
                'Referer': f'https://www.kuwo.cn/search/list?key={encoded_kw}',
            }
            if cookie_str:
                headers['Cookie'] = cookie_str
            resp = requests.get(url, headers=headers, timeout=12, verify=False)
            text = resp.text.strip()
            try:
                data = json.loads(text)
            except json.JSONDecodeError:
                try:
                    data = ast.literal_eval(text)
                except Exception as e:
                    print(f"[Kuwo] Parse error: {e}")
                    break

            abslist = data.get('abslist', [])
            if not abslist:
                break
            for item in abslist:
                if len(results) >= max_results:
                    break
                song_name = _decode_escapes(html_mod.unescape(item.get('SONGNAME', '')))
                artist = _decode_escapes(html_mod.unescape(item.get('ARTIST', '')))
                rid = item.get('DC_TARGETID', '')
                # 优先使用新版详情页链接
                song_url = f'https://www.kuwo.cn/play_detail/{rid}' if rid else ''
                results.append({
                    'song_name': song_name,
                    'performer': artist,
                    'platform': '酷我音乐',
                    'platform_code': 'kuwo',
                    'song_url': song_url,
                    'album': _decode_escapes(html_mod.unescape(item.get('ALBUM', ''))),
                    'lyricist': _decode_escapes(html_mod.unescape(item.get('LYRICIST', '') or item.get('WORDS', ''))),
                    'composer': _decode_escapes(html_mod.unescape(item.get('COMPOSER', '') or item.get('MUSIC', ''))),
                    'collection_count': None,
                    'listening_count': None,
                    'comment_count': None,
                    'share_count': None,
                    'record_label': None,
                    '_rid': rid,
                    # v4.25.12：时长维度（秒）。酷我搜索返回 DURATION 字段（秒）。
                    'duration': _safe_int(item.get('DURATION')),
                })
        if results:
            # 用轻量 HTTP 歌词接口补充词曲作者（替代慢速 Playwright，~0.3s）
            # ⚠️ 兜底覆盖「全部」结果（不再限前 5）：
            # 搜索词不限定歌手时，目标歌曲可能排第 6 之后，旧逻辑只抓前 5 条会导致
            # 后排歌曲词曲永远为空（如「可能」搜索下「大头针的可能」）。
            # 仅对有缺失的记录发请求，已齐全的会被 _fetch_kuwo_lyricist_http 内部跳过。
            _need_lyricist = [r for r in results if not (r.get('lyricist') and r.get('composer'))]
            if _need_lyricist:
                _fetch_kuwo_lyricist_http(_need_lyricist)
            _fetch_kuwo_comment_counts(results[:SEARCH_ENRICH_CAP])
            return results
    except Exception as e:
        print(f"[Kuwo] Web API error: {e}")

    # 方案二：旧版 r.s API（兜底）
    token = _kuwo_token()
    try:
        page_size = 20
        max_pages = (max_results + page_size - 1) // page_size
        for page in range(max_pages):
            if len(results) >= max_results:
                break
            url = 'http://search.kuwo.cn/r.s'
            params = {
                'all': keyword, 'ft': 'music', 'itemset': 'web_2013',
                'client': 'kt', 'pn': page, 'rn': page_size,
                'rformat': 'json', 'encoding': 'utf8',
            }
            headers = {'User-Agent': COMMON_UA, 'Referer': 'http://www.kuwo.cn/'}
            if cookie_str:
                headers['Cookie'] = cookie_str
            resp = requests.get(url, params=params, headers=headers, timeout=10, verify=False)
            text = resp.text.strip()
            try:
                data = json.loads(text)
            except json.JSONDecodeError:
                data = ast.literal_eval(text)

            abslist = data.get('abslist', [])
            if not abslist:
                break
            for item in abslist:
                if len(results) >= max_results:
                    break
                song_name = _decode_escapes(html_mod.unescape(item.get('SONGNAME', '')))
                artist = _decode_escapes(html_mod.unescape(item.get('ARTIST', '')))
                rid = item.get('DC_TARGETID', '')
                results.append({
                    'song_name': song_name,
                    'performer': artist,
                    'platform': '酷我音乐',
                    'platform_code': 'kuwo',
                    'song_url': f'http://www.kuwo.cn/play_detail/{rid}',
                    'album': _decode_escapes(html_mod.unescape(item.get('ALBUM', ''))),
                    'lyricist': _decode_escapes(html_mod.unescape(item.get('LYRICIST', '') or item.get('WORDS', ''))),
                    'composer': _decode_escapes(html_mod.unescape(item.get('COMPOSER', '') or item.get('MUSIC', ''))),
                    'collection_count': None,
                    'listening_count': None,
                    'comment_count': None,
                    'share_count': None,
                    'record_label': None,
                    '_rid': rid,
                    # v4.25.12：时长维度（秒）。酷我 r.s 兜底通道 DURATION 字段（秒）。
                    'duration': _safe_int(item.get('DURATION')),
                })
        if results:
            _fetch_kuwo_details(results[:SEARCH_ENRICH_CAP], token, cookie_str)
            _fetch_kuwo_comment_counts(results[:SEARCH_ENRICH_CAP])
            return results
    except Exception as e:
        print(f"[Kuwo] Old API error: {e}")

    return results


def _fetch_kuwo_lyricist_http(results):
    """轻量 HTTP 抓取酷我词曲作者（替代慢速 Playwright，~0.1s/条）

    酷我 musicInfo 登录态仍返回 illegal，拿不到收藏/评论；
    但歌词接口 m.kuwo.cn/newh5/singles/songinfoandlrc 无需 CSRF，
    歌词文本中含「词：xxx / 曲：xxx」，正则即可提取。
    """
    if not results:
        return

    def fetch_one(r):
        rid = r.get('_rid')
        if not rid:
            return
        if r.get('lyricist') and r.get('composer'):
            return
        # 酷我歌词接口偶发返回 null，加重试
        for attempt in range(3):
            try:
                url = f'http://m.kuwo.cn/newh5/singles/songinfoandlrc?musicId={rid}'
                resp = requests.get(url, headers={'User-Agent': COMMON_UA}, timeout=6, verify=False)
                ldata = resp.json()
                lrclist = ldata.get('data', {}).get('lrclist', [])
                if not lrclist:
                    continue
                lyric_text = '\n'.join(item.get('lineLyric', '') for item in lrclist)
                lyricist, composer = _parse_lyricist_composer(lyric_text)
                if lyricist and not r.get('lyricist'):
                    r['lyricist'] = lyricist
                if composer and not r.get('composer'):
                    r['composer'] = composer
                if r.get('lyricist') and r.get('composer'):
                    break
            except Exception:
                pass
            time.sleep(0.4)

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        list(executor.map(fetch_one, results))


def _fetch_kuwo_comment_counts(results):
    """抓取酷我评论数（comment.kuwo.cn/com.s，无需签名，游客即可）

    与开源 music-link-filler 的 getKuwoStats 同款；
    返回 JSON 顶层 total 即为评论总数。
    """
    if not results:
        return

    def fetch_one(r):
        rid = r.get('_rid')
        if not rid or r.get('comment_count'):
            return
        params = {
            'type': 'get_comment', 'f': 'web', 'page': '1', 'rows': '1',
            'digest': '15', 'sid': rid, 'uid': '0', 'prod': 'newWeb', 'httpsStatus': '1',
        }
        try:
            resp = requests.get(
                'https://comment.kuwo.cn/com.s', params=params,
                headers={'User-Agent': COMMON_UA, 'Referer': 'https://www.kuwo.cn/'},
                timeout=10, verify=False,
            )
            j = resp.json()
            cnt = _safe_int(j.get('total'))
            if cnt:
                r['comment_count'] = cnt
        except Exception as e:
            print(f"[Kuwo] comment fetch error: {e}")

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        list(executor.map(fetch_one, results))


def _fetch_kuwo_details(results, token, cookie_str=''):
    """获取酷我详情：评论数、词曲作者（通过旧版 API）。

    注意：酷我收藏量不抓取——其 musicInfo 接口无可靠的收藏数字段，
    favCount 缺失时会回退到 playCnt（播放量），导致数据严重不准。
    为保持「绝对准确」，酷我 collection_count 始终留空（前端显示 —）。
    """

    def fetch_one(r):
        rid = r.get('_rid')
        if not rid:
            return
        headers = {
            'User-Agent': COMMON_UA,
            'Referer': 'https://www.kuwo.cn/',
            'Cookie': f'kw_token={token}',
            'csrf': token,
        }
        if cookie_str:
            headers['Cookie'] = f'{cookie_str}; kw_token={token}'
        try:
            url = f'https://www.kuwo.cn/api/www/music/musicInfo?mid={rid}&httpsStatus=1'
            resp = requests.get(url, headers=headers, timeout=6, verify=False)
            data = resp.json()
            info = data.get('data', {})
            r['comment_count'] = _safe_int(info.get('commentCount'))
            if not r.get('lyricist'):
                r['lyricist'] = info.get('lyricist') or info.get('wordAuthor')
            if not r.get('composer'):
                r['composer'] = info.get('composer') or info.get('musicAuthor')
        except:
            pass

        # 尝试歌词接口获取词曲作者
        if not r.get('lyricist') or not r.get('composer'):
            try:
                lyric_url = f'http://m.kuwo.cn/newh5/singles/songinfoandlrc?musicId={rid}'
                resp = requests.get(lyric_url, headers={'User-Agent': COMMON_UA}, timeout=6)
                ldata = resp.json()
                lrclist = ldata.get('data', {}).get('lrclist', [])
                lyric_text = '\n'.join(item.get('lineLyric', '') for item in lrclist)
                song_info = ldata.get('data', {}).get('songInfo', {})
                if not r.get('lyricist'):
                    r['lyricist'] = song_info.get('lyricist') or song_info.get('wordAuthor')
                if not r.get('composer'):
                    r['composer'] = song_info.get('composer') or song_info.get('musicAuthor')
                if lyric_text and (not r.get('lyricist') or not r.get('composer')):
                    lyricist, composer = _parse_lyricist_composer(lyric_text)
                    if lyricist and not r.get('lyricist'):
                        r['lyricist'] = lyricist
                    if composer and not r.get('composer'):
                        r['composer'] = composer
            except:
                pass

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_one, r): r for r in results}
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except:
                pass


def _fetch_kuwo_details_playwright(results):
    """使用 Playwright 访问酷我详情页提取词曲作者"""
    if not results:
        return
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return

    try:
        with sync_playwright() as p:
            try:
                browser = p.chromium.launch(headless=True, channel='chrome')
            except Exception:
                browser = p.chromium.launch(headless=True)

            context = browser.new_context(
                user_agent=COMMON_UA,
                viewport={'width': 1280, 'height': 800},
                locale='zh-CN',
            )
            page = context.new_page()
            page.route('**/*.{png,jpg,jpeg,gif,svg,webp,woff,woff2,mp4,webm,mp3}', lambda route: route.abort())

            for r in results:
                rid = r.get('_rid')
                if not rid:
                    continue
                try:
                    url = f'https://www.kuwo.cn/play_detail/{rid}'
                    page.goto(url, timeout=25000, wait_until='domcontentloaded')
                    page.wait_for_timeout(3000)
                    html = page.content()

                    # 从歌词区域提取，例如：词：林夕 / 曲：陈小霞

                    # 先尝试带 "词：" 和 "曲：" 的格式
                    m_lyricist = re.search(r'词[：:]\s*([^<\n]+)', html)
                    m_composer = re.search(r'曲[：:]\s*([^<\n]+)', html)

                    if m_lyricist and not r.get('lyricist'):
                        r['lyricist'] = html_mod.unescape(m_lyricist.group(1).strip())
                    if m_composer and not r.get('composer'):
                        r['composer'] = html_mod.unescape(m_composer.group(1).strip())
                except Exception as e:
                    print(f"[Kuwo] Detail page error for {rid}: {e}")

            browser.close()
    except Exception as e:
        print(f"[Kuwo] Playwright detail fetch error: {e}")


# ═══════════════════════════════════════════════════
#  网易云音乐
# ═══════════════════════════════════════════════════

# 网易云限流闸门（2026-08-04）：批量搜索并发 8 时，同一时刻会有 8 个请求打向
# music.163.com，网易云直接返回空 result（不报错、不 429），导致 68/79 首「无结果」。
# 单独调用完全正常（0.2s 返回），确认是并发触发的软限流。这里把打向网易云的
# 并发压到 2，配合空结果重试，实测命中率从 6/79 恢复到正常水平。
_NETEASE_SEM = threading.Semaphore(2)
# 网易云「操作频繁」(code=405) 限流时间戳；命中后立即停手并在批量结果里标「限流」。
_NETEASE_BLOCKED_AT = 0.0


def _netease_availability(song):
    """判断网易云单曲的可用状态。

    /api/search/get 返回的 song 字段：
      status: 0 正常；< 0（-1/-200）已下架不可用
      fee:    0 免费 / 1 VIP 专享 / 4 数字专辑 / 8 普通付费（非会员可听低音质）
    """
    try:
        status = int(song.get('status', 0) or 0)
    except (TypeError, ValueError):
        status = 0
    if status < 0:
        return '已下架'
    try:
        fee = int(song.get('fee', 0) or 0)
    except (TypeError, ValueError):
        fee = 0
    if fee in (1, 4):
        return '在架(VIP)'
    return '在架'


def search_netease(keyword, max_results=30, light=False):
    """网易云音乐搜索（支持多页抓取）

    light=True：跳过详情抓取，只返回列表 + song_id/album_id（监控建档用）。
    """
    results = []
    page_size = 30
    max_pages = (max_results + page_size - 1) // page_size
    cookie_str = get_cookie_string('netease')
    try:
        for page in range(max_pages):
            if len(results) >= max_results:
                break
            offset = page * page_size
            url = 'https://music.163.com/api/search/get'
            post_data = {'s': keyword, 'type': 1, 'limit': page_size, 'offset': offset}
            headers = {'User-Agent': COMMON_UA, 'Referer': 'https://music.163.com/'}
            if cookie_str:
                headers['Cookie'] = cookie_str

            # 限流 + 空结果重试：并发高时网易云会静默返回空 result（HTTP 200、无报错）
            songs = []
            cur_headers = headers
            for attempt in range(3):
                with _NETEASE_SEM:
                    resp = requests.post(url, data=post_data, headers=cur_headers, timeout=10)
                    try:
                        data = resp.json()
                    except Exception:
                        data = {}
                # 网易云「操作频繁」(code=405)：实测是**账号级**限流（只有带 Cookie 才触发，
                # 同一秒匿名请求照样正常返回）。搜索/红心数接口本身免登录，所以先摘掉 Cookie
                # 匿名重试一次；匿名也被打回，才算真的 IP 级封禁，这时才熔断。
                if data.get('code') == 405 or '操作频繁' in str(data.get('message', '')):
                    if cur_headers.get('Cookie'):
                        cur_headers = {k: v for k, v in headers.items() if k != 'Cookie'}
                        cookie_str = ''  # 后续详情抓取一并转匿名，避免再撞 405
                        time.sleep(0.2)
                        continue
                    global _NETEASE_BLOCKED_AT
                    _NETEASE_BLOCKED_AT = time.time()
                    songs = []
                    break
                songs = (data.get('result') or {}).get('songs') or []
                if songs or page > 0:
                    break  # 拿到了，或翻页到尽头（第 2 页起为空属正常）
                time.sleep(0.6 * (attempt + 1))  # 退避后再试，避开限流窗口
            if not songs:
                break
            for song in songs:
                if len(results) >= max_results:
                    break
                artists = '/'.join(a.get('name', '') for a in song.get('artists', []))
                song_id = song.get('id', 0)
                album_info = song.get('album', {})
                album_id = album_info.get('id', 0)
                results.append({
                    'song_name': song.get('name', ''),
                    'performer': artists,
                    'platform': '网易云音乐',
                    'platform_code': 'netease',
                    # status < 0（-1/-200）= 已下架；fee: 0免费 1VIP 4数字专辑 8普通付费
                    'availability': _netease_availability(song),
                    'song_url': f'https://music.163.com/#/song?id={song_id}',
                    'album': album_info.get('name', ''),
                    'release_date': _ts_to_date(album_info.get('publishTime'), is_ms=True),  # 搜索结果 album.publishTime（毫秒时间戳）
                    'lyricist': None,
                    'composer': None,
                    'collection_count': None,
                    'listening_count': None,
                    'comment_count': None,
                    'share_count': None,
                    'record_label': None,
                    '_song_id': song_id,
                    '_album_id': album_id,
                    # v4.25.12：时长维度（秒）。网易云搜索返回 duration 为毫秒。
                    'duration': (_safe_int(song.get('duration')) or 0) // 1000 or None,
                })

        if not light:
            _fetch_netease_details(results[:SEARCH_ENRICH_CAP], cookie_str)
        return results
    except Exception as e:
        print(f"[NetEase] Error: {e}")
        return results


def _fetch_netease_details(results, cookie_str=''):
    """获取网易云详情：收藏量(红心数/eapi)、评论数、唱片公司、词曲作者"""

    # 批量获取收藏量（eapi 加密接口，免登录；优先纯 Python，回退 Node）
    red_map = {}
    ids = [r.get('_song_id') for r in results if r.get('_song_id')]
    if ids:
        red_map = _netease_eapi_red_count(ids)

    def fetch_one(r):
        song_id = r.get('_song_id')
        album_id = r.get('_album_id')
        headers = {'User-Agent': COMMON_UA, 'Referer': 'https://music.163.com/'}
        if cookie_str:
            headers['Cookie'] = cookie_str

        # 0. 收藏量（来自批量 eapi 结果）
        # ⚠️ 同 QQ/酷狗 _safe_int 吞 0 隐患：原 `if red and red.get('count')` 把 count=0
        #（冷门歌合法 0 红心）当 falsy 跳过，且 `_safe_int(0)` 也会吞成 None → 误标「未查到」。
        # 改用 _safe_int_keep_zero 显式保留 0。v4.25.13 补修。
        rid = str(song_id) if song_id else ''
        red = red_map.get(rid)
        cc = _safe_int_keep_zero(red.get('count') if red else None)
        if cc is not None:
            r['collection_count'] = cc
            r['collection_text'] = red.get('countDesc')

        # 1. 评论数
        if song_id:
            try:
                url = f'https://music.163.com/api/v1/resource/comments/R_SO_4_{song_id}?limit=1&offset=0'
                resp = requests.get(url, headers=headers, timeout=6)
                data = resp.json()
                total = data.get('total', 0)
                r['comment_count'] = _safe_int(total)
            except:
                pass

        # 2. 唱片公司
        if album_id:
            try:
                url = f'https://music.163.com/api/album/{album_id}'
                resp = requests.get(url, headers=headers, timeout=6)
                data = resp.json()
                company = data.get('album', {}).get('company')
                if company:
                    r['record_label'] = company
            except:
                pass

        # 3. 词曲作者（通过歌词API）
        if song_id and (not r.get('lyricist') or not r.get('composer')):
            try:
                url = f'https://music.163.com/api/song/lyric?id={song_id}&lv=1&tv=-1'
                resp = requests.get(url, headers=headers, timeout=6)
                data = resp.json()
                lyric_text = data.get('lrc', {}).get('lyric', '')
                if lyric_text:
                    lyricist, composer = _parse_lyricist_composer(lyric_text)
                    # v4.28.x 根因修复：解析歌词演唱者，供 picker ②g 认回改名艺人
                    _lps = _parse_lyric_performer(lyric_text)
                    if _lps:
                        r['_lyric_performers'] = _lps
                    if lyricist:
                        r['lyricist'] = lyricist
                    if composer:
                        r['composer'] = composer
            except:
                pass

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_one, r): r for r in results}
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except:
                pass


# ═══════════════════════════════════════════════════
#  汽水音乐（Playwright 浏览器自动化）
# ═══════════════════════════════════════════════════

def search_qishui(keyword, max_results=30):
    """汽水音乐搜索。

    主路径：直接调用 Mac 客户端接口 api.qishui.com/luna/pc/search/all
    （免登录、免签名头，由 Charles 抓包获得，实测稳定）。
    失败回退：保留的「浏览器登录」实时会话 / Playwright（需 Cookie）。
    """
    # 主路径：App 接口
    try:
        results = _search_qishui_api(keyword, max_results)
        if results:
            PLATFORM_WARNINGS.pop('qishui', None)
            return results
        print('[Qishui] App 接口返回空，尝试回退通道')
    except Exception as e:
        print(f'[Qishui] App 接口异常，回退: {e}')

    # 回退一：实时已登录会话（需用户曾「浏览器登录」并保持窗口）
    try:
        live = _search_qishui_live(keyword, max_results)
        if live:
            return live
    except Exception as e:
        print(f'[Qishui] 实时会话搜索异常: {e}')

    # 回退二：Playwright 无头（需 Cookie）
    cookie_str = get_cookie_string('qishui')
    if not cookie_str:
        if not PLATFORM_WARNINGS.get('qishui'):
            PLATFORM_WARNINGS['qishui'] = (
                '汽水音乐 App 接口未返回结果且未配置 Cookie。'
                'App 接口本应免登录可用，若持续为空请检查网络/代理设置后重试；'
                '或到「Cookie设置」用「浏览器登录」抖音并保持窗口以启用回退通道。'
            )
        return []
    try:
        results = _search_qishui_playwright(keyword, max_results, cookie_str)
        if results:
            return results
    except Exception as e:
        print(f'[Qishui] Playwright error: {e}')

    if not PLATFORM_WARNINGS.get('qishui'):
        PLATFORM_WARNINGS['qishui'] = (
            '汽水音乐搜索未返回结果。App 接口与回退通道均无数据，'
            '请检查网络后重试，或在「Cookie设置」用「浏览器登录」抖音并保持窗口。'
        )
    return []


def _parse_qishui_cookie(cookie_str):
    """解析汽水音乐/抖音 Cookie，并补充 domain/path"""
    cookies = []
    for pair in cookie_str.split(';'):
        pair = pair.strip()
        if '=' in pair:
            name, value = pair.split('=', 1)
            name = name.strip()
            value = value.strip()
            if not name:
                continue
            cookies.append({
                'name': name,
                'value': value,
                'domain': '.douyin.com',
                'path': '/',
            })
    return cookies


def _qishui_music_stats(music):
    """从抖音 music 对象提取 收藏 / 评论 / 分享。返回 (collection_count, comment_count, share_count)。"""
    if not isinstance(music, dict):
        return None, None, None
    stat = music.get('stat') or {}
    raw = (music.get('user_count') or music.get('use_count') or music.get('collect_count')
           or stat.get('collect_count') or stat.get('use_count') or stat.get('user_count'))
    collection = _safe_int(raw) if raw else None
    comment = _safe_int(stat.get('comment_count') or music.get('comment_count'))
    share = _safe_int(stat.get('share_count') or music.get('share_count'))
    return collection, comment, share


def _extract_qishui_song_maker(music):
    """从汽水/抖音 music 对象提取 (lyricist, composer) 字符串。

    汽水 track 通过 song_maker_team 结构化返回词曲作者（composers/lyricists 数组），
    比解析歌词更可靠（之前误判汽水无词曲作者即因漏看此字段）。
    抖音 web 回退路径可能用顶层 composers/lyricists，一并兼容。
    """
    if not isinstance(music, dict):
        return None, None
    smt = music.get('song_maker_team') or {}
    comps = smt.get('composers') or music.get('composers') or []
    lyrs = smt.get('lyricists') or music.get('lyricists') or []
    composer = ' / '.join(c.get('name') for c in comps
                           if isinstance(c, dict) and c.get('name')) or None
    lyricist = ' / '.join(l.get('name') for l in lyrs
                          if isinstance(l, dict) and l.get('name')) or None
    return lyricist, composer


def _is_video_original_sound(music):
    """启发式判断一个 music 对象是否为抖音视频原声（非汽水正式歌曲）。"""
    if not isinstance(music, dict):
        return True
    title = (music.get('title') or '').strip()
    duration = music.get('duration') or 0
    owner = (music.get('owner_nickname') or '').strip()
    is_original = music.get('is_original')

    # 标题带「原声」「创作的原声」直接判为视频原声
    if '原声' in title or '创作的原声' in title:
        return True
    # 模板音乐/系统音效
    if title in ('模板音乐', '模板创作人') or '模板' in title:
        return True
    # 短片段（<60s）且带创作者昵称的，大概率是用户上传原声
    if duration and int(duration) < 60 and owner:
        return True
    # 明确标记非原创作曲且创作者就是昵称
    if is_original is False and owner:
        return True
    return False


def _make_qishui_result(music, seen):
    """从 music/qishui_music 对象构造统一结果字典；若已存在或无效返回 None。"""
    if not isinstance(music, dict):
        return None
    title = (music.get('title') or '').strip()
    if not title or title in seen:
        return None
    seen.add(title)
    author = (music.get('author') or music.get('author_name') or music.get('artist') or '')
    if isinstance(author, dict):
        author = author.get('nickname') or author.get('name') or ''
    mid = music.get('id_str') or music.get('mid') or music.get('id') or ''
    collection_count, comment_count, share_count = _qishui_music_stats(music)
    lyricist, composer = _extract_qishui_song_maker(music)
    return {
        'song_name': title,
        'performer': author,
        'platform': '汽水音乐',
        'platform_code': 'qishui',
        'song_url': f'https://music.douyin.com/qishui/share/track?track_id={mid}' if mid else '',
        'album': (music.get('album') or '').strip() if isinstance(music.get('album'), str) else '',
        'lyricist': lyricist,
        'composer': composer,
        'collection_count': collection_count,
        'listening_count': None,
        'comment_count': comment_count,
        'share_count': share_count,
        'record_label': None,
    }


def _parse_qishui_music_response(data, max_results, music_only=False):
    """从抖音搜索响应解析汽水音乐正式歌曲。

    优先级：
      1. item['qishui_music_list'] —— 汽水平台正式歌曲（最准）
      2. item['music_info_list']  —— 音乐信息列表
      3. item['music'] / aweme_info.music —— 仅当不像视频原声时才保留
    music_only=True 时额外跳过明确含 aweme_info/video/aweme 字段的视频项。
    """
    results = []
    seen = set()
    candidates = []
    if isinstance(data.get('data'), list):
        candidates.extend(data['data'])
    if isinstance(data.get('aweme_list'), list):
        candidates.extend(data['aweme_list'])
    if isinstance(data.get('music_list'), list):
        candidates.extend(data['music_list'])

    for item in candidates:
        if not isinstance(item, dict):
            continue

        # 明确视频项（带 aweme_info/video/aweme）且 music_only 模式下直接跳过
        is_video_item = bool(item.get('aweme_info') or item.get('video') or item.get('aweme'))
        if music_only and is_video_item:
            continue

        # 1) 优先取汽水音乐列表（正式歌曲）
        qishui_list = item.get('qishui_music_list') if isinstance(item.get('qishui_music_list'), list) else []
        for qm in qishui_list:
            r = _make_qishui_result(qm, seen)
            if r:
                results.append(r)
                if len(results) >= max_results:
                    return results

        # 2) 音乐信息列表
        info_list = item.get('music_info_list') if isinstance(item.get('music_info_list'), list) else []
        for im in info_list:
            r = _make_qishui_result(im, seen)
            if r:
                results.append(r)
                if len(results) >= max_results:
                    return results

        # 3) 顶层 music / aweme_info.music，过滤视频原声
        for src_key in ('music', 'aweme_info'):
            src = item.get(src_key) if isinstance(item.get(src_key), dict) else {}
            if src_key == 'aweme_info' and not src:
                continue
            music = src.get('music') if isinstance(src.get('music'), dict) else src
            if not isinstance(music, dict) or not music.get('title'):
                continue
            if _is_video_original_sound(music):
                continue
            r = _make_qishui_result(music, seen)
            if r:
                results.append(r)
                if len(results) >= max_results:
                    return results

    return results


def _qishui_is_blocked(d):
    """判断抖音搜索响应是否被风控/要求登录。"""
    if not isinstance(d, dict):
        return False
    nil = d.get('search_nil_info') or {}
    if isinstance(nil, dict) and nil.get('search_nil_type') == 'verify_check':
        return True
    sm = d.get('status_msg') or ''
    if '登录' in sm or 'login' in sm.lower():
        return True
    if d.get('status_code') in (2483, 8, 10000):  # 2483=请先登录；8/10000=风控
        return True
    return False


# ── 汽水音乐 Mac 客户端搜索接口（Charles 抓包获得）─────────────────
# 实测：无需登录 Cookie、无需 X-Bogus/签名头，裸 GET 即返回 200 与正式歌曲列表。
# 搜索结果位于 result_groups 的 'tracks' 分组；接口为 SPA，字段稳定。
QISHUI_SEARCH_BASE = (
    "https://api.qishui.com/luna/pc/search/all"
    "?aid=386088&app_name=luna_pc&region=cn&geo_region=cn&os_region=cn&sim_region="
    "&device_id=139010146316592&cdid=&iid=3665111206344379&version_name=3.5.1"
    "&version_code=30050100&channel=official&build_mode=master&network_carrier="
    "&ac=wifi&tz_name=Asia%2FShanghai&resolution=&device_platform=mac&device_type=MacOS"
    "&os_version=Darwin+Kernel+Version+24.6.0%3A+Wed+Nov++5+21%3A33%3A59+PST+2025%3B+root%3Axnu-11417.140.69.705.2%7E1%2FRELEASE_ARM64_T8112"
    "&fp=139010146316592"
)


def _qishui_get(url, timeout=15):
    """请求 api.qishui.com，兼容 Charles 代理（自签证书）与直连两种环境。"""
    last = None
    # 1) 直连 + 验证证书
    try:
        return requests.get(url, timeout=timeout, verify=True,
                            proxies={'http': None, 'https': None})
    except Exception as e:
        last = e
    # 2) 直连 + 跳过验证（Charles 设为某代理但系统代理未生效时的兜底）
    try:
        return requests.get(url, timeout=timeout, verify=False,
                            proxies={'http': None, 'https': None})
    except Exception as e:
        last = e
    # 3) 走系统代理 + 跳过验证（Charles 设为 macOS 系统代理时）
    try:
        return requests.get(url, timeout=timeout, verify=False)
    except Exception as e:
        last = e
    raise last


def _qishui_album_pclines(album_id):
    """根据汽水 album_id 调专辑详情接口，取 pclines（厂牌/版权信息）。"""
    if not album_id:
        return None
    try:
        m = re.match(r'(https://[^/]+)(/[^?]+)?(\?.*)', QISHUI_SEARCH_BASE)
        if not m:
            return None
        url = m.group(1) + f'/luna/pc/albums/{album_id}' + m.group(3) + '&ignore_tracks=false'
        r = _qishui_get(url, timeout=12)
        if r.status_code != 200:
            return None
        album = r.json().get('album_info') or {}
        pclines = album.get('pclines') or []
        if isinstance(pclines, list):
            for line in pclines:
                if line and isinstance(line, str):
                    # 去掉 ℗ / © 等版权前缀及多余空格（可能同时出现 © ℗）
                    label = re.sub(r'^[©℗\s]+', '', line.strip()).strip()
                    return label if label else None
        return None
    except Exception as e:
        print(f"[Qishui] album detail error: {e}")
        return None


def _search_qishui_api(keyword, max_results=30, cookie_str=''):
    """汽水音乐搜索：直接调用 Mac 客户端接口 api.qishui.com/luna/pc/search/all。

    该接口免登录、免签名头，实测稳定可用（由 Charles 抓包获得）。
    对每首结果补充专辑详情 pclines 作为厂牌/版权信息。
    返回 list（可能为空）；异常时返回 []。
    """
    import uuid
    try:
        sid = str(uuid.uuid4())
        url = (QISHUI_SEARCH_BASE
               + "&q=" + urllib.parse.quote(keyword)
               + "&cursor=0&search_id=" + sid
               + "&search_method=input&debug_params=&from_search_id=&search_scene=")
        resp = _qishui_get(url, timeout=15)
        if resp.status_code != 200:
            print(f"[Qishui] app API HTTP {resp.status_code}")
            return []
        data = resp.json()
        groups = {g.get('id'): g for g in data.get('result_groups', [])}
        tracks = (groups.get('tracks') or {}).get('data', [])
        results = []
        seen = set()
        for it in tracks:
            tr = (it.get('entity') or {}).get('track') or {}
            name = tr.get('name') or (tr.get('album') or {}).get('name') or ''
            if not name or name in seen:
                continue
            seen.add(name)
            artists = [a.get('name') for a in tr.get('artists', []) if a.get('name')]
            performer = ' / '.join(artists)
            album_obj = tr.get('album') or {}
            album = album_obj.get('name') or ''
            album_id = album_obj.get('id') or ''
            stats = tr.get('stats') or {}
            tid = tr.get('id') or ''
            lyricist, composer = _extract_qishui_song_maker(tr)
            record_label = _qishui_album_pclines(album_id)
            cover = ''
            uc = album_obj.get('url_cover') or {}
            if uc.get('uri'):
                cover = "https://p3-luna.douyinpic.com/img/" + uc['uri']
            results.append({
                'song_name': name,
                'performer': performer,
                'platform': '汽水音乐',
                'platform_code': 'qishui',
                'song_url': ('https://music.douyin.com/qishui/share/track?track_id=' + tid) if tid
                            else ('https://www.douyin.com/search/' + urllib.parse.quote(keyword)),
                'album': album,
                'lyricist': lyricist,
                'composer': composer,
                # ⚠️ 不能用 `_safe_int(stats.get('count_collected'))`：同 QQ/酷狗/网易云吞 0 隐患，
                # 0 收藏量（冷门歌合法值）会被吞成 None → 误标「未查到」。v4.25.13 补修。
                'collection_count': _safe_int_keep_zero(stats.get('count_collected')),
                'listening_count': _safe_int(stats.get('count_played')),
                'comment_count': _safe_int(stats.get('count_comment')),
                'share_count': _safe_int(stats.get('count_shared')),
                'record_label': record_label,
                'cover': cover,
                # v4.25.12：时长维度（秒）。汽水 track 自带 duration（毫秒）。
                'duration': (_safe_int(tr.get('duration')) or 0) // 1000 or None,
            })
            if len(results) >= max_results:
                break
        return results
    except Exception as e:
        print(f"[Qishui] app API error: {e}")
        return []


def _filter_qishui_playwright_results(items):
    """过滤 Playwright 提取结果中的视频原声/模板音乐。"""
    out = []
    seen = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        name = (item.get('song_name') or '').strip()
        if not name or name in seen:
            continue
        if '原声' in name or '创作的原声' in name:
            continue
        if name in ('模板音乐', '模板创作人') or '模板' in name:
            continue
        seen.add(name)
        out.append(item)
    return out


def _search_qishui_playwright(keyword, max_results=30, cookie_str=''):
    """使用 Playwright（系统 Chrome）搜索汽水/抖音音乐。

    关键修复：抖音 web 搜索接口强制要求 X-Bogus 签名，纯 requests 无签名必被拒
    （返回 2483「请先登录」，即登录态未生效）。浏览器自身发出的请求由抖音前端自动
    签名，因此本函数通过 page.on('response') 拦截抖音前端「带签名」的搜索接口响应，
    直接解析 JSON，彻底绕开手写签名的脆弱性。DOM 提取仅作最后兜底。
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("[Qishui] Playwright not installed")
        return []

    results = []
    captured = {}

    def _on_response(response):
        url = response.url
        if 'aweme/v1/web/search/item/' in url or 'general/search/single' in url:
            try:
                body = response.body()
                if body:
                    captured.setdefault('payload', body)
            except Exception:
                pass

    with sync_playwright() as p:
        # 优先使用系统 Chrome，避免下载 Chromium
        try:
            browser = p.chromium.launch(headless=True, channel='chrome')
        except Exception:
            try:
                browser = p.chromium.launch(headless=True)
            except Exception as e:
                print(f"[Qishui] Browser launch failed: {e}")
                return []

        context = browser.new_context(
            user_agent=COMMON_UA,
            viewport={'width': 1280, 'height': 900},
            locale='zh-CN',
        )

        # 注入 Cookie
        if cookie_str:
            cookies = _parse_qishui_cookie(cookie_str)
            if cookies:
                try:
                    context.add_cookies(cookies)
                except Exception as e:
                    print(f"[Qishui] Cookie injection warning: {e}")

        page = context.new_page()
        page.on('response', _on_response)
        # 拦截不必要的资源加速加载
        page.route('**/*.{png,jpg,jpeg,gif,svg,webp,woff,woff2,mp4,webm,mp3}', lambda route: route.abort())

        # 进入抖音音乐搜索页，触发前端签名请求
        try:
            page.goto(f'https://www.douyin.com/search/{urllib.parse.quote(keyword)}?type=music',
                      timeout=20000, wait_until='domcontentloaded')
        except Exception as e:
            print(f'[Qishui] navigate error: {e}')

        # 等待并解析被拦截的「带签名」响应
        waited = 0
        while waited < 12000 and 'payload' not in captured:
            try:
                page.wait_for_timeout(300)
            except Exception:
                pass
            waited += 300

        data = None
        if 'payload' in captured:
            try:
                import json as _json
                data = _json.loads(captured['payload'])
            except Exception as e:
                print(f'[Qishui] payload parse error: {e}')

        if data is not None:
            if _qishui_is_blocked(data):
                PLATFORM_WARNINGS['qishui'] = (
                    '汽水音乐（抖音）在无头浏览器中登录态未生效：注入的 Cookie 被抖音判为未登录并弹出登录墙，'
                    '因此搜索无法返回结果。这是抖音反爬/反自动化的技术限制，并非你的账号被风控。'
                    '请改用「浏览器登录」并保持登录窗口不关闭，即可复用真实已登录会话直接搜索。'
                )
                browser.close()
                return results
            parsed = _parse_qishui_music_response(data, max_results, music_only=True)
            if parsed:
                results = parsed

        # 兜底：未捕获到签名响应时，用 DOM/SSR 提取（复用统一兜底函数）
        if not results:
            print('[Qishui] 未捕获签名响应，回退 DOM/SSR 提取')
            results = _run_qishui_dom_fallback(page, keyword, max_results)

        # 尽力而为：抓取详情页统计（使用/收藏/评论）
        if results:
            try:
                _scrape_qishui_detail_stats(context, results)
            except Exception as e:
                print(f"[Qishui] detail scrape skip: {e}")

        browser.close()

    return results


# 抖音音乐搜索页 DOM/SSR 兜底提取脚本（过滤视频原声/模板音乐）
_QISHUI_DOM_JS = r"""() => {
    const results = [];
    const seen = new Set();
    function isOriginalSound(t) {
        return !t || t.includes('原声') || t.includes('创作的原声') || t === '模板音乐' || t.includes('模板');
    }
    try {
        const state = window.__INITIAL_STATE__ || window._SSR_INITIAL_DATA_ || window.__SSR_RENDER_DATA__;
        if (state) {
            const sd = state.search || state.Search || {};
            const vl = sd.data || sd.videoList || sd.video_list || [];
            for (const v of vl) {
                const m = v.music || v.aweme_info?.music || v.video?.music;
                if (m && m.title && !isOriginalSound(m.title) && !seen.has(m.title)) {
                    seen.add(m.title);
                    results.push({
                        song_name: m.title,
                        performer: m.author || '',
                        song_url: m.mid ? 'https://music.douyin.com/qishui/share/track?track_id=' + m.mid : '',
                        collection_count: m.user_count || m.use_count || m.collect_count || (m.stat && (m.stat.collect_count || m.stat.use_count)) || null,
                        comment_count: (m.stat && m.stat.comment_count) || null,
                    });
                }
            }
        }
    } catch(e) {}
    if (results.length === 0) {
        const items = document.querySelectorAll('[data-e2e="search-video-list"] > ul > li, li[data-e2e="search-result-item"]');
        items.forEach((item) => {
            try {
                const link = item.querySelector('a[href*="/music/"]');
                if (link) {
                    const t = link.innerText.trim();
                    if (t && !isOriginalSound(t) && !seen.has(t)) {
                        seen.add(t);
                        results.push({ song_name: t, performer: '', song_url: link.href, collection_count: null });
                    }
                }
            } catch(e) {}
        });
    }
    return results;
}"""


def _run_qishui_dom_fallback(page, keyword, max_results):
    """DOM/SSR 兜底：在已加载的抖音音乐搜索页上提取歌曲（过滤视频原声）。"""
    try:
        js_results = page.evaluate(_QISHUI_DOM_JS)
    except Exception as e:
        print(f'[Qishui] DOM fallback error: {e}')
        return []
    js_results = _filter_qishui_playwright_results(js_results or [])
    seen = set()
    out = []
    for item in js_results[:max_results]:
        name = item.get('song_name', '')
        if name and name not in seen:
            seen.add(name)
            out.append({
                'song_name': name,
                'performer': item.get('performer', ''),
                'platform': '汽水音乐',
                'platform_code': 'qishui',
                'song_url': item.get('song_url', '') or f'https://www.douyin.com/search/{keyword}',
                'album': '',
                'lyricist': None,
                'composer': None,
                # ⚠️ 同款吞 0 隐患：DOM 兜底若取到 0 收藏量也会被吞成 None。v4.25.13 补修。
                'collection_count': _safe_int_keep_zero(item.get('collection_count')),
                'listening_count': None,
                'comment_count': None,
                'share_count': None,
                'record_label': None,
            })
    return out


def _search_qishui_live(keyword, max_results=30):
    """复用「浏览器登录」保留的实时已登录会话搜索汽水音乐。

    关键：该会话是用户真实登录、指纹一致、非无头的浏览器，抖音识别为已登录，
    不会出现登录墙，搜索接口（前端自动签名）能正常返回数据。
    返回：list（可能为空）= 已用实时会话；None = 无可用实时会话，调用方应回退。
    """
    # 1) 取出会话引用（短暂持锁，避免与登录/关闭竞争）
    with LIVE_SESSIONS_LOCK:
        sess = LIVE_SESSIONS.get('qishui')
        if not sess:
            return None
        browser = sess.get('browser')
        context = sess.get('context')
        page = sess.get('page')
        if browser is None or context is None or page is None:
            _close_live_session('qishui')
            return None
        # 校验会话还活着（窗口被用户关闭会抛异常）
        try:
            _ = page.url
        except Exception:
            _close_live_session('qishui')
            return None

    # 2) 用实时会话搜索（串行化，避免并发操作同一 page）
    with QISHUI_LIVE_LOCK:
        captured = {}
        def _on_response(response):
            url = response.url
            if 'aweme/v1/web/search/item/' in url or 'general/search/single' in url:
                try:
                    body = response.body()
                    if body:
                        captured.setdefault('payload', body)
                except Exception:
                    pass

        try:
            page.on('response', _on_response)
            page.goto(f'https://www.douyin.com/search/{urllib.parse.quote(keyword)}?type=music',
                      timeout=30000, wait_until='domcontentloaded')
            waited = 0
            while waited < 15000 and 'payload' not in captured:
                try:
                    page.wait_for_timeout(300)
                except Exception:
                    pass
                waited += 300
            data = None
            if 'payload' in captured:
                try:
                    data = json.loads(captured['payload'])
                except Exception as e:
                    print(f'[Qishui] live payload parse error: {e}')
            try:
                page.remove_listener('response', _on_response)
            except Exception:
                pass

            results = []
            if data is not None:
                if _qishui_is_blocked(data):
                    print('[Qishui] 实时会话已失效（登录墙/风控），回退')
                    with LIVE_SESSIONS_LOCK:
                        _close_live_session('qishui')
                    PLATFORM_WARNINGS['qishui'] = (
                        '汽水音乐（抖音）实时登录会话已失效（抖音判定为未登录/风控）。'
                        '请重新在「Cookie设置」点「浏览器登录」抖音并保持窗口打开后重试。'
                    )
                    return []
                parsed = _parse_qishui_music_response(data, max_results, music_only=True)
                if parsed:
                    results = parsed

            if not results:
                print('[Qishui] 实时会话未捕获签名响应，回退 DOM 提取')
                results = _run_qishui_dom_fallback(page, keyword, max_results)

            if results:
                try:
                    _scrape_qishui_detail_stats(context, results)
                except Exception as e:
                    print(f"[Qishui] live detail scrape skip: {e}")
                return results
            return []
        except Exception as e:
            print(f'[Qishui] live search error: {e}')
            return None  # 异常则回退到无头路径


def _scrape_qishui_detail_stats(context, results):
    """尽力而为：逐个打开汽水音乐详情页，抓取 使用次数/收藏/评论。出错不阻断。"""
    def parse_count(text):
        if not text:
            return None
        m = re.search(r'([\d.]+)\s*(万|亿|w|k)?', text)
        if not m:
            return None
        n = float(m.group(1))
        u = (m.group(2) or '').lower()
        if u in ('万', 'w'):
            n *= 10000
        elif u == '亿':
            n *= 1e8
        elif u == 'k':
            n *= 1000
        return int(round(n))

    targets = [r for r in results
               if r.get('song_url') and 'douyin.com/music/' in r['song_url']
               and r.get('collection_count') is None][:8]
    if not targets:
        return
    pages = []
    for r in targets:
        try:
            p = context.new_page()
            p.route('**/*.{png,jpg,jpeg,gif,svg,webp,woff,woff2,mp4,webm,mp3}',
                    lambda route: route.abort())
            p.goto(r['song_url'], wait_until='domcontentloaded', timeout=15000)
            pages.append((p, r))
        except Exception:
            pass
    for p, r in pages:
        try:
            p.wait_for_timeout(2500)
            body = p.inner_text('body') or ''
            m_use = re.search(r'([\d.]+)\s*(万|亿)?\s*使用', body)
            if m_use:
                r['collection_count'] = parse_count(m_use.group(0))
            m_collect = re.search(r'([\d.]+)\s*(万|亿)?\s*收藏', body)
            if m_collect:
                r['collection_count'] = r.get('collection_count') or parse_count(m_collect.group(0))
            m_comment = re.search(r'([\d.]+)\s*(万|亿)?\s*评论', body)
            if m_comment:
                r['comment_count'] = parse_count(m_comment.group(0))
            m_share = re.search(r'([\d.]+)\s*(万|亿)?\s*分享', body)
            if m_share:
                r['share_count'] = parse_count(m_share.group(0))
        except Exception:
            pass
        finally:
            try:
                p.close()
            except Exception:
                pass

# ═══════════════════════════════════════════════════
#  统一搜索入口
# ═══════════════════════════════════════════════════

PLATFORM_FUNCS = {
    'qq': search_qq,
    'kugou': search_kugou,
    'kuwo': search_kuwo,
    'netease': search_netease,
    'qishui': search_qishui,
}

# v4.27.33：详情补全（词曲/发行方/收藏量等）的条数上限。
# 用户选每平台 100/500 时，抓取量随 limit 放大，但补全只做「最相关的前 N 条」——
# 既让大数量搜索返回有用的卡片（而非一堆空白），又把耗时框死在上限内，避免无限拉长。
SEARCH_ENRICH_CAP = 100

# ── 搜索进度注册表（v4.27.34）──
# 搜索是一次性 POST，选每平台 100/500 时要跑 40~90 秒。以前前端只能显示「已等待 N 秒」，
# 用户分不清「还在跑」和「已经卡死」。这里在内存里记录每次搜索的真实进度（每个平台已抓到
# 多少条、平台任务完成几个、当前处于抓取/补全/聚合哪个阶段），前端带 search_id 发起搜索后
# 每秒轮询 /api/search_progress 读取。
# 前提：服务是 app.run(threaded=True) 单进程多线程，轮询请求与搜索请求在不同线程共享本字典，
# 所以进度能在搜索返回前被读到；所有写入走 _search_progress_lock。
# 只存内存不落库；新建时顺手清掉超过 TTL 的旧记录，避免长跑进程里堆积。
SEARCH_PROGRESS = {}
SEARCH_PROGRESS_TTL = 600
_search_progress_lock = threading.Lock()


def _sp_start(sid, keyword_count=1, platform_count=None):
    """注册一次搜索的进度。keyword_count = 主关键词 + alt_keywords 个数。"""
    if not sid:
        return
    now = time.time()
    n_plat = platform_count if platform_count is not None else len(PLATFORM_FUNCS)
    with _search_progress_lock:
        for k, v in list(SEARCH_PROGRESS.items()):
            if now - v.get('updated', 0) > SEARCH_PROGRESS_TTL:
                SEARCH_PROGRESS.pop(k, None)
        SEARCH_PROGRESS[sid] = {
            'stage': 'search',
            'platforms': {},   # platform_code -> 累计条数（跨关键词累加，去重前的原始条数）
            'done_tasks': 0,   # 已完成的「平台 × 关键词」任务数
            'total_tasks': max(1, keyword_count) * max(1, n_plat),
            'total': 0,
            'started': now,
            'updated': now,
            'finished': False,
        }


def _sp_platform_done(sid, code, count):
    """某平台某关键词的抓取任务结束（失败也要调，count 传 0），累加条数。"""
    if not sid:
        return
    with _search_progress_lock:
        p = SEARCH_PROGRESS.get(sid)
        if not p:
            return
        p['platforms'][code] = p['platforms'].get(code, 0) + int(count or 0)
        p['done_tasks'] += 1
        p['total'] = sum(p['platforms'].values())
        p['updated'] = time.time()


def _sp_stage(sid, stage, **extra):
    if not sid:
        return
    with _search_progress_lock:
        p = SEARCH_PROGRESS.get(sid)
        if not p:
            return
        p['stage'] = stage
        p.update(extra)
        p['updated'] = time.time()


def _sp_finish(sid):
    _sp_stage(sid, 'done', finished=True)


LABEL_PLATFORMS = {'netease'}

# 平台级告警（搜索时由各平台函数写入，api_search 读取后清空）
PLATFORM_WARNINGS = {}

# 强鉴权 Cookie：出现这些才算真正登录态（用于浏览器登录捕获时等待其落盘）
STRONG_AUTH_COOKIES = {
    'netease': ['MUSIC_U', '__csrf'],
}


def search_all(keyword, per_platform_limit=30, alt_keywords=None, platforms=None,
               progress_sid=None):
    """并发搜索所有平台。

    主关键词 = keyword；alt_keywords 是辅助关键词列表（用来弥补带副标题的歌名
    在严匹配平台漏抓的情况——例如「记得你是女人（忘记你是女人）」整段作为关键词
    时，网易云会要求副标题也在歌里才匹配，而汽水会整体错位）。

    各关键词独立搜索，结果按平台内部 ID 去重合并；先返回主关键词的结果，
    同一首歌在不同关键词下命中时仅保留一份。
    """
    keywords = []
    for kw in ([keyword] if keyword else []) + list(alt_keywords or []):
        kw = (kw or '').strip()
        if kw and kw not in keywords:
            keywords.append(kw)
    if not keywords:
        return []

    _sp_start(progress_sid, keyword_count=len(keywords),
              platform_count=len(platforms) if platforms else len(PLATFORM_FUNCS))

    def _run_one(kw):
        _funcs = PLATFORM_FUNCS
        if platforms:
            _funcs = {k: v for k, v in PLATFORM_FUNCS.items() if k in platforms}
        # v4.27.33：用户选多少抓多少，不再打折。100/500 都如实抓取；
        # 补全条数由 SEARCH_ENRICH_CAP(100) 框死，耗时可控（见下方 results[:SEARCH_ENRICH_CAP]）。
        # 输入上界由 api_search 的 min(limit,1000) 兜底，这里不再二次缩量。
        fetch_limit = per_platform_limit
        with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, len(_funcs))) as executor:
            futures = {
                executor.submit(func, kw, fetch_limit): code
                for code, func in _funcs.items()
            }
            sub = []
            for future in concurrent.futures.as_completed(futures):
                code = futures[future]
                try:
                    # v4.27.33：单平台超时 70s → 120s。选 500 时单平台抓 500 + 补全前 100 条，
                    # 多关键词并发下最慢的单平台-关键词任务可能逼近 90s，放宽到 120s 避免大数量搜索被截断丢结果；
                    # 前端软超时 150s 仍留足余量（真卡死时用户可点取消）。
                    items = future.result(timeout=120)
                    sub.extend(items)
                    _sp_platform_done(progress_sid, code, len(items))
                except Exception as e:
                    print(f"[{code}] Failed: {e}")
                    # 失败也要计数，否则前端进度分母永远差一，看着像卡住
                    _sp_platform_done(progress_sid, code, 0)
        return sub

    all_results = []
    seen_ids = set()  # (platform_code, internal_id) — 用于跨关键词去重同一首歌

    def _key_of(r):
        code = r.get('platform_code') or ''
        # 各平台内部 ID 字段（任一非空即视为同一条）
        ids = (
            r.get('_songid'), r.get('_songmid'),
            r.get('_hash'), r.get('_album_hash'),
            r.get('_rid'), r.get('_album_id'),
            r.get('_song_id'),
            r.get('id'), r.get('video_id'),
        )
        # 取首个非空
        iid = next((v for v in ids if v), None)
        return (code, iid)

    # v4.27.32：多关键词（主关键词 + alt_keywords）改为并发执行，避免串行叠加超时。
    # 例：「嚣张 en」会生成 ['嚣张','嚣张 en'] 两个关键词，旧版串行 → 最坏 2×50s ≈ 100s+ 触发
    # 前端 120s 软超时；并发后墙钟时间 = 最慢单个关键词的耗时（≤35s），不再随关键词数线性叠加。
    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, len(keywords))) as kw_exec:
        kw_futures = {kw_exec.submit(_run_one, kw): kw for kw in keywords}
        for kw_fut in concurrent.futures.as_completed(kw_futures):
            kw = kw_futures[kw_fut]
            try:
                sub = kw_fut.result(timeout=80)
            except Exception as e:
                print(f"[search_all] 关键词「{kw}」搜索失败: {e}")
                sub = []
            is_alt = kw != keyword
            for r in sub:
                k = _key_of(r)
                if k[1] and k in seen_ids:
                    continue
                if k[1]:
                    seen_ids.add(k)
                # 记录来自哪个关键词（评分时可加权原 keyword 命中的）
                r['_matched_keyword'] = kw
                r['_is_alt_keyword'] = is_alt
                all_results.append(r)

    # QQ 详情补全（收藏/评论/在听/词曲作者），与 api_search 保持一致。
    # 批量搜索若不补全，QQ 的收藏数/评论数全是 None——这是历史 bug（2026-08-04 发现）。
    # 每平台前 100 条，避免大批量请求被风控。
    _sp_stage(progress_sid, 'enrich')
    qq_results = [r for r in all_results if r.get('platform_code') == 'qq'][:100]
    if qq_results:
        try:
            qq_cookie = get_cookie_string('qq')
            _fetch_qq_details(qq_results, qq_cookie)
            _fetch_qq_album_labels(qq_results, qq_cookie)
        except Exception as e:
            print(f"[QQ] batch enrich error: {e}")

    # Discogs 通用厂牌补全（第三方源，免 Kugou 凭证；token 缺失时自动跳过）
    try:
        from discogs_meta import enrich as _discogs_enrich
        _discogs_enrich(all_results)
    except Exception as e:
        print(f"[Discogs] enrich error: {e}")

    return all_results


# ═══════════════════════════════════════════════════
#  授权核对辅助函数
# ═══════════════════════════════════════════════════

def _normalize(text):
    if not text:
        return ''
    text = html_mod.unescape(text)
    text = re.sub(r'[\(（\[【].*?[\)）\]】]', '', text)
    text = re.sub(r'[\s\-_/&＆·•,，、;；:：.。]+', '', text)
    return text.lower().strip()


def _name_match(auth_name, result_name):
    a = _normalize(auth_name)
    b = _normalize(result_name)
    if not a or not b:
        return False
    return a == b or a in b or b in a


def _performer_match(auth_artist, result_artist):
    a = _normalize(auth_artist)
    b = _normalize(result_artist)
    if not a or not b:
        return False
    if a == b or a in b or b in a:
        return True
    for part in re.split(r'[/&,，、]', auth_artist):
        p = _normalize(part)
        if p and (p in b or b in p):
            return True
    for part in re.split(r'[/&,，、]', result_artist):
        p = _normalize(part)
        if p and (p in a or a in p):
            return True
    return False


def _label_match(auth_label, result_label):
    a = _normalize(auth_label)
    b = _normalize(result_label)
    if not a or not b:
        return None
    return a == b or a in b or b in a


# ═══════════════════════════════════════════════════
#  搜索结果匹配度评分（歌曲名相似度为核心）
# ═══════════════════════════════════════════════════

def _normalize_match(text):
    """用于匹配的标准化：去 HTML 转义、括号、空格、标点、英文大小写统一。

    字符白名单：ASCII 字母数字 + 中文（CJK 统一汉字）+ 日文（平假名/片假名）+ 韩文（Hangul 音节）。
    不能只保留 a-zA-Z0-9\u4e00-\u9fa5，否则韩文/日文假名会被全过滤掉，
    导致所有韩日文歌曲的红心匹配全部失效。
    """
    if not text:
        return ''
    text = html_mod.unescape(str(text))
    # 去掉括号及其内容（Live/DJ/Remix 版本号等）
    text = re.sub(r'[\(（\[【].*?[\)）\]】]', '', text)
    # 统一常见变体字符
    text = text.replace('·', '').replace('•', '').replace(' ', '')
    # 只过滤标点/空白：保留 ASCII + CJK + Hangul + 日文假名
    text = re.sub(r'[^\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff\uac00-\ud7afa-zA-Z0-9]', '', text)
    return text.lower().strip()


def _normalize_exact(text):
    """精确标准化：**保留括号及其内容**（伴奏 / DJ版 / 粤语版 / 2023新版 等版本信息）。

    与 _normalize_match 的唯一区别：不删除括号内容。用途分工必须严格区分——
      * _normalize_match（模糊）：搜索评分、红心匹配。搜「笑柄」要能搜出「笑柄(伴奏)」，
        所以必须剥掉版本后缀，否则搜不到。
      * _normalize_exact（精确）：标记 key、结果聚合分组。版权方需要「原版 / 伴奏 / DJ版」
        各自独立成条一对一标记，剥掉括号会把 N 个版本塌缩成 1 条，标记信息丢失。

    仍然吃掉的只是「格式差异」而非「内容差异」：全角/半角括号归一、大小写归一、
    空格与间隔符去除。这样跨平台写法差异（如 QQ「笑柄 (DJ阿卓版)」vs 酷狗「笑柄（dj阿卓版）」）
    依旧能匹配到一起，只有真正不同的版本才会分开。
    """
    if not text:
        return ''
    text = html_mod.unescape(str(text))
    # 各类括号归一为半角 ( )，保留内容
    text = re.sub(r'[\(（\[【〔｛]', '(', text)
    text = re.sub(r'[\)）\]】〕｝]', ')', text)
    text = text.replace('·', '').replace('•', '').replace(' ', '').replace('\u3000', '')
    # 白名单同 _normalize_match，额外保留括号本身
    text = re.sub(r'[^\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff\uac00-\ud7afa-zA-Z0-9()]', '', text)
    return text.lower().strip()


def _song_name_score(query, result_name):
    """计算歌名相似度得分 0-100，以及匹配类型。

    优先级（v4.17.1 改版——藏语版优先）：
    1. **用户输入的精确版本（保留括号）作为最强匹配信号**
       1a. 两者 _normalize_exact 完全相等                              → 100 'exact'
       1b. 用户输入完整出现在结果名里（搜「野人（藏语版）」命中更长同名）→ 96  'exact_fullword'
       1c. 结果名完整出现在用户输入里（搜「野人」命中「野人（藏语版）」）→ 95  'contains'
    2. **剥括号模糊匹配降级**：搜「野人（藏语版）」命中无括号的「野人」原本被打 100 分 exact，
       导致用户输入与对版完全一致但被无括号的同名冒泡到首位（v4.17 之前都是这毛病）。
       现统一降为 85 'approx_dropped_suffix'，仅当用户未提供辅助字段且平台只返回
       同名无括号版本时排在精确版之后。
    3. 结果是查询的子串（搜「블랙밤」命中「블랙」）→ 35 'contained'，弱相关标签，防冒泡。
    4. 字符合覆盖度递增 / None。
    """
    # 先用 _normalize_exact 做保留括号的精确比较
    q_exact = _normalize_exact(query)
    r_exact = _normalize_exact(result_name)
    has_q_paren = bool(re.search(r'[(\[【〔（）]', str(query) or ''))
    has_r_paren = bool(re.search(r'[(\[【〔（）]', str(result_name) or ''))
    if not q_exact or not r_exact:
        return 0, None

    # 1a. 完全相等（保留括号标准化后）
    if q_exact == r_exact:
        return 100, 'exact'

    # 1b. 用户输入比结果名短，结果名带括号（用户没明示版本，平台恰好有某个版本的扩展同名）
    #     典型：搜「笑柄」命中「笑柄(伴奏)」、「笑柄(Live版)」。平台稍有版本号，但用户没要求具体版本，所以给 95。
    if r_exact.startswith(q_exact) and has_r_paren and not has_q_paren and len(r_exact) > len(q_exact):
        return 95, 'contains_with_version'

    # 1c. 用户输入带括号（明示要某个版本），结果名无括号 → 平台多半没收录对版，降级
    if r_exact in q_exact and has_q_paren and not has_r_paren:
        return 85, 'approx_dropped_suffix'

    # 1d. 用户输入是结果名的完整前缀子串（用户已包含最完整版本信息，平台命中带扩展名版本）。
    #     v4.24 修：要求 r.startswith(q)，避免「在中间」的子串误判——
    #     比如「你」不该被当成「经典你」的开头命中。
    if q_exact in r_exact and r_exact.startswith(q_exact):
        return 95, 'contains_with_version'

    # 1e. 结果名是用户输入的完整前缀子串（用户比结果长，但没什么带括号语义）。
    #     v4.24 修：① 要求 q.startswith(r)（前缀，不是任意 in）；
    #     ② 要求 len(r_exact) >= 3（避免「说」二字误中「在你说再见之前」这种中间子串，
    #     短结果本来就跟很多长查询字面撞，盲目给 90 分会把枯木逢春的《你说》错挂给任意
    #     包含「你说」二字的长歌名）。
    if r_exact in q_exact and q_exact.startswith(r_exact) and len(r_exact) >= 3:
        return 90, 'contains'

    # 2. 剥括号的近似匹配——降级为 85 分近似，不再算精确
    q_match = _normalize_match(query)
    r_match = _normalize_match(result_name)
    if q_match and q_match == r_match:
        return 85, 'approx_dropped_suffix'

    # 3. 剥括号后单边包含（v4.24 修：要求前缀包含，且短结果<3字不算）
    if q_match and q_match in r_match and r_match.startswith(q_match):
        if len(q_match) >= 4:
            return 88, 'approx_with_suffix'
        return 78, 'contains_match'
    if r_match and r_match in q_match and q_match.startswith(r_match) and len(r_match) >= 3:
        return 30, 'contained_match'

    # 4. 字符覆盖度（按字符计算交集）
    q_set = set(q_match)
    r_set = set(r_match)
    if q_set and r_set:
        intersection = q_set & r_set
        cover = len(intersection) / max(len(q_set), len(r_set))
        if cover >= 0.8:
            return 60, 'fuzzy_high'
        if cover >= 0.5:
            return 40, 'fuzzy_medium'
        if cover >= 0.3:
            return 20, 'fuzzy_low'

    return 0, None


def _person_match(query, result_val):
    """人名匹配：支持 A/B/C 与 A/B 这种组合形式。"""
    if not query or not result_val:
        return False
    a = _normalize_match(query)
    b = _normalize_match(result_val)
    if not a or not b:
        return False
    if a == b or a in b or b in a:
        return True
    # 拆分多表演者 / 多作者
    for part in re.split(r'[/&,，、]', str(query)):
        p = _normalize_match(part)
        if p and (p in b or b in p):
            return True
    for part in re.split(r'[/&,，、]', str(result_val)):
        p = _normalize_match(part)
        if p and (p in a or a in p):
            return True
    return False


def _artist_match_score(query, result_val):
    """歌手匹配「分级打分」（0~100），替代 _person_match 的粗暴布尔判断。

    为什么必须分级：网易云/酷狗搜索结果里挤满了蹭原唱名字的 UGC 翻唱账号，
    比如「周杰伦./Asasblue」「陈奕迅-/INKK」「B-KLl/周杰伦、」。
    用布尔命中的话，它们和真正的「陈奕迅」完全等价，排序只能靠运气 ——
    实测「孤勇者」就抽中了红心 173 的翻唱，而原版有 2245 万，
    数据直接差 13 万倍，拿去做曲库决策是灾难性的。

    分档：
      100  归一化后完全相等                → 原版，可信
       92  多艺人串里存在完全相等的一项      （如「陈奕迅/王菲」查「陈奕迅」）
       70  仅子串包含                     → 十有八九是蹭名字的翻唱号，存疑
        0  不沾边
    """
    if not query or not result_val:
        return 0
    a = _normalize_match(query)
    b = _normalize_match(result_val)
    if not a or not b:
        return 0
    if a == b:
        return 100

    q_parts = _split_artist_parts(query)
    r_parts = _split_artist_parts(result_val)

    # ⓪ 悬空分隔符 = 蹭名铁证。
    #    正经歌手字段不会以「、」「/」「&」结尾，也不会出现「//」这种空档。
    #    UGC 翻唱号的标准手法就是「自己ID/原唱名、」，靠这个把原唱名塞进搜索索引。
    #    实测「晴天」在网易云被「B-KLl/周杰伦、」抢到，红心 11846，而 QQ 原版 5600 万。
    _rv = str(result_val or '').strip()
    _dangling = bool(re.search(r'[/&,，、]\s*$', _rv) or re.search(r'[/&,，、]\s*[/&,，、]', _rv))

    # ① 原文级严格相等：「陈奕迅/王菲」查「陈奕迅」、「冯沁苑(买辣椒也用券)」查
    #    「买辣椒也用券」—— 都是正主本人（合唱 / 本名+艺名），给 92。
    if {p.strip() for p in q_parts} & {p.strip() for p in r_parts}:
        return 70 if _dangling else 92

    # ② 只有「归一化之后」才相等：说明原文带了装饰字符，典型如
    #    「周杰伦.」「陈奕迅-」—— 蹭名 UGC 账号的标准手法（原唱名后缀一个标点，
    #    再跟自己的 ID）。绝不能和正主同级，压到 70 = 存疑。
    q_norm = {_normalize_match(p) for p in q_parts}
    r_norm = {_normalize_match(p) for p in r_parts}
    q_norm.discard(''); r_norm.discard('')
    if q_norm & r_norm:
        return 70

    # ③ 纯子串沾边，可信度最低
    if a in b or b in a:
        return 70
    for p in q_norm:
        if any(p in x or x in p for x in r_norm):
            return 70
    return 0


# ── 超级大艺人智能排除（v4.25.2 新增）──
# 用户曲库全是独立/小众艺人，不可能有这些「超级无敌大艺人」的歌。
# 当搜索结果返回的歌手命中以下名单、而曲库本歌的表演者并非该大艺人时，
# 视为「同名撞车」的误配 → 直接判未收录，绝不把大艺人的收藏量算到小艺人头上。
# 若曲库本歌表演者就是该大艺人（如《晴天》→周杰伦），则正常保留（「智能」所在）。
_MEGA_ARTISTS = {
    # 华语超级巨星
    '周杰伦', 'jaychou', 'jay',
    '薛之谦', 'xuezhiqian',
    '林俊杰', 'jjlin', 'jj', '林俊傑',
    '邓紫棋', 'gemyao',
    '周深', 'zhoushen',
    '毛不易', 'maobuyi',
    '李荣浩', 'lironghao',
    '张杰', 'zhangjie', '张傑',
    '五月天', 'mayday',
    '华晨宇', 'huachenyu',
    '鹿晗', 'luhan',
    '蔡徐坤', 'caixukun',
    '刘德华', 'liudehua', 'andylui', '劉德華',
    '张学友', 'zhangxueyou', '張學友',
    '陈奕迅', 'chenyixun', 'eason', '陳奕迅',
    '王菲', 'wangfei', 'faye',
    '李健', 'lijian',
    '朴树', 'pushu',
    '赵雷', 'zhanglei',
    '陈粒', 'chenli',
    '孙燕姿', 'sunyanzi',
    '梁静茹', 'liangjingru', 'fishleong',
    '周华健', 'zhouhuajian', 'emilchau',
    '任贤齐', 'renxianqi', 'richiejen',
    '刀郎', 'daolang',
    '汪峰', 'wangfeng',
    '许巍', 'xuwei',
    '张惠妹', 'zhanghuimei', 'amei', 'amit',
    '泰勒斯威夫特', 'taylorswift',
    '黄家驹', 'beyond',
}


def _artist_core_strip(s):
    """剥掉歌手名尾部的账号/Latin/数字装饰，取「人核」。

    例：「蒋雪儿Snow.J」→「蒋雪儿」；「周杰伦.」→「周杰伦」；「陈奕迅-/INKK」→「陈奕迅-」
    （保留 '-' 以便识别蹭名）。纯英文艺名（TaylorSwift）不剥。
    用于批量查歌时把「带装饰的小艺人同名账号」认回曲库里的原艺人。
    """
    s = (s or '').strip()
    if not s:
        return ''
    n = _normalize_match(s)
    if not n:
        return ''
    # 仅当含 CJK 时，剥掉尾部连续 Latin/数字尾缀（账号装饰，如 Snow.J / Asasblue / 99）
    if re.search(r'[\u4e00-\u9fff]', n):
        n = re.sub(r'[a-z0-9]+$', '', n)
    return n


def _alias_variants(s):
    """歌手「写法变体」归一到同一核，让 dollyy / dollyy99 / Dollyy / DOLLYy99 /
    jaychou_01 这类账号序号/大小写/分隔符变体认回为同一人。

    与 _artist_core_strip 的分工（单一真相源，避免两套逻辑各剥各的）：
      - 含 CJK 的歌手名 → 直接交给 _artist_core_strip（它已剥尾部 latin/数字装饰，
        如「蒋雪儿Snow.J」→「蒋雪儿」），这里不再重复处理。
      - 纯英文/数字（无 CJK）→ 归一成「小写字母核」：剥掉尾部数字序号与尾部分隔符，
        如 dollyy99 → dollyy、Dollyy-99 → dollyy、DOLLYy_01 → dollyy。
    返回小写归一核；空或归一后为空返回 ''。
    """
    s = (s or '').strip()
    if not s:
        return ''
    # 含 CJK：交给 _artist_core_strip（单一真相源）
    if re.search(r'[\u4e00-\u9fff]', s):
        return _artist_core_strip(s)
    n = _normalize_match(s)
    if not n:
        return ''
    # 剥尾部分隔符（- _ . 空格）再剥尾部数字序号，取纯字母核。
    # 注意 _normalize_match 已吞掉 - _ . 空格，这里主要对付尾随数字序号。
    core = re.sub(r'[-_.\s]+$', '', n)
    core = re.sub(r'\d+$', '', core)
    return core or n


def _load_performer_aliases():
    """启动时加载「艺人变体白名单」，文件不存在则初始化默认（钦觉 变体表）。"""
    global _PERFORMER_ALIASES, _performer_aliases_loaded
    if _performer_aliases_loaded:
        return
    defaults = {
        '钦觉': {'钦觉呀', '钦觉(微唱)'},
        '胡艾彤': {'胡艾彤喂猪吗'},
        # v4.25.x 根因修复·「本名 + 个人后缀/别名/乐队」自动认回（同艺人多账号/改名发歌）。
        # 这些账号名 = 曲库艺名 + 个人化尾缀（真实姓名/昵称/乐队/口号），本人无疑，升 asc=100 →
        # 标签「匹配(艺人变体:xxx)」。每对均来自脏数据自查中逐条确认的「本人账号」，非放开任意 CJK 尾缀。
        'TEn': {'TEn李祎麟'},
        'Yusee': {'Yusee西'},
        'yusee': {'Yusee西'},
        '王靖雯': {'王靖雯不胖'},
        '上官承诺': {'上官承诺AKADa橙子'},
        '皮卡丘多多': {'皮卡丘多多尹子桐'},
        '火火': {'火火带你去流浪'},
        '小黑': {'小黑同学'},
        'Holy Grail': {'HolyGrail猴里贵'},
        '李悟': {'李悟小礼物', '李悟小礼物、泡泡留声'},
        '偲雯': {'偲雯是我'},
        'rango': {'RanGo/可爱多'},
        '陈嘉俊': {'陈嘉俊是这个嘉'},
        '张浩': {'张浩晨'},
        '鹏鹏': {'鹏鹏音乐'},
        '锦零': {'锦零真是可爱死了'},
        'ZTt': {'ZTt璕'},
        'Gibb-Z': {'Gibb-Z黄泽'},
        'Sugar': {'Sugar张惠晴'},
        '阿修罗': {'阿修罗乐队'},
    }
    loaded = {}
    try:
        if os.path.exists(_PERFORMER_ALIASES_FILE):
            with open(_PERFORMER_ALIASES_FILE, encoding='utf-8') as f:
                raw = json.load(f)
            for k, v in (raw or {}).items():
                if isinstance(k, str) and isinstance(v, (list, tuple)):
                    loaded[k] = {x for x in v if isinstance(x, str) and x}
    except Exception:
        loaded = {}
    # 合并默认值（已存在的条目以文件为主，仅补缺失的）
    for k, v in defaults.items():
        loaded.setdefault(k, set(v))
    _PERFORMER_ALIASES = {k: v for k, v in loaded.items()}
    _performer_aliases_loaded = True


def _save_performer_aliases():
    """把当前白名单写回磁盘（线程安全）。"""
    try:
        with _performer_aliases_lock:
            payload = {k: sorted(v) for k, v in _PERFORMER_ALIASES.items()}
            tmp = _PERFORMER_ALIASES_FILE + '.tmp'
            with open(tmp, 'w', encoding='utf-8') as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            os.replace(tmp, _PERFORMER_ALIASES_FILE)
    except Exception as e:
        print(f'[performer_aliases] save error: {e}')


def _performer_is_alias(perf, res_perf):
    """判断 res_perf 是否是 perf 的已知变体（同艺人不同账号）。

    返回 True / False；perf / res_perf 为空或 perf 不在白名单里都返回 False。
    picker 仅在外部已调用 _load_performer_aliases() 后才有效；启动期 _autoload
    触发后整个进程内一致可用。
    """
    if not perf or not res_perf:
        return False
    aliases = _PERFORMER_ALIASES.get(perf)
    if not aliases:
        return False
    return (res_perf or '').strip() in aliases


def _all_performer_aliases():
    """只读快照：返回 {base: sorted_aliases} 供 admin 端点回传。"""
    return {k: sorted(v) for k, v in _PERFORMER_ALIASES.items()}


# ── 艺人变体「待确认建议」队列（v4.25.x 根因修复配套）──
# picker 在严格匹配下发现「歌手名相似 / 歌名强匹配但未被白名单或作者铁证认回」的候选时，
# 记录一条建议（base=曲库名, alias=平台名），admin 后台一键批准即可并入白名单。
# 这样「同一艺人改名/多账号」无需再逐首手工脚本，系统自动发现、人工一眼审。
_PERFORMER_ALIAS_SUG_FILE = os.path.join(COOKIE_DIR, 'performer_alias_suggestions.json')
_performer_alias_sug_lock = threading.Lock()
_performer_alias_sug_loaded = False
_performer_alias_sugs = {}


def _load_alias_suggestions():
    global _performer_alias_sugs, _performer_alias_sug_loaded
    if _performer_alias_sug_loaded:
        return
    try:
        if os.path.exists(_PERFORMER_ALIAS_SUG_FILE):
            with open(_PERFORMER_ALIAS_SUG_FILE, encoding='utf-8') as f:
                raw = json.load(f)
            _performer_alias_sugs = {k: v for k, v in (raw or {}).items() if isinstance(v, dict)}
        else:
            _performer_alias_sugs = {}
    except Exception:
        _performer_alias_sugs = {}
    _performer_alias_sug_loaded = True
    # v4.27.9：清理旧 picker 污染的多表演者别名建议（幂等）
    try:
        n = _clean_dirty_alias_suggestions()
        if n:
            print(f'[performer_alias_sugs] cleaned {n} dirty entries (multi-performer strings)')
    except Exception as e:
        print(f'[performer_alias_sugs] cleanup error: {e}')


def _save_alias_suggestions():
    try:
        with _performer_alias_sug_lock:
            tmp = _PERFORMER_ALIAS_SUG_FILE + '.tmp'
            with open(tmp, 'w', encoding='utf-8') as f:
                json.dump(_performer_alias_sugs, f, ensure_ascii=False, indent=2)
            os.replace(tmp, _PERFORMER_ALIAS_SUG_FILE)
    except Exception as e:
        print(f'[performer_alias_sugs] save error: {e}')


def _record_alias_suggestion(base, alias, song_name='', ns=0):
    """记录一条待确认别名建议（去重：同 base||alias 只记一条，累加命中次数）。"""
    base = (base or '').strip()
    alias = (alias or '').strip()
    if not base or not alias or base == alias:
        return
    _load_alias_suggestions()
    key = f"{base}||{alias}"
    now = int(time.time())
    cur = _performer_alias_sugs.get(key)
    if cur:
        cur['count'] = cur.get('count', 0) + 1
        cur['last_seen'] = now
        if ns and ns > cur.get('ns', 0):
            cur['ns'] = ns
    else:
        _performer_alias_sugs[key] = {
            'base': base, 'alias': alias, 'song': song_name or '',
            'ns': ns, 'count': 1, 'first_seen': now, 'last_seen': now,
        }
    _save_alias_suggestions()


# v4.27.9：多表演者感知的别名建议。把 perf 和 res_perf 各自按 / & , ，、 拆，对每对
# 归一化相似的 (perf_seg, res_seg) 单独记一条，避开「Kui Kui / 徐梦圆」被整体当成
# Kui Kui 的别名候选入库。同一 perf_seg 取最长的 res_seg（更接近平台展示形态）。
def _record_alias_suggestion_split(perf, res_perf, song_name='', ns=0):
    perf_parts = _split_artist_parts(perf)
    res_parts = _split_artist_parts(res_perf)
    if len(perf_parts) <= 1 and len(res_parts) <= 1:
        # 都是单表演者：直接走老逻辑
        _record_alias_suggestion(perf_parts[0], res_parts[0], song_name, ns)
        return
    for pp in perf_parts:
        pp_n = _normalize_match(pp)
        if not pp_n:
            continue
        best_rp = None
        for rp in res_parts:
            rp_n = _normalize_match(rp)
            if not rp_n:
                continue
            if pp_n == rp_n or pp_n in rp_n or rp_n in pp_n:
                if best_rp is None or len(rp) > len(best_rp):
                    best_rp = rp
        if best_rp:
            _record_alias_suggestion(pp, best_rp, song_name, ns)


# v4.27.9：清理被污染的旧别名建议（base 或 alias 含 / & , ，、 ; 的多表演者串）。
# 旧 picker 直接整串入库，导致 1207/1512（80%）的候选都是脏的。幂等：无脏数据返回 0。
_DIRTY_ALIAS_SEP_RE = re.compile(r'[/&,，、;]')


def _clean_dirty_alias_suggestions():
    _load_alias_suggestions()
    dirty_keys = []
    for key, item in list(_performer_alias_sugs.items()):
        if not isinstance(item, dict):
            continue
        base = item.get('base', '') or ''
        alias = item.get('alias', '') or ''
        if _DIRTY_ALIAS_SEP_RE.search(base) or _DIRTY_ALIAS_SEP_RE.search(alias):
            dirty_keys.append(key)
    if not dirty_keys:
        return 0
    for k in dirty_keys:
        _performer_alias_sugs.pop(k, None)
    _save_alias_suggestions()
    return len(dirty_keys)


_ANNOT_LABEL_RE = re.compile(
    r'^\s*(?:艺名|别名|英文名|aka|艺人|名称|歌手|演唱|原名|本名|组合|乐队|'
    r'昵称|曾用名|工作室|团队|曲名|歌名|曲目)[:：\s]+(.+)$', re.I)


def _strip_annotation_label(text):
    """剥掉『标签：内容』式说明文字，返回干净内容。

    「艺人：小嘉玲」→「小嘉玲」；「名称：你之外」→「你之外」；
    「徐嘉玲（艺名：小嘉玲）」里的「艺名：小嘉玲」→「小嘉玲」。
    仅在「标签 + 冒号/空格 + 内容」形态时触发，不会误伤正常的人名/歌名
    （如歌名叫《歌手》不会被误剥，必须带『歌手：』前缀才处理）。
    """
    if not text:
        return ''
    m = _ANNOT_LABEL_RE.match(str(text).strip())
    return m.group(1).strip() if m else str(text).strip()


def _split_artist_parts(val):
    """把歌手字段拆成候选人名列表，括号内的别名也单独拎出来。

    「冯沁苑(买辣椒也用券)」→ ['冯沁苑', '买辣椒也用券']（本名 + 艺名）
    「徐嘉玲（艺名：小嘉玲）」→ ['徐嘉玲', '小嘉玲']（剥掉「艺名：」前缀）
    「艺人：小嘉玲」        → ['小嘉玲']（整段就是『标签：内容』，剥标签留真名）
    「陈奕迅-/INKK」        → ['陈奕迅-', 'INKK']（注意保留尾部的 '-'，
                              它正是识别蹭名账号的关键证据，不能提前抹掉）
    v4.26.3 增强：标签前缀集合扩到 艺名/别名/英文名/aka/艺人/名称/歌手/演唱/原名/
    本名/组合/乐队/昵称/曾用名/工作室/团队/曲名/歌名/曲目，且无论写在括号内还是
    整段独立（如『艺人：小嘉玲』），都剥掉只留真名，避免带说明文字的字段搜不到。
    """
    s = str(val or '')
    out = []
    # v4.27.9 补：分隔符集合加 `;`（平台字段如「A1 TRIP;大籽;Lil E」常见）。
    # 不拆分会把整串当成 Kui Kui 的别名入库，正是用户截图反馈的 bug。
    # 含 `\s*` 在分隔符外是冗余；后续 seg.strip() 已经处理。
    for seg in re.split(r'[/&,，、;]', s):
        seg = seg.strip()
        if not seg:
            continue
        inner = re.findall(r'[(（\[【]([^)）\]】]+)[)）\]】]', seg)
        outer = re.sub(r'[(（\[【][^)）\]】]*[)）\]】]', '', seg).strip()
        # 整段就是「标签：内容」（如『艺人：小嘉玲』）→ 剥标签留真名
        outer_clean = _strip_annotation_label(outer)
        if outer_clean and outer_clean != outer:
            outer = outer_clean
        if outer:
            out.append(outer)
        for x in inner:
            x = x.strip()
            if not x:
                continue
            # 括号内也可能是「标签：内容」，同样剥标签只留真名/别名
            x_clean = _strip_annotation_label(x)
            out.append(x_clean if x_clean else x)
    return out or [s.strip()]


def _person_set_strict(query_val, result_val):
    """严格子集匹配：用户输入的每个作者都必须出现在结果集合里。

    用于「歌单输入带词曲作者 → 精准匹配」。拆分规则支持 / & , ，、 空格 等多人分隔符。
    返回 True/False；query_val 为空返回 True（不约束）。

    v4.4 智能归一化（兼顾中英文）：
      - 先把「xxx@yyy」「xxxJiaHua」「xxx feat. yyy」「xxx & yyy」这类噪声剥掉
      - 再按中文核心 / 英文核心分别比对
      - 例：「张荣昊」 vs 「张荣昊JiaHua@半杯音乐」 → 剥后缀后「张荣昊」 vs 「张荣昊」 → 命中
      - 例：「Jay Chou」 vs 「Jay Chou@杰威尔」 → 剥后缀后「Jay Chou」 vs 「Jay Chou」 → 命中
    """
    if not query_val:
        return True
    if not result_val:
        return False
    # 先把每个 person 做「智能归一化」：剥 @xxx / JiaHua / feat. / with / & 后取核心
    def _core(s):
        s = (s or '').strip()
        if not s:
            return ''
        # 1) 按 @ 切，取 @ 前那段
        if '@' in s:
            s = s.split('@', 1)[0]
        # 2) 剥常见合作/账号后缀（不区分大小写）
        for sep in [' feat. ', ' feat ', ' featuring ', ' with ', '/', '&', ',', '，', '、', ';', '；']:
            if sep.lower() in s.lower():
                # 实际切分要按原 case 切（sep 已是 lower-friendly）
                idx = s.lower().find(sep.lower())
                s = s[:idx]
                break
        # 3) 剥 JiaHua/JiaXXX 账号后缀（大小写不敏感，紧贴中文或英文名字后）
        s = re.sub(r'Jia[Hh]ua[A-Za-z]*$', '', s, flags=re.IGNORECASE)
        # 4) 去掉空白
        return s.strip()
    def _split_persons(s):
        # 多个 person 走 _normalize_match 标准化
        out = set()
        for p in re.split(r'[/&,，、\s]+', str(s)):
            n = _normalize_match(p)
            if n:
                out.add(n)
        return out
    def _split_cores(s):
        # 多个 person 走 _core 剥后缀再标准化
        out = set()
        for p in re.split(r'[/&,，、\s]+', str(s)):
            c = _core(p)
            if c:
                out.add(_normalize_match(c))
        return out
    # Phase 1: 走 _normalize_match 严格子集（兼容旧行为）
    q1 = _split_persons(query_val)
    r1 = _split_persons(result_val)
    if q1 and q1.issubset(r1):
        return True
    # Phase 2: 走 _core 剥后缀后严格子集（处理「张荣昊 vs 张荣昊JiaHua@半杯音乐」这类）
    q2 = _split_cores(query_val)
    r2 = _split_cores(result_val)
    if q2 and q2.issubset(r2):
        return True
    return False


def _dur_tier(dur):
    """v4.25.12：时长维度分级（供匹配器「片段救援」）。

    把歌曲时长（秒）映射为排序权重，权重越高越像「完整正版」：
        2  正常完整版（90~600 秒）—— 大多数流行/原声曲的合法区间，最优先
        1  超长（>600 秒）—— 现场/Live/串烧，确为完整但一般不作为首选
        0  未知（接口没给时长）—— 中性，不参与偏好
       -1  偏短（60~90 秒）—— 疑似剪辑/短版
       -2  片段/铃声（<60 秒）—— 试听片段、手机铃声，明确避开
    ⚠️ 仅作「同歌多版本」的偏好与兜底，绝不在歌名/歌手不匹配时强行替换。
    """
    d = _safe_int(dur) or 0
    if d <= 0:
        return 0
    if d < 60:
        return -2
    if d < 90:
        return -1
    if d <= 600:
        return 2
    return 1


def _score_result(result, song_name, lyricist='', composer='', performer=''):
    """给单个结果打分并生成匹配标签。分数范围 0-100。

    评分策略：
    - 歌名相似度占主体（0-100）。
    - 辅助字段（词作者、曲作者、表演者）作为加分项，命中越多越靠前。
    - 当用户明确输入了辅助字段且全部命中时，即便歌名带「Live/摇滚版」等后缀，
      也视为「精准匹配」，避免 ICE 杨长青《野人》排在 郑志宏《野人（摇滚版）》前面。
    - v4.17.4：把「歌名强匹配 + 所有用户提供的辅助字段全命中」记成 complete_match，
      写入 result['_complete_match']，作为 _sort_by_relevance 的首要排序键。
      修复 bug：旧逻辑下歌名 exact=100 时 min(100, 100+bonus)=100，
      「完整命中」和「歌名精准但辅助未命中」被同一上限拉平，导致用户填了字段
      却看不到组合匹配的结果排第一。
    """
    score, match_type = _song_name_score(song_name, result.get('song_name', ''))

    # 辅助字段加权：表演者权重最高，词曲作者次之
    perf_hit = bool(performer and _person_match(performer, result.get('performer', '')))
    lyr_hit = bool(lyricist and _person_match(lyricist, result.get('lyricist', '')))
    comp_hit = bool(composer and _person_match(composer, result.get('composer', '')))

    bonus = 0
    if perf_hit:
        bonus += 18
    if lyr_hit:
        bonus += 14
    if comp_hit:
        bonus += 14

    score = min(100, score + bonus)

    # 判断「所有用户提供的辅助字段」是否全部命中
    provided_aux = []
    if performer:
        provided_aux.append(perf_hit)
    if lyricist:
        provided_aux.append(lyr_hit)
    if composer:
        provided_aux.append(comp_hit)
    all_aux_hit = all(provided_aux) if provided_aux else False
    aux_hits = sum(provided_aux) if provided_aux else 0
    aux_total = len(provided_aux)
    # v4.21：把「命中维数」和「用户填的维数」挂到 result，供排序和前端展示用
    result['_aux_hits'] = aux_hits
    result['_aux_total'] = aux_total

    # name_strong 集合（v4.17.1 扩展 — contains_with_version 也算强匹配，便于辅助字段命中时升级）
    name_strong = match_type in (
        'exact', 'contains_with_version', 'contains',
        'approx_dropped_suffix', 'approx_with_suffix',
        'fuzzy_high',
    )

    # v4.17.4：完整命中标志（歌名强匹配 + 所有用户提供的辅助字段全命中）
    complete_match = bool(name_strong and all_aux_hit)
    result['_complete_match'] = complete_match

    # 匹配标签（v4.21 重写：四维严格制 + 歌名优先）
    # 核心语义：
    #   1. 歌名强匹配 + 用户填的辅助字段全中 → 精准匹配（绿）
    #   2. 歌名强匹配 + 部分命中 → 高近似/近似匹配（黄/蓝）
    #   3. 歌名强匹配 + 用户填了辅助但 0 命中 → 近似匹配（不标"精准"，避免误导）
    #   4. 歌名不强但演唱者/词曲命中 → "相关结果"（推荐其他作品）
    #   5. 完全不沾边 → 低相关
    name_strong = match_type in (
        'exact', 'contains_with_version', 'contains',
        'approx_dropped_suffix', 'approx_with_suffix',
        'fuzzy_high',
    )
    if complete_match:
        # 歌名强匹配 + 所有辅助字段命中 → 即便带版本后缀也视为精准（v4.17.4 合并）
        label = '精准匹配'
    elif aux_total >= 1 and not all_aux_hit and name_strong:
        # 歌名强匹配，用户填了辅助但没全中 → 按命中维数细分
        if aux_hits >= max(2, aux_total - 1) and aux_total >= 2:
            label = '高近似(%d/%d维)' % (aux_hits, aux_total)
        elif aux_hits >= 1:
            label = '近似匹配(%d/%d维)' % (aux_hits, aux_total)
        else:
            label = '近似匹配'  # 歌名强 + 用户填的辅助字段 0 命中（不能叫"精准"）
    elif match_type in ('exact', 'contains_with_version') and not provided_aux:
        # 歌名完全一致，或用户搜不带括号的基础歌名命中带括号扩展名
        # 后者括号内容只是补充信息（如父子合唱版），不是功能性版本差异，视为精准匹配
        label = '精准匹配'
    elif match_type in ('approx_dropped_suffix', 'approx_with_suffix'):
        # v4.17.1：剥括号近似的版本差异显式标注
        label = '近似匹配(版本差异)'
    elif aux_hits >= 1 and not name_strong:
        # 歌名不对但演唱者/词曲命中 → 推荐该歌手其他作品
        label = '相关结果'
    elif score >= 90 or match_type == 'exact_with_suffix':
        label = '近似匹配'
    elif score >= 55 or match_type in ('contains', 'fuzzy_high'):
        label = '近似匹配'
    elif score >= 20:
        label = '相关结果'
    else:
        label = '低相关'

    return score, label


def _sort_by_relevance(results):
    """按匹配相关性降序排列。
    v4.17.4 排序键（修复「完整命中」被同分淹没）：
      1. 完整命中优先（歌名强匹配 + 用户提供的所有辅助字段全命中） — bool 键
      2. 同命中等级下，匹配分数降序
      3. 同分下，平台顺序升序
      4. 同分同台下，收藏数降序
    v4.21 升级排序键：
      1. 完整命中优先（4 维全中）
      2. 命中维数多者优先（用户填了 4 维但只中 3 维，比只中歌名排前）
      3. 命中维数相同下，匹配分数降序
      4. 同分下，平台顺序升序
      5. 同分同台下，收藏数降序
    v4.25.12 新增时长维度（片段救援）：
      6. 同分（同歌名命中）下，优先「正常完整版」（时长 90~600s），自动避开
         <60s 的试听片段/铃声；无时长数据者中性，不影响排序。
    """
    order = {code: i for i, code in enumerate(PLATFORM_ORDER)}
    def key(r):
        complete = r.get('_complete_match', False)
        aux_hits = r.get('_aux_hits', 0)
        score = r.get('match_score', 0)
        plat_idx = order.get(r.get('platform_code', ''), 99)
        coll = r.get('collection_count') or 0
        # not True=0 排前；命中维数降序；负分降序；时长维度越完整越前；平台升序；收藏数降序
        return (not complete, -aux_hits, -score, -_dur_tier(r.get('duration')), plat_idx, -coll)
    return sorted(results, key=key)


def _cross_platform_fill(results):
    """跨平台回填歌曲级元数据。

    唱片公司、词曲作者属于「这首歌本身」的信息，不是平台私有计数。
    当 A 平台能抓到、B 平台没抓到时，用 A 平台的数据补全 B 平台，
    避免同一首歌在不同平台出现「平台无数据」的误标。
    """
    groups = {}
    for r in results:
        key = (
            _normalize_match(r.get('song_name', '')),
            _normalize_match(r.get('performer', '')),
            _normalize_match(r.get('album', '')),
        )
        groups.setdefault(key, []).append(r)

    for grp in groups.values():
        for field in ('record_label', 'lyricist', 'composer'):
            # 收集该组里所有有效值（含来源平台）
            candidates = []
            for r in grp:
                val = r.get(field)
                if val:
                    candidates.append((val, r.get('platform_code'), r.get('_detail_status', {}).get(field)))
            if not candidates:
                continue
            # 优先选 status=ok 的值；同状态按平台优先级
            order_idx = {code: i for i, code in enumerate(PLATFORM_ORDER)}
            candidates.sort(key=lambda x: (0 if x[2] == 'ok' else (1 if x[2] == 'missing' else 2),
                                           order_idx.get(x[1], 99)))
            best_val, best_plat, _ = candidates[0]
            for r in grp:
                if not r.get(field):
                    r[field] = best_val
                    r.setdefault('_detail_status', {})[field] = 'ok'
                    r.setdefault('_source_platform', {})[field] = best_plat


def _split_artist_names(name):
    """把歌手/表演者字符串拆成标准化的人名集合（忽略顺序和分隔符）。

    支持：顿号、逗号、斜杠、&、and、空格 等常见分隔符。
    返回排序后的元组，可用于 dict key。
    """
    if not name:
        return ()
    name = html_mod.unescape(str(name))
    # 统一常见分隔符：中文顿号、逗号、斜杠、反斜杠、&、and、分号、空格
    parts = re.split(r'[,，/&、\\;；\s]+', name)
    cleaned = []
    for p in parts:
        p = p.strip()
        # 去掉英文 and 大小写变体
        if re.fullmatch(r'(?i)and', p):
            continue
        if p:
            cleaned.append(p.lower())
    return tuple(sorted(cleaned))


def _dedup_results_by_platform_id(results):
    """v4.9 修复：同一平台、同一首歌（平台歌曲 ID 相同）被搜索接口以多种标题写法
    返回多次时，只保留一条，避免「同一首歌在界面里反复重复显示」。

    典型场景：QQ 把 mid=002K6yLY2sRmY0 的同一首歌以
      『记得你是女人』『记得你是女人 (忘记你是女人)』『记得你是女人 (伴奏)』
    三种标题各返回一次 → 旧逻辑按「歌名(去括号)+歌手+专辑」分组会得到 3~4 个 group，
    界面就出现 3 张指向同一链接的 QQ 卡。本函数按 (platform_code, 平台ID) 去重，
    只留 match_score 最高 / 标题最干净（无版本括号）的一条。

    没有平台 ID 的结果无法判定是否同一首，全部保留（各自独立桶，不会误合并）。
    """
    PLAT_ID_FIELDS = {
        'qq': ['_songid', '_songmid'],
        'kugou': ['_hash', '_mixsongid'],
        'netease': ['_song_id'],
        'kuwo': ['_songid', '_songmid'],
        'qishui': ['_song_id', '_hash'],
    }

    def _pid(r):
        code = r.get('platform_code')
        fields = PLAT_ID_FIELDS.get(code)
        if not fields:
            return None
        for fld in fields:
            v = r.get(fld)
            if v:
                return (code, str(v))
        return None

    def _score(r):
        s = float(r.get('match_score', 0) or 0)
        # 标题带版本括号（伴奏/DJ版…）的轻微降权，让更干净的主标题优先保留
        if _has_version_suffix(r.get('song_name', '')):
            s -= 0.5
        return s

    buckets = {}   # pid -> result（保留最优）
    order = []     # 维持首次出现顺序
    for r in results:
        pid = _pid(r)
        if pid is None:
            pid = ('__noid__', id(r))   # 每条无 ID 结果独占一桶，永不互相合并
        if pid not in buckets:
            order.append(pid)
            buckets[pid] = r
        elif _score(r) > _score(buckets[pid]):
            buckets[pid] = r
    return [buckets[pid] for pid in order]


def _repair_paren(name):
    """修复残缺括号：平台返回的歌名偶发「爱旧爱 (女声」（左括号无右括号）之类，
    展示前按括号类型在末尾补全，保证歌曲名完整。"""
    if not name:
        return name
    name = str(name)
    pairs = [('(', ')'), ('（', '）'), ('[', ']'), ('【', '】'), ('〔', '〕')]
    for l, r in pairs:
        lc = name.count(l)
        rc = name.count(r)
        if lc > rc:
            name += r * (lc - rc)
    return name


_RE_PLAIN_PARENS = re.compile(r'[（(][^）)]*[）)]')


def _canonical_performer(name):
    """返回 name 的规范艺人名：若 name 是某 base 的已知别名，返回其 base；否则原样返回。

    用于把「菲菲公主」「陆绮菲」这类同一艺人的不同写法归一到同一个分组 key（v4.28.x）。
    """
    if not name:
        return name
    if name in _PERFORMER_ALIASES:
        return name
    for base, aliases in _PERFORMER_ALIASES.items():
        if name in aliases:
            return base
    return name


def _artist_group_key(performer):
    """艺人分组 key（v4.27.31 + v4.28.x 括号内外别名归一）。

    背景：QQ/酷狗/网易云等平台对同一艺人常同时返回两种写法 ——「en」与「en (王翊恩)」。
    若不归一，同一首歌被分成两行，用户得手动勾合并。**先剥括号再调
    _split_artist_names**（「林俊杰 (JJ Lin)」剥括号成「林俊杰 」，split 不会把括号内
    内容当多人拆开），然后按多艺人集合归一，「en (王翊恩)」和「en」会落到同一 group。

    v4.28.x 新增：外面人名统一经 _canonical_performer 规范。只要曲库艺名与平台写法在
    performer_aliases 白名单里互为别名（如「菲菲公主 → 陆绮菲」），「菲菲公主（陆绮菲）」
    剥括号得「菲菲公主」、「陆绮菲（菲菲公主）」剥括号得「陆绮菲」、再经别名归一都落到
    同一 canonical，从而五大平台正确合并（括号内外顺序反转也视为一人）。未配别名的括号
    内容仍按原「剥括号取外面」行为忽略，保证「en」「林俊杰 (JJ Lin)」等场景零退化。
    """
    if not performer:
        return ()
    # 先剥中英文括号（连内容整段拿掉），避免括号里的空格被当成多人分隔符
    raw = _RE_PLAIN_PARENS.sub('', performer).strip()
    if not raw:
        return ()
    # 外面人名统一走别名规范，使同一艺人的不同写法归一到同一 key
    return tuple(sorted(_canonical_performer(n) for n in _split_artist_names(raw)))


def _group_results(results):
    """按 歌名+表演者 聚合，把五大平台的数据合并到一行（v4.27.31）。

    进入分组前先按平台 ID 去重（v4.9）：同一平台同一首歌被多次返回的只留一条，
    避免「同一首歌重复显示」。

    关键变更：v4.27.31 把「专辑」移出 group key。同一首歌在不同平台常以空专辑
    / 完整专辑名 / 略有差异的写法出现，按专辑分桶会把 QQ 36M 那条与酷狗/网易
    那几条拆成两行（实测「嚣张 / en (王翊恩)」就被 album='嚣张' vs album=''
    拆开，用户需手动合并）。专辑信息在展示层用 _pick_common 选最常见的展示，
    每平台实际专辑仍保留在 platform_data[p]['album'] 中可看明细。
    艺人归一用 _artist_group_key（剥括号 → 拆分隔符），让「en」和「en (王翊恩)」
    等同 key。综合两条规则，合并「歌名+艺人」相同的所有平台数据为一行，专辑仅作展示字段。
    """
    results = _dedup_results_by_platform_id(results)

    groups = {}
    for r in results:
        key = (
            _normalize_exact(r.get('song_name', '')),
            _artist_group_key(r.get('performer', '')),
        )
        groups.setdefault(key, []).append(r)

    grouped = []
    order_idx = {code: i for i, code in enumerate(PLATFORM_ORDER)}

    def _pick_common(grp, field):
        """优先选 status=ok 的值，再按平台顺序兜底。"""
        vals = []
        for r in grp:
            val = r.get(field)
            if val:
                vals.append((val, r.get('platform_code'), r.get('_detail_status', {}).get(field)))
        if not vals:
            return None, None
        vals.sort(key=lambda x: (0 if x[2] == 'ok' else (1 if x[2] == 'missing' else 2),
                                 order_idx.get(x[1], 99)))
        return vals[0][0], vals[0][1]

    def _collect_field_per_platform(grp, field):
        """收集每个平台实际抓到的字段值（去重，按 PLATFORM_ORDER 保序）。

        用途：当前端展示聚合值时，能看出 5 个平台数据是否一致，不一致时
        用 tooltip 展开各家明细（例如「网易云=酽逸 / QQ/酷狗=名决」）。
        """
        seen = {}
        for code in PLATFORM_ORDER:
            for r in grp:
                if r.get('platform_code') != code:
                    continue
                v = r.get(field)
                if not v:
                    continue
                # 同平台多条时只取第一条（其它逻辑已保证 grp 里每平台通常只一条）
                if code not in seen:
                    seen[code] = v
                    break
        return seen

    def _common_status(grp, field):
        """聚合后的字段状态：有 ok 则 ok；无 ok 有 error 则 error；否则 missing。"""
        has_ok = any(r.get('_detail_status', {}).get(field) == 'ok' for r in grp)
        if has_ok:
            return 'ok'
        has_err = any(r.get('_detail_status', {}).get(field) == 'error' for r in grp)
        return 'error' if has_err else 'missing'

    for grp in groups.values():
        platform_data = {}
        for code in PLATFORM_ORDER:
            candidates = [r for r in grp if r.get('platform_code') == code]
            if not candidates:
                continue
            # v4.21：每平台 best 也要「完整命中优先 + 命中维数 + 分数」兜底，
            # 否则同分时 _complete_match=False 的版本会盖住 True 的版本
            best = max(candidates, key=lambda r: (
                1 if r.get('_complete_match') else 0,
                r.get('_aux_hits', 0),
                r.get('match_score', 0),
                -(order_idx.get(code, 99)),
            ))
            platform_data[code] = best
        if not platform_data:
            continue

        # 基本信息：歌名/歌手取出现次数最多的原始值，避免把「伴奏版」等后缀当成主版本。
        # v4.25：次数相同时优先选更长/更完整的原始歌名（平台写法差异如半角/全角括号
        # 归一后同组，选信息量更大的那条展示），并对残缺括号歌名（如「爱旧爱 (女声」缺右括号）
        # 做修复，保证搜索页展示完整歌曲名。
        from collections import Counter
        name_counts = Counter(r.get('song_name', '') for r in grp if r.get('song_name'))
        if name_counts:
            song_name = max(name_counts.items(), key=lambda kv: (kv[1], len(kv[0])))[0]
        else:
            song_name = ''
        song_name = _repair_paren(song_name)
        artist_counts = Counter(r.get('performer', '') for r in grp if r.get('performer'))
        performer = artist_counts.most_common(1)[0][0] if artist_counts else ''

        album, album_src = _pick_common(grp, 'album')
        release_date, _ = _pick_common(grp, 'release_date')
        lyricist, lyr_src = _pick_common(grp, 'lyricist')
        composer, comp_src = _pick_common(grp, 'composer')
        record_label, label_src = _pick_common(grp, 'record_label')

        # 5 平台各自抓到的词曲作者明细——前端用来展示各家差异（同名歌不同平台
        # 词曲不一致时主列只显示一个聚合值，差异提示 + tooltip 展开看各家）
        lyricist_pp = _collect_field_per_platform(grp, 'lyricist')
        composer_pp = _collect_field_per_platform(grp, 'composer')

        # v4.17.6：把「完整命中」标志透传到聚合层。
        # 单条结果在 _score_result 里写了 _complete_match（歌名强匹配 + 用户填的
        # 辅助字段全命中），但旧版聚合时没带上，导致 _sort_by_relevance 精心设计的
        # 「完整命中排第一」在分组后完全失效（前端拿到的 grouped 里 complete=None）。
        # 组内任一平台完整命中，即认为这一组是完整命中。
        group_complete = any(r.get('_complete_match') for r in grp)
        # v4.21：聚合层也透传「命中维数」（取组内最大值 = 真正命中的版本）
        group_aux_hits = max((r.get('_aux_hits', 0) for r in grp), default=0)
        group_aux_total = max((r.get('_aux_total', 0) for r in grp), default=0)

        # best_r 也要优先取完整命中的那条，否则 match_label 可能被同分的
        # 非完整命中结果覆盖（显示「近似匹配」而实际是精准命中）。
        best_r = max(platform_data.values(),
                     key=lambda r: (1 if r.get('_complete_match') else 0,
                                    r.get('match_score', 0)))
        grouped.append({
            'song_name': song_name,
            'performer': performer,
            'album': album or '',
            'release_date': release_date or '',
            'lyricist': lyricist,
            'composer': composer,
            'record_label': record_label,
            '_detail_status': {
                'lyricist': _common_status(grp, 'lyricist'),
                'composer': _common_status(grp, 'composer'),
                'record_label': _common_status(grp, 'record_label'),
            },
            '_source_platform': {
                'lyricist': lyr_src,
                'composer': comp_src,
                'record_label': label_src,
            },
            # 5 平台各自抓到的实际值（去重保序，按 PLATFORM_ORDER）；前端用
            # 来判断各家是否一致 + 在 tooltip 展开明细
            '_per_platform': {
                'lyricist': lyricist_pp,
                'composer': composer_pp,
            },
            'platform_data': platform_data,
            'platform_count': len(platform_data),
            'match_score': best_r.get('match_score', 0),
            'match_label': best_r.get('match_label', ''),
            '_complete_match': group_complete,
            '_aux_hits': group_aux_hits,
            '_aux_total': group_aux_total,
            # v4.22：供前端「合并歌曲」功能标识每条结果（与标记系统同一套 key）
            'mark_key': _mark_key(song_name, performer, album),
        })

    # v4.17.6 排序键：完整命中优先 → 匹配分 → 覆盖平台数 → 歌名
    # v4.21 升级：完整命中 → 命中维数 → 匹配分 → 覆盖平台数 → 歌名
    # （not True=0 排前面，与 _sort_by_relevance 保持一致的写法）
    grouped.sort(key=lambda g: (not g.get('_complete_match'),
                                -g.get('_aux_hits', 0),
                                -g['match_score'],
                                -g['platform_count'],
                                g['song_name']))

    # ── v4.15 反向匹配已固定歌曲：按 _version_key 与当前用户的 pin 列表比对 ──
    # 命中则挂 _pinned={pin_id,pinned_at,version_key} 给前端加🔖金标 + 自动展开📌按钮
    try:
        pins = _load_pins() or []
        if pins:
            pin_by_key = {}
            for p in pins:
                k = p.get('version_key')
                if k:
                    pin_by_key.setdefault(k, []).append(p)
            for g in grouped:
                try:
                    gk = _version_key(g)
                except Exception:
                    gk = ''
                hits = pin_by_key.get(gk) if gk else None
                if hits:
                    g['_pinned'] = {
                        'pin_id': hits[0].get('version_key', ''),
                        'pinned_at': hits[0].get('pinned_at', ''),
                        'version_key': gk,
                    }
                else:
                    g['_pinned'] = False
        else:
            for g in grouped:
                g['_pinned'] = False
    except Exception:
        for g in grouped:
            g['_pinned'] = False

    return grouped


# ═══════════════════════════════════════════════════
#  路由
# ═══════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('index.html')


# 注：monitor.html / evidence.html 是首页页签片段（index.html 已 {% include %}），
# 不是独立页面，因此没有 /monitor、/evidence 顶层路由 —— 访问入口是首页顶部的页签。

# ── 平台表现监控模块（独立子包 monitor/，挂载失败不影响主程序）──────
try:
    from monitor import routes as _monitor_routes
    _monitor_routes.register(app)
    print('[MusicFinder] 监控模块已挂载 → /api/monitor/*')
except Exception as _mon_err:                                # pragma: no cover
    print(f'[MusicFinder] 监控模块挂载失败（主功能不受影响）: {_mon_err}')

# ── 音乐证据监测模块（独立子包 evidence/，挂载失败不影响主程序）──────
try:
    from evidence import routes as _evidence_routes
    _evidence_routes.register(app)
    print('[MusicFinder] 证据监测模块已挂载 → /api/evidence/*')
except Exception as _ev_err:                                # pragma: no cover
    print(f'[MusicFinder] 证据监测模块挂载失败（主功能不受影响）: {_ev_err}')

# ── 批量任务 v2（独立模块 batch_v2.py，SQLite 持久化 + 后台 worker + 断点续跑）──────
# 解决原 /api/batch_search 一锅端跑 2 万首会超时/丢结果/不能关浏览器的痛点
# 范围精简：仅 3 平台（QQ/酷狗/网易云）+ 2 字段（收藏量+单曲链接）
try:
    import batch_v2 as _batch_v2
    # ⚠️ Flask use_reloader=True 时父进程也会完整执行本文件一遍。
    # 若不加判定，批量引擎会被创建两份 → 补跑线程 ×2、续跑任务 ×2、
    # 对 QQ/酷狗的请求量凭空翻倍，是触发平台限流的隐形推手。
    # 只有「真正对外服务的那个进程」才启动后台 worker：
    #   · reloader 子进程：WERKZEUG_RUN_MAIN == 'true'
    #   · 打包冻结版（无 reloader）：sys._MEIPASS 存在
    #   · 其它情况（reloader 父进程 / 被当模块 import）只注册路由，不起线程
    _bv2_frozen = getattr(sys, '_MEIPASS', None) is not None
    _bv2_serving = (
        os.environ.get('WERKZEUG_RUN_MAIN') == 'true'
        or (_bv2_frozen and __name__ == '__main__')
        or os.environ.get('MF_BATCH_WORKERS') == '1'
    )
    _batch_v2.register(app, start_workers=_bv2_serving)
    print(f'[MusicFinder] 批量任务 v2 已挂载 → /api/batch_v2_*（后台worker={"开" if _bv2_serving else "关"}）')
except Exception as _bv2_err:                               # pragma: no cover
    print(f'[MusicFinder] 批量任务 v2 挂载失败（旧批量功能不受影响）: {_bv2_err}')

# 标记云端历史同步与 blob 同步的启动线程已移至文件末尾 `if __name__ == '__main__':` 块，
# 因为这两个函数定义在本段之后，模块顶层 import 时直接求值 target= 会 NameError。


# ═══════════════════════════════════════════════════
#  歌曲手动标记（本地持久化，跨版本更新保留）
#  数据存于用户目录（包外），升级 app 不会丢失
# ═══════════════════════════════════════════════════

# 与 Cookie 同处用户目录 ~/.musicfinder，打包后位于 .app 之外，升级不丢
def _resolve_marks_dir():
    """标记存储目录可配置，支持指向云盘同步文件夹实现多设备/多人共享。

    优先级：环境变量 MUSICFINDER_MARKS_DIR > ~/.musicfinder/config.json 的 marks_dir > 默认 ~/.musicfinder
    """
    env = os.environ.get('MUSICFINDER_MARKS_DIR', '').strip()
    if env:
        return os.path.expanduser(env)
    cfg_path = os.path.join(os.path.expanduser('~/.musicfinder'), 'config.json')
    try:
        if os.path.exists(cfg_path):
            with open(cfg_path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            d = (cfg.get('marks_dir') or '').strip()
            if d:
                return os.path.expanduser(d)
    except Exception:
        logger.exception('读取 config.json 失败')
    return os.path.expanduser('~/.musicfinder')


_MARKS_DIR = _resolve_marks_dir()
_MARKS_FILE = os.path.join(_MARKS_DIR, 'song_marks.json')
_marks_lock = threading.Lock()

# ── 标记云端同步：本地优先 + 后台异步推送（避免云端慢拖垮本地打标） ──
_cloud_push_pool = concurrent.futures.ThreadPoolExecutor(max_workers=1, thread_name_prefix='marks-cloud')


def _push_marks_to_cloud(marks):
    """后台异步把最新 marks 推到云端；失败只记日志，本地已落盘不丢。"""
    try:
        _save_marks_cloudbase(dict(marks))
    except Exception:
        logger.exception('云端标记同步失败（本地已保存，稍后重试）')


def _pull_cloud_marks_once():
    """后台拉一次云端标记合并进本地，保证历史/多设备标记不丢（只做一次）。

    v4.7.1：已按 username 隔离，登录/未登录都安全执行（未登录取 legacy 行）。
    """
    try:
        if not _cloudbase_cfg():
            return
        cloud = _load_marks_cloudbase()
        if not cloud:
            return
        with _marks_lock:
            local = _load_marks_local()
            merged = _merge_marks(local, cloud)
            if merged != local:
                _save_marks_local(merged, merge_with_disk=False)
                logger.info('启动拉取云端标记 %s 条，合并进本地', len(cloud))
    except Exception:
        logger.exception('启动拉取云端标记失败，使用本地标记')


def _marks_backend_cfg():
    """读取标记存储后端配置。

    返回 {'backend': 'local'|'supabase'|'leancloud'|'cloudbase', 'supabase': {...}, 'leancloud': {...}, 'cloudbase': {...}}。
    配置来源：~/.musicfinder/config.json 的 marks_backend / supabase / leancloud / cloudbase 字段。
    默认 local（本地文件），未配置或不合法时回落 local。

    v4.26 零配置 fallback：未显式声明其他后端（如 local/supabase/leancloud）时，
    若能从 config.json cloudbase 字段 / 环境变量 / config.py 默认值 取到云端凭据，
    则自动启用 cloudbase 后端——实现「新设备零配置自动上云」
    （标记、歌单、红心、固定等全部同步到云端）。仅在用户显式写 marks_backend='local' 等时才尊重纯本地。
    """
    backend = 'local'
    supabase = {}
    leancloud = {}
    cloudbase = {}
    explicit = False
    cfg_path = os.path.join(os.path.expanduser('~/.musicfinder'), 'config.json')
    try:
        if os.path.exists(cfg_path):
            with open(cfg_path, 'r', encoding='utf-8') as f:
                d = json.load(f)
            explicit = bool((d.get('marks_backend') or '').strip())
            backend = (d.get('marks_backend') or 'local').strip().lower()
            supabase = d.get('supabase') or {}
            leancloud = d.get('leancloud') or {}
            cloudbase = d.get('cloudbase') or {}
    except Exception:
        logger.exception('读取标记后端配置失败，回落 local')
    if backend not in ('local', 'supabase', 'leancloud', 'cloudbase'):
        backend = 'local'
    # 零配置 fallback：未显式选其他后端时，尝试用云端默认值补全凭据并启用 cloudbase
    if backend == 'local' and not explicit:
        url = (cloudbase.get('url') or '').strip().rstrip('/') \
            or (getattr(config, 'CLOUDBASE_URL', '') or '').strip().rstrip('/') \
            or (os.environ.get('MUSICFINDER_CLOUDBASE_URL') or '').strip().rstrip('/')
        token = (cloudbase.get('token') or '').strip() \
            or (getattr(config, 'CLOUDBASE_TOKEN', '') or '').strip() \
            or (os.environ.get('MUSICFINDER_CLOUDBASE_TOKEN') or '').strip()
        if url and token:
            backend = 'cloudbase'
            cloudbase = {'url': url, 'token': token}
    return {'backend': backend, 'supabase': supabase, 'leancloud': leancloud, 'cloudbase': cloudbase}


def _effective_marks_paths():
    """返回 (marks_file, marks_dir)，按当前登录用户隔离；未登录走旧路径（向后兼容）。"""
    ud = _user_dir()
    if ud:
        return os.path.join(ud, 'song_marks.json'), ud
    return _MARKS_FILE, _MARKS_DIR


def _read_marks_raw():
    """只读磁盘文件，不做 migrate（避免递归）；损坏/缺失返回 {}。供写回合并用。"""
    try:
        mf, _ = _effective_marks_paths()
        if not os.path.exists(mf):
            return {}
        with open(mf, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data.get('marks', {}) or {}
    except Exception:
        logger.exception('读取标记文件失败')
    return {}


def _merge_marks(base, incoming):
    """键级合并（支持云盘多设备/多人协作防覆盖）。

    base = 磁盘最新版；incoming = 本次改动。同 key 按 updated_at 取较新，不同 key 全部保留。
    """
    merged = dict(base or {})
    for k, v in (incoming or {}).items():
        if k in merged:
            if str(v.get('updated_at', '')) >= str(merged[k].get('updated_at', '')):
                merged[k] = v
        else:
            merged[k] = v
    return merged


def _mark_key(song_name, performer, album):
    """标记用的稳定 key，与聚合分组逻辑保持一致（人名集合匹配）。

    用 _normalize_exact（保留括号内容），使「笑柄 / 笑柄(伴奏) / 笑柄(DJ阿卓版)」
    成为三条互不干扰的独立标记 —— 版权方按版本逐一确权，不能塌缩合并。
    改动前用的是 _normalize_match，会把 71 个版本合并成 26 条。
    """
    return json.dumps({
        'song': _normalize_exact(song_name or ''),
        'artist': list(_split_artist_names(performer or '')),
        'album': _normalize_exact(album or ''),
    }, ensure_ascii=False)


# 标记 key 的版本号。v1 = 旧口径（_normalize_match，剥括号，多版本塌缩）；
# v2 = 新口径（_normalize_exact，保留括号，版本一对一）。
# 文件里 key_version 不等于当前值时自动重算全部 key，避免老标记失联。
_MARK_KEY_VERSION = 2


def _migrate_mark_keys(marks):
    """用记录自身的 song_name/performer/album 重算 key，把旧口径标记迁到新口径。

    极少数情况下两条旧记录会算出同一个新 key（例如内容完全相同），
    保留 updated_at 较新的那条，不静默丢标签。
    """
    rebuilt = {}
    for _old_key, v in (marks or {}).items():
        if not isinstance(v, dict) or not v.get('song_name'):
            continue
        nk = _mark_key(v.get('song_name', ''), v.get('performer', ''), v.get('album', ''))
        prev = rebuilt.get(nk)
        if prev and str(prev.get('updated_at', '')) >= str(v.get('updated_at', '')):
            continue
        rebuilt[nk] = v
    return rebuilt


def _load_marks_local():
    try:
        mf, _ = _effective_marks_paths()
        if not os.path.exists(mf):
            return {}
        with open(mf, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {}
        marks = data.get('marks', {}) or {}
        if data.get('key_version') != _MARK_KEY_VERSION:
            migrated = _migrate_mark_keys(marks)
            logger.info('标记 key 升级至 v%s：%s 条 → %s 条',
                        _MARK_KEY_VERSION, len(marks), len(migrated))
            _save_marks_local(migrated, merge_with_disk=False)  # migrate 不需再合并旧 key
            return migrated
        return marks
    except Exception:
        logger.exception('读取标记文件失败')
        return {}


# ── v4.22 歌曲合并映射（同一首歌因艺名写法不同被拆成多条时，手动合并成一条）──
# 持久层：~/.musicfinder/merge_map.json
#   groups: { canonical_mark_key: { canonical, aliases:[...], created_at, note, alias_marks_backup } }
# 标记策略：以「正主」(canonical) 为准——合并时把各 alias 的独立标记迁移到 canonical
#           （若 canonical 无标记则继承 alias 的；若已有则保留 canonical），并备份 alias
#           原标记到 alias_marks_backup，便于「撤销合并」时还原。
_MERGE_MAP_FILE = os.path.join(_MARKS_DIR, 'merge_map.json')
_merge_map_lock = threading.Lock()


def _load_merge_map():
    try:
        if not os.path.exists(_MERGE_MAP_FILE):
            return {'version': 1, 'groups': {}}
        with open(_MERGE_MAP_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {'version': 1, 'groups': {}}
        return {'version': data.get('version', 1), 'groups': data.get('groups', {}) or {}}
    except Exception:
        logger.exception('读取合并映射失败')
        return {'version': 1, 'groups': {}}


def _save_merge_map(mm):
    try:
        os.makedirs(os.path.dirname(_MERGE_MAP_FILE), exist_ok=True)
        with open(_MERGE_MAP_FILE, 'w', encoding='utf-8') as f:
            json.dump(mm, f, ensure_ascii=False, indent=2)
    except Exception:
        logger.exception('写入合并映射失败')


def _now_str():
    return time.strftime('%Y-%m-%d %H:%M:%S')


def _do_merge(canonical_key, alias_keys, note=''):
    """把若干 alias 合并到 canonical（以正主为准迁移标记）。返回 (ok, msg)。"""
    if not canonical_key or not alias_keys:
        return False, '参数缺失'
    alias_keys = [k for k in alias_keys if k and k != canonical_key]
    if not alias_keys:
        return False, '至少需要一个待合并项'
    with _merge_map_lock:
        mm = _load_merge_map()
        # 收集所有涉及 key 的已有 aliases（打破旧分组，重组成新组，避免嵌套合并）
        involved = set([canonical_key] + alias_keys)
        for gk, g in list(mm['groups'].items()):
            allk = set([gk] + list(g.get('aliases', [])))
            if allk & involved:
                involved |= allk
        new_aliases = sorted(involved - set([canonical_key]))
        # 标记迁移（以正主为准）
        with _marks_lock:
            marks = _load_marks_local()
        backup = {}
        for ak in new_aliases:
            am = marks.get(ak)
            if am:
                backup[ak] = am
                if not marks.get(canonical_key):
                    nm = dict(am)
                    nm['merged_from'] = ak
                    nm['updated_at'] = nm.get('updated_at') or _now_str()
                    marks[canonical_key] = nm
                marks.pop(ak, None)  # 不论 canonical 是否已有标记，都删 alias 独立标记
        if backup:
            with _marks_lock:
                _save_marks_local(marks, merge_with_disk=False)
        mm['groups'][canonical_key] = {
            'canonical': canonical_key,
            'aliases': new_aliases,
            'created_at': _now_str(),
            'note': note or '',
            'alias_marks_backup': backup,
        }
        _save_merge_map(mm)
    return True, 'ok'


def _do_unmerge(key):
    """撤销合并。key 为 canonical 或 alias 均可。返回 (ok, msg)。"""
    if not key:
        return False, '参数缺失'
    with _merge_map_lock:
        mm = _load_merge_map()
        target = None
        for gk, g in mm['groups'].items():
            if gk == key or key in g.get('aliases', []):
                target = g
                break
        if not target:
            return False, '未找到合并组'
        # 还原 alias 标记（防丢失）
        backup = target.get('alias_marks_backup', {}) or {}
        with _marks_lock:
            marks = _load_marks_local()
        for ak, am in backup.items():
            if am:
                marks[ak] = am
        if key == target['canonical']:
            mm['groups'].pop(target['canonical'], None)
        else:
            target['aliases'] = [a for a in target['aliases'] if a != key]
            if not target['aliases']:
                mm['groups'].pop(target['canonical'], None)
            else:
                mm['groups'][target['canonical']] = target
        _save_marks_local(marks, merge_with_disk=False)
        _save_merge_map(mm)
    return True, 'ok'


def _apply_merge_map(grouped):
    """搜索结果展示层应用合并：把 alias group 的平台数据/字段并入 canonical group。"""
    try:
        mm = _load_merge_map()
        groups = mm.get('groups') or {}
        if not groups:
            return grouped
        by_key = {}
        for g in grouped:
            k = g.get('mark_key')
            if k:
                by_key.setdefault(k, g)
        removed = set()
        out = []
        for g in grouped:
            k = g.get('mark_key')
            if k in removed:
                continue
            grp = groups.get(k)
            if not grp:
                out.append(g)
                continue
            canon = g
            merged_from = []
            for ak in grp.get('aliases', []):
                ag = by_key.get(ak)
                if not ag or ag is canon:
                    continue
                pd = canon.setdefault('platform_data', {})
                for code, r in (ag.get('platform_data') or {}).items():
                    if code not in pd:
                        pd[code] = r
                for fld in ('performer', 'lyricist', 'composer', 'album', 'record_label', 'release_date'):
                    if not canon.get(fld) and ag.get(fld):
                        canon[fld] = ag[fld]
                if ag.get('hearts'):
                    hv = canon.get('hearts') or []
                    for h in ag['hearts']:
                        if h not in hv:
                            hv.append(h)
                    canon['hearts'] = hv
                pp = canon.setdefault('_per_platform', {})
                for fld in ('lyricist', 'composer'):
                    src = (ag.get('_per_platform') or {}).get(fld)
                    if src and fld not in pp:
                        pp[fld] = src
                merged_from.append('%s - %s' % (ag.get('song_name', ''), ag.get('performer', '')))
                removed.add(ak)
            if merged_from:
                canon['_merged'] = True
                canon['_merged_from'] = merged_from
                canon['_merge_canonical_key'] = k
                canon['platform_count'] = len(canon.get('platform_data') or {})
            out.append(canon)
        return out
    except Exception:
        logger.exception('应用合并映射失败')
        return grouped


def _save_marks_local(marks, merge_with_disk=True):
    """写回前先与磁盘最新版合并（支持云盘多设备/多人协作，避免互相覆盖）。

    merge_with_disk=False 用于 key 升级 migrate（此时 marks 已是重算后的真理，不应再混入旧 key）。

    v4.13 改动：mark_type 不再强制归一为 4 类，直接保存用户原值（已确认/已标记/
    正版确认/疑似盗版/待核实/我的精选/已排除/未打盗版/已打盗版/正版@xx/盗版@xx），
    搜索页和已标记页保持显示一致。tags[] 附加标签按列表存。
    """
    # 1. 字段轻清洗（mark_type 兜底、tags 强转 list）
    for k, v in (marks or {}).items():
        if not isinstance(v, dict):
            continue
        v['mark_type'] = _normalize_mark_type(v.get('mark_type', ''))
        # tags 字段：list 化（兼容老数据缺失/字符串）
        tags = v.get('tags')
        if tags is None:
            v['tags'] = []
        elif isinstance(tags, str):
            v['tags'] = [t.strip() for t in tags.split(',') if t.strip()] if tags else []
        elif not isinstance(tags, list):
            v['tags'] = []
        # 清洗每个 tag（去空白、去空串、去重保序）
        seen = set()
        cleaned = []
        for t in v['tags']:
            t = (t or '').strip()
            if not t or t in seen:
                continue
            seen.add(t)
            cleaned.append(t)
        v['tags'] = cleaned
    # 2. 与磁盘合并（如需）
    mf, md = _effective_marks_paths()
    os.makedirs(md, exist_ok=True)
    if merge_with_disk:
        disk_marks = _read_marks_raw()
        if disk_marks:
            marks = _merge_marks(disk_marks, marks)
            # 合并进来的磁盘数据也走一次字段清洗
            for k, v in (marks or {}).items():
                if not isinstance(v, dict):
                    continue
                v['mark_type'] = _normalize_mark_type(v.get('mark_type', ''))
                tags = v.get('tags')
                if tags is None:
                    v['tags'] = []
                elif isinstance(tags, str):
                    v['tags'] = [t.strip() for t in tags.split(',') if t.strip()] if tags else []
                elif not isinstance(tags, list):
                    v['tags'] = []
                seen = set()
                cleaned = []
                for t in v['tags']:
                    t = (t or '').strip()
                    if not t or t in seen:
                        continue
                    seen.add(t)
                    cleaned.append(t)
                v['tags'] = cleaned
    tmp = mf + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump({'key_version': _MARK_KEY_VERSION, 'marks': marks},
                  f, ensure_ascii=False, indent=2)
    os.replace(tmp, mf)  # 原子替换，避免写一半损坏


# ── Supabase 云端后端（方案2：多设备/多人实时共享标记） ──
def _supabase_cfg():
    cfg = _marks_backend_cfg()
    if cfg['backend'] != 'supabase':
        return None
    sb = cfg.get('supabase') or {}
    url = (sb.get('url') or '').strip().rstrip('/')
    # 兼容：用户可能粘了带 /rest/v1/ 的 REST 端点地址，统一剥掉避免拼出 /rest/v1/rest/v1
    if url.endswith('/rest/v1'):
        url = url[:-len('/rest/v1')].rstrip('/')
    key = (sb.get('key') or '').strip()
    table = (sb.get('table') or 'song_marks').strip()
    if not url or not key:
        logger.warning('supabase 后端已选但未配置 url/key，回落本地')
        return None
    return {'url': url, 'key': key, 'table': table}


def _load_marks_supabase():
    sb = _supabase_cfg()
    if not sb:
        return _load_marks_local()
    url = '%s/rest/v1/%s?select=mark_key,data' % (sb['url'], sb['table'])
    headers = {'apikey': sb['key'], 'Authorization': 'Bearer %s' % sb['key']}
    # 与项目内 QQ/网易云等外部 API 一致：绕过沙箱代理（supabase 走代理会 502）
    try:
        r = requests.get(url, headers=headers, timeout=15,
                         proxies={'http': None, 'https': None})
        r.raise_for_status()
        out = {}
        for row in r.json():
            k = row.get('mark_key')
            d = row.get('data') or {}
            if k and isinstance(d, dict):
                out[k] = d
        return out
    except Exception:
        logger.exception('从 Supabase 读取标记失败，回落本地')
        return _load_marks_local()


def _save_marks_supabase(marks):
    sb = _supabase_cfg()
    if not sb:
        _save_marks_local(marks, merge_with_disk=False)
        return
    headers = {
        'apikey': sb['key'],
        'Authorization': 'Bearer %s' % sb['key'],
        'Content-Type': 'application/json',
        'Prefer': 'resolution=merge-duplicates',
    }
    url = '%s/rest/v1/%s' % (sb['url'], sb['table'])
    rows = [{
        'mark_key': k,
        'song_name': v.get('song_name', ''),
        'performer': v.get('performer', ''),
        'album': v.get('album', ''),
        'mark_type': v.get('mark_type', ''),
        'note': v.get('note', ''),
        'data': v,
    } for k, v in marks.items()]
    # 与项目内其他外部 API 一致：绕过沙箱代理
    no_proxy = {'http': None, 'https': None}
    try:
        resp = requests.post(url, headers=headers, json=rows, timeout=20,
                             proxies=no_proxy)
        resp.raise_for_status()
    except Exception:
        logger.exception('写入 Supabase 失败，回落本地保存')
        _save_marks_local(marks, merge_with_disk=False)


def _delete_one_mark_supabase(mark_key):
    """单独删除云端一条 mark（只用于用户显式删除，避免与并发写互相清空）。"""
    sb = _supabase_cfg()
    if not sb:
        return
    headers = {
        'apikey': sb['key'],
        'Authorization': 'Bearer %s' % sb['key'],
    }
    url = '%s/rest/v1/%s?mark_key=eq.%s' % (sb['url'], sb['table'], mark_key)
    try:
        requests.delete(url, headers=headers, timeout=15,
                        proxies={'http': None, 'https': None})
    except Exception:
        logger.exception('从 Supabase 删除单条 mark 失败')


# ── LeanCloud 云端后端（国内版，大陆网可直连；方案2 的国内替代） ──
def _leancloud_cfg():
    cfg = _marks_backend_cfg()
    if cfg['backend'] != 'leancloud':
        return None
    lc = cfg.get('leancloud') or {}
    app_id = (lc.get('app_id') or '').strip()
    app_key = (lc.get('app_key') or '').strip()
    api_base = (lc.get('api_base') or '').strip().rstrip('/')
    # 兼容：只给 app_id 时自动拼默认国内域名；也可在配置里写自定义备案域名
    if not api_base and app_id:
        api_base = 'https://%s.api.lncld.net/1.1' % app_id
    cls = (lc.get('class') or 'song_marks').strip()
    if not app_id or not app_key:
        logger.warning('leancloud 后端已选但未配置 app_id/app_key，回落本地')
        return None
    return {'app_id': app_id, 'app_key': app_key, 'api_base': api_base, 'class': cls}


def _leancloud_headers(lc):
    return {
        'X-LC-Id': lc['app_id'],
        'X-LC-Key': lc['app_key'],
        'Content-Type': 'application/json',
    }


def _leancloud_query(lc, where=None):
    """查询 LeanCloud 存储类，返回 results 列表（含 objectId）。"""
    url = '%s/classes/%s' % (lc['api_base'], lc['class'])
    params = {'limit': 1000}
    if where:
        params['where'] = json.dumps(where, ensure_ascii=False)
    r = requests.get(url, headers=_leancloud_headers(lc), params=params, timeout=15,
                     proxies={'http': None, 'https': None})
    r.raise_for_status()
    return r.json().get('results', [])


def _load_marks_leancloud():
    lc = _leancloud_cfg()
    if not lc:
        return _load_marks_local()
    try:
        rows = _leancloud_query(lc)
        out = {}
        for row in rows:
            k = row.get('mark_key')
            d = row.get('data') or {}
            if k and isinstance(d, dict):
                out[k] = d
        return out
    except Exception:
        logger.exception('从 LeanCloud 读取标记失败，回落本地')
        return _load_marks_local()


def _save_marks_leancloud(marks):
    """全量同步：upsert 本地集合到云端（云端即真理），并删除本地已取消的多余项。

    写只 upsert / 删只单点，避免多设备并发互相清空（同 supabase 修复思路）。
    用 LeanCloud 批量接口（/1.1/batch）分批发，规避免费版请求频率限制。
    """
    lc = _leancloud_cfg()
    if not lc:
        _save_marks_local(marks, merge_with_disk=False)
        return
    headers = _leancloud_headers(lc)
    no_proxy = {'http': None, 'https': None}
    base = lc['api_base']
    cls = lc['class']
    try:
        # 1. 拉云端现有 objectId 映射
        rows = _leancloud_query(lc)
        cloud_by_key = {row.get('mark_key'): row.get('objectId')
                        for row in rows if row.get('mark_key')}
        cloud_keys = set(cloud_by_key.keys())
        # 2. 构造 batch 操作（upsert + 删除多余）
        ops = []
        for k, v in marks.items():
            payload = {
                'mark_key': k,
                'song_name': v.get('song_name', ''),
                'performer': v.get('performer', ''),
                'album': v.get('album', ''),
                'mark_type': v.get('mark_type', ''),
                'note': v.get('note', ''),
                'data': v,
            }
            if k in cloud_by_key:
                ops.append({'method': 'PUT',
                             'path': '/1.1/classes/%s/%s' % (cls, cloud_by_key[k]),
                             'body': payload})
            else:
                ops.append({'method': 'POST',
                             'path': '/1.1/classes/%s' % cls,
                             'body': payload})
        for ck in cloud_keys - set(marks.keys()):
            ops.append({'method': 'DELETE',
                        'path': '/1.1/classes/%s/%s' % (cls, cloud_by_key[ck])})
        # 3. 分批发送（每批 20，避免超限）
        for i in range(0, len(ops), 20):
            batch = ops[i:i + 20]
            r = requests.post('%s/batch' % base, headers=headers, json=batch,
                              timeout=30, proxies=no_proxy)
            r.raise_for_status()
    except Exception:
        logger.exception('写入 LeanCloud 失败，回落本地保存')
        _save_marks_local(marks, merge_with_disk=False)


def _delete_one_mark_leancloud(mark_key):
    """单独删除云端一条 mark（只用于用户显式删除，避免与并发写互相清空）。"""
    lc = _leancloud_cfg()
    if not lc:
        return
    try:
        rows = _leancloud_query(lc, where={'mark_key': mark_key})
        for row in rows:
            oid = row.get('objectId')
            if oid:
                requests.delete('%s/classes/%s/%s' % (lc['api_base'], lc['class'], oid),
                                headers=_leancloud_headers(lc), timeout=15,
                                proxies={'http': None, 'https': None})
    except Exception:
        logger.exception('从 LeanCloud 删除单条 mark 失败')


# ===== 腾讯云 CloudBase 后端（国内节点，免备案，免费额度） =====
# 接入方式：用户创建 CloudBase 环境 + 一个云函数 + HTTP 触发器；云函数用 @cloudbase/node-sdk
# 操作 song_marks 集合（自带 token 校验，只让 MusicFinder 调）。MusicFinder 这边只 POST
# action + payload 到云函数的 HTTP 触发 URL，鉴权通过 url 头里的 shared token。
# 字段：
#   url: 形如 https://{envId}.app.tcloudbase.com/{path}（HTTP 触发器自动分配的访问路径）
#   token: 用户在云函数里写死的共享密钥（避免他人随便调你的云函数）
#   class: 集合名，默认 song_marks

def _cloudbase_cfg():
    cfg = _marks_backend_cfg()
    if cfg['backend'] != 'cloudbase':
        return None
    cb = cfg.get('cloudbase') or {}
    url = (cb.get('url') or '').strip().rstrip('/')
    token = (cb.get('token') or '').strip()
    cls = (cb.get('class') or 'song_marks').strip()
    if not url or not token:
        logger.warning('cloudbase 后端已选但未配置 url/token，回落本地')
        return None
    return {'url': url, 'token': token, 'class': cls}


def _cloudbase_call(cb, action, **kwargs):
    """POST 一个 action 到云函数；云函数代理 CRUD。"""
    body = {'action': action, 'token': cb['token'], 'class': cb['class']}
    body.update(kwargs)
    r = requests.post(cb['url'], json=body, timeout=6,
                      proxies={'http': None, 'https': None})
    r.raise_for_status()
    res = r.json()
    if res.get('code') not in (0, None):
        raise RuntimeError(
            'CloudBase 云函数返回错误: %s' % (
                res.get('msg') or res.get('message') or res.get('error') or '未知错误'))
    return res.get('data')


def _load_marks_cloudbase():
    cb = _cloudbase_cfg()
    if not cb:
        return _load_marks_local()
    u = _cur_user() or 'legacy'
    try:
        items = _cloudbase_call(cb, 'get_all') or []
        out = {}
        for item in items:
            if item.get('datatype'):  # 非标记类 blob，跳过
                continue
            if (item.get('username') or 'legacy') != u:  # 只取当前用户的行
                continue
            # 云函数把整条 m 当 data 存，mark_key 真实落在 _id（或 data.mark_key）里
            k = item.get('_id') or (item.get('data') or {}).get('mark_key')
            d = (item.get('data') or {}).get('data')
            if d is None:
                d = item.get('data')  # 兜底：整条当作 data
            if k and isinstance(d, dict):
                out[k] = d
        return out
    except Exception:
        logger.exception('从 CloudBase 读取标记失败，回落本地')
        return _load_marks_local()


def _save_marks_cloudbase(marks):
    """按人隔离全量同步：云端每行带 username，只动当前用户的行（杜绝多用户互删）。"""
    cb = _cloudbase_cfg()
    if not cb:
        _save_marks_local(marks, merge_with_disk=False)
        return
    u = _cur_user() or 'legacy'
    try:
        # 1. 拉云端现有记录，只算当前用户的 mark_keys（跳过 blob 类、其他用户）
        items = _cloudbase_call(cb, 'get_all') or []
        cloud_keys = set()
        for item in items:
            if item.get('datatype'):  # blob 类，不算标记
                continue
            if (item.get('username') or 'legacy') != u:
                continue
            kk = item.get('_id') or (item.get('data') or {}).get('mark_key')
            if kk:
                cloud_keys.add(kk)
        local_keys = set(marks.keys())
        # 2. 构造 upsert 批量（每行带 username）
        payloads = []
        for k, v in marks.items():
            payloads.append({
                'mark_key': k,
                'username': u,
                'song_name': v.get('song_name', ''),
                'performer': v.get('performer', ''),
                'album': v.get('album', ''),
                'mark_type': v.get('mark_type', ''),
                'note': v.get('note', ''),
                'data': v,
            })
        # 分批 50 条一次，避免云函数 body 过大或单次超限
        for i in range(0, len(payloads), 50):
            _cloudbase_call(cb, 'batch_upsert', payloads=payloads[i:i + 50])
        # 3. 只删「当前用户」云端多出的（其他设备已取消的），绝不碰别人的行
        for extra in (cloud_keys - local_keys):
            try:
                _cloudbase_call(cb, 'delete', mark_key=extra)
            except Exception:
                logger.exception('CloudBase 单点删除失败：%s' % extra)
    except Exception:
        logger.exception('写入 CloudBase 失败，回落本地保存')
        _save_marks_local(marks, merge_with_disk=False)


def _delete_one_mark_cloudbase(mark_key):
    """用户显式删除时单点删除云端这一条（不与并发 upsert 互踩）。"""
    cb = _cloudbase_cfg()
    if not cb:
        return
    try:
        _cloudbase_call(cb, 'delete', mark_key=mark_key)
    except Exception:
        logger.exception('从 CloudBase 删除单条 mark 失败')


# ── 通用云端 blob 存储（v4.7.1：歌单/固定/回收站/红心 跨设备同步） ──
# 复用 CloudBase 云函数，每行 doc = {mark_key:'blob:<user>:<datatype>', username, datatype, payload}
# payload 为任意 JSON；读取时按 username+datatype 过滤，各用户互不串台。
_cloud_blob_pool = concurrent.futures.ThreadPoolExecutor(max_workers=1, thread_name_prefix='blob-cloud')


def _cloud_blob_key(datatype):
    return 'blob:%s:%s' % (_cur_user() or 'legacy', datatype)


def _cloud_blob_save(datatype, payload):
    """把一类个人数据推到云端（按人隔离）；失败只记日志，本地已落盘不丢。"""
    cb = _cloudbase_cfg()
    if not cb:
        return
    try:
        key = _cloud_blob_key(datatype)
        doc = {'mark_key': key, 'username': _cur_user() or 'legacy',
               'datatype': datatype, 'payload': payload}
        _cloudbase_call(cb, 'batch_upsert', payloads=[doc])
    except Exception:
        logger.exception('云端保存 %s 失败（本地已存，稍后重试）', datatype)


def _cloud_blob_load(datatype):
    """从云端读一类个人数据；无则返回 None。"""
    cb = _cloudbase_cfg()
    if not cb:
        return None
    u = _cur_user() or 'legacy'
    key = _cloud_blob_key(datatype)
    try:
        items = _cloudbase_call(cb, 'get_all') or []
        for it in items:
            d = it.get('data') or it
            dt = it.get('datatype') or d.get('datatype')
            un = it.get('username') or d.get('username') or 'legacy'
            mk = it.get('mark_key') or d.get('mark_key')
            if dt == datatype and un == u and mk == key:
                return it.get('payload') or d.get('payload')
    except Exception:
        logger.exception('云端读取 %s 失败', datatype)
    return None


def _push_blob_async(datatype, payload):
    try:
        _cloud_blob_pool.submit(_cloud_blob_save, datatype, payload)
    except Exception:
        logger.exception('提交 blob 云端任务失败')


def _pull_cloud_blobs_once():
    """登录/启动时把云端个人数据拉回本地（云端为准，单用户 last-push-wins）。"""
    try:
        if not _cloudbase_cfg():
            return
        pl = _cloud_blob_load('playlist')
        if isinstance(pl, dict) and pl.get('songs') is not None:
            _save_playlist(pl.get('songs') or [])
        pr = _cloud_blob_load('playlist_results')
        if isinstance(pr, dict) and pr.get('items') is not None:
            _save_playlist_results(pr.get('items') or [])
        pins = _cloud_blob_load('pins')
        if isinstance(pins, list):
            _save_pins(pins)
        dl = _cloud_blob_load('deleted')
        if isinstance(dl, dict):
            _save_deleted(set(dl.get('keys', [])), dl.get('log', []))
        hb = _cloud_blob_load('hearts')
        if isinstance(hb, dict):
            _save_hearts(hb)
    except Exception:
        logger.exception('拉取个人云端数据失败')


def _migrate_legacy_cloud_marks(username):
    """首次登录时，把云端无 username 的 legacy 标记归属到该账号（只跑一次，幂等）。"""
    cb = _cloudbase_cfg()
    if not cb:
        return
    try:
        items = _cloudbase_call(cb, 'get_all') or []
        legacy = [it for it in items if not it.get('datatype') and not it.get('username')]
        if not legacy:
            return
        payloads = []
        for it in legacy:
            k = it.get('_id') or (it.get('data') or {}).get('mark_key')
            d = (it.get('data') or {}).get('data') or it.get('data') or {}
            if not k or not isinstance(d, dict):
                continue
            payloads.append({
                'mark_key': k, 'username': username,
                'song_name': d.get('song_name', ''), 'performer': d.get('performer', ''),
                'album': d.get('album', ''), 'mark_type': d.get('mark_type', ''),
                'note': d.get('note', ''), 'data': d,
            })
        for i in range(0, len(payloads), 50):
            _cloudbase_call(cb, 'batch_upsert', payloads=payloads[i:i + 50])
        logger.info('迁移 %d 条 legacy 云端标记到账号 %s', len(payloads), username)
    except Exception:
        logger.exception('迁移 legacy 云端标记失败')


def _load_marks():
    """本地优先加载；云端仅作历史同步源（启动时后台拉一次），不阻塞每次读写。"""
    return _load_marks_local()


def _save_marks(marks, merge_with_disk=True):
    """本地优先写（即时、快）；云端同步丢后台线程，绝不阻塞调用方。

    v4.7.1：登录用户也走云端，按 username 隔离（互删 bug 已修），支持跨设备同步。
    """
    _save_marks_local(marks, merge_with_disk=merge_with_disk)
    try:
        _cloud_push_pool.submit(_push_marks_to_cloud, marks)
    except Exception:
        logger.exception('提交云端同步任务失败')


def _load_marks_cloudbase_all():
    """云端全网视图：admin 专用，不过滤 username，返回 [{...mark, 'owner': 'xxx'}, ...]。
    非管理员调此函数会被拒绝（在路由层 403 拦截，本函数不再冗余检查）。
    """
    cb = _cloudbase_cfg()
    if not cb:
        # 没配云端 → 退化到本地全部（只能看本机的 legacy/自己的）
        with _marks_lock:
            marks = _load_marks()
        return [{'owner': _cur_user() or 'legacy', 'key': k, **v} for k, v in marks.items()
                if not (isinstance(v, dict) and v.get('datatype'))]
    try:
        items = _cloudbase_call(cb, 'get_all') or []
    except Exception:
        logger.exception('云端全网拉取失败')
        return []
    out = []
    for it in items:
        if it.get('datatype'):
            continue
        owner = it.get('username') or (it.get('data') or {}).get('username') or 'legacy'
        k = it.get('_id') or (it.get('data') or {}).get('mark_key')
        d = (it.get('data') or {}).get('data')
        if d is None:
            d = it.get('data') or {}
        if k and isinstance(d, dict) and d.get('song_name'):
            out.append({'owner': owner, 'key': k, **d})
    return out


# ── v4.25.7 共享视图 + 跨账号管理（细粒度权限） ──
def _load_marks_cloudbase_shared():
    """共享视图：聚合所有 view_all_marks=True 的用户标记（默认全员可读）。
    每行带 owner；legacy（无 username 的老数据，属管理员的）始终可见。
    没配云端时退化成只看自己本地标记（无法跨账号看，降级但不报错）。
    """
    cb = _cloudbase_cfg()
    if not cb:
        u = _cur_user() or 'legacy'
        with _marks_lock:
            marks = _load_marks_local()
        return [{'owner': u, 'key': k, **v} for k, v in marks.items()
                if isinstance(v, dict) and v.get('song_name')]
    users = _load_users_auth()
    visible = set()
    for nm, u in users.items():
        pu = u.get('perms')
        if not isinstance(pu, dict):
            pu = _def_perms(nm)
        if pu.get('view_all_marks', True):
            visible.add(nm)
    try:
        items = _cloudbase_call(cb, 'get_all') or []
    except Exception:
        logger.exception('云端共享视图拉取失败')
        return []
    out = []
    for it in items:
        if it.get('datatype'):
            continue
        owner = it.get('username') or (it.get('data') or {}).get('username') or 'legacy'
        if owner != 'legacy' and owner not in visible:
            continue  # 该用户隐藏了自己的标记
        k = it.get('_id') or (it.get('data') or {}).get('mark_key')
        d = (it.get('data') or {}).get('data')
        if d is None:
            d = it.get('data') or {}
        if k and isinstance(d, dict) and d.get('song_name'):
            out.append({'owner': owner, 'key': k, **d})
    return out


def _find_cloud_mark_owner(key):
    """在云端找某 mark_key 的 owner（共享视图改他人标时保留归属）。无则 None。"""
    cb = _cloudbase_cfg()
    if not cb:
        return None
    try:
        items = _cloudbase_call(cb, 'get_all') or []
    except Exception:
        return None
    for it in items:
        if it.get('datatype'):
            continue
        k = it.get('_id') or (it.get('data') or {}).get('mark_key')
        if k == key:
            return it.get('username') or (it.get('data') or {}).get('username') or 'legacy'
    return None


def _get_shared_mark(owner, key):
    """取共享视图里某条已存在的标记完整 dict（用于 PATCH 增量前读旧值）。无则 None。"""
    cb = _cloudbase_cfg()
    if not cb:
        return None
    try:
        items = _cloudbase_call(cb, 'get_all') or []
    except Exception:
        return None
    for it in items:
        if it.get('datatype'):
            continue
        k = it.get('_id') or (it.get('data') or {}).get('mark_key')
        if k != key:
            continue
        o = it.get('username') or (it.get('data') or {}).get('username') or 'legacy'
        if o != owner:
            continue
        d = (it.get('data') or {}).get('data')
        if d is None:
            d = it.get('data') or {}
        return d if isinstance(d, dict) else None
    return None


def _write_owner_local_mark(owner, key, mark_data):
    """直接改某用户本地的 song_marks.json（共享视图下授权用户改他人标时，保持本机一致）。
    mark_data=None 表示删除该 key。返回是否成功落盘。"""
    try:
        ud = _user_dir(owner)
        if not ud:
            return False
        mf = os.path.join(ud, 'song_marks.json')
        raw = {}
        if os.path.exists(mf):
            try:
                with open(mf, 'r', encoding='utf-8') as f:
                    raw = json.load(f)
            except Exception:
                raw = {}
        if not isinstance(raw, dict):
            raw = {}
        marks = raw.get('marks')
        if not isinstance(marks, dict):
            marks = {}
        if mark_data is None:
            if key not in marks:
                return True
            marks.pop(key, None)
        else:
            marks[key] = mark_data
        raw['marks'] = marks
        raw['key_version'] = _MARK_KEY_VERSION
        tmp = mf + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(raw, f, ensure_ascii=False, indent=2)
        os.replace(tmp, mf)
        return True
    except Exception:
        logger.exception('直写用户 %s 本地标记失败', owner)
        return False


def _cloudbase_write_mark(owner, key, mark_data):
    """直接 upsert 一条云端标记（指定 owner，用于共享视图下授权用户改他人标）。"""
    cb = _cloudbase_cfg()
    if not cb:
        return False
    try:
        payloads = [{
            'mark_key': key, 'username': owner,
            'song_name': mark_data.get('song_name', ''), 'performer': mark_data.get('performer', ''),
            'album': mark_data.get('album', ''), 'mark_type': mark_data.get('mark_type', ''),
            'note': mark_data.get('note', ''), 'data': mark_data,
        }]
        _cloudbase_call(cb, 'batch_upsert', payloads=payloads)
        return True
    except Exception:
        logger.exception('云端直写标记失败')
        return False


def _cloudbase_delete_mark_by_key(key):
    cb = _cloudbase_cfg()
    if not cb:
        return False
    try:
        _cloudbase_call(cb, 'delete', mark_key=key)
        return True
    except Exception:
        logger.exception('云端直删标记失败')
        return False


def _apply_shared_mark_write(op, owner, key, mark_data):
    """单条跨账号写：同时落「对方本地文件 + 云端该行」，尽力保持一致。
    op='set'|'delete'。返回是否至少一处成功。"""
    ok = True
    if op == 'delete':
        if not _write_owner_local_mark(owner, key, None):
            ok = False
        if not _cloudbase_delete_mark_by_key(key):
            ok = False
    else:
        if not _write_owner_local_mark(owner, key, mark_data):
            ok = False
        if not _cloudbase_write_mark(owner, key, mark_data):
            ok = False
    return ok


def _apply_shared_mark_writes(ops):
    """批量跨账号写：ops=[(op, owner, key, mark_data), ...]。
    本地逐个落盘；云端 payload 合并成一次 batch_upsert + 逐条 delete。"""
    ups = []
    dels = []
    for op, owner, key, mark_data in ops:
        if op == 'delete':
            _write_owner_local_mark(owner, key, None)
            dels.append(key)
        else:
            _write_owner_local_mark(owner, key, mark_data)
            ups.append({
                'mark_key': key, 'username': owner,
                'song_name': mark_data.get('song_name', ''), 'performer': mark_data.get('performer', ''),
                'album': mark_data.get('album', ''), 'mark_type': mark_data.get('mark_type', ''),
                'note': mark_data.get('note', ''), 'data': mark_data,
            })
    cb = _cloudbase_cfg()
    if cb:
        try:
            for i in range(0, len(ups), 50):
                _cloudbase_call(cb, 'batch_upsert', payloads=ups[i:i + 50])
            for dk in dels:
                try:
                    _cloudbase_call(cb, 'delete', mark_key=dk)
                except Exception:
                    logger.exception('批量删云端标失败 %s', dk)
        except Exception:
            logger.exception('批量云端共享写失败')


@app.route('/api/marks', methods=['GET'])
def api_get_marks():
    """标记列表：?scope=mine|shared（shared 为默认，全员可读）
       - shared（默认）：聚合所有 view_all_marks=True 的账号标记；返回 list[{owner,key,...data}]
       - mine：仅当前用户本地标记；返回 dict{key:data}
       - all：向后兼容别名，等同 shared（不再要求 admin）
    """
    cur = _cur_user() or 'legacy'
    scope = (request.args.get('scope') or 'shared').strip().lower()
    if scope == 'all':
        scope = 'shared'
    can_manage = bool(_can_user(cur, 'manage_marks'))
    if scope == 'shared':
        items = _load_marks_cloudbase_shared()
        return jsonify({'marks': items, 'scope': 'shared', 'is_admin': _is_admin(cur), 'can_manage': can_manage})
    with _marks_lock:
        marks = _load_marks()
    return jsonify({'marks': marks, 'scope': 'mine', 'is_admin': _is_admin(cur), 'can_manage': can_manage})


@app.route('/api/marks/stats', methods=['GET'])
def api_marks_stats():
    """标签分布统计：?scope=mine|shared（shared 为默认，全员可读）
    返回 {'scope': ..., 'by_label': {label: count}, 'total': N, 'unique_songs': N, 'tags_total': N}
    - by_label：主标签 (mark_type) 的计数；"未标记"也算一类（空值归到这里），方便筛选
    - tags_total：所有附加标签的总出现次数（用于趋势参考）
    """
    cur = _cur_user() or 'legacy'
    scope = (request.args.get('scope') or 'shared').strip().lower()
    if scope == 'all':
        scope = 'shared'
    if scope == 'shared':
        items = _load_marks_cloudbase_shared()
    else:
        with _marks_lock:
            marks = _load_marks()
        items = list(marks.values())
    by_label = {}
    unique_songs = 0
    tags_total = 0
    for v in items:
        if not isinstance(v, dict):
            continue
        unique_songs += 1
        mt = (v.get('mark_type') or '').strip() or '未标记'
        by_label[mt] = by_label.get(mt, 0) + 1
        for t in (v.get('tags') or []):
            if isinstance(t, str) and t.strip():
                tags_total += 1
    return jsonify({
        'scope': scope,
        'by_label': by_label,
        'total': sum(by_label.values()),
        'unique_songs': unique_songs,
        'tags_total': tags_total,
    })


# 8 类规范主标签（弹窗可选值 + 颜色映射 + 过滤/统计）
_VALID_MARK_TYPES = {
    '已确认', '已标记', '正版确认', '疑似盗版',
    '待核实', '我的精选', '已排除', '未打盗版',
    '已打盗版',  # 标记完成后改的，规范值
}
# 歌单页自动打标的平台标签（不在弹窗里，但允许出现）
_PLATFORM_MARK_PREFIX = ('正版@', '盗版@')


def _normalize_mark_type(raw_type):
    """mark_type 轻清洗（不再压成 4 类）：
    - None → 返回 ''（显式清空，clear 模式专用，不填回默认值）
    - 空值/纯空白 → 返回 '已确认' 作为默认值（弹窗默认选项）
    - 否则原样保留用户原值（已确认/已标记/正版确认/疑似盗版/待核实/
      我的精选/已排除/未打盗版/已打盗版/正版@xx/盗版@xx 全部保留）
    - 非规范值（历史脏数据）也原样保留——搜索页/已标记页会按原值显示，
      不再强制压成"自定义备注"。原值如被新版替换，回填逻辑会从 note
      「原类型：xxx」里取回。
    """
    if raw_type is None:
        return ''  # 显式清空（clear 模式），与默认 '已确认' 区分
    s = (raw_type or '').strip()
    if not s:
        return '已确认'
    return s


def _migrate_mark_types_once():
    """启动时跑一次：修复 v4.11 buggy 迁移造成的 mark_type 破坏。

    v4.11 用有 bug 的正则（`+?` 非贪婪）做归一时，把 note 里的"原类型：xxx"反向
    提取成 mark_type 时只取了一个字（如"正版确认" → "正"），并把 note 里的"原类型"
    标记清掉了，导致 mark_type 字段被破坏成单字（如"已"/"正"/"疑"）。

    本函数：
    1. 从黄金备份 `~/.musicfinder/song_marks.json.bak`（v4.11 迁移前的 8/5 完整快照）
       按 (song, perf, album) 三元组匹配，修复 BROKEN 值（自定义备注/已/正/疑）的 mark_type
       和 note（如果当前 note 是 buggy 残留的"版确认"/"标记"/"确认"）。
    2. tags 字段缺省补 []（v4.13 新增，老数据无 tags 自动兼容）。
    3. 残留 BROKEN 值且 .bak 里没找到的：保留原样（v4.11 之后用户新加的，不动）。

    幂等：只覆盖 mark_type 在 BROKEN 集合里的记录；已正确的值不动。
    """
    try:
        import re
        # 1. 找黄金备份
        bak_path = os.path.join(_MARKS_DIR, 'song_marks.json.bak')
        bak_idx = {}  # (song, perf, album) → (mark_type, note)
        if os.path.exists(bak_path):
            try:
                with open(bak_path, 'r', encoding='utf-8') as f:
                    bak_raw = json.load(f)
                bak_marks = bak_raw.get('marks') if isinstance(bak_raw, dict) and 'marks' in bak_raw else bak_raw
                if isinstance(bak_marks, dict):
                    for k, v in bak_marks.items():
                        if not isinstance(v, dict):
                            continue
                        s = (v.get('song_name') or '').strip().lower()
                        p = (v.get('performer') or '').strip().lower()
                        a = (v.get('album') or '').strip().lower()
                        s = re.sub(r'[\s\u3000]+', '', s)
                        p = re.sub(r'[\s\u3000]+', '', p)
                        a = re.sub(r'[\s\u3000]+', '', a)
                        bak_idx[(s, p, a)] = (v.get('mark_type', ''), v.get('note', ''))
                    logger.info('[marks] 已加载黄金备份 %s: %d 条', bak_path, len(bak_idx))
            except Exception:
                logger.exception('加载黄金备份 %s 失败', bak_path)

        # 2. v4.11 buggy 造成的破坏值集合
        BROKEN_MT = {'自定义备注', '已', '正', '疑'}
        # 残留 note（v4.11 buggy 提取后的"单字"残留）
        BROKEN_NOTE = {'版确认', '标记', '确认'}

        targets = []
        if os.path.exists(_MARKS_FILE):
            targets.append(_MARKS_FILE)
        users_root = os.path.join(_MARKS_DIR, 'users')
        if os.path.isdir(users_root):
            for name in sorted(os.listdir(users_root)):
                mf = os.path.join(users_root, name, 'song_marks.json')
                if os.path.exists(mf):
                    targets.append(mf)
        grand_changed = 0
        grand_samples = {}
        for mf in targets:
            try:
                with open(mf, 'r', encoding='utf-8') as f:
                    raw = json.load(f)
                if not isinstance(raw, dict):
                    continue
                marks = raw.get('marks') or {}
                file_changed = 0
                for k, v in marks.items():
                    if not isinstance(v, dict):
                        continue
                    # 1) tags 缺省补 []
                    if 'tags' not in v:
                        v['tags'] = []
                        file_changed += 1
                    # 2) 用黄金备份修复 BROKEN mark_type
                    cur_mt = (v.get('mark_type') or '').strip()
                    if cur_mt in BROKEN_MT and bak_idx:
                        s = (v.get('song_name') or '').strip().lower()
                        p = (v.get('performer') or '').strip().lower()
                        a = (v.get('album') or '').strip().lower()
                        s = re.sub(r'[\s\u3000]+', '', s)
                        p = re.sub(r'[\s\u3000]+', '', p)
                        a = re.sub(r'[\s\u3000]+', '', a)
                        hit = bak_idx.get((s, p, a))
                        if hit:
                            bak_mt, bak_note = hit
                            v['mark_type'] = bak_mt
                            # 还原 note：当前 note 是 buggy 残留则替换
                            cur_note = (v.get('note') or '').strip()
                            if cur_note in BROKEN_NOTE or not cur_note:
                                v['note'] = bak_note
                            file_changed += 1
                            if len(grand_samples) < 8:
                                grand_samples.setdefault(mf, []).append((cur_mt, bak_mt))
                if file_changed:
                    tmp = mf + '.tmp'
                    with open(tmp, 'w', encoding='utf-8') as f:
                        json.dump({'key_version': _MARK_KEY_VERSION, 'marks': marks},
                                  f, ensure_ascii=False, indent=2)
                    os.replace(tmp, mf)
                    logger.info('[marks] 启动修复 %s: %s 条', mf, file_changed)
                    grand_changed += file_changed
            except Exception:
                logger.exception('mark_type 启动修复 %s 失败', mf)
        if grand_changed:
            logger.info('[marks] mark_type 启动修复总计：%s 条。样本：%s',
                        grand_changed, {k: v[:3] for k, v in grand_samples.items()})
        else:
            logger.info('[marks] mark_type 启动修复：无变化')
    except Exception:
        logger.exception('mark_type 启动修复失败')


@app.route('/api/marks', methods=['POST'])
def api_set_mark():
    data = request.get_json(silent=True) or {}
    song_name = (data.get('song_name') or '').strip()
    performer = (data.get('performer') or '').strip()
    album = (data.get('album') or '').strip()
    raw_type = (data.get('mark_type') or '').strip()
    mark_type = _normalize_mark_type(raw_type)
    note = (data.get('note') or '').strip()
    # 附加标签 tags[]：可空数组，list 化 + 去重保序（复用工具函数，与 PATCH/batch_update 同行为）
    tags = _normalize_tags(data.get('tags'))
    if not song_name:
        return jsonify({'error': 'song_name 必填'}), 400
    key = _mark_key(song_name, performer, album)
    scope = (data.get('scope') or 'mine').strip().lower()
    if scope == 'all':
        scope = 'shared'
    cur = _cur_user() or 'legacy'
    mark_data = {
        'song_name': song_name, 'performer': performer, 'album': album,
        'mark_type': mark_type, 'tags': tags, 'note': note,
        'updated_at': time.strftime('%Y-%m-%d %H:%M:%S'),
    }
    # 共享视图下：若 key 已属他人，则视为「改他人标」→ 需 manage_marks 授权
    owner = cur
    if scope == 'shared':
        existing_owner = _find_cloud_mark_owner(key)
        if existing_owner and existing_owner != cur:
            if not (_can_user(cur, 'manage_marks') or _is_admin(cur)):
                return jsonify({'error': '只有被授权（管理他人标记）的账号才能修改他人的标记'}), 403
            owner = existing_owner  # 保留原归属
    if owner == cur:
        with _marks_lock:
            marks = _load_marks()
            marks[key] = mark_data
            _save_marks(marks)
        return jsonify({'ok': True, 'mark': mark_data, 'key': key})
    # 跨账号写：直接落对方本地 + 云端（保留 owner 归属）
    _apply_shared_mark_write('set', owner, key, mark_data)
    return jsonify({'ok': True, 'mark': mark_data, 'key': key, 'owner': owner})


def _normalize_tags(raw_tags):
    """把传入值规整为附加标签 list：支持 list / 字符串（逗号分隔），去重保序。"""
    if isinstance(raw_tags, str):
        raw_tags = [t.strip() for t in raw_tags.split(',') if t.strip()]
    elif not isinstance(raw_tags, list):
        return []
    seen = set()
    out = []
    for t in raw_tags:
        t = (t or '').strip()
        if not t or t in seen:
            continue
        seen.add(t)
        out.append(t)
    return out


def _build_set_from_payload(payload):
    """从 {mark_type?, tags?, note?, mode?} 构造应用到一个 mark dict 上的字段变化 set 字典。

    mode:
      - 'replace'（默认）：用 set 里的值覆盖现有字段
      - 'append_tags'：tags 追加（不去重，set.tags 是相对增量；mark_type/note 仍覆盖）
      - 'clear'：把所有字段清空（删 mark_type / tags / note；song_name 等元信息保留）

    返回 {apply_func: callable(mark_dict) -> bool_changed}。返回 None 表示无变化。
    """
    mark_type_in = payload.get('mark_type')
    tags_in = payload.get('tags')
    note_in = payload.get('note')
    mode = (payload.get('mode') or 'replace').strip().lower()

    has_mt = 'mark_type' in payload
    has_tags = 'tags' in payload
    has_note = 'note' in payload
    if mode == 'clear':
        # clear 模式：把所有字段清空（保留 song_name/performer/album/updated_at）；
        # mark_type='' 不被 _normalize_mark_type 默认填为「已确认」——清空就是清空
        def apply_clear(d):
            changed = False
            if has_mt and d.get('mark_type') not in (None, ''):
                # 设 None 让 _save_marks_local 的 _normalize_mark_type 识别为「真清空」
                # （_normalize_mark_type 对 None 返回 ''，不再填回「已确认」）
                d['mark_type'] = None
                changed = True
            if has_tags and d.get('tags'):
                d['tags'] = []
                changed = True
            if has_note and d.get('note'):
                d['note'] = ''
                changed = True
            return changed
        if not (has_mt or has_tags or has_note):
            return None
        return apply_clear

    if not (has_mt or has_tags or has_note):
        return None

    norm_mt = _normalize_mark_type(mark_type_in) if has_mt else None
    norm_tags = _normalize_tags(tags_in) if has_tags else None
    norm_note = (note_in or '').strip() if has_note else None

    if mode == 'append_tags':
        def apply_append(d):
            before = list(d.get('tags') or [])
            extras = norm_tags or []
            merged = before + [t for t in extras if t not in before]
            if merged != before:
                d['tags'] = merged
                if has_mt:
                    d['mark_type'] = norm_mt
                if has_note:
                    d['note'] = norm_note
                return True
            changed = False
            if has_mt and d.get('mark_type') != norm_mt:
                d['mark_type'] = norm_mt
                changed = True
            if has_note and d.get('note') != norm_note:
                d['note'] = norm_note
                changed = True
            return changed
        return apply_append

    # 默认 replace
    def apply_replace(d):
        changed = False
        if has_mt and d.get('mark_type') != norm_mt:
            d['mark_type'] = norm_mt
            changed = True
        if has_tags:
            new_tags = norm_tags or []
            if d.get('tags') != new_tags:
                d['tags'] = new_tags
                changed = True
        if has_note and d.get('note') != norm_note:
            d['note'] = norm_note
            changed = True
        return changed
    return apply_replace


def _apply_mark_update(mark, set_builder):
    """对单个 mark dict 应用一个 set_builder（_build_set_from_payload 返回的 callable）。"""
    if set_builder is None:
        return False
    return bool(set_builder(mark))


@app.route('/api/marks', methods=['PATCH'])
def api_patch_mark():
    """单条内联改类型/标签（点标签直接改走这个接口）。
    Body: {song_name, performer, album, set: {mark_type?, tags?, note?, mode?}}
    返回 {ok, mark, key}；mark 不存在返回 404。
    """
    data = request.get_json(silent=True) or {}
    song_name = (data.get('song_name') or '').strip()
    performer = (data.get('performer') or '').strip()
    album = (data.get('album') or '').strip()
    if not song_name:
        return jsonify({'error': 'song_name 必填'}), 400
    set_payload = data.get('set') or {}
    set_builder = _build_set_from_payload(set_payload)
    if set_builder is None:
        return jsonify({'error': 'set 字段为空（至少传 mark_type / tags / note 之一）'}), 400
    key = _mark_key(song_name, performer, album)
    scope = (data.get('scope') or 'mine').strip().lower()
    if scope == 'all':
        scope = 'shared'
    cur = _cur_user() or 'legacy'
    owner_param = (data.get('owner') or '').strip()
    owner = cur
    if scope == 'shared' and owner_param and owner_param != cur:
        # 改他人标 → 需 manage_marks 授权（或 admin）
        if not (_can_user(cur, 'manage_marks') or _is_admin(cur)):
            return jsonify({'error': '只有被授权（管理他人标记）的账号才能修改他人的标记'}), 403
        owner = owner_param
    if owner == cur:
        with _marks_lock:
            marks = _load_marks()
            if key not in marks:
                return jsonify({'error': '该歌未标记（无 mark 记录）', 'key': key}), 404
            if _apply_mark_update(marks[key], set_builder):
                marks[key]['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
                _save_marks(marks, merge_with_disk=False)  # 单条改直接落盘（不合并，避免覆盖并发改）
        return jsonify({'ok': True, 'mark': marks.get(key) or {}, 'key': key})
    # 跨账号 PATCH：读旧值 → 增量 → 写共享（保留 owner 归属）
    old = _get_shared_mark(owner, key)
    if not old:
        return jsonify({'error': '该歌未标记（无 mark 记录）', 'key': key}), 404
    new_mark = dict(old)
    if _apply_mark_update(new_mark, set_builder):
        new_mark['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
        _apply_shared_mark_write('set', owner, key, new_mark)
    return jsonify({'ok': True, 'mark': new_mark, 'key': key, 'owner': owner})


@app.route('/api/marks/batch_update', methods=['POST'])
def api_batch_update_marks():
    """批量改主标签/附加标签/备注。
    Body: {keys: [{song_name, performer, album, owner?}, ...],  set: {mark_type?, tags?, note?, mode?}}

    - 一次最多 500 条
    - mine 视图：所有 key 都改当前用户（不传 owner）
    - all 视图（admin）：每行可传 owner，后端校验 owner == _cur_user()，否则该条失败
    - 返回 {ok, updated: N, skipped: [{reason, song_name}, ...], failed: [{reason, song_name}, ...]}
    """
    data = request.get_json(silent=True) or {}
    keys = data.get('keys') or []
    if not isinstance(keys, list) or not keys:
        return jsonify({'error': 'keys 必填非空数组'}), 400
    if len(keys) > 500:
        return jsonify({'error': '一次最多 500 条'}), 400
    set_payload = data.get('set') or {}
    set_builder = _build_set_from_payload(set_payload)
    if set_builder is None:
        return jsonify({'error': 'set 字段为空'}), 400
    cur = _cur_user() or 'legacy'
    scope = (data.get('scope') or 'mine').strip().lower()
    if scope == 'all':
        scope = 'shared'
    is_shared = (scope == 'shared')

    updated = 0
    skipped = []
    failed = []
    shared_ops = []

    # 共享视图下：若有改他人标的请求，一次性拉全集建索引，避免逐条读云端
    shared_index = {}
    if is_shared and any((e.get('owner') or '').strip() and (e.get('owner') or '').strip() != cur for e in keys):
        try:
            shared_index = {(it.get('owner'), it.get('key')): it
                            for it in _load_marks_cloudbase_shared()}
        except Exception:
            shared_index = {}

    with _marks_lock:
        marks = _load_marks()
        any_changed = False
        for entry in keys:
            if not isinstance(entry, dict):
                failed.append({'reason': '参数格式错误', 'song_name': '(invalid)'})
                continue
            song_name = (entry.get('song_name') or '').strip()
            performer = (entry.get('performer') or '').strip()
            album = (entry.get('album') or '').strip()
            owner = (entry.get('owner') or '').strip()
            if not song_name:
                failed.append({'reason': 'song_name 缺失', 'song_name': '(empty)'})
                continue
            key = _mark_key(song_name, performer, album)
            # 跨账号（共享视图下改他人标）：需 manage_marks 授权
            if owner and owner != cur:
                if not (_can_user(cur, 'manage_marks') or _is_admin(cur)):
                    failed.append({'reason': f'无权改 owner={owner} 的标（当前={cur}）',
                                   'song_name': song_name})
                    continue
                old = shared_index.get((owner, key)) or _get_shared_mark(owner, key)
                if isinstance(old, dict):
                    old = {k: v for k, v in old.items() if k not in ('owner', 'key')}
                if not old:
                    skipped.append({'reason': '未标记', 'song_name': song_name})
                    continue
                new_mark = dict(old)
                if _apply_mark_update(new_mark, set_builder):
                    new_mark['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
                    shared_ops.append(('set', owner, key, new_mark))
                    updated += 1
                else:
                    skipped.append({'reason': '值未变化（跳过写入）', 'song_name': song_name})
                continue
            # 自己的标：原路径
            if key not in marks:
                skipped.append({'reason': '未标记', 'song_name': song_name})
                continue
            if _apply_mark_update(marks[key], set_builder):
                marks[key]['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
                any_changed = True
                updated += 1
            else:
                skipped.append({'reason': '值未变化（跳过写入）', 'song_name': song_name})
        if any_changed:
            _save_marks(marks, merge_with_disk=False)  # 批量改直接落盘

    if shared_ops:
        _apply_shared_mark_writes(shared_ops)

    return jsonify({'ok': True, 'updated': updated, 'skipped': skipped, 'failed': failed})


@app.route('/api/marks', methods=['DELETE'])
def api_delete_mark():
    data = request.get_json(silent=True) or {}
    song_name = (data.get('song_name') or '').strip()
    performer = (data.get('performer') or '').strip()
    album = (data.get('album') or '').strip()
    key = _mark_key(song_name, performer, album)
    owner = (data.get('owner') or '').strip()
    scope = (data.get('scope') or 'mine').strip().lower()
    if scope == 'all':
        scope = 'shared'
    cur = _cur_user() or 'legacy'
    # 删他人标 → 需 manage_marks 授权（或 admin），随后同时落对方本地 + 云端
    if owner and owner != cur:
        if not (_can_user(cur, 'manage_marks') or _is_admin(cur)):
            return jsonify({'error': '只有被授权（管理他人标记）的账号才能删除他人的标记'}), 403
        _apply_shared_mark_write('delete', owner, key, None)
        return jsonify({'ok': True, 'removed': True, 'owner': owner})
    with _marks_lock:
        marks = _load_marks()
        removed = marks.pop(key, None)
        # 关键：删除必须 merge_with_disk=False，直接落盘「已移除该 key」的精确状态。
        # 若用默认 merge_with_disk=True，写盘前会重新读磁盘（仍含该 key）并 _merge_marks
        # 把刚 pop 掉的 key 加回来 → 删除永远不生效（v4.7「防云端覆盖」的副作用）。
        # 云端仍由 _save_marks 后台按人隔离同步（_save_marks_cloudbase 会删掉当前用户
        # 云端多出的 key），本地与云端都正确清除。
        _save_marks(marks, merge_with_disk=False)
    return jsonify({'ok': True, 'removed': removed is not None})


@app.route('/api/marks/export', methods=['GET'])
def api_export_marks():
    """导出全部标记为 JSON 文件（含元信息），可跨设备/跨版本迁移。

    返回 attachment，浏览器会直接下载，不污染页面。
    """
    with _marks_lock:
        marks = _load_marks()
    export_data = {
        'app': 'MusicFinder',
        'format': 'song-marks',
        'version': 1,
        'exported_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        'marks': marks,
    }
    body = json.dumps(export_data, ensure_ascii=False, indent=2)
    filename = 'musicfinder_marks_%s.json' % time.strftime('%Y%m%d')
    resp = make_response(body)
    resp.headers['Content-Type'] = 'application/json; charset=utf-8'
    resp.headers['Content-Disposition'] = 'attachment; filename="%s"' % filename
    return resp


@app.route('/api/marks/import', methods=['POST'])
def api_import_marks():
    """从导出的 JSON 文件合并标记。

    合并策略：导入的文件与本地标记按 key 合并，key 冲突时以导入文件为准
    （用户通常拿自己的备份覆盖），本地独有标记保留。
    """
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        # 允许直接上传纯 marks 字典（无元信息包装）
        try:
            raw = request.get_data(as_text=True)
            data = json.loads(raw) if raw else None
        except Exception:
            data = None
    if not isinstance(data, dict):
        return jsonify({'error': '无效的标记文件：应为 JSON 对象'}), 400

    # 兼容两种格式：{"marks": {...}} 包装 / 纯 marks 字典
    incoming = data.get('marks') if ('marks' in data and isinstance(data.get('marks'), dict)) else data

    # 基本校验：value 必须像标记对象
    # key 一律由服务端按 song_name/performer/album 重算，忽略文件里自带的 key。
    # 这样外部工具（Excel/CSV 转出的 JSON）无需复刻内部 key 算法也能 100% 正确落位，
    # 且旧版本导出的 v1 口径备份导进来会自动升到 v2 口径。
    cleaned = {}
    for _k, v in incoming.items():
        if not isinstance(v, dict) or 'song_name' not in v:
            continue
        song_name = str(v.get('song_name', '')).strip()
        if not song_name:
            continue
        performer = str(v.get('performer', '')).strip()
        album = str(v.get('album', '')).strip()
        raw_t = str(v.get('mark_type', '')).strip()
        nt = _normalize_mark_type(raw_t)
        nn = str(v.get('note', '')).strip()
        if raw_t and raw_t != nt and not nn:
            nn = '原类型：%s' % raw_t
    # 附加标签 tags[]：可空数组，list 化 + 去重保序
    raw_tags_in = v.get('tags') or []
    if isinstance(raw_tags_in, str):
        raw_tags_in = [t.strip() for t in raw_tags_in.split(',') if t.strip()]
    elif not isinstance(raw_tags_in, list):
        raw_tags_in = []
    seen = set()
    tags_in = []
    for t in raw_tags_in:
        t = (t or '').strip()
        if not t or t in seen:
            continue
        seen.add(t)
        tags_in.append(t)
    cleaned[_mark_key(song_name, performer, album)] = {
        'song_name': song_name,
        'performer': performer,
        'album': album,
        'mark_type': nt,
        'tags': tags_in,
        'note': nn,
        'updated_at': str(v.get('updated_at', time.strftime('%Y-%m-%d %H:%M:%S'))),
    }
    if not cleaned:
        return jsonify({'error': '标记文件里没有可用的标记记录'}), 400

    with _marks_lock:
        marks = _load_marks()
        added = 0
        updated = 0
        for k, v in cleaned.items():
            if k in marks:
                updated += 1
            else:
                added += 1
            marks[k] = v
        _save_marks(marks)
    return jsonify({
        'ok': True,
        'imported': len(cleaned),
        'added': added,
        'updated': updated,
        'total': len(marks),
    })



# ═══════════════════════════════════════════════════
#  平台红心歌单（识别各平台「我喜欢 / 收藏」的歌，搜索结果打标）
#  数据存于用户目录（包外 ~/.musicfinder/），升级 app 不会丢失；
#  先支持 QQ 音乐，其余平台预留扩展位。
# ═══════════════════════════════════════════════════

# 与标记/Cookie 同处用户目录，打包后位于 .app 之外，升级不丢
_HEARTS_DIR = os.path.expanduser('~/.musicfinder')
_HEARTS_FILE = os.path.join(_HEARTS_DIR, 'platform_hearts.json')
_hearts_lock = threading.Lock()

# 红心抓取用的登录 Cookie，单独存一份（与「默认搜索」用的 cookies.json 完全隔离）。
# 这样：默认搜索登录态（cookies.json）和「抓取我喜欢歌单」的登录态（hearts_cookies.json）
# 互不干扰，用户可分别为搜索 / 红心配置不同账号，也避免一处改动污染另一处。
_HEART_COOKIE_FILE = os.path.join(_HEARTS_DIR, 'hearts_cookies.json')
_heart_cookie_lock = threading.Lock()


def _effective_hearts_paths():
    """返回 (hearts_file, heart_cookie_file, base_dir)，按当前登录用户隔离；未登录走旧路径。"""
    ud = _user_dir()
    if ud:
        return (os.path.join(ud, 'platform_hearts.json'),
                os.path.join(ud, 'hearts_cookies.json'), ud)
    return _HEARTS_FILE, _HEART_COOKIE_FILE, _HEARTS_DIR


def _load_heart_cookies():
    hf, _, _ = _effective_hearts_paths()
    try:
        if os.path.exists(hf):
            with open(hf, 'r', encoding='utf-8') as f:
                return json.load(f) or {}
    except Exception:
        logger.exception('读取红心 Cookie 文件失败')
    return {}


def _save_heart_cookies(data):
    _, cf, cd = _effective_hearts_paths()
    os.makedirs(cd, exist_ok=True)
    tmp = cf + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, cf)  # 原子替换


# 各平台红心徽章样式（先上 QQ，其余留作扩展）
HEART_PLATFORMS = {
    'qq': {'name': 'QQ音乐', 'badge': 'Q音红心', 'cls': 'qq'},
    'netease': {'name': '网易云', 'badge': '云村红心', 'cls': 'netease'},
    'kugou': {'name': '酷狗', 'badge': '酷狗红心', 'cls': 'kugou'},
    'kuwo': {'name': '酷我', 'badge': '酷我红心', 'cls': 'kuwo'},
    'qishui': {'name': '汽水', 'badge': '汽水红心', 'cls': 'qishui'},
}


def _load_hearts():
    hf, _, _ = _effective_hearts_paths()
    try:
        if not os.path.exists(hf):
            return {}
        with open(hf, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        logger.exception('读取红心歌单文件失败')
        return {}


def _save_hearts(data):
    _, _, hd = _effective_hearts_paths()
    os.makedirs(hd, exist_ok=True)
    tmp = os.path.join(hd, 'platform_hearts.json') + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, os.path.join(hd, 'platform_hearts.json'))  # 原子替换，避免写一半损坏
    _push_blob_async('hearts', data)  # 跨设备同步


def _hearts_platform(data, plat):
    p = data.get(plat)
    return p if isinstance(p, dict) else {}


def _parse_pasted_list(text):
    """把用户粘贴的歌单文本解析成 [(song, artist), ...]。

    兼容常见格式：「歌名 - 歌手」「歌名—歌手」「歌名 | 歌手」「歌名 歌手」，
    也兼容纯「歌名」一行一首。自动去除行首序号(1. / 01 )。
    """
    out = []
    for raw in (text or '').splitlines():
        line = raw.strip()
        if not line:
            continue
        # 去除行首序号：1. 01、1、
        line = re.sub(r'^\s*\d+[\.、)]\s*', '', line)
        # 常见分隔符：- – — 丨 | ；多个空格
        parts = re.split(r'\s*[-–—丨|]\s*|\s{2,}', line)
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) >= 2:
            song, artist = parts[0], '、'.join(parts[1:])
        elif len(parts) == 1:
            song, artist = parts[0], ''
        else:
            continue
        # 过滤掉表头类无效行
        if song in ('歌名', '歌曲', 'song') and artist in ('歌手', 'artist'):
            continue
        out.append((song, artist))
    return out


def _build_hearts_index(plat_data):
    """把某平台的红心歌单构建成匹配索引：[(norm_song, artist_set, norm_album, norm_release), ...]。

    同时兼容历史数据结构：
      - 新格式 dict: {'song':..., 'artist':..., 'album':..., 'release':...}
      - 历史 dict: {'song':..., 'artist':...}（album/release 为空）
      - 历史 tuple: (song, artist) 或 (song, artist, album, release)
    专辑/发行时间在「红心记录为空」时跳过严格匹配，向后兼容。
    """
    idx = []
    for s in plat_data.get('songs', []) or []:
        if isinstance(s, dict):
            song = s.get('song', '') or ''
            artist = s.get('artist', '') or ''
            album = s.get('album', '') or ''
            release = s.get('release', '') or ''
        elif isinstance(s, (list, tuple)):
            song = s[0] if len(s) > 0 else ''
            artist = s[1] if len(s) > 1 else ''
            album = s[2] if len(s) > 2 else ''
            release = s[3] if len(s) > 3 else ''
        else:
            continue
        ns = _normalize_match(song)
        if not ns:
            continue
        aset = set(_split_artist_names(artist))
        an = _normalize_match(album)
        rn = _normalize_release(release)
        idx.append((ns, aset, an, rn))
    return idx


def _normalize_release(text):
    """把发行时间标准化到 yyyy-mm-dd（去空白、去中点符号）。空返回 ''。"""
    if not text:
        return ''
    s = str(text).strip()
    # 兼容 yyyy-mm-dd / yyyy.mm.dd / yyyy/mm/dd / yyyy-mm / 时间戳
    m = re.match(r'(\d{4})[\.\-/年](\d{1,2})(?:[\.\-/月](\d{1,2}))?', s)
    if m:
        y, mo, d = m.group(1), int(m.group(2)), int(m.group(3) or 0)
        if d:
            return '%s-%02d-%02d' % (y, mo, d)
        return '%s-%02d' % (y, mo)
    # 时间戳
    try:
        ts = float(s)
        if ts > 1e12:
            ts = ts / 1000.0
        return _ts_to_date(ts) or ''
    except Exception:
        return s[:10] if len(s) >= 10 else s


def _match_hearts(song_name, performer, album, release, hearts_index_map):
    """返回命中的平台代码列表（如 ['qq']）。

    匹配规则（按严格度递进）：
    1. 歌名标准化后必须相等；
    2. 红心记录含歌手时，结果也必须有歌手，且两者有交集；
    3. 红心为独唱、结果出现 3 个及以上不同歌手（合集 / 群星翻唱辑）时，视为不命中，
       避免一首「别再为他流泪 - 梁静茹」给所有同名翻唱合集都打上红心。
    4. 红心记录有专辑时，结果的专辑必须标准化后相等；
       避免同名歌曲（如「聊天」收录在不同专辑）误匹配红心。
    5. 红心记录有发行时间时，结果的发行时间必须一致（精度到月）；
       进一步区分同一专辑下的不同版本（如 live / 翻唱）。

    向后兼容：旧数据无专辑/发行时间时跳过 4、5 步。
    """
    if not hearts_index_map:
        return []
    ns = _normalize_match(song_name)
    if not ns:
        return []
    artists = set(_split_artist_names(performer))
    an = _normalize_match(album)
    rn = _normalize_release(release)
    matched = []
    for plat, idx in hearts_index_map.items():
        for (hsong, hartists, halbum, hrelease) in idx:
            if hsong != ns:
                continue
            # 红心记录有歌手时，结果不能没歌手
            if hartists and not artists:
                continue
            # 必须有交集
            if hartists and artists and not (hartists & artists):
                continue
            # 独唱红心 vs 群星/合集结果：过滤掉误匹配
            if hartists and len(hartists) == 1 and len(artists) >= 3:
                continue
            # 专辑精确匹配：红心有专辑时，结果必须也匹配（专辑相同才算同一首）
            if halbum:
                if not an or an != halbum:
                    continue
            # 发行时间精确匹配：红心有发行时间时，结果必须匹配到月
            if hrelease:
                if not rn:
                    continue
                # 精度到月：yyyy-mm-dd 比 yyyy-mm 比较时取前 7 位
                hr_prefix = hrelease[:7]
                rn_prefix = rn[:7]
                if hr_prefix != rn_prefix:
                    continue
            matched.append(plat)
            break
    return matched


def _qq_uin(cookie_str):
    for pat in (r'uin=([^;]+)', r'qqmusic_uin=([^;]+)'):
        m = re.search(pat, cookie_str or '')
        if m:
            return m.group(1)
    return ''


def _qq_gtk(cookie_str):
    m = re.search(r'qqmusic_key=([^;]+)', cookie_str or '')
    key = m.group(1) if m else (cookie_str or '')
    h = 5381
    for ch in key:
        h += (h << 5) + ord(ch)
    return h & 0x7fffffff


def fetch_qq_fav_songs(cookie_str):
    """拉取 QQ 音乐「我喜欢」歌单，返回 [(song, artist, album, release), ...]。

    依赖登录 cookie（含 qqmusic_key / uin）。cookie 失效会抛异常。

    历史接口 music.favorite.FavSongList 已失效（返回 500003），
    现改用 fcg_musiclist_getmyfav 拿全部 songid，再分批用
    fcg_play_single_song 换取歌名/歌手/专辑/发行时间详情。

    返回 4 元组以支持精确匹配（同一首歌的不同专辑版本不算同一首红心）。
    """
    raw_uin = _qq_uin(cookie_str)
    # 部分接口只需要纯数字 QQ 号
    num_uin = raw_uin.replace('o0', '') if raw_uin.startswith('o0') else raw_uin
    g_tk = _qq_gtk(cookie_str)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36',
        'Referer': 'https://y.qq.com/',
        'Cookie': cookie_str,
    }

    # 1) 拿「我喜欢」全部 songid
    song_ids = []
    sin = 0
    page_size = 2000
    fav_url = 'https://c.y.qq.com/splcloud/fcgi-bin/fcg_musiclist_getmyfav.fcg'
    while True:
        params = {
            'cid': 205360956,
            'uin': num_uin,
            'sin': sin,
            'ein': sin + page_size - 1,
            'g_tk': g_tk,
            'loginUin': num_uin,
            'hostUin': num_uin,
            'format': 'json',
            'inCharset': 'utf8',
            'outCharset': 'utf-8',
            'notice': 0,
            'platform': 'yqq.json',
            'needNewCode': 0,
        }
        r = requests.get(fav_url, params=params, headers=headers, timeout=20)
        r.raise_for_status()
        resp = r.json()
        if resp.get('code', -1) != 0:
            raise RuntimeError('QQ 我喜欢列表接口错误码 %s' % resp.get('code'))
        part = list(resp.get('map', {}).keys())
        song_ids.extend(part)
        total = resp.get('totalNum') or 0
        if sin + page_size >= total:
            break
        sin += page_size
        if not part:
            break

    # 2) 分批换取歌曲详情（该接口一次最多返回约 50 首）
    out = []
    detail_url = 'https://c.y.qq.com/v8/fcg-bin/fcg_play_single_song.fcg'
    batch_size = 50
    for i in range(0, len(song_ids), batch_size):
        batch = song_ids[i:i + batch_size]
        r2 = requests.get(detail_url, params={
            'songid': ','.join(batch),
            'format': 'json',
            'inCharset': 'utf8',
            'outCharset': 'utf-8',
        }, headers=headers, timeout=30)
        r2.raise_for_status()
        resp2 = r2.json()
        if resp2.get('code', -1) != 0:
            print('[QQ] fav detail batch %s error: %s' % (i, resp2.get('code')))
            continue
        for song in (resp2.get('data') or []):
            name = song.get('name') or song.get('title') or ''
            singers = song.get('singer') or []
            if isinstance(singers, list):
                artist = '、'.join(s.get('name', '') for s in singers if isinstance(s, dict) and s.get('name'))
            else:
                artist = str(singers)
            album_obj = song.get('album') or {}
            if isinstance(album_obj, dict):
                album = album_obj.get('name') or album_obj.get('title') or ''
                release = album_obj.get('time_public') or album_obj.get('pubtime') or ''
            else:
                album = ''
                release = ''
            if name:
                out.append((name, artist, album, release))
    return out


def _favs_to_dict_list(favs):
    """把抓取结果统一转 dict 列表（支持 2/4 元组两种格式）。"""
    out = []
    for item in favs:
        if len(item) >= 4:
            s, a, album, release = item[0], item[1], item[2] or '', item[3] or ''
            out.append({
                'song': s, 'artist': a, 'album': album, 'release': release,
                'raw': '%s - %s' % (s, a),
            })
        else:
            s, a = item[0], item[1] if len(item) > 1 else ''
            out.append({'song': s, 'artist': a, 'raw': '%s - %s' % (s, a)})
    return out


# ── 各平台「我喜欢 / 收藏」抓取实现 ──
# QQ 已验证；网易云按 account → playlist → detail 流程，需 MUSIC_U Cookie；
# 酷狗 / 酷我 / 汽水的收藏接口都需要签名或登录态加密（标准库难以复刻），
# 暂时返回友好提示，引导用「手动粘贴」导入，后续可按需逐个补全真实抓取。

def fetch_fav_songs(plat, cookie):
    """按平台分发到对应的收藏歌单抓取函数。"""
    if plat == 'qq':
        return fetch_qq_fav_songs(cookie)
    if plat == 'netease':
        return fetch_netease_fav_songs(cookie)
    if plat == 'kugou':
        return fetch_kugou_fav_songs(cookie)
    if plat == 'kuwo':
        return fetch_kuwo_fav_songs(cookie)
    if plat == 'qishui':
        return fetch_qishui_fav_songs(cookie)
    raise RuntimeError('暂不支持的平台：%s' % plat)


def _netease_fetch_songs_by_ids(cookie, song_ids):
    """网易云 /song/detail：按歌曲 ID 批量取完整歌曲信息（带专辑、发行时间）。

    用途：当 /playlist/detail 返回的 tracks 字段为空（常见于「我喜欢的音乐」大歌单），
    可以从 trackIds 里拿 ID，再来这里补齐完整歌曲信息。
    """
    song_ids = [int(x) for x in song_ids if x]
    if not song_ids:
        return []
    headers = {
        'User-Agent': COMMON_UA,
        'Referer': 'https://music.163.com/',
        'Origin': 'https://music.163.com',
        'Cookie': cookie or '',
    }
    out = []
    # 网易云 /v3/song/detail 单次最多 150 个 ID（实测 >150 直接返回空）
    batch = 100
    for i in range(0, len(song_ids), batch):
        chunk = song_ids[i:i + batch]
        # /api/song/detail 用 ids=[id1,id2,...]（方括号）参数
        c_param = json.dumps([{'id': x} for x in chunk])
        try:
            r = requests.get('https://music.163.com/api/v3/song/detail',
                             params={'c': c_param, 'ids': json.dumps(chunk)},
                             headers=headers, timeout=30)
            r.raise_for_status()
            d = r.json() or {}
        except Exception:
            # 退化：试老接口
            try:
                r2 = requests.get('https://music.163.com/api/song/detail',
                                  params={'ids': json.dumps(chunk)},
                                  headers=headers, timeout=30)
                r2.raise_for_status()
                d = r2.json() or {}
            except Exception:
                continue
        songs = d.get('songs') or d.get('data', {}).get('songs') or []
        out.extend(songs)
    return out


def _netease_parse_tracks(tracks):
    """把网易云 track dict 列表解析成 [(name, artist, album, release), ...]。

    同时兼容两种接口返回：
    - /api/song/detail（老）：artists / album.name / album.publishTime
    - /api/v3/song/detail（新）：ar / al.name / 顶层 publishTime
    """
    out = []
    for t in tracks:
        if not isinstance(t, dict):
            continue
        name = t.get('name', '') or ''
        # 歌手：旧字段 artists，新字段 ar
        artists = t.get('artists') or t.get('ar') or []
        artist = '、'.join(a.get('name', '') for a in artists if isinstance(a, dict) and a.get('name'))
        # 专辑：旧字段 album，新字段 al
        album_obj = t.get('album') or t.get('al') or {}
        if isinstance(album_obj, dict):
            album = album_obj.get('name', '') or ''
            # 发行时间：新接口在歌顶层，老接口在 album 里
            pub_ts = t.get('publishTime') or album_obj.get('publishTime')
            release = _ts_to_date(pub_ts, is_ms=True) if pub_ts else ''
        else:
            album = ''
            release = t.get('publishTime') and _ts_to_date(t.get('publishTime'), is_ms=True) or ''
        if name:
            out.append((name, artist, album, release))
    return out


def fetch_netease_fav_songs(cookie):
    """拉取网易云「我喜欢的音乐」歌单。依赖登录 Cookie（含 MUSIC_U）。

    返回 [(song, artist, album, release), ...] 4 元组；
    album/release 取自 tracks[].album.name / publishTime，便于精确匹配。
    """
    headers = {
        'User-Agent': COMMON_UA,
        'Referer': 'https://music.163.com/',
        'Origin': 'https://music.163.com',
        'Cookie': cookie,
    }
    csrf = ''
    m = re.search(r'__csrf=([^;]+)', cookie or '')
    if m:
        csrf = m.group(1)
    # 1. 取账号 uid
    try:
        r = requests.get('https://music.163.com/api/nuser/account/get', headers=headers, timeout=20)
        r.raise_for_status()
        acc = r.json() or {}
    except Exception as e:
        raise RuntimeError('获取网易云账号失败（Cookie 可能失效）：%s' % e)
    uid = (acc.get('account') or {}).get('id')
    if not uid:
        raise RuntimeError('未从 Cookie 解析到网易云账号(uid)，Cookie 可能已失效')
    # 2. 找「我喜欢的音乐」歌单（specialType == 5）
    try:
        r2 = requests.get('https://music.163.com/api/user/playlist',
                          params={'uid': uid, 'offset': 0, 'limit': 1000, 'csrf': csrf},
                          headers=headers, timeout=20)
        r2.raise_for_status()
        pls = (r2.json() or {}).get('playlist', [])
    except Exception as e:
        raise RuntimeError('获取网易云歌单列表失败：%s' % e)
    fav_pid = None
    for pl in pls:
        if pl.get('specialType') == 5 or pl.get('name') == '我喜欢的音乐':
            fav_pid = pl.get('id')
            break
    if not fav_pid:
        raise RuntimeError('未找到「我喜欢的音乐」歌单')
    # 3. 取歌单内歌曲（用 /api/v6/playlist/detail——这是唯一在登录态下能拿到完整 trackIds 的接口）
    try:
        r3 = requests.get('https://music.163.com/api/v6/playlist/detail',
                          params={'id': fav_pid, 'n': 1000, 'csrf': csrf},
                          headers=headers, timeout=30)
        r3.raise_for_status()
        d3 = r3.json() or {}
        pl = d3.get('playlist') or {}
    except Exception as e:
        raise RuntimeError('获取网易云歌单详情失败：%s' % e)
    # 优先看 result 字段（老接口），否则看 playlist（v6 接口）
    if not pl and d3.get('result'):
        pl = d3.get('result') or {}
    tracks = pl.get('tracks') or pl.get('trackList') or []
    track_count = pl.get('trackCount') or 0
    # 兜底：tracks 字段可能只给前 1000 首；用完整 trackIds 把剩下补全
    track_ids_seen = {t.get('id') for t in tracks if isinstance(t, dict) and t.get('id')}
    track_ids = []
    for ti in (pl.get('trackIds') or []):
        if isinstance(ti, dict):
            tid = ti.get('id')
        else:
            tid = ti
        if tid and tid not in track_ids_seen:
            track_ids.append(tid)
            track_ids_seen.add(tid)
    if track_ids:
        # 批量去拉详情
        fetched = _netease_fetch_songs_by_ids(cookie, track_ids)
        tracks = list(tracks) + fetched
    out = _netease_parse_tracks(tracks)
    if not out:
        if track_count > 0:
            raise RuntimeError('「我喜欢的音乐」歌单内应有 %d 首歌曲但接口未能拉取到，请稍后重试；如持续失败可改用「歌单链接抓取」方式（把歌单设为公开后用分享链接导入）' % track_count)
        raise RuntimeError('「我喜欢的音乐」歌单内 0 首歌曲——可能你网易云账号还没收藏任何歌曲，请先在网易云 App 加几首红心再试')
    return out


def fetch_kugou_fav_songs(cookie):
    """酷狗「我喜欢」收藏歌单（Cookie 自动抓取，需登录态）。
    流程：MD5 签名 -> 用户歌单列表 -> 找「我喜欢」-> 拉歌曲。
    接口细节随平台版本可能变化，失败请用「歌单链接抓取」方式。
    """
    userid = _kugou_userid(cookie)
    clienttime = int(time.time())
    params = {
        'r': 'my/getUserPlaylistList',
        'appid': '1014',
        'clientver': '20000',
        'clienttime': clienttime,
        'mid': str(clienttime),
        'uuid': str(clienttime),
        'platid': '4',
        'srcappid': '2919',
        'dfid': '-',
        'userid': userid,
        'token': _kugou_token(cookie),
    }
    params['signature'] = _kugou_sign(params)
    headers = {'User-Agent': COMMON_UA, 'Referer': 'https://www.kugou.com/', 'Cookie': cookie}
    try:
        r = requests.get('https://www.kugou.com/mykugou/v2/index.php', params=params, headers=headers, timeout=20)
        r.raise_for_status()
    except Exception as e:
        raise RuntimeError('酷狗「我喜欢」歌单列表接口请求失败：%s；建议改用「歌单链接抓取」或「手动粘贴」' % e)
    if 'Access Deny' in r.text or 'text/html' in (r.headers.get('Content-Type') or ''):
        raise RuntimeError('酷狗接口拒绝访问（反爬限制），自动抓取暂不可用，请改用「歌单链接」或「手动粘贴」方式')
    try:
        d = r.json()
    except ValueError:
        raise RuntimeError('酷狗接口返回异常，自动抓取暂不可用，请改用「歌单链接」或「手动粘贴」方式')
    fav_listid = None
    for pl in (d.get('data', {}).get('plist', []) or d.get('data', {}).get('list', []) or []):
        if pl.get('name') == '我喜欢' or pl.get('is_favorite'):
            fav_listid = pl.get('listid') or pl.get('specialid')
            break
    if not fav_listid:
        raise RuntimeError('未在歌单列表找到「我喜欢」（Cookie 可能失效或格式不全），可改用「歌单链接」或「手动粘贴」方式')
    return fetch_kugou_playlist(fav_listid)


def fetch_kuwo_fav_songs(cookie):
    """酷我「我喜欢」收藏歌单（Cookie 自动抓取，需登录态 kw_token）。"""
    s = requests.Session()
    s.headers.update({'User-Agent': COMMON_UA, 'Referer': 'https://www.kuwo.cn'})
    if cookie:
        s.headers['Cookie'] = cookie
    s.get('https://www.kuwo.cn', timeout=20)
    kw_token = s.cookies.get('kw_token', '')
    s.headers['csrf'] = kw_token
    try:
        r = s.get('https://www.kuwo.cn/api/www/user/playlist', params={'pn': 1, 'rn': 30}, timeout=20)
        r.raise_for_status()
        d = r.json()
    except Exception as e:
        raise RuntimeError('酷我「我喜欢」歌单列表接口请求失败：%s；建议改用「歌单链接抓取」或「手动粘贴」' % e)
    fav_pid = None
    for pl in ((d.get('data') or {}).get('list') or []):
        if pl.get('name') == '我喜欢' or pl.get('isFavorite'):
            fav_pid = pl.get('id')
            break
    if not fav_pid:
        raise RuntimeError('未找到「我喜欢」歌单（Cookie 可能失效），可改用「歌单链接」或「手动粘贴」方式')
    return fetch_kuwo_playlist(fav_pid)


def fetch_qishui_fav_songs(cookie):
    # 汽水音乐为抖音系，接口需 X-Gorgon/X-Khronos 强签名，标准库无法复刻
    raise RuntimeError('汽水音乐不支持 Cookie 自动抓取（抖音签名限制）。请把「我喜欢」歌单设为公开，复制分享链接，用「歌单链接抓取」方式导入。')


# ── 歌单链接抓取（全平台通用，不需苛刻 Cookie，最稳）──
# 五大平台均可把「我喜欢」歌单设为公开 -> 复制分享链接 -> 自动抓取。
def _kugou_userid(cookie_str):
    for pat in (r'(?:^|;)\s*userid=(\d+)', r'(?:^|;)\s*USERID=(\d+)', r'(?:^|;)\s*kg_userid=(\d+)'):
        m = re.search(pat, cookie_str or '')
        if m:
            return m.group(1)
    return '0'


def _kugou_token(cookie_str):
    m = re.search(r'(?:^|;)\s*token=([^;]+)', cookie_str or '')
    return m.group(1) if m else ''


def _kugou_sign(params):
    """酷狗 MD5 签名：密钥 + 按 key 升序的 key=value 拼接 + 密钥。"""
    secret = 'NVPh5oo715z5DIWAeQlhMDsWXXQV4hwt'
    arr = ['%s=%s' % (k, params[k]) for k in sorted(params.keys())]
    raw = secret + ''.join(arr) + secret
    return hashlib.md5(raw.encode('utf-8')).hexdigest()


def _resolve_playlist_link(link):
    """识别歌单分享链接所属平台与歌单 ID/短码。"""
    link = (link or '').strip()
    if not link:
        raise RuntimeError('请粘贴歌单链接')
    low = link.lower()
    if 'qq.com' in low:
        m = re.search(r'(?:disstid|id)=(\d+)', link) or re.search(r'/playlist/(\d+)', link)
        if not m:
            raise RuntimeError('无法从链接解析 QQ 歌单 ID')
        return 'qq', m.group(1)
    if '163.com' in low or '163cn' in low:
        m = re.search(r'id=(\d+)', link)
        if not m:
            raise RuntimeError('无法从链接解析网易云歌单 ID')
        return 'netease', m.group(1)
    if 'kugou.com' in low:
        # 优先匹配 gcid（分享链接常见，如 gcid_xxx），其次数字 specialid/listid
        m = (re.search(r'gcid[_=]?([A-Za-z0-9]+)', link)
             or re.search(r'(?:specialid|listid|playlist_id)=(\d+)', link)
             or re.search(r'/songlist/([A-Za-z0-9]+)', link)
             or re.search(r'/(\d+)\.html', link))
        if not m:
            raise RuntimeError('无法从链接解析酷狗歌单 ID')
        return 'kugou', m.group(1)
    if 'kuwo.cn' in low:
        m = re.search(r'(?:pid=|detail/)(\d+)', link) or re.search(r'/(\d+)\.html', link)
        if not m:
            raise RuntimeError('无法从链接解析酷我歌单 ID')
        return 'kuwo', m.group(1)
    if 'qishui' in low or 'douyin.com' in low:
        m = re.search(r'/s/([A-Za-z0-9]+)', link)
        if m:
            return 'qishui', m.group(1)
        m = re.search(r'playlist[/_-]?id=?(\d+)', link) or re.search(r'/playlist/(\d+)', link)
        if m:
            return 'qishui', m.group(1)
        raise RuntimeError('无法从链接解析汽水歌单（抖音签名限制，建议手动粘贴）')
    raise RuntimeError('无法识别该链接属于哪个平台')


def fetch_qq_playlist(pid):
    url = 'https://c.y.qq.com/v8/fcg-bin/fcg_v8_playlist_cp.fcg'
    params = {'type': 1, 'disstid': pid, 'utf8': 1, 'format': 'json', 'new_format': 1}
    headers = {'User-Agent': COMMON_UA, 'Referer': 'https://y.qq.com/'}
    r = requests.get(url, params=params, headers=headers, timeout=20)
    r.raise_for_status()
    d = r.json()
    cdlist = d.get('cdlist') or []
    if not cdlist:
        data = d.get('data') or {}
        cdlist = data.get('cdlist') or []
        if not cdlist and isinstance(data.get('songlist'), list):
            cdlist = [{'songlist': data['songlist']}]
    out = []
    for cd in cdlist:
        for song in (cd.get('songlist') or []):
            name = song.get('songname') or song.get('name') or ''
            singers = song.get('singer') or []
            artist = '、'.join(s.get('name', '') for s in singers if isinstance(s, dict) and s.get('name'))
            album_obj = song.get('album') or {}
            if isinstance(album_obj, dict):
                album = album_obj.get('name') or album_obj.get('title') or ''
                release = album_obj.get('time_public') or song.get('time_public') or ''
            else:
                album = song.get('albumname') or ''
                release = song.get('time_public') or ''
            if name:
                out.append((name, artist, album, release))
    return out


def fetch_netease_playlist(pid):
    headers = {'User-Agent': COMMON_UA, 'Referer': 'https://music.163.com/', 'Origin': 'https://music.163.com'}
    r = requests.get('https://music.163.com/api/v6/playlist/detail', params={'id': pid, 'n': 1000}, headers=headers, timeout=30)
    r.raise_for_status()
    d = r.json() or {}
    # v6 接口用 playlist；老接口可能用 result
    result = d.get('playlist') or d.get('result') or {}
    tracks = result.get('tracks') or result.get('songs') or result.get('trackList') or []
    track_ids_seen = {t.get('id') for t in tracks if isinstance(t, dict) and t.get('id')}
    track_ids = []
    for ti in (result.get('trackIds') or []):
        if isinstance(ti, dict):
            tid = ti.get('id')
        else:
            tid = ti
        if tid and tid not in track_ids_seen:
            track_ids.append(tid)
            track_ids_seen.add(tid)
    if track_ids:
        fetched = _netease_fetch_songs_by_ids('', track_ids)
        tracks = list(tracks) + fetched
    return _netease_parse_tracks(tracks)


def fetch_kugou_playlist(pid):
    """从酷狗移动端歌单页(H5)解析歌曲，绕过被反爬挡死的签名 API。

    酷狗 PC 接口 playlist/getInfo 需要 dfid+签名，返回 Access Deny；
    但移动端 H5 页面（m.kugou.com/songlist/<id>/）把歌单数据直接 SSR 进
    页面的 "songs":[...] JSON 数组里，无需签名即可解析。

    注意：移动端 H5 只渲染首屏约 30 首（歌单总量可能上千）；超长歌单如需完整
    列表，后续版本再补 H5 翻页。红心匹配本就按歌名+歌手逐首比对，首屏够用。
    """
    url = 'https://m.kugou.com/songlist/%s/' % pid
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Mobile/15E148 Safari/604.1',
        'Referer': 'https://m.kugou.com/',
    }
    last_err = None
    html = ''
    # 酷狗对高频请求会静默返回空 body（限流），偶发时重试一次可解；
    # 持续限流则明确提示用户稍后重试或手动粘贴，避免误判为「链接失效」。
    for _attempt in range(2):
        try:
            r = requests.get(url, headers=headers, timeout=20)
        except Exception as e:
            last_err = e
            time.sleep(2)
            continue
        if r.status_code != 200:
            raise RuntimeError('酷狗歌单页返回 HTTP %s，链接可能失效' % r.status_code)
        html = r.text or ''
        if '"songs"' in html:
            break
        last_err = RuntimeError('酷狗返回空页面（可能被临时限流），请稍后重试，或用「手动粘贴」方式')
        time.sleep(3)
    else:
        raise last_err or RuntimeError('酷狗歌单抓取失败')
    m = re.search(r'"songs"\s*:\s*\[', html)
    if not m:
        raise RuntimeError('酷狗歌单页面未包含歌曲数据（链接可能失效或需登录），请改用「手动粘贴」方式')
    arr_start = m.end() - 1  # 指向 [
    depth = 0
    arr_end = None
    for j in range(arr_start, len(html)):
        c = html[j]
        if c == '[':
            depth += 1
        elif c == ']':
            depth -= 1
            if depth == 0:
                arr_end = j + 1
                break
    if arr_end is None:
        raise RuntimeError('酷狗歌单数据解析失败，请改用「手动粘贴」方式')
    try:
        songs_data = json.loads(html[arr_start:arr_end])
    except ValueError:
        raise RuntimeError('酷狗歌单数据解析失败（JSON 异常），请改用「手动粘贴」方式')
    out = []
    for s in songs_data:
        if not isinstance(s, dict):
            continue
        name = s.get('name') or ''
        # name 形如 "歌手 - 歌名 (后缀)"，拆分出歌手与歌名
        if ' - ' in name:
            performer, song = name.split(' - ', 1)
        else:
            performer, song = '', name
        album = (s.get('albuminfo') or {}).get('name') or ''
        release = ''  # H5 不暴露发行时间
        if song:
            out.append((song, performer, album, release))
    if not out:
        raise RuntimeError('酷狗歌单为空或解析失败，请改用「手动粘贴」方式')
    return out


def fetch_kuwo_playlist(pid):
    s = requests.Session()
    s.headers.update({'User-Agent': COMMON_UA, 'Referer': 'https://www.kuwo.cn'})
    s.get('https://www.kuwo.cn', timeout=20)
    kw_token = s.cookies.get('kw_token', '')
    s.headers['csrf'] = kw_token
    s.headers['Cookie'] = 'kw_token=%s' % kw_token
    out = []
    rn = 30
    for pn in range(1, 60):
        r = s.get('https://www.kuwo.cn/api/www/playlist/playListInfo', params={'pid': pid, 'pn': pn, 'rn': rn}, timeout=20)
        r.raise_for_status()
        d = r.json()
        songs = (d.get('data') or {}).get('musicList') or []
        if not songs:
            break
        for t in songs:
            name = t.get('name', '')
            artist = t.get('artist', '') or ''
            album_obj = t.get('album') or ''
            if isinstance(album_obj, dict):
                album = album_obj.get('name') or album_obj.get('title') or ''
            else:
                album = str(album_obj) if album_obj else ''
            release = t.get('releaseDate') or t.get('publishTime') or t.get('pubtime') or ''
            if name:
                out.append((name, artist, album, release))
        if len(songs) < rn:
            break
    return out


def fetch_qishui_playlist(code):
    """从汽水音乐分享短链抓取歌单（绕过抖音签名墙）。

    汽水音乐分享页 qishui.douyin.com/s/CODE 会 302 跳转到
    music.douyin.com/qishui/share/playlist?playlist_id=...，
    服务端把歌单歌曲**直接 SSR 进 _ROUTER_DATA** 的
    loaderData.playlist_page.medias 数组里（每首含 name/artists/album/release_date），
    无需 X-Gorgon/X-Khronos 签名即可解析，实测可拿完整列表。
    """
    url = 'https://qishui.douyin.com/s/%s' % code
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Mobile/15E148 Safari/604.1',
        'Referer': 'https://qishui.douyin.com/',
    }
    r = requests.get(url, headers=headers, timeout=20, allow_redirects=True)
    r.raise_for_status()
    html = r.text
    i = html.find('_ROUTER_DATA')
    if i < 0:
        raise RuntimeError('汽水歌单页面未包含歌曲数据，请改用「手动粘贴」方式')
    start = html.find('{', i)
    depth = 0
    end = None
    for j in range(start, len(html)):
        c = html[j]
        if c == '{':
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                end = j + 1
                break
    if end is None:
        raise RuntimeError('汽水歌单数据解析失败，请改用「手动粘贴」方式')
    try:
        data = json.loads(html[start:end])
    except Exception:
        raise RuntimeError('汽水歌单数据解析失败（JSON 异常），请改用「手动粘贴」方式')
    page = (data.get('loaderData') or {}).get('playlist_page') or {}
    medias = page.get('medias') or []
    out = []
    for media in medias:
        tr = (media.get('entity') or {}).get('track') or {}
        name = tr.get('name') or tr.get('title') or ''
        arts = tr.get('artists') or tr.get('artist') or tr.get('authors') or tr.get('author') or []
        if isinstance(arts, list):
            artist = '、'.join((a.get('name', '') if isinstance(a, dict) else str(a)) for a in arts if a)
        else:
            artist = str(arts)
        album_obj = tr.get('album') or {}
        album = album_obj.get('name') or '' if isinstance(album_obj, dict) else ''
        release = ''
        if isinstance(album_obj, dict) and album_obj.get('release_date'):
            try:
                release = time.strftime('%Y-%m-%d', time.localtime(int(album_obj['release_date'])))
            except Exception:
                release = ''
        if name:
            out.append((name, artist, album, release))
    if not out:
        raise RuntimeError('汽水歌单为空或解析失败，请改用「手动粘贴」方式')
    return out


def fetch_hearts_by_link(plat, link):
    """按平台用歌单链接抓取收藏歌曲列表。"""
    pid = _resolve_playlist_link(link)[1]
    if plat == 'qq':
        return fetch_qq_playlist(pid)
    if plat == 'netease':
        return fetch_netease_playlist(pid)
    if plat == 'kugou':
        return fetch_kugou_playlist(pid)
    if plat == 'kuwo':
        return fetch_kuwo_playlist(pid)
    if plat == 'qishui':
        return fetch_qishui_playlist(pid)
    raise RuntimeError('暂不支持的平台：%s' % plat)


@app.route('/api/hearts', methods=['GET'])
def api_get_hearts():
    plat = (request.args.get('platform') or 'qq').strip()
    with _hearts_lock:
        data = _load_hearts()
    with _heart_cookie_lock:
        hcookies = _load_heart_cookies()
    if plat == 'all':
        result = {}
        for code in PLATFORM_ORDER:
            p = _hearts_platform(data, code)
            result[code] = {
                'platform': code,
                'name': PLATFORM_NAMES.get(code, code),
                'count': len(p.get('songs', []) or []),
                'updated_at': p.get('updated_at'),
                'source': p.get('source'),
                'has_cookie': bool((hcookies.get(code) or {}).get('cookie')),
            }
        return jsonify({'platforms': result})
    p = _hearts_platform(data, plat)
    songs = p.get('songs', []) or []
    ck = (hcookies.get(plat) or {}).get('cookie')
    return jsonify({
        'platform': plat,
        'count': len(songs),
        'updated_at': p.get('updated_at'),
        'source': p.get('source'),
        'has_cookie': bool(ck),
        'sample': [{'song': s.get('song', ''), 'artist': s.get('artist', '')} for s in songs[:20]],
    })


@app.route('/api/hearts/fetch', methods=['POST'])
def api_fetch_hearts():
    data = request.get_json(silent=True) or {}
    plat = (data.get('platform') or 'qq').strip()
    if plat not in PLATFORM_ORDER:
        return jsonify({'error': '未知平台：%s' % plat}), 400
    # Cookie 来源：请求体优先；否则用本机单独保存的红心 Cookie（与搜索 Cookie 分离）
    cookie = (data.get('cookie') or '').strip()
    if not cookie:
        with _heart_cookie_lock:
            hcookies = _load_heart_cookies()
        cookie = (hcookies.get(plat) or {}).get('cookie', '')
    if not cookie:
        return jsonify({'error': '请先在红心管理页保存该平台的登录 Cookie'}), 400
    try:
        favs = fetch_fav_songs(plat, cookie)
    except Exception as e:
        logger.exception('抓取 %s 红心歌单失败' % plat)
        return jsonify({'error': '抓取失败：%s' % e}), 502
    if not favs:
        return jsonify({'error': '抓取成功但未解析到歌曲，Cookie 可能已失效或「我喜欢」为空'}), 502
    with _hearts_lock:
        data_all = _load_hearts()
        p = _hearts_platform(data_all, plat)
        p['songs'] = _favs_to_dict_list(favs)
        p['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
        p['source'] = 'cookie'
        data_all[plat] = p
        _save_hearts(data_all)
    # 同步保存红心 Cookie 到独立文件（与搜索 Cookie 完全隔离）
    with _heart_cookie_lock:
        hcookies = _load_heart_cookies()
        hcookies[plat] = {'cookie': cookie, 'updated_at': time.strftime('%Y-%m-%d %H:%M:%S')}
        _save_heart_cookies(hcookies)
    return jsonify({'ok': True, 'count': len(favs), 'updated_at': p['updated_at']})


@app.route('/api/hearts/fetch_link', methods=['POST'])
def api_fetch_hearts_link():
    """用歌单分享链接自动抓取（全平台通用，不需苛刻 Cookie，最稳）。"""
    data = request.get_json(silent=True) or {}
    plat = (data.get('platform') or '').strip()
    link = (data.get('link') or '').strip()
    if plat not in PLATFORM_ORDER:
        return jsonify({'error': '未知平台：%s' % plat}), 400
    if not link:
        return jsonify({'error': '请粘贴该平台的「我喜欢」歌单分享链接'}), 400
    try:
        favs = fetch_hearts_by_link(plat, link)
    except Exception as e:
        logger.exception('歌单链接抓取 %s 失败' % plat)
        return jsonify({'error': '抓取失败：%s' % e}), 502
    if not favs:
        return jsonify({'error': '抓取成功但未解析到歌曲，请确认链接是「我喜欢」公开歌单'}), 502
    with _hearts_lock:
        data_all = _load_hearts()
        p = _hearts_platform(data_all, plat)
        p['songs'] = _favs_to_dict_list(favs)
        p['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
        p['source'] = 'link'
        data_all[plat] = p
        _save_hearts(data_all)
    return jsonify({'ok': True, 'count': len(favs), 'updated_at': p['updated_at'], 'source': 'link'})


@app.route('/api/hearts/refresh', methods=['POST'])
def api_refresh_hearts():
    data = request.get_json(silent=True) or {}
    plat = (data.get('platform') or 'qq').strip()
    if plat not in PLATFORM_ORDER:
        return jsonify({'error': '未知平台：%s' % plat}), 400
    # 刷新用本机单独保存的红心 Cookie（与搜索 Cookie 分离）
    with _heart_cookie_lock:
        hcookies = _load_heart_cookies()
    cookie = (hcookies.get(plat) or {}).get('cookie', '')
    if not cookie:
        return jsonify({'error': '没有可刷新的 Cookie，请先在红心管理页保存该平台的登录 Cookie'}), 400
    try:
        favs = fetch_fav_songs(plat, cookie)
    except Exception as e:
        logger.exception('刷新 %s 红心歌单失败' % plat)
        return jsonify({'error': '刷新失败（Cookie 可能已过期）：%s' % e}), 502
    with _hearts_lock:
        data_all = _load_hearts()
        p = _hearts_platform(data_all, plat)
        p['songs'] = _favs_to_dict_list(favs)
        p['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
        p['source'] = 'cookie'
        data_all[plat] = p
        _save_hearts(data_all)
    return jsonify({'ok': True, 'count': len(favs), 'updated_at': p['updated_at']})


@app.route('/api/hearts/manual', methods=['POST'])
def api_manual_hearts():
    data = request.get_json(silent=True) or {}
    plat = (data.get('platform') or 'qq').strip()
    text = data.get('text') or ''
    parsed = _parse_pasted_list(text)
    if not parsed:
        return jsonify({'error': '没有解析到任何歌曲，请检查粘贴内容（每行一首：歌名 - 歌手）'}), 400
    with _hearts_lock:
        data_all = _load_hearts()
        p = _hearts_platform(data_all, plat)
        existing = {(s.get('song', ''), s.get('artist', '')) for s in p.get('songs', []) or []}
        added = 0
        for s, a in parsed:
            key = (s, a)
            if key in existing:
                continue
            existing.add(key)
            p.setdefault('songs', []).append({'song': s, 'artist': a, 'raw': '%s - %s' % (s, a)})
            added += 1
        p['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
        p['source'] = p.get('source') or 'manual'
        # 手动粘贴不清空已存 cookie（若之前用 cookie 抓过，刷新仍可用）
        data_all[plat] = p
        _save_hearts(data_all)
    return jsonify({'ok': True, 'added': added, 'total': len(p.get('songs', []))})


@app.route('/api/hearts/clear', methods=['POST'])
def api_clear_hearts():
    data = request.get_json(silent=True) or {}
    plat = (data.get('platform') or 'qq').strip()
    with _hearts_lock:
        data_all = _load_hearts()
        p = _hearts_platform(data_all, plat)
        p['songs'] = []
        p['source'] = None
        p['updated_at'] = None
        # 保留 cookie 以便后续刷新
        data_all[plat] = p
        _save_hearts(data_all)
    return jsonify({'ok': True})


@app.route('/api/hearts/clear_cookie', methods=['POST'])
def api_clear_hearts_cookie():
    data = request.get_json(silent=True) or {}
    plat = (data.get('platform') or 'qq').strip()
    if plat not in PLATFORM_ORDER:
        return jsonify({'error': '未知平台'}), 400
    # 红心 Cookie 已独立存储，从 hearts_cookies.json 清除（不影响搜索用 cookies.json）
    with _heart_cookie_lock:
        hcookies = _load_heart_cookies()
        hcookies.pop(plat, None)
        _save_heart_cookies(hcookies)
    return jsonify({'ok': True})


@app.route('/api/hearts/cookie', methods=['GET'])
def api_get_heart_cookie():
    plat = (request.args.get('platform') or '').strip()
    if plat not in PLATFORM_ORDER:
        return jsonify({'error': '未知平台'}), 400
    with _heart_cookie_lock:
        hcookies = _load_heart_cookies()
    ck = (hcookies.get(plat) or {}).get('cookie')
    return jsonify({'platform': plat, 'has_cookie': bool(ck)})


@app.route('/api/hearts/cookie', methods=['POST'])
def api_save_heart_cookie():
    data = request.get_json(silent=True) or {}
    plat = (data.get('platform') or '').strip()
    cookie = (data.get('cookie') or '').strip()
    if plat not in PLATFORM_ORDER:
        return jsonify({'error': '未知平台'}), 400
    if not cookie:
        return jsonify({'error': '请填入 Cookie'}), 400
    with _heart_cookie_lock:
        hcookies = _load_heart_cookies()
        hcookies[plat] = {'cookie': cookie, 'updated_at': time.strftime('%Y-%m-%d %H:%M:%S')}
        _save_heart_cookies(hcookies)
    return jsonify({'ok': True, 'has_cookie': True})


@app.route('/api/hearts/cookie', methods=['DELETE'])
def api_delete_heart_cookie():
    data = request.get_json(silent=True) or {}
    plat = (data.get('platform') or '').strip()
    if plat not in PLATFORM_ORDER:
        return jsonify({'error': '未知平台'}), 400
    with _heart_cookie_lock:
        hcookies = _load_heart_cookies()
        hcookies.pop(plat, None)
        _save_heart_cookies(hcookies)
    return jsonify({'ok': True})


@app.route('/api/search', methods=['POST'])
def api_search():
    data = request.get_json(silent=True) or {}
    song_name = (data.get('song_name') or '').strip()
    lyricist = (data.get('lyricist') or '').strip()
    composer = (data.get('composer') or '').strip()
    performer = (data.get('performer') or '').strip()
    # 兼容旧版前端仍传 artist 字段的情况
    legacy_artist = (data.get('artist') or '').strip()
    if legacy_artist and not performer:
        performer = legacy_artist
    per_platform_limit = min(int(data.get('limit', 30)), 1000)
    # v4.27.34：前端生成的搜索标识，用于让 /api/search_progress 报告本次搜索的真实进度
    progress_sid = (data.get('search_id') or '').strip()[:64]

    if not song_name:
        return jsonify({'error': '请输入歌曲名称'}), 400

    try:
        return jsonify(_search_core(song_name, lyricist, composer, performer,
                                    per_platform_limit, progress_sid=progress_sid))
    finally:
        # 无论成功/异常都收尾，避免前端一直轮询到「进行中」
        _sp_finish(progress_sid)


@app.route('/api/search_progress')
def api_search_progress():
    """查询某次搜索的实时进度（v4.27.34）。

    前端每秒轮询一次；sid 未注册（搜索刚发出、还没进 search_all）时返回 found=False，
    前端按「准备中」显示即可，不算错误。
    """
    sid = (request.args.get('sid') or '').strip()
    with _search_progress_lock:
        p = SEARCH_PROGRESS.get(sid)
        if not p:
            return jsonify({'found': False})
        snapshot = {
            'found': True,
            'stage': p['stage'],
            'total': p['total'],            # 各平台原始条数累加（含跨关键词重复）
            'unique': p.get('unique') or 0,  # 去重后的条数（进入补全/聚合阶段才有）
            'done_tasks': p['done_tasks'],
            'total_tasks': p['total_tasks'],
            'finished': p['finished'],
            'elapsed': round(time.time() - p['started'], 1),
            'platforms': [
                {'platform_code': c, 'platform': PLATFORM_NAMES.get(c, c), 'count': n}
                for c, n in sorted(p['platforms'].items(), key=lambda kv: -kv[1])
            ],
        }
    return jsonify(snapshot)


@app.route('/api/lookup_url', methods=['POST'])
def api_lookup_url():
    """粘贴平台歌曲分享链接/短链，解析出精确歌曲 ID 并拉取完整信息。

    支持：QQ音乐（短链 c6.y.qq.com/...、直链 y.qq.com/...）、酷狗、酷我、网易云、汽水。
    返回结构同单平台搜索结果，可直接渲染 / 打标。
    """
    data = request.get_json(silent=True) or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'error': '请提供歌曲链接'}), 400

    try:
        parsed = _parse_song_url(url)
    except Exception as e:
        return jsonify({'error': '链接解析失败：%s' % e}), 400
    if not parsed:
        return jsonify({'error': '无法识别该链接所属平台或歌曲 ID，请确认是 QQ/酷狗/酷我/网易云/汽水的歌曲分享链接'}), 400

    platform_code = parsed['platform_code']
    try:
        if platform_code == 'qq':
            result = _lookup_qq_by_id(parsed.get('songid'), parsed.get('songmid'))
        elif platform_code == 'kugou':
            result = _lookup_kugou_by_hash(parsed.get('hash'))
        elif platform_code == 'kuwo':
            result = _lookup_kuwo_by_rid(parsed.get('rid'))
        elif platform_code == 'netease':
            result = _lookup_netease_by_id(parsed.get('song_id'))
        elif platform_code == 'qishui':
            result = _lookup_qishui_by_id(parsed.get('id'))
        else:
            result = None
    except Exception as e:
        print('[lookup_url][%s] 抓取失败: %s' % (platform_code, e))
        return jsonify({'error': '%s 单曲详情抓取失败：%s' % (PLATFORM_NAMES.get(platform_code, platform_code), e)}), 500

    if not result:
        return jsonify({'error': '未能从链接获取歌曲信息（%s 可能无此歌曲或接口受限）' % PLATFORM_NAMES.get(platform_code, platform_code)}), 404

    # 统一附标记 / 红心（与搜索结果一致）
    with _marks_lock:
        _marks_map = _load_marks()
    _attach_marks_to_result(result, _marks_map)
    return jsonify({'platform': result.get('platform'), 'platform_code': platform_code, 'result': result})


def _parse_song_url(url):
    """解析歌曲分享链接，返回 {platform_code, songid/songmid/hash/rid/song_id/id, song_url} 或 None。"""
    import re
    raw = (url or '').strip()
    if not raw:
        return None
    # 跟随重定向拿最终 URL（处理短链，如 c6.y.qq.com/base/fcgi-bin/u?__=XXX）
    final = raw
    try:
        resp = requests.get(raw, headers={'User-Agent': COMMON_UA, 'Referer': 'https://y.qq.com/'},
                            allow_redirects=True, timeout=15)
        if resp.url:
            final = resp.url
    except Exception:
        pass

    # QQ音乐
    if 'qq.com' in final or 'qq.com' in raw:
        songid = None
        m = re.search(r'songid=(\d+)', final) or re.search(r'songDetail/(\d+)', final)
        if m:
            songid = m.group(1)
        songmid = None
        m2 = re.search(r'songmid=([0-9A-Za-z]{8,})', final) or re.search(r'songDetail/([0-9A-Za-z]{8,})', final)
        if m2:
            songmid = m2.group(1)
        if songid or songmid:
            return {'platform_code': 'qq', 'songid': songid, 'songmid': songmid,
                    'song_url': 'https://y.qq.com/n/ryqq/songDetail/%s' % (songmid or songid)}
    # 酷狗
    if 'kugou.com' in final:
        m = re.search(r'hash=([0-9A-Fa-f]{32})', final)
        if m:
            return {'platform_code': 'kugou', 'hash': m.group(1),
                    'song_url': 'https://www.kugou.com/song/#hash=%s' % m.group(1)}
    # 酷我
    if 'kuwo.cn' in final:
        m = re.search(r'play_detail/(\d+)', final) or re.search(r'rid=(\d+)', final)
        if m:
            return {'platform_code': 'kuwo', 'rid': m.group(1),
                    'song_url': 'https://www.kuwo.cn/play_detail/%s' % m.group(1)}
    # 网易云
    if '163.com' in final or '163cn.tv' in final:
        m = re.search(r'id=(\d+)', final) or re.search(r'song/(\d+)', final)
        if m:
            return {'platform_code': 'netease', 'song_id': m.group(1),
                    'song_url': 'https://music.163.com/song?id=%s' % m.group(1)}
    # 汽水（抖音系）
    if 'qishui.douyin.com' in final or 'music.douyin.com' in final or 'douyin.com' in final:
        m = re.search(r'[?/]s=([A-Za-z0-9]+)', final)
        if m:
            return {'platform_code': 'qishui', 'id': m.group(1),
                    'song_url': 'https://qishui.douyin.com/s/%s' % m.group(1)}
    return None


def _lookup_qq_by_id(songid, songmid):
    """QQ音乐：按 songid/songmid 拉取单曲完整信息（基础信息 + 收藏量/词曲/评论/在听）。"""
    headers = {'User-Agent': COMMON_UA, 'Referer': 'https://y.qq.com/'}
    params = {'format': 'json', 'inCharset': 'utf8', 'outCharset': 'utf-8', 'platform': 'yqq.json'}
    if songid:
        params['songid'] = songid
    elif songmid:
        params['songmid'] = songmid
    else:
        return None
    resp = requests.get('https://c.y.qq.com/v8/fcg-bin/fcg_play_single_song.fcg',
                        params=params, headers=headers, timeout=15)
    d = resp.json()
    song = d.get('data')
    if isinstance(song, list):
        song = song[0] if song else {}
    if not song or not song.get('name'):
        return None
    singers = song.get('singer') or []
    performer = '/'.join(x.get('name', '') for x in singers)
    album = song.get('album')
    if isinstance(album, dict):
        album = album.get('name', '')
    result = {
        'platform': 'QQ音乐', 'platform_code': 'qq',
        'song_name': song.get('name', ''),
        'performer': performer,
        'album': album or '',
        'release_date': _ts_to_date(song.get('time_public')) if song.get('time_public') else '',
        'subtitle': song.get('subtitle', ''),
        'song_url': 'https://y.qq.com/n/ryqq/songDetail/%s' % (songmid or songid),
        'availability': '在架',
        '_songmid': song.get('mid') or songmid,
        '_songid': int(song.get('id') or songid) if (song.get('id') or songid) else None,
        'lyricist': None, 'composer': None,
        'collection_count': None, 'listening_count': None, 'comment_count': None,
        'share_count': None, 'record_label': None,
        'match_score': 100, 'match_label': '精准匹配',
    }
    # 补全收藏量 / 词曲 / 评论 / 在听
    cookie = get_cookie_string('qq')
    try:
        _fetch_qq_details([result], cookie)
    except Exception as e:
        print('[lookup_url][qq] detail fetch warn: %s' % e)
    return result


def _lookup_kugou_by_hash(hash_val):
    """酷狗：按 hash 拉单曲基础信息 + 词曲（detail fetch）。

    注意：旧的 user_get_songinfo(openapi/v3) 端点已从本沙箱 IP 返回 404/空，
    改用 m.kugou.com/app/i/getSongInfo.php（按 hash 反查真实歌名/歌手/时长，稳定可用）。
    v4.25.14 修正——否则任何"酷狗按 hash 反查"都会失败（含片段重匹配扫描的酷狗分支）。
    """
    if not hash_val:
        return None
    cookie = get_cookie_string('kugou')
    headers = {'User-Agent': COMMON_UA, 'Referer': 'https://www.kugou.com/'}
    if cookie:
        headers['Cookie'] = cookie
    try:
        info_url = 'https://m.kugou.com/app/i/getSongInfo.php'
        resp = requests.get(info_url, params={'hash': hash_val, 'cmd': 'playInfo'}, headers=headers, timeout=12)
        d = resp.json()
        authors = d.get('authors') or []
        performer = '/'.join(a.get('author_name', '') for a in authors if a.get('author_name')) or d.get('singerName', '')
        result = {
            'platform': '酷狗音乐', 'platform_code': 'kugou',
            'song_name': d.get('songName', ''),
            'performer': performer,
            'album': d.get('album_name', ''),
            'song_url': 'https://www.kugou.com/song/#hash=%s' % hash_val,
            'availability': '在架',
            '_hash': hash_val,
            'timelength': _safe_int(d.get('timelength') or d.get('duration')),
            'lyricist': None, 'composer': None,
            'collection_count': None, 'listening_count': None, 'comment_count': None,
            'share_count': None, 'record_label': None,
            'match_score': 100, 'match_label': '精准匹配',
        }
        try:
            _fetch_kugou_details([result], get_cookie_string('kugou'))
        except Exception as e:
            print('[lookup_url][kugou] detail warn: %s' % e)
        return result if result.get('song_name') else None
    except Exception as e:
        print('[lookup_url][kugou] base fetch fail: %s' % e)
        return None


def _lookup_kuwo_by_rid(rid):
    """酷我：按 rid 拉单曲基础信息。"""
    if not rid:
        return None
    headers = {'User-Agent': COMMON_UA, 'Referer': 'https://www.kuwo.cn/'}
    try:
        info_url = 'https://www.kuwo.cn/api/www/music/musicInfo'
        resp = requests.get(info_url, params={'mid': rid}, headers=headers, timeout=12)
        d = resp.json().get('data', {})
        if not d:
            return None
        singer = d.get('artist') or d.get('singer') or ''
        if isinstance(singer, list):
            singer = '/'.join(x.get('name', '') for x in singer)
        result = {
            'platform': '酷我音乐', 'platform_code': 'kuwo',
            'song_name': d.get('name', ''),
            'performer': singer,
            'album': (d.get('album') or {}).get('name', '') if isinstance(d.get('album'), dict) else (d.get('album') or ''),
            'song_url': 'https://www.kuwo.cn/play_detail/%s' % rid,
            'availability': '在架',
            '_rid': rid,
            'lyricist': None, 'composer': None,
            'collection_count': None, 'listening_count': None, 'comment_count': None,
            'share_count': None, 'record_label': None,
            'match_score': 100, 'match_label': '精准匹配',
        }
        return result if result.get('song_name') else None
    except Exception as e:
        print('[lookup_url][kuwo] base fetch fail: %s' % e)
        return None


def _lookup_netease_by_id(song_id):
    """网易云：按 id 拉单曲基础信息。"""
    if not song_id:
        return None
    headers = {'User-Agent': COMMON_UA, 'Referer': 'https://music.163.com/'}
    try:
        info_url = 'https://music.163.com/api/song/detail/'
        resp = requests.get(info_url, params={'id': song_id, 'ids': '[%s]' % song_id},
                            headers=headers, timeout=12)
        d = resp.json()
        songs = d.get('songs') or []
        if not songs:
            return None
        s = songs[0]
        artists = s.get('artists') or s.get('ar') or []
        performer = '/'.join(x.get('name', '') for x in artists)
        al = s.get('album') or s.get('al') or {}
        result = {
            'platform': '网易云音乐', 'platform_code': 'netease',
            'song_name': s.get('name', ''),
            'performer': performer,
            'album': al.get('name', '') if isinstance(al, dict) else '',
            'song_url': 'https://music.163.com/song?id=%s' % song_id,
            'availability': '在架',
            '_song_id': song_id,
            'lyricist': None, 'composer': None,
            'collection_count': None, 'listening_count': None, 'comment_count': None,
            'share_count': None, 'record_label': None,
            'match_score': 100, 'match_label': '精准匹配',
        }
        return result if result.get('song_name') else None
    except Exception as e:
        print('[lookup_url][netease] base fetch fail: %s' % e)
        return None


def _lookup_qishui_by_id(sid):
    """汽水（抖音系）：链接直查基础信息（接口复杂，先解析 ID 返回骨架）。"""
    if not sid:
        return None
    # 汽水单曲详情需渲染页 + 签名，成本高；先返回可打开的骨架，引导用户在搜索页用歌名补全
    return {
        'platform': '汽水音乐', 'platform_code': 'qishui',
        'song_name': '', 'performer': '', 'album': '',
        'song_url': 'https://qishui.douyin.com/s/%s' % sid,
        'availability': '在架',
        '_qishui_id': sid,
        'lyricist': None, 'composer': None,
        'collection_count': None, 'listening_count': None, 'comment_count': None,
        'share_count': None, 'record_label': None,
        'match_score': 100, 'match_label': '精准匹配',
        '_lookup_partial': True,
    }


def _attach_marks_to_result(result, marks_map):
    """把手动标记 / 红心附加到单个直查结果（与搜索结果处理逻辑一致）。"""
    try:
        key = _mark_key(result.get('song_name', ''), result.get('performer', ''), result.get('album', ''))
        if key and key in marks_map:
            m = marks_map[key]
            result['mark'] = {
                'mark_type': m.get('mark_type', ''),
                'tags': m.get('tags', []),
                'note': m.get('note', ''),
                'updated_at': m.get('updated_at', ''),
            }
        # 红心（若有）
        hearts = marks_map.get('_hearts', {}).get(key)
        if hearts:
            result['hearts'] = hearts
    except Exception:
        pass


def _search_core(song_name, lyricist='', composer='', performer='',
                 per_platform_limit=30, progress_sid=None):
    """搜索主流程（五平台搜索 → 评分 → 详情补全 → 跨平台回填 → 聚合 → 附标记/红心）。

    /api/search 与歌词识别的「补全平台数据」共用此函数，保证两处字段完全一致：
    收藏量 / 在听 / 评论 / 唱片公司（发行公司）/ 发行时间 / 链接 / 标记 / 红心。
    必须在 Flask 请求上下文内调用（get_cookie_string 依赖 g）。
    """
    # 搜索关键词以歌名为主，不把词曲作者/表演者塞进查询，避免平台按艺人返回一堆无关结果
    # v4.26.3：歌名字段也可能带「名称：/歌名：」等说明文字，先剥掉再当主关键词
    song_name_clean = _strip_annotation_label(song_name)
    keyword = song_name_clean
    # 同时给一份"剥括号"的关键词作为兜底（带副标题的"歌名（副标题）"会让网易云/汽水等
    # 严匹配平台卡死或错位；剥括号版本可以命中真正的歌——结果按平台内部 ID 去重合并，
    # 不会重复打分）。
    stripped_keyword = _normalize_match(song_name_clean)
    alt_keywords = []
    if stripped_keyword and stripped_keyword != keyword:
        alt_keywords.append(stripped_keyword)
    # 组合搜索托底（v4.17.6 + v4.25 增强）：用户提供歌手/词曲作者时，额外用
    # 「歌名 + 作者」组合关键词再搜一遍。小众歌（如崔金水版《将军》在 QQ 搜"将军"
    # 结果里排第 3 页、前 30 条抓不到）在「将军 崔金水」的搜索结果里会直接排到第 1 条，
    # 从而进入候选池。组合结果按平台内部 ID 与主关键词结果去重合并，不会重复打分。
    # v4.25 增强：
    #   a) 表演者用 _split_artist_parts 拆解「本名/艺名/别名」（如「徐嘉玲（艺名：小嘉玲）」
    #      → [徐嘉玲, 小嘉玲]），每个名字各生成一个组合，避免整串艺名去搜平台找不到；
    #   b) 用户提供词作者/曲作者时，生成「歌名+词作者」「歌名+曲作者」组合，
    #      让「歌名+词曲都填对」的版本直接进候选池。
    def _push_combo(base, person):
        """把「base + person」组合加入 alt_keywords（去重，限制关键词数量防搜索爆炸）。"""
        if not base or not person:
            return
        combo = (base + ' ' + person).strip()
        if combo and combo not in alt_keywords and len(alt_keywords) < 9:
            alt_keywords.append(combo)

    if performer:
        # 表演者：只用拆解后的干净名字（本名/艺名/别名）各生成组合，
        # 不再把整串带注解的原文（如「徐嘉玲（艺名：小嘉玲）」）拿去搜，避免平台搜不到
        for pname in _split_artist_parts(performer)[:3]:
            pname = pname.strip()
            if not pname:
                continue
            _push_combo(song_name, pname)
            if stripped_keyword:
                _push_combo(stripped_keyword, pname)
    for field_name, field_val in (('词作者', lyricist), ('曲作者', composer)):
        if field_val:
            # 词曲作者也可能多人/带括号（如「32°（副歌）」），拆出候选名逐个组合
            for pname in _split_artist_parts(field_val)[:2]:
                pname = pname.strip()
                if not pname:
                    continue
                _push_combo(song_name, pname)
                if stripped_keyword:
                    _push_combo(stripped_keyword, pname)
    alt_keywords = alt_keywords if alt_keywords else None
    results = search_all(keyword, per_platform_limit, alt_keywords=alt_keywords,
                         progress_sid=progress_sid)

    # 初始化详情字段抓取状态：搜索已带回的字段记 ok，缺失的先记 missing。
    # 之后 QQ 详情补全会按真实抓取结果覆盖为 ok / missing / error。
    for r in results:
        st = r.setdefault('_detail_status', {})
        for f in DETAIL_FIELDS:
            if f not in st:
                st[f] = 'ok' if r.get(f) not in (None, '', []) else 'missing'

    # 计算匹配度（歌曲名相似度为核心，辅助字段仅作加权）
    for r in results:
        score, label = _score_result(r, song_name, lyricist, composer, performer)
        r['match_score'] = score
        r['match_label'] = label

    # 按匹配度重新排序
    results = _sort_by_relevance(results)

    # 详情补全（必须在排序之后做，确保用户看到的前 N 条结果有完整信息）。
    # v4.27.33：补全条数随用户每平台上限放大，但框死在 SEARCH_ENRICH_CAP(100) 以内，
    # 选 100/500 时补齐前 100 条的词曲/发行方等，其余长尾结果保留搜索接口自带的基础字段。
    DETAIL_TOP_N = min(per_platform_limit, SEARCH_ENRICH_CAP)
    # 注意：这里报 unique（跨关键词去重后的条数）而不是覆盖 total，
    # 否则前端会看到「已抓到 808 条」突然变 706 条，像是丢结果。
    _sp_stage(progress_sid, 'enrich', unique=len(results))
    qq_results = [r for r in results if r.get('platform_code') == 'qq'][:DETAIL_TOP_N]
    if qq_results:
        qq_cookie = get_cookie_string('qq')
        _fetch_qq_details(qq_results, qq_cookie)
        _fetch_qq_album_labels(qq_results, qq_cookie)

    # 跨平台回填：同一首歌的唱片公司/词曲作者，可由其他平台补全
    _cross_platform_fill(results)

    # 关键：详情和跨平台回填后，词曲作者/表演者信息才完整。
    # 用完整信息重新计算匹配度并排序，确保「歌名+词曲作者」都命中的结果排在最前面。
    for r in results:
        score, label = _score_result(r, song_name, lyricist, composer, performer)
        r['match_score'] = score
        r['match_label'] = label
    results = _sort_by_relevance(results)

    # 统计各平台数量
    platform_stats = {}
    for r in results:
        code = r['platform_code']
        if code not in platform_stats:
            platform_stats[code] = {
                'platform': r['platform'],
                'platform_code': code,
                'count': 0,
                'status': 'success',
            }
        platform_stats[code]['count'] += 1

    for code in PLATFORM_FUNCS:
        if code not in platform_stats:
            platform_stats[code] = {
                'platform': PLATFORM_NAMES[code],
                'platform_code': code,
                'count': 0,
                'status': 'no_result',
            }

    # 收集平台级告警（如汽水 Cookie 失效），读取后清空
    warnings = []
    for code, msg in PLATFORM_WARNINGS.items():
        if msg:
            warnings.append({'platform_code': code, 'message': msg})
    PLATFORM_WARNINGS.clear()

    # 按 歌名+表演者+专辑 聚合，一行展示五大平台数据
    _sp_stage(progress_sid, 'group', unique=len(results))
    grouped_results = _group_results(results)

    # 附加上次手动标记的标记（从用户目录读取，跨版本/跨打开保留）
    with _marks_lock:
        _marks_map = _load_marks()
    for g in grouped_results:
        key = _mark_key(g.get('song_name', ''), g.get('performer', ''), g.get('album', ''))
        if key in _marks_map:
            g['mark'] = _marks_map[key]

    # 附加上次识别的平台红心歌单（QQ 音乐等），命中即在结果打标
    with _hearts_lock:
        _hearts_data = _load_hearts()
    _hearts_index = {}
    for plat in HEART_PLATFORMS:
        p = _hearts_platform(_hearts_data, plat)
        if p.get('songs'):
            _hearts_index[plat] = _build_hearts_index(p)
    for g in grouped_results:
        g['hearts'] = _match_hearts(g.get('song_name', ''), g.get('performer', ''),
                                       g.get('album', ''), g.get('release_date', ''),
                                       _hearts_index)

    # 为前端「重抓」功能保留必要的平台内部 ID，其余清理掉避免污染响应
    _REFETCH_ID_FIELDS = {
        'qq': ('_songid', '_songmid', '_albummid'),
        'kugou': ('_hash', '_album_hash'),
        'kuwo': ('_rid', '_album_id'),
        'netease': ('_song_id',),
        'qishui': (),
    }
    for r in results:
        code = r.get('platform_code')
        refetch_ids = {}
        for f in _REFETCH_ID_FIELDS.get(code, ()):
            val = r.get(f)
            if val:
                refetch_ids[f] = val
        if refetch_ids:
            r['_refetch_ids'] = refetch_ids

    # 清理其它内部字段（⚠️ _rid 不再清掉：保留酷我歌曲 ID 用于「重抓」/跨次搜索去重）
    _INTERNAL_FIELDS = ('_hash', '_album_id', '_song_id', '_mixsongid', '_album_hash')
    for r in results:
        for f in _INTERNAL_FIELDS:
            r.pop(f, None)
    for g in grouped_results:
        for r in g.get('platform_data', {}).values():
            for f in _INTERNAL_FIELDS:
                r.pop(f, None)

    # v4.22：应用合并映射——用户手动合并的歌曲在展示层并成一首（跨 group 聚合）
    grouped_results = _apply_merge_map(grouped_results)

    # v4.23：过滤「完全无匹配」的结果。
    # 平台搜索接口经常会带「猜你喜欢」「兜底推荐」式的无关结果（典型：搜「回家」时
    # QQ/酷狗/酷我会顺带返回「有爱就不怕 · 庄心妍」「秋风吹起」之类 score=0 的歌）。
    # 评分系统已判 0 分并打「低相关」标签，但旧版照样塞进结果流——既浪费用户时间，
    # 又把「精准/高近似/相关」真正有用的结果淹到视野外。
    # 折叠规则：完全命中 / 命中≥1 维（即使歌名不对：同艺人其他作品、找翻唱找同人有
    # 用）/ 歌名匹配≥20 分 → 主结果流展示；其余（0 命中 + score<20）→ 默认折叠。
    qualified_results = []
    filtered_irrelevant = []
    for g in grouped_results:
        is_complete = bool(g.get('_complete_match'))
        aux_hits = int(g.get('_aux_hits', 0) or 0)
        score = int(g.get('match_score', 0) or 0)
        if is_complete or aux_hits >= 1 or score >= 20:
            qualified_results.append(g)
        else:
            # 标记一下被折叠的原因，便于前端做"为何展示"提示
            g['_irrelevant_reason'] = '平台返回的兜底推荐结果，与搜索词无关联'
            filtered_irrelevant.append(g)

    return {
        'total': len(results),
        'total_songs': len(grouped_results),
        'qualified_count': len(qualified_results),
        'filtered_irrelevant_count': len(filtered_irrelevant),
        # 给前端最多 20 条用于"展开"折叠区——再多了也没人翻
        'filtered_irrelevant': filtered_irrelevant[:20],
        'keyword': keyword,
        'platform_stats': list(platform_stats.values()),
        'warnings': warnings,
        'results': results,
        'grouped_results': qualified_results,
    }


@app.route('/api/merge_songs', methods=['POST'])
def api_merge_songs():
    """v4.22：手动合并若干首歌（因艺名写法不同被拆成多条）为一条。

    以「正主」(canonical_key) 为准：合并后展示层并成一首，标记迁移到正主。
    请求体：{canonical_key, alias_keys:[...], note}
    """
    try:
        data = request.get_json(force=True, silent=True) or {}
        ck = (data.get('canonical_key') or '').strip()
        aks = data.get('alias_keys') or []
        aks = [str(k).strip() for k in aks if k]
        note = data.get('note') or ''
        ok, msg = _do_merge(ck, aks, note)
        return jsonify({'ok': ok, 'msg': msg})
    except Exception as e:
        logger.exception('合并歌曲失败')
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route('/api/unmerge_songs', methods=['POST'])
def api_unmerge_songs():
    """v4.22：撤销合并。请求体：{canonical_key} 或 {key} 或 {alias_key}。"""
    try:
        data = request.get_json(force=True, silent=True) or {}
        key = (data.get('canonical_key') or data.get('key') or data.get('alias_key') or '').strip()
        ok, msg = _do_unmerge(key)
        return jsonify({'ok': ok, 'msg': msg})
    except Exception as e:
        logger.exception('撤销合并失败')
        return jsonify({'ok': False, 'msg': str(e)}), 500


# ── v4.24.1：本机自重启接口（免 sudo，供 WorkBuddy 沙箱 / 前端按钮触发）──
# 57074 由系统级 LaunchDaemon 以 root 运行，沙箱（toya）无法 sudo / kill / launchctl，
# 故改为「进程自己重启自己」：收到本地 + token 正确的请求后，原地 execv 重新加载磁盘
# 上的最新 app.py（launchd 跟踪的 PID 不变，只是镜像被替换）。execv 失败时退化为
# SIGTERM 让 launchd 的 KeepAlive 拉起新进程。从此代码更新后无需人工 sudo kickstart。
_RESTART_TOKEN_FILE = os.path.join(_MARKS_DIR, 'restart_token.txt')


def _get_restart_token():
    """读/生成本机重启 token（仅本机可用，权限 600）。"""
    p = _RESTART_TOKEN_FILE
    try:
        if os.path.exists(p):
            with open(p, 'r') as f:
                t = f.read().strip()
                if t:
                    return t
    except Exception:
        pass
    t = secrets.token_hex(16)
    try:
        with open(p, 'w') as f:
            f.write(t)
        os.chmod(p, 0o600)
    except Exception:
        pass
    return t


@app.route('/api/restart_token', methods=['GET'])
def api_restart_token():
    """仅本机可获取重启 token（供前端按钮 / 自动化脚本两步触发自重启）。"""
    if (request.remote_addr or '') not in ('127.0.0.1', '::1'):
        return jsonify({'error': 'only local'}), 403
    return jsonify({'token': _get_restart_token()})


@app.route('/api/self_restart', methods=['POST'])
def api_self_restart():
    """本机自重启：仅限 127.0.0.1 + 正确 token，免 sudo。

    机制（适配 57074 由系统 LaunchDaemon 托管、KeepAlive=true 的架构）：
    给当前进程发 SIGTERM → launchd 检测到退出后在 ~10s 内拉起**全新**进程，
    端口被完全释放后由新进程干净绑定，并加载磁盘上的最新代码。
    注意：不能用 os.execv 原地换镜像——execv 会继承旧监听 socket，触发「端口守卫」
    判定「同版本已占用」而直接退出，反而把服务搞挂。SIGTERM + launchd 重启最稳。"""
    if (request.remote_addr or '') not in ('127.0.0.1', '::1'):
        return jsonify({'error': 'only local'}), 403
    auth = request.headers.get('X-Restart-Token') or (request.get_json(silent=True) or {}).get('token') or ''
    expected = _get_restart_token()
    if not auth or not secrets.compare_digest(auth, expected):
        return jsonify({'error': 'unauthorized'}), 403

    def _do_restart():
        import time as _t
        import signal as _sig
        _t.sleep(0.4)
        # 交给 launchd 的 KeepAlive 重启（生产 57074 由系统 LaunchDaemon 托管）
        try:
            os.kill(os.getpid(), _sig.SIGTERM)
        except Exception:
            pass

    threading.Thread(target=_do_restart, daemon=True).start()
    return jsonify({'ok': True, 'msg': 'restarting', 'version': APP_VERSION})


@app.route('/api/refetch', methods=['POST'])
def api_refetch():
    """对单条平台记录强制补全详情（用于用户看到「平台无数据」或「词曲为空」时手动触发重抓）。

    通用策略（5 平台：QQ/酷狗/酷我/网易云/汽水）：
      1) 优先用 row._refetch_ids 调详情接口补全（QQ/酷狗/酷我/网易云）
      2) 没 ID 或详情接口失败时，用 song_name+performer 反查平台搜索，匹配后用新结果替换
      3) 完全找不到 → 返回 400 + 建议「重搜一次」
    """
    data = request.get_json(silent=True) or {}
    code = data.get('platform_code')
    row = data.get('row') or {}
    refetch_ids = row.get('_refetch_ids') or {}

    if not code or not row:
        return jsonify({'error': '缺少平台或记录信息'}), 400

    if code not in ('qq', 'kugou', 'kuwo', 'netease', 'qishui'):
        return jsonify({'error': f'未知平台：{code}'}), 400

    # ── 1) 先用 ID 尝试详情补全（QQ/酷狗/酷我/网易云；汽水无独立详情接口）──
    enriched = _try_fetch_detail_by_id(code, row, refetch_ids)

    # ── 2) 详情接口不可用 / 失败 / 词曲仍空 → 反查搜索兜底 ──
    need_fallback = (
        not enriched
        or (
            (row.get('lyricist') or '').strip() and (row.get('composer') or '').strip()
            and enriched.get('lyricist') == row.get('lyricist')
            and enriched.get('composer') == row.get('composer')
            and not enriched.get('_refetch_ids')
        )
    )
    # 简化判定：详情补全后仍缺词曲 OR 根本没详情接口 → 走搜索兜底
    if (not enriched) or (not (enriched.get('lyricist') and enriched.get('composer'))
                          and code in ('qq', 'kugou', 'kuwo', 'netease')):
        fallback_row = _refetch_via_search(code, row)
        if fallback_row:
            enriched = fallback_row

    if not enriched:
        return jsonify({'error': f'未能补全{PLATFORM_NAMES.get(code, code)}数据：建议重搜一次'}), 400

    # 保留/补全 _refetch_ids 方便再次重抓
    if not enriched.get('_refetch_ids'):
        existing = dict(refetch_ids or {})
        # 把 row 上的 ID 字段也保留进来（如果有的话）
        for f in ('_songid', '_songmid', '_albummid', '_hash', '_album_hash',
                  '_rid', '_album_id', '_song_id', '_mixsongid'):
            v = enriched.get(f) or row.get(f)
            if v:
                existing[f] = v
        if existing:
            enriched['_refetch_ids'] = existing

    return jsonify({'row': enriched})


def _try_fetch_detail_by_id(code, row, refetch_ids):
    """用 row._refetch_ids 调平台详情接口补全。返回新 row 或 None。"""
    try:
        if code == 'qq':
            songid = refetch_ids.get('_songid') or row.get('_songid')
            songmid = refetch_ids.get('_songmid') or row.get('_songmid')
            albummid = refetch_ids.get('_albummid') or row.get('_albummid')
            if not (songid or songmid):
                return None
            tmp = {**row,
                   '_songid': songid, '_songmid': songmid, '_albummid': albummid,
                   '_detail_status': row.get('_detail_status', {})}
            qq_cookie = get_cookie_string('qq')
            _fetch_qq_details([tmp], qq_cookie)
            _fetch_qq_album_labels([tmp], qq_cookie)
            tmp['_refetch_ids'] = {**refetch_ids, '_songid': songid, '_songmid': songmid}
            if albummid:
                tmp['_refetch_ids']['_albummid'] = albummid
            return tmp

        if code == 'kugou':
            hash_val = refetch_ids.get('_hash') or row.get('_hash')
            album_hash = refetch_ids.get('_album_hash') or row.get('_album_hash')
            if not hash_val:
                return None
            tmp = {**row, '_hash': hash_val, '_album_hash': album_hash,
                   '_detail_status': row.get('_detail_status', {})}
            cookie_str = get_cookie_string('kugou')
            _fetch_kugou_details([tmp], cookie_str)
            tmp['_refetch_ids'] = {**refetch_ids, '_hash': hash_val}
            if album_hash:
                tmp['_refetch_ids']['_album_hash'] = album_hash
            return tmp

        if code == 'kuwo':
            rid = refetch_ids.get('_rid') or row.get('_rid')
            # 兜底 1：从 song_url 抽
            if not rid:
                url = (row.get('song_url') or row.get('link') or '').strip()
                m = re.search(r'/play_detail/([A-Za-z0-9_-]+)', url)
                if m:
                    rid = m.group(1)
                    print(f"[Kuwo][refetch] 从 song_url 抽出 rid={rid}")
            if not rid:
                return None
            tmp = {**row, '_rid': rid, '_detail_status': row.get('_detail_status', {})}
            _fetch_kuwo_lyricist_http([tmp])
            if not (tmp.get('lyricist') and tmp.get('composer')):
                try:
                    token = _kuwo_token()
                    cookie_str = get_cookie_string('kuwo')
                    _fetch_kuwo_details([tmp], token, cookie_str)
                except Exception as e:
                    print(f"[Kuwo][refetch] details error: {e}")
            tmp['_refetch_ids'] = {**refetch_ids, '_rid': rid}
            return tmp

        if code == 'netease':
            song_id = refetch_ids.get('_song_id') or row.get('_song_id')
            if not song_id:
                return None
            tmp = {**row, '_song_id': song_id,
                   '_detail_status': row.get('_detail_status', {})}
            cookie_str = get_cookie_string('netease')
            _fetch_netease_details([tmp], cookie_str)
            tmp['_refetch_ids'] = {**refetch_ids, '_song_id': song_id}
            return tmp

        # 汽水：搜索结果本身已含词曲，无独立详情接口
        if code == 'qishui':
            return None
    except Exception as e:
        print(f"[Refetch][{code}] 详情补全失败: {e}")
        return None
    return None


def _refetch_via_search(code, row):
    """无 ID 或详情补全失败时，用 song_name+performer 反查平台搜索，匹配后用新结果替换。

    5 平台通用。匹配策略：
      ① 歌名相同 + performer 包含 query.performer → 最佳
      ② performer 不命中但歌名相同 → 次选
      ③ 仍无 → 返回 None
    """
    sn = (row.get('song_name') or '').strip()
    pf = (row.get('performer') or '').strip()
    if not sn:
        return None
    try:
        if code == 'qq':
            hits = search_qq(sn)
        elif code == 'kugou':
            hits = search_kugou(sn, light=False)
        elif code == 'kuwo':
            hits = search_kuwo(sn)
        elif code == 'netease':
            hits = search_netease(sn, light=False)
        elif code == 'qishui':
            hits = search_qishui(sn)
        else:
            return None
    except Exception as e:
        print(f"[Refetch][{code}] 反查搜索失败: {e}")
        return None

    if not hits:
        return None

    # 匹配策略：
    #   ① 歌名相同 + performer 包含 row.performer → 最佳
    #   ② 歌名相同（performer 略不同）→ 次选，但保留 row.performer 作为正解
    #   ③ 仍无 → 返回 None
    def _song_ok(h):
        from_search = (h.get('song_name') or '').strip()
        return from_search == sn or from_search.replace(' ', '') == sn.replace(' ', '')

    hit = next((h for h in hits if _song_ok(h) and pf and pf in (h.get('performer') or '')), None)
    if not hit:
        hit = next((h for h in hits if _song_ok(h)), None)
    if not hit:
        return None

    # 若 performer 略不同，row.performer 优先（用户认定的正解）
    pf_resolved = pf if pf else (hit.get('performer') or '').strip()
    print(f"[Refetch][{code}] 反查命中: {hit.get('song_name')} - {hit.get('performer')} | 词={hit.get('lyricist')} 曲={hit.get('composer')}{' (performer 被 row 覆盖)' if pf and pf not in (hit.get('performer') or '') else ''}")

    # 把 row 字段用新命中结果覆盖（保留 _refetch_ids 用于再次重抓）
    new_row = {**row}
    for k in ('song_name', 'album', 'lyricist', 'composer',
              'song_url', 'link', 'duration', 'release_date',
              'record_label', 'collection_count', 'listening_count', 'comment_count',
              'availability', '_songid', '_songmid', '_albummid',
              '_hash', '_album_hash', '_rid', '_album_id',
              '_song_id', '_mixsongid'):
        v = hit.get(k)
        if v is not None:
            new_row[k] = v
    # performer 优先用 row 的
    new_row['performer'] = pf_resolved
    # 收集 _refetch_ids
    rids = {}
    for f in ('_songid', '_songmid', '_albummid', '_hash', '_album_hash',
              '_rid', '_album_id', '_song_id', '_mixsongid'):
        v = hit.get(f)
        if v:
            rids[f] = v
    if rids:
        new_row['_refetch_ids'] = rids
    return new_row


# ═══════════════════════════════════════════════════
#  批量歌单工具（独立页面 /batch）
# ═══════════════════════════════════════════════════
# 单轮搜索超时。2026-08-04 从 30 提到 60：search_all 末尾新增 QQ 详情补全
# （粉丝数/评论数/在听/词曲）后单条耗时升到 12-20 秒，并发 6 时更慢，
# 30 秒会让「刘德华/信乐团/罗大佑」这类结果多的老歌批量超时被误判成「查询失败」。
BATCH_SONG_TIMEOUT = 60
BATCH_PLATFORMS = ['qq', 'kugou', 'kuwo', 'netease', 'qishui']
BATCH_PLAT_NAMES = {'qq': 'QQ音乐', 'kugou': '酷狗', 'kuwo': '酷我', 'netease': '网易云', 'qishui': '汽水'}

_INVALID_LABELS = {'', '未确定', '未明确', '未知', '暂无', '无', 'none', 'na', 'n/a', 'tbd', '待定', 'null'}

# 艺人「装饰尾缀」白名单（v4.25.x 根因修复·②f 自动折叠用）。
# 仅这些固定小尾缀算「同一艺人的账号装饰」，绝不放开任意 CJK 尾缀（防任意蹭名冒名误并；
# 个别已确认本人但带 CJK 尾缀的账号如 胡艾彤→胡艾彤喂猪吗 走 ②c 显式白名单认回）。
# 注意：经 _artist_core_strip 归一后括号已被吞，故「(微唱)」在核里是「微唱」；大小写归一。
_DECOR_ARTIST_SUFFIXES = {
    '呀', '啊', '阿', '呢', '呐', '喔', '哦', '呦', '哟',
    '微唱', '翻自', '翻唱', 'cover', 'live',
}


def _enrich_result(r, code):
    """picker 的按需详情补全分发器（词曲/歌词演唱者），按平台调用对应详情接口。

    batch 路径的 _pick_best_for_batch 之前一直没传 enrich，导致 ②d/②g 的
    「搜索期未富集到歌词 → 按需补抓」兜底在批量下完全失效。接上后，目标候选
    在 ns>=85 且缺歌词字段时会现场补抓一次，改名/多账号歌得以认回。
    """
    try:
        cookie = get_cookie_string(code)
    except Exception:
        cookie = ''
    if code == 'qq':
        _fetch_qq_details([r], cookie)
    elif code == 'kugou':
        _fetch_kugou_details([r], cookie)
    elif code == 'netease':
        _fetch_netease_details([r], cookie)


def _valid_label(lab):
    if not lab:
        return False
    return str(lab).strip().lower() not in _INVALID_LABELS


def _pick_best_for_batch(results, code, song_name, performer,
                         lyricist='', composer='', enrich=None):
    """某平台结果中：优先选取【表演者命中（目标艺人）】的版本（组内取收藏量最高）；
    仅在无任何目标艺人版本时，才退而取歌名匹配的其它版本。返回 (result_or_None, label)。

    v4.25.2 强化（解决「其他:xxx / 存疑 数据全错」）：
      - 歌手名归一：结果歌手尾部带账号装饰（蒋雪儿Snow.J / 周杰伦.）→ 剥掉后与曲库表演者
        核比对，同名小艺人账号认回为「精准匹配(目标艺人)」。
      - 超级大艺人智能排除：结果歌手命中大艺人名单、而曲库本歌表演者并非该大艺人时，
        视为同名撞车误配 → 跳过（判未收录），绝不把大艺人收藏量算到小艺人头上。
        若曲库本就标了大艺人（如《晴天》→周杰伦），则正常保留（「智能」所在）。
      - 四维校验（best-effort）：对「歌手名相似但未精确(asc 70~99)」的最佳候选，若曲库
        已填词+曲作者，则补抓该候选的词曲详情做四维比对；词曲均命中 → 升「精准匹配(四项全中)」。
      v4.25.x 起 艺人变体白名单：同艺人用不同账号发布（钦觉 / 钦觉呀 / 钦觉(微唱)）
        这类场景，_artist_core_strip 因 CJK 尾缀严格不动会漏。允许通过 admin 端点
        维护「艺人变体表」显式认回，picker 命中后升 asc=100、标签改为「匹配(艺人变体:xxx)」，
        既不与 ②b 变体写法归一冲突，也不破坏 v4.25.16 CJK 尾缀防线（白名单是用户显式确认）。
    """
    cands = [r for r in results if r.get('platform_code') == code]
    if not cands:
        return None, '无结果'
    perf = (performer or '').strip()
    perf_core = _artist_core_strip(perf) if perf else ''
    perf_alias = _alias_variants(perf) if perf else ''
    lyr = (lyricist or '').strip()
    comp = (composer or '').strip()

    def _mega_hit(val):
        return _artist_core_strip(val) in _MEGA_ARTISTS

    scored = []
    for r in cands:
        ns, _ = _song_name_score(song_name, r.get('song_name', ''))
        res_perf = (r.get('performer') or '').strip()
        # ① 歌手分级打分（含归一升级）
        asc = _artist_match_score(perf, res_perf) if perf else 0
        # 归一升级：剥掉账号尾缀后，结果歌手核 == 曲库表演者核 → 视精准
        if perf and asc < 100 and _artist_core_strip(res_perf) == perf_core:
            asc = 100
        # ② 超级大艺人智能排除：曲库非大艺人 + 结果撞大艺人 → 同名误配，跳过
        if perf and asc < 100 and _mega_hit(res_perf) and perf_core not in _MEGA_ARTISTS:
            continue
        # ②b 变体写法归一（v4.25.15）：dollyy / dollyy99 / Dollyy 认作同一人。
        # 仅当「歌名强匹配(ns>=80)」或「歌名>=60 且词/曲作者任一严格命中」时才升精准，
        # 防「同变体壳子但其实是另一首歌/另一人」误合并（用户明示：歌手变体+歌名+作者 三维）。
        # 必须放在大艺人排除之后，避免把大艺人账号序号变体也认回。
        if perf and asc < 100 and perf_alias:
            res_alias = _alias_variants(res_perf)
            if res_alias and res_alias == perf_alias:
                _author_hit = (bool(lyr) and _person_set_strict(lyr, r.get('lyricist', '') or '')) or \
                              (bool(comp) and _person_set_strict(comp, r.get('composer', '') or ''))
                if ns >= 80 or (ns >= 60 and _author_hit):
                    asc = 100
        # ②c 艺人变体白名单（v4.25.x）：用户显式确认「同一艺人多账号发布」（钦觉/钦觉呀/
        # 钦觉(微唱)）的场景。CJK 尾缀（呀/啊/呢/哦）被 ②b 严守不放行，但人工确认后可放开。
        # 硬性约束：必须「歌名匹配(ns>=60) + 时长 >=60s（防试听片段冒名）」，避免把短片段
        # 顶替到主歌上；放在大艺人排除之后，避免蹭大艺人变体被认回。
        if perf and asc < 100 and not _mega_hit(res_perf) and _performer_is_alias(perf, res_perf):
            _dur = _safe_int(r.get('duration')) or 0
            if ns >= 60 and (_dur == 0 or _dur >= 60):
                asc = 100
        # ②f 装饰尾缀自动折叠（Case A：同艺人 + 中文装饰尾缀，如 钦觉→钦觉呀 / 钦觉(微唱) / 钦觉啊）
        # 用「小而固定的装饰尾缀白名单」判定，绝不放行任意 CJK 尾缀（已确认本人带 CJK 尾缀者走 ②c 白名单）。
        # 仅当「结果核 = 目标核 + 已知装饰尾缀」且歌名能搭上(ns>=60) 才认回，安全且无需手工白名单。
        if perf and asc < 100 and not _mega_hit(res_perf) and ns >= 60:
            res_core = _artist_core_strip(res_perf)
            if res_core != perf_core and res_core.startswith(perf_core):
                if res_core[len(perf_core):] in _DECOR_ARTIST_SUFFIXES:
                    asc = 100
        # ②d 作者铁证自动认（v4.25.x 根因修复：艺人改名/多账号发歌，无需手动白名单）
        # 当候选「歌名强匹配(ns>=85) + 真实词曲作者与曲库已知作者一致」即证明
        # 「同一首歌=同一人」，升 asc=100 自动认回。覆盖两类：
        #   - 中文装饰尾缀（钦觉/钦觉呀，原本 asc=70 子串沾边）
        #   - 彻底改艺名（叶落→沐野/优秀少年好好→好好（张轩睿），原本 asc=0 完全不沾边）
        # 冒名账号（歌不对/作者不对/悬空分隔符蹭名）自然不满足 → 仍判未收录或被建议队列捕获。
        # 优先用搜索结果里已有的词曲（kugou search 普遍带 lyr/comp），无需 enrich；只有
        # 搜索没带时再 enrich 补抓（QQ/网易云）。覆盖了批量流程不传 enrich 的场景。
        if perf and asc < 100 and not _mega_hit(res_perf) and ns >= 85:
            _dangling = bool(re.search(r'[/&,，、;]\s*$', res_perf) or re.search(r'[/&,，、;]\s*[/&,，、;]', res_perf))
            if not _dangling:
                # ②d-pre：搜索结果里已有词曲 -> 直接比对（kugou 走这条）
                _res_lyr = (r.get('lyricist') or '').strip()
                _res_cmp = (r.get('composer') or '').strip()
                l_ok_pre = bool(lyr) and bool(_res_lyr) and _person_set_strict(lyr, _res_lyr)
                c_ok_pre = bool(comp) and bool(_res_cmp) and _person_set_strict(comp, _res_cmp)
                if (lyr and comp and l_ok_pre and c_ok_pre) \
                        or (lyr and not comp and l_ok_pre) \
                        or (comp and not lyr and c_ok_pre):
                    asc = 100
                # ②d-enrich：搜索结果没带 -> enrich 补抓再比（QQ/网易云）
                if asc < 100 and callable(enrich):
                    try:
                        enr = enrich(r, code)
                        if enr:
                            l_ok = (bool(lyr) and _person_set_strict(lyr, enr.get('lyricist', '') or ''))
                            c_ok = (bool(comp) and _person_set_strict(comp, enr.get('composer', '') or ''))
                            if (lyr and comp and l_ok and c_ok) or (lyr and not comp and l_ok) or (comp and not lyr and c_ok):
                                r = enr  # 采纳补抓后含真实作者的候选，升精准
                                asc = 100
                    except Exception:
                        pass
        # ②g 歌词演唱者认回（v4.28.x 根因修复）：歌词头部「歌名 - 歌手」或制作信息行
        # 内嵌的 演唱/原唱/和声 名，直接揭示该曲真实演唱者。当曲库表演者与歌词揭示的
        # 演唱者一致（且歌名强匹配 ns>=85），即可认回改名/多账号艺人——既不依赖别名白名单，
        # 也不依赖 ②d 词曲解析成功。是 ②d 失效（歌词词曲未抓到/未解析出）时的关键兜底。
        # 典型：曲库写「任芯冉」，平台结果歌手显示「焦七七」，但歌词首行「醉花间 - 任芯冉」
        # 与「和声：任芯冉/化十」坐实同一人 → 升精准。
        if perf and asc < 100 and not _mega_hit(res_perf) and ns >= 85:
            _lps = r.get('_lyric_performers') or []
            if not _lps and callable(enrich):
                # 搜索期未富集到歌词时，按需补抓（与 ②d-enrich 同机制）
                try:
                    enrich(r, code)
                    _lps = r.get('_lyric_performers') or []
                except Exception:
                    pass
            for _lp in _lps:
                if _lp and _performer_match(perf, _lp):
                    asc = 100
                    break
        # ②e 待确认建议记录：歌手名沾边(asc==70) 或 两者互为子串（彻底改名也含此）
        # 但既没被白名单、也没被作者铁证认回 → 记一条建议，admin 后台一键批准并入白名单。
        # 不自动认：未进白名单、也无作者铁证的「CJK 尾缀」仍交人工审（防冒名）；
        # 已确认本人(如 胡艾彤→胡艾彤喂猪吗) 走 ②c 白名单自动认回，不在此分支。
        # v4.27.9：走多表演者拆分版，不再把「Kui Kui / 徐梦圆」整体当成 Kui Kui 的别名候选
        if perf and asc < 100 and not _mega_hit(res_perf) and res_perf != perf and not _performer_is_alias(perf, res_perf):
            _dangling = bool(re.search(r'[/&,，、;]\s*$', res_perf) or re.search(r'[/&,，、;]\s*[/&,，、;]', res_perf))
            if not _dangling:
                np_ = _normalize_match(perf)
                nr_ = _normalize_match(res_perf)
                _similar = (asc == 70) or (np_ and nr_ and (np_ in nr_ or nr_ in np_))
                if _similar:
                    _record_alias_suggestion_split(perf, res_perf, song_name, ns)
        # 目标艺人版本放宽到 ns>=60；其它版本要求 ns>=80
        if perf:
            ok = ns >= 60 if asc else ns >= 80
        else:
            ok = ns >= 80
            # 曲库无表演者时仍要防大艺人同名撞车（如库空艺人却搜到周杰伦）
            if ok and _mega_hit(res_perf):
                continue
        if not ok:
            continue
        if perf:
            # v4.25 严格口径：默认只认「目标艺人本人(asc=100)」或「多艺人串含正主(asc>=92)」，
            # 低于此视为未收录，绝不写脏数据。但若有「四维校验」能力（enrich 且词+曲齐全），
            # 则把 asc∈[70,92) 的「存疑」候选保留下来，交给后面的词曲补抓确认能否升「四项全中」。
            if asc < 92:
                _can_4d = callable(enrich) and 70 <= asc < 100 and lyr and comp
                if not _can_4d:
                    continue
        scored.append((r, asc, ns))
    if not scored:
        # v4.26 放宽批量查询筛选：无目标艺人版本时，若候选池存在「歌名强匹配 +
        # 平台标记为已下架」的版本，降级返回并标「已下架」，而不是判「未收录」。
        # 解决「平台有这首但版权下架，搜索接口仍能返回（如 QQ alertid=11）却被
        # 严格艺人口径拒掉」的场景（典型：《明日世界终结时》QQ 灰色不可播）。
        # 只对已下架版本兜底、且要求歌名强匹配(ns>=80)，绝不把「在架」的
        # 同名他唱顶替给目标艺人。
        for r in cands:
            ns, _ = _song_name_score(song_name, r.get('song_name', ''))
            if ns >= 80 and (r.get('availability') or '') == '已下架':
                res_perf = (r.get('performer') or '').strip()
                label = f'已下架(仅歌名匹配:{res_perf})' if res_perf else '已下架(仅歌名匹配)'
                return r, label
        return None, ('未收录(非目标艺人版本)' if perf else '低相关(无匹配)')
    # 排序优先级：① 歌手匹配等级 > ② 歌名相似度 > ③ 时长维度(完整版优先) > ④ 收藏量
    scored.sort(key=lambda x: (x[1], x[2], _dur_tier(x[0].get('duration')), x[0].get('collection_count') or 0), reverse=True)
    best, best_asc, _best_ns = scored[0]

    # v4.25.12 片段救援：最佳候选明显是试听片段(<60s)时，若候选池里存在
    # 「歌名强匹配(ns>=80) + 正常完整版时长(>=90s)」的其它版本（即便歌手名是恶搞/
    # 别名变体，如『胡艾彤 (Live)』重传的完整版），则越过歌手名差异优先选完整版，
    # 避免把 30s 片段当正主写入曲库。无完整版可选时保持原选择（不误伤真·短歌）。
    # ⚠️ 必须在「全部候选 cands」里找（含被上面严格口径过滤掉的别名完整版），
    # 而非只在 scored 里找——别名完整版在 strict 过滤阶段已被排除，scored 里看不到它。
    # v4.25.16 收紧 alias_ok：候选歌手核默认仅认「等于目标核」或「目标核 + 仅装饰尾缀
    # (拉丁/数字/标点/空格)」；但「白名单显式确认的同艺人账号」(如 胡艾彤 / 胡艾彤喂猪吗) 例外认回，
    # 既防任意 CJK 尾缀冒名错配，又让已确认本人的别名完整版能顶替片段。
    best_dur = _safe_int(best.get('duration')) or 0
    if best_dur and best_dur < 60:
        target_core = _artist_core_strip(perf) if perf else ''
        for r in cands:
            if r is best:
                continue
            ns, _ = _song_name_score(song_name, r.get('song_name', ''))
            d = _safe_int(r.get('duration')) or 0
            if d < 90 or ns < 80:
                continue
            cand_core = _artist_core_strip(r.get('performer') or '')
            if not target_core:
                alias_ok = True
            elif cand_core == target_core:
                # 同核（如『胡艾彤 (Live)』归一成『胡艾彤』）→ 视为本人重传完整版
                alias_ok = True
            elif _performer_is_alias(perf, r.get('performer') or ''):
                # v4.25.x：白名单「同一艺人多账号」(如 胡艾彤 / 胡艾彤喂猪吗) 视为本人别名完整版，
                # 越过歌手名差异优先选完整版；白名单为人工显式确认，不破坏 CJK 尾缀通用防线。
                alias_ok = True
            elif cand_core.startswith(target_core):
                # v4.25.16：仅允许「目标核 + 装饰尾缀」，CJK 尾缀视为不同人（白名单已覆盖的例外在上分支）
                leftover = cand_core[len(target_core):]
                alias_ok = bool(leftover) and all(ch.isascii() or ch in '・·. ' for ch in leftover)
            else:
                alias_ok = False
            if alias_ok:
                best = r
                best_asc = _artist_match_score(perf, r.get('performer') or '') if perf else 0
                _best_ns = ns
                break

    # ③ 四维校验（best-effort）：歌手已相似(70~99)且曲库有词+曲 → 补抓候选词曲比对
    best_is_4d = False
    if (lyr and comp) and 70 <= best_asc < 100 and callable(enrich):
        try:
            enr = enrich(best, code)
            if enr:
                best = enr
                l_ok = _person_set_strict(lyr, best.get('lyricist', '') or '')
                c_ok = _person_set_strict(comp, best.get('composer', '') or '')
                if l_ok and c_ok:
                    best_asc = 100
                    best_is_4d = True
        except Exception:
            best_is_4d = False

    res_perf = (best.get('performer') or '').strip()
    if best_asc >= 100:
        # v4.25.x：best_asc 命中艺人变体（人工白名单 ②c / 装饰尾缀 ②f / 作者铁证 ②d 任一认回的
        # 改名·多账号变体）统一标注「匹配(艺人变体:xxx)」，便于导出表一眼分辨「账号变体 vs 直名命中」。
        if best_is_4d:
            label = '精准匹配(四项全中)'
        elif performer.strip() and _artist_core_strip(res_perf) != _artist_core_strip(performer):
            label = f'匹配(艺人变体:{res_perf})'
        else:
            label = '精准匹配(目标艺人)'
    elif best_asc >= 92:
        label = f'匹配(多艺人/别名:{res_perf})'
    elif best_asc >= 70:
        label = f'存疑(疑似翻唱:{res_perf})'
    else:
        label = f'其他:{res_perf}' if res_perf else '近似(无艺人)'
    # v4.25.16：最终最佳候选仍是试听片段(<60s)时打标，便于导出表一眼识别「非完整原版」
    _final_dur = _safe_int(best.get('duration')) or 0
    if _final_dur and _final_dur < 60:
        label += '[片段<60s]'
    return best, label


def _select_best_batch_v2(candidates, code, song_name, performer, lyricist='', composer=''):
    """四项匹配选择器（v4.20 新增，严格贴合用户批量查歌需求）。

    入参 candidates 必须已完成 enrich（携带 lyricist / composer / collection_count）。
    规则：
      1) 歌名 + 歌手名 + 词作者 + 曲作者 四项完全匹配 → 记为「全中」；
      2) 多个「全中」版本 → 取收藏量(collection_count)最高的登记；
      3) 无「全中」→ 退回最佳「歌名+歌手」匹配，标签注明词曲未全中，便于人工复核。
    返回 (winner_result_or_None, label)。
    """
    if not candidates:
        return None, '无结果'
    perf = (performer or '').strip()
    lyr = (lyricist or '').strip()
    comp = (composer or '').strip()

    def _name_ok(ns, singer_ok):
        # 歌手命中放宽到 60；否则要求歌名较强(>=80)避免同名误取
        return (ns >= 60) if singer_ok else (ns >= 80)

    scored = []
    for r in candidates:
        ns, _ = _song_name_score(song_name, r.get('song_name', ''))
        singer_ok = _person_set_strict(perf, r.get('performer', '')) if perf else True
        lyr_ok = _person_set_strict(lyr, r.get('lyricist', '')) if lyr else True
        comp_ok = _person_set_strict(comp, r.get('composer', '')) if comp else True
        if not _name_ok(ns, singer_ok):
            continue
        full = bool(singer_ok and lyr_ok and comp_ok)
        scored.append({
            'r': r, 'ns': ns, 'singer_ok': singer_ok,
            'lyr_ok': lyr_ok, 'comp_ok': comp_ok, 'full': full,
            'fav': _safe_int(r.get('collection_count')) or 0,
            'asc': _artist_match_score(perf, r.get('performer', '')) if perf else 0,
        })

    if not scored:
        return None, '低相关(无匹配)'

    fulls = [s for s in scored if s['full']]
    if fulls:
        # 多项全中 → 取收藏量最高（用户明确要求：同四项多首取最高收藏量版本）
        fulls.sort(key=lambda x: x['fav'], reverse=True)
        best = fulls[0]['r']
        n = len(fulls)
        label = '精准匹配(四项全中)'
        if n > 1:
            label += f'(共{n}版取最高收藏量)'
        return best, label

    # 无四项全中 → 退回最佳歌名+歌手，并标注词曲核验情况（v4.25.12：时长维度参与排序）
    scored.sort(key=lambda x: (x['asc'], x['ns'], _dur_tier(x['r'].get('duration')), x['fav']), reverse=True)
    best = scored[0]
    r = best['r']
    miss = []
    if lyr and not best['lyr_ok']:
        miss.append('词')
    if comp and not best['comp_ok']:
        miss.append('曲')
    asc = best['asc']
    if asc >= 100:
        base = '精准匹配(目标艺人)'
    elif asc >= 92:
        base = f'匹配(多艺人/别名:{r.get("performer")})'
    elif asc >= 70:
        base = f'存疑(疑似翻唱:{r.get("performer")})'
    else:
        base = f'其他:{r.get("performer")}' if r.get('performer') else '近似(无艺人)'
    if miss:
        base += f'[词曲未全中:{"".join(miss)}不符]'
    if lyr and r.get('lyricist') in (None, ''):
        base += '[词曲未核验]'
    if comp and r.get('composer') in (None, ''):
        base += '[词曲未核验]'
    return r, base


def _pick_label_for_batch(results, code, song_name, performer):
    """某平台厂牌：目标艺人版本优先取真实厂牌；过滤未确定等占位。"""
    cands = [r for r in results if r.get('platform_code') == code]
    if not cands:
        return '', '无结果'
    perf = (performer or '').strip()
    zq = [r for r in cands if perf and _person_match(perf, r.get('performer', ''))]
    pool = zq if zq else cands
    for r in pool:
        lab = r.get('record_label')
        if _valid_label(lab):
            src = (r.get('record_label_source') or ('精准匹配(目标艺人)' if r in zq else '其他版本'))
            return lab, src
    return '', ('精准匹配(目标艺人)无厂牌' if zq else '无厂牌')


def _resolve_record_label(code, results, song_name, performer, best):
    """发行公司多源兜底（修复「best 版本无厂牌即整列空白」）。

    回退顺序：① best 自身厂牌 → ② 同平台目标艺人其它有厂牌版本 →
    ③ 同平台任意有厂牌版本。均缺则返回 ''（交由 Discogs enrich 补充或如实留空）。
    """
    lab = best.get('record_label')
    if _valid_label(lab):
        return lab
    cands = [r for r in results if r.get('platform_code') == code]
    perf = (performer or '').strip()
    if perf:
        for r in cands:
            if _person_match(perf, r.get('performer', '')) and _valid_label(r.get('record_label')):
                return r['record_label']
    for r in cands:
        if _valid_label(r.get('record_label')):
            return r['record_label']
    return ''


def _search_one_timeout(keyword, timeout=BATCH_SONG_TIMEOUT):
    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(search_all, keyword, 30)
        try:
            return fut.result(timeout=timeout)
        except Exception:
            fut.cancel()
            return None


def _search_two_pass(song_name, performer):
    """两级搜索策略（2026-08-04 新增）。

    第 1 轮用「歌名 + 表演者」搜——关键词更具体，能把同名歌的干扰压下去，
    对「世界末日」「重逢」这类大众歌名尤其重要。
    第 2 轮只在有平台没搜到目标艺人版本时才触发，用纯歌名补搜，
    只合并那些缺失平台的结果（某些平台对长关键词的分词较差，纯歌名反而搜得到）。

    返回合并去重后的结果列表；两轮都超时/失败才返回 None。
    """
    performer = (performer or '').strip()
    # 多艺人组合时只用第一个艺人名，避免关键词过长导致 QQ 搜索分词失败
    if performer:
        _first_artist = re.split(r'[/&,，、;；|]+', performer)[0].strip()
        kw1 = f'{song_name} {_first_artist}'.strip() if _first_artist else song_name
    else:
        kw1 = song_name
    res = _search_one_timeout(kw1)
    first_failed = res is None
    res = res or []

    missing = []
    for code in BATCH_PLATFORMS:
        best, label = _pick_best_for_batch(res, code, song_name, performer, enrich=_enrich_result)
        # 没搜到、或只搜到别人的版本 → 该平台值得用纯歌名再试一次
        if not best or label != '精准匹配(目标艺人)':
            missing.append(code)

    if missing and kw1 != song_name:
        res2 = _search_one_timeout(song_name)
        if res2 is None and first_failed:
            return None  # 两轮都挂了，才算真失败
        seen = {(r.get('platform_code'), r.get('song_url')) for r in res}
        for r in (res2 or []):
            if r.get('platform_code') not in missing:
                continue
            sig = (r.get('platform_code'), r.get('song_url'))
            if sig in seen:
                continue
            seen.add(sig)
            res.append(r)

    if first_failed and not res:
        return None
    return res


def _empty_song_result(item):
    plats = {}
    for code in BATCH_PLATFORMS:
        plats[code] = {
            'collection_count': None, 'listening_count': None,
            'comment_count': None, 'record_label': '', 'match_label': '查询失败',
            'song_name': '', 'performer': '', 'link': '', 'availability': '',
        }
    return {
        'song_name': item['song_name'], 'performer': item.get('performer', ''),
        'lyricist': item.get('lyricist', ''), 'composer': item.get('composer', ''),
        'platforms': plats, 'status': 'error',
    }


@app.route('/batch')
def batch_page():
    """批量歌单数据导出独立页（见 templates/batch.html）"""
    try:
        return render_template('batch.html')
    except Exception as _e:
        logger.exception('batch_page render failed: %s', _e)
        return redirect(url_for('index'))


# 镜像路由 /batch.html（防历史链接 + 老版本 index.html 误指）
@app.route('/batch.html')
def batch_page_html():
    from flask import redirect
    return redirect(url_for('batch_page'))


@app.route('/api/batch_search', methods=['POST'])
def api_batch_search():
    data = request.get_json(silent=True) or {}
    raw = data.get('songs') or []
    clean = []
    for s in raw:
        if not isinstance(s, dict):
            continue
        name = (s.get('song_name') or s.get('name') or '').strip()
        if not name:
            continue
        clean.append({
            'song_name': name,
            'performer': (s.get('performer') or '').strip(),
            'lyricist': (s.get('lyricist') or '').strip(),
            'composer': (s.get('composer') or '').strip(),
        })
    if not clean:
        return jsonify({'error': '请提供至少一首歌（需包含歌名）'}), 400

    max_workers = min(int(data.get('concurrency', 6)), 10)

    def worker(item):
        # 2026-08-04：改为「歌名+表演者」两级搜索，纯歌名会被同名歌淹没
        res = _search_two_pass(item['song_name'], item['performer'])
        if res is None:
            return _empty_song_result(item)
        plats = {}
        for code in BATCH_PLATFORMS:
            best, label = _pick_best_for_batch(res, code, item['song_name'], item['performer'], enrich=_enrich_result)
            if best:
                # 发行公司：多源兜底，避免「best 版本无厂牌即整列空白」
                # ① best 自身 → ② 同平台目标艺人其它有厂牌版本 → ③ 同平台任意有厂牌版本
                record_label = _resolve_record_label(code, res, item['song_name'], item['performer'], best)
                plats[code] = {
                    'collection_count': best.get('collection_count'),
                    'listening_count': best.get('listening_count'),
                    'comment_count': best.get('comment_count'),
                    'record_label': record_label,
                    'match_label': label,
                    'song_name': best.get('song_name'),
                    'performer': best.get('performer'),
                    'link': best.get('song_url'),  # 2026-08-04 修复：平台搜索返回 song_url，不是 link
                    # QQ/网易云能给出精确状态（正常/需付费/已下架）；
                    # 酷狗/酷我/汽水接口无下架标志，搜到即视为「在架」
                    'availability': best.get('availability') or '在架',
                }
            else:
                # 搜遍两轮都没有目标艺人版本 → 该平台未收录（或已彻底下架）
                # 网易云若命中「操作频繁」限流，空结果应标「限流」而非「未收录」（避免误判无版权）
                _ne_blocked = (code == 'netease' and (time.time() - _NETEASE_BLOCKED_AT) < 900)
                plats[code] = {
                    'collection_count': None, 'listening_count': None,
                    'comment_count': None, 'record_label': '', 'match_label': label,
                    'song_name': '', 'performer': '', 'link': '',
                    'availability': '限流' if _ne_blocked else ('未收录' if label == '无结果' else ''),
                }
        return {
            'song_name': item['song_name'], 'performer': item.get('performer', ''),
            'lyricist': item.get('lyricist', ''), 'composer': item.get('composer', ''),
            'platforms': plats, 'status': 'ok',
        }

    out = [None] * len(clean)
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = {ex.submit(worker, item): i for i, item in enumerate(clean)}
        for f in concurrent.futures.as_completed(futs):
            i = futs[f]
            try:
                out[i] = f.result()
            except Exception:
                out[i] = _empty_song_result(clean[i])

    return jsonify({'total': len(clean), 'results': out})


@app.route('/api/batch_export', methods=['POST'])
def api_batch_export():
    """接收 batch_search 的结果，生成 xlsx 返回下载（不重算）。"""
    from flask import send_file
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment
    data = request.get_json(silent=True) or {}
    results = data.get('results') or []
    if not results:
        return jsonify({'error': '无数据'}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '歌单各平台数据'
    headers = ['歌名', '表演者', '词作者', '曲作者']
    for code in BATCH_PLATFORMS:
        n = BATCH_PLAT_NAMES[code]
        headers += [
            f'{n}_状态', f'{n}_收藏', f'{n}_在听', f'{n}_评论',
            f'{n}_发行公司', f'{n}_匹配', f'{n}_链接',
        ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def g(v):
        return v if v not in (None, '', 'None') else ''

    for r in results:
        plats = r.get('platforms', {})
        row = [g(r.get('song_name')), g(r.get('performer')), g(r.get('lyricist')), g(r.get('composer'))]
        for code in BATCH_PLATFORMS:
            p = plats.get(code, {})
            row += [
                g(p.get('availability')),
                g(p.get('collection_count')), g(p.get('listening_count')),
                g(p.get('comment_count')), g(p.get('record_label')), g(p.get('match_label')),
                g(p.get('link')),
            ]
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = '歌单各平台数据导出.xlsx'
    try:
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    except TypeError:
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, attachment_filename=fname)


@app.route('/api/batch_upload', methods=['POST'])
def api_batch_upload():
    """接收 Excel(.xlsx/.xlsm) / Word(.docx) 文件，解析为歌单数组返回（不查询）。"""
    import io, zipfile, xml.etree.ElementTree as ET
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': '未收到文件'}), 400
    fname = f.filename.lower()
    try:
        if fname.endswith(('.xlsx', '.xlsm')):
            import openpyxl
            wb = openpyxl.load_workbook(f.stream, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = []
            for r in ws.iter_rows(values_only=True):
                rows.append([('' if v is None else str(v)) for v in r])
        elif fname.endswith('.docx'):
            rows = _parse_docx_rows(f.read())
        elif fname.endswith('.xls'):
            return jsonify({'error': '暂不支持旧版 .xls，请用 .xlsx 另存后上传'}), 400
        elif fname.endswith('.doc'):
            return jsonify({'error': '暂不支持旧版 .doc，请用 .docx 另存后上传'}), 400
        else:
            return jsonify({'error': '仅支持 .xlsx / .xlsm / .docx 文件'}), 400
    except Exception as e:
        return jsonify({'error': '文件解析失败：' + str(e)}), 400

    songs = _extract_songs_from_rows(rows)
    if not songs:
        return jsonify({'error': '未在文件中识别到歌单（需包含歌名列，如「歌名/歌曲」）'}), 400
    # v4.27.4：去掉 600 上限，整文件完整识别。
    # 后续走 /api/batch_v2_submit（最多 10 万首），上游由 v2 worker 后台分批持久化入库，
    # 关浏览器/电脑也不丢。≥500 首会在前端自动切 v2 模式锁定。
    return jsonify({
        'file_name': f.filename,
        'total': len(songs),
        'truncated': False,
        'songs': songs,
    })


# ══════════════════════════════════════════════════════════════════════════
# v4.27.11：迁移老数据 — 把库里所有「歌名-歌手」但 artist 为空的歌一次性拆分。
# 场景：v4.27.10 之前的版本导入时没自动 split，导致库里 1000+ 条
# 「歌名=「歌名-歌手」/ artist="" 」的脏数据。批量监控仍能跑（仅匹配会模糊），
# 但点开歌名查看时会看到"歌名带歌手"且歌手列空，影响人工校核。
# 这是一次性的数据规范化，可反复调用（已拆分的不再重复动）。
# ══════════════════════════════════════════════════════════════════════════
@app.route('/api/admin/migrate_split_titles', methods=['POST'])
def api_admin_migrate_split_titles():
    """把所有 song_archive 中 artist=='' 且 song_name 含 '歌名-歌手' 格式的歌
    自动拆出 artist 字段。前端先调 GET dry-run 预览，POST 真正执行。"""
    from monitor import db as _mon_db
    try:
        _mon_db.init_db()
        conn = _mon_db.get_conn()
    except Exception as e:
        return jsonify({'error': f'db init fail: {e}'}), 500
    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get('dry_run', True))
    rows = conn.execute(
        "SELECT id, song_name, artist FROM song_archive "
        "WHERE (artist='' OR artist IS NULL) AND song_name LIKE '%-%'"
    ).fetchall()
    fixes = []
    for r in rows:
        name = (r['song_name'] or '').strip()
        if not name or not ('-' in name or '·' in name or ' - ' in name
                            or '–' in name or '—' in name):
            continue
        new_name, new_artist = _split_mixed_title(name)
        if new_artist and new_name != name:
            fixes.append({
                'id': r['id'], 'old_name': name, 'new_name': new_name,
                'new_artist': new_artist
            })
    if dry_run:
        return jsonify({
            'ok': True, 'dry_run': True,
            'matched': len(rows), 'wants_split': len(fixes),
            'samples': fixes[:30]
        })
    n = 0
    try:
        with _mon_db.tx() as c:
            for f in fixes:
                c.execute(
                    "UPDATE song_archive SET song_name=?, artist=?, updated_at=? WHERE id=?",
                    (f['new_name'], f['new_artist'], _mon_db.now_str(), f['id'])
                )
                n += 1
    except Exception as e:
        return jsonify({'error': f'update fail: {e}'}), 500
    return jsonify({'ok': True, 'dry_run': False, 'updated': n, 'fixed': fixes[:30]})


def _parse_docx_rows(data):
    """从 docx 二进制解析：优先取第一个表格的二维数据；无表格则按段落逐行。"""
    import io, zipfile, xml.etree.ElementTree as ET
    W = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
    z = zipfile.ZipFile(io.BytesIO(data))
    xml = z.read('word/document.xml')
    root = ET.fromstring(xml)
    tables = list(root.iter('{%s}tbl' % W))
    if tables:
        rows = []
        for tr in tables[0].iter('{%s}tr' % W):
            cells = []
            for tc in tr.iter('{%s}tc' % W):
                texts = [t.text for t in tc.iter('{%s}t' % W) if t.text]
                cells.append(''.join(texts).strip())
            rows.append(cells)
        if len(rows) > 1:
            return rows
    # 回退：段落文本（每行一首，按 逗号/制表符/顿号/空格 拆列）
    import re as _re_docx
    paras = []
    for p in root.iter('{%s}p' % W):
        texts = [t.text for t in p.iter('{%s}t' % W) if t.text]
        line = ''.join(texts).strip()
        if not line:
            continue
        parts = _re_docx.split(r'[,，\t、]+', line)
        if len(parts) <= 1:
            parts = _re_docx.split(r'\s+', line)
        paras.append([x.strip() for x in parts if x.strip()])
    return paras


_SONG_HEADERS = {
    'song_name': ['歌曲名称', '歌名', '歌曲名', '歌曲', '曲名', '歌目', '曲目', 'title', 'song', 'name'],
    'performer': ['表演者', '歌手', '演唱', '艺人', '演唱者', '原唱', 'artist', 'performer', 'singer'],
    'lyricist': ['词作者', '作词', '填词', 'lyricist', '词'],
    'composer': ['曲作者', '作曲', '谱曲', 'composer', '曲'],
}

# 这些列即使包含 song/performer 等关键字，也应直接跳过，避免「歌曲ID」被误识别为歌名
_SKIP_HEADERS = ['id', '编号', '序号', '编码', 'track', '歌曲id']


def _split_mixed_title(title):
    """从 '歌名 - 表演者' / '歌名·表演者' 这类混合标题中拆分歌名与表演者。
    未识别到明确分隔符时原样返回。

    v4.27.11 增强：
    - 即使右侧歌手很短（≥1 字符，如「en」/「周杰伦」），也会尝试拆分。
    - 尾随版本括号/标注如「（DJ版）」/「(现场版)」会被干净剥离到歌名侧。
    - 启发式：拆分前先剥掉尾部 "(...)" / "(现场版)" 这种纯版本括号。
    """
    if not title:
        return title, ''
    t = title.strip()
    if not t:
        return t, ''

    # 先剥掉纯版本括号（只剥括号，不影响歌名）
    # 例：Love Song（DJ版）- 张学友 → Love Song - 张学友
    t = re.sub(r'[\(（][^)）]*[\)）]\s*$', '', t).strip()

    # v4.27.11：用第一个出现的较粗分隔符；split 后若右侧还含同名较细分隔符，
    # 把右侧进一步细分，最后一段作歌手，前面整体作歌名。
    # 处理：「不止-派-王不怂」→ 歌名=不止-派, 歌手=王不怂 （用户截图痛点）
    seps = [' - ', ' – ', ' — ', ' -', '- ', '–', '—', '·', '•']
    for sep in seps:
        if sep in t:
            parts = t.split(sep)
            # 第一个 sep 后再按更细分隔符细分最后一段
            if len(parts) >= 2:
                *song_parts, artist = parts
                if song_parts and artist and artist.strip() and not re.fullmatch(r'[\s\W_]+', artist):
                    # 用原 sep 拼回
                    full_song = sep.join(song_parts).strip()
                    return full_song, artist.strip()
    # 退路：单个 `-`（无空格）
    for sep in ['-', '／', '/', '｜', '|']:
        if sep in t:
            parts = t.split(sep)
            if len(parts) >= 2:
                *song_parts, artist = parts
                if song_parts and artist and artist.strip() and not re.fullmatch(r'[\s\W_]+', artist):
                    full_song = sep.join(song_parts).strip()
                    return full_song, artist.strip()
    return t, ''


def _maybe_split_title_into_performer(name, perf, has_performer_col):
    """v4.27.11：无论是否识别出独立的歌手列，只要歌手内容为空，
    都会尝试把歌名列的 '歌名 - 歌手' 格式拆分到歌手里。"""
    if not name:
        return name, perf
    if perf and perf.strip():
        return name, perf  # 有歌手内容就保留
    new_name, new_perf = _split_mixed_title(name)
    if new_perf:
        return new_name, new_perf
    return name, perf


def _extract_songs_from_rows(rows):
    """二维单元格 -> 歌单。首行若为表头则智能识别列，否则按列序假设。
    当没有独立表演者列时，自动从歌名列的 '歌名 - 表演者' 格式中拆分表演者。"""
    if not rows:
        return []
    header = rows[0]
    colmap = {}
    used = set()
    for i, c in enumerate(header):
        txt = (c or '').strip().lower()
        if not txt or i in used:
            continue
        # 先排除 ID/编号/序号等干扰列
        if any(skip in txt for skip in _SKIP_HEADERS):
            used.add(i)
            continue
        for fld, kws in _SONG_HEADERS.items():
            if fld in colmap:
                continue
            if any(k in txt for k in kws):
                colmap[fld] = i
                used.add(i)
                break
    data_rows = rows[1:] if colmap else rows
    order = ['song_name', 'performer', 'lyricist', 'composer']
    defaults = {'song_name': 0, 'performer': 1, 'lyricist': 2, 'composer': 3}
    free = [i for i in range(20) if i not in used]
    fi = 0
    resolved = {}
    for fld in order:
        if fld in colmap:
            resolved[fld] = colmap[fld]
        elif defaults[fld] not in used and defaults[fld] not in resolved.values():
            resolved[fld] = defaults[fld]
        else:
            while fi < len(free) and free[fi] in resolved.values():
                fi += 1
            resolved[fld] = free[fi] if fi < len(free) else defaults[fld]
            fi += 1

    has_performer_col = 'performer' in colmap
    songs = []
    for r in data_rows:
        if not r or all((c or '').strip() == '' for c in r):
            continue
        name = (r[resolved['song_name']] if resolved['song_name'] < len(r) else '').strip()
        if not name:
            continue
        perf = (r[resolved['performer']] if resolved['performer'] < len(r) else '').strip()
        lyr = (r[resolved['lyricist']] if resolved['lyricist'] < len(r) else '').strip()
        comp = (r[resolved['composer']] if resolved['composer'] < len(r) else '').strip()
        # v4.27.11：无论是否有独立表演者列，只要表演者列为空就尝试从歌名列的
        # '歌名 - 歌手' / '歌名·歌手' 格式拆分表演者（满足用户「歌名单只是一列
        # 歌名-歌手名」的工作流）。
        name, perf = _maybe_split_title_into_performer(name, perf, has_performer_col)
        songs.append({'song_name': name, 'performer': perf, 'lyricist': lyr, 'composer': comp})
    return songs


# ══════════════════════════════════════════════════════════════════════════
#  我的歌单（设备本地）— 原版/翻唱版 多平台数据，缓存自动展示
#  数据存 ~/.musicfinder/（每台设备独立），不进云端；云端只同步手动标记。
#  v4.1: 词曲作者精准匹配；v4.2: SSE 边搜边推
# ══════════════════════════════════════════════════════════════════════════
_PLAYLIST_FILE = os.path.join(_MARKS_DIR, 'my_playlist.json')
_PLAYLIST_RESULTS_FILE = os.path.join(_MARKS_DIR, 'my_playlist_results.json')


def _collect_person_names(g, field):
    """从 group + 全部平台聚合一个人名字段（lyricist/composer）的所有非空值，去重保序。

    同时支持「/」与「&」多人分隔（不同平台的格式不一样）。
    """
    values = []
    seen = set()
    top = (g.get(field) or '') if isinstance(g, dict) else ''
    for v in re.split(r'[/&]', top):
        v = v.strip()
        if v and v.lower() not in seen:
            seen.add(v.lower()); values.append(v)
    pd = (g.get('platform_data') or {}) if isinstance(g, dict) else {}
    if isinstance(pd, dict):
        for code, p in pd.items():
            if not isinstance(p, dict): continue
            v = (p.get(field) or '')
            for part in re.split(r'[/&]', str(v)):
                part = part.strip()
                if part and part.lower() not in seen:
                    seen.add(part.lower()); values.append(part)
    return values
# RLock（可重入）：导入路由会在持锁状态下调用 _save_playlist（其内部也加同一把锁），
# 用非可重入 Lock 会自我死锁导致导入永久卡住。RLock 让同一线程可重复加锁。
_playlist_lock = threading.RLock()

_VERSION_RE = re.compile(r'[\(（\[【]([^（）()\[\]【】]*?)[\)）\]】]')


def _extract_version_tag(name):
    """从歌名里提取版本括号内容，如『晴天(DJ版)』→『DJ版』。"""
    m = _VERSION_RE.search(name or '')
    return m.group(1).strip() if m else ''


def _has_version_suffix(name):
    return bool(_VERSION_RE.search(name or ''))


def _alt_name_diff(plat_name, query_name):
    """v4.3: 计算「原版歌曲的另一种叫法」用于旁注。

    当 plat_name 和 query_name 剥括号标准化后等价但字符不同时,返回两个里
    **更长**的那个(通常带括号副标题,信息量更大)。前端把它展示在原版 tag 旁。
    典型场景:
      query_name = '记得你是女人(忘记你是女人)'  剥括号='记得你是女人'
      plat_name  = '记得你是女人'                  剥括号='记得你是女人'
      → 返回 query_name '记得你是女人(忘记你是女人)' (带副标题,信息量更大)

      query_name = '记得你是女人'                  剥括号='记得你是女人'
      plat_name  = '记得你是女人(忘记你是女人)'  剥括号='记得你是女人'
      → 返回 plat_name '记得你是女人(忘记你是女人)' (带副标题)

      query_name = plat_name = '记得你是女人'
      → 字符完全相同,alt_name=''
    """
    pn = (plat_name or '').strip()
    qn = (query_name or '').strip()
    if not pn or not qn:
        return ''
    if _normalize_match(pn) != _normalize_match(qn):
        return ''  # 标准化后不等,根本不是同一首,别称无意义
    if pn == qn:
        return ''
    # 标准化等价但字符不同 → 返回更长的那个（带括号副标题或版本号）
    return pn if len(pn) >= len(qn) else qn


def _trim_platform(p):
    if not isinstance(p, dict):
        return None
    return {
        'platform_code': p.get('platform_code'),
        'platform': p.get('platform'),
        'song_name': p.get('song_name'),
        'performer': p.get('performer'),
        'collection_count': p.get('collection_count'),
        'listening_count': p.get('listening_count'),
        'comment_count': p.get('comment_count'),
        'record_label': p.get('record_label'),
        'release_date': p.get('release_date'),
        'lyricist': p.get('lyricist'),
        'composer': p.get('composer'),
        'link': p.get('song_url') or p.get('link'),
        'availability': p.get('availability') or '在架',
        'match_label': p.get('match_label'),
        'match_score': p.get('match_score'),
        # v4.6：保留原始平台 ID 字段，供「固定监测 / 刷新固定数据」按 ID 直接回填最新数据
        '_songid': p.get('_songid'),
        '_songmid': p.get('_songmid'),
        '_hash': p.get('_hash'),
        '_mixsongid': p.get('_mixsongid'),
        '_song_id': p.get('_song_id'),
        '_album_id': p.get('_album_id'),
        '_album_hash': p.get('_album_hash'),
        '_rid': p.get('_rid'),
    }


def _trim_playlist_group(g):
    pd = {}
    # v4.3：优先展示 _classify_playlist_song 按 stripped-key 合并后的 platform_data，
    # 这样「记得你是女人」和「记得你是女人 (忘记你是女人)」等带括号的同歌版本
    # 5 平台都能在原版行展示；未合并时回退原始 platform_data。
    src_pd = g.get('_merged_platform_data') or g.get('platform_data') or {}
    for code, p in src_pd.items():
        t = _trim_platform(p)
        if t:
            pd[code] = t
    return {
        'song_name': g.get('song_name'),
        'performer': g.get('performer'),
        'album': g.get('album') or '',
        'release_date': g.get('release_date') or '',
        'lyricist': g.get('lyricist'),
        'composer': g.get('composer'),
        'record_label': g.get('record_label'),
        'match_score': g.get('match_score', 0),
        'match_label': g.get('match_label'),
        'platform_data': pd,
        # v4.15：把 mark 透传给歌单页，前端用 mark_type pill 与搜索页/已标记页
        # 保持同一份视觉（markClass 配色 + 同一文本），实现三页标签逻辑统一。
        # tags[] 也带上，弹窗"附加标签"用。
        'mark': g.get('mark') or None,
        'tags': g.get('tags') or [],
        # 顺手把 _pinned 透传（搜索页反向匹配用的就是它）
        '_pinned': g.get('_pinned') or False,
    }


def _classify_playlist_song(query, grouped_results):
    """将原版与翻唱/其它版本分开。

    规则（与用户约定一致）：
      - 歌名(去版本括号后) + 表演者 精确对上 → 原版
      - 否则（歌名带 (DJ版)/(默涵版) 等版本括号，或表演者不同）→ 翻唱/其它版本
      - 用户输入带词曲作者时：原版必须词曲作者也对得上（_person_match 已支持多人/分隔符）

    v4.3: 在分类前先按 stripped-key (剥括号歌名 + 艺人 + 剥括号专辑) 把同名的
    「不同平台括号写法」group 合并成一份共享 platform_data，让 5 平台一排都能展示
    数据。分类仍按精确 key 走（不影响 v4.1 标记 key 的精度——标记按精确 key 算）。
    """
    # ---- 新增：按 stripped-key 合并 platform_data（仅展示用，不影响 _mark_key / 分类）----
    stripped_buckets = {}  # stripped_key → list[idx]
    for idx, g in enumerate(grouped_results):
        sk = (
            _normalize_match(g.get('song_name', '')),
            _split_artist_names(g.get('performer', '')),
            _normalize_match(g.get('album', '')),
        )
        stripped_buckets.setdefault(sk, []).append(idx)
    # 每个 stripped_key 桶的合并 platform_data：每个 platform_code 取 score 最高的
    merged_pd_per_stripped = {}
    for sk, idxs in stripped_buckets.items():
        merged = {}
        for idx in idxs:
            g = grouped_results[idx]
            for code, p in (g.get('platform_data') or {}).items():
                cur = merged.get(code)
                if cur is None or (p.get('match_score', 0) or 0) > (cur.get('match_score', 0) or 0):
                    merged[code] = p
        merged_pd_per_stripped[sk] = merged
    # 给每个 group 写入 _merged_platform_data（分类器优先用这个做展示）
    for idx, g in enumerate(grouped_results):
        sk = (
            _normalize_match(g.get('song_name', '')),
            _split_artist_names(g.get('performer', '')),
            _normalize_match(g.get('album', '')),
        )
        merged = merged_pd_per_stripped.get(sk) or {}
        if len(merged) > len(g.get('platform_data') or {}):
            g['_merged_platform_data'] = merged

    q_name = (query.get('song_name') or '').strip()
    q_perf = (query.get('performer') or '').strip()
    q_lyr = (query.get('lyricist') or '').strip()
    q_com = (query.get('composer') or '').strip()
    q_base = _normalize_match(q_name)
    # _split_artist_names 返回的是 tuple，分类时要用集合交集(&)，统一转 set
    q_artists = set(_split_artist_names(q_perf)) if q_perf else set()
    has_lyr = bool(q_lyr)
    has_com = bool(q_com)
    classified = []
    for g in grouped_results:
        gname = g.get('song_name', '')
        gperf = g.get('performer', '')
        name_ok = bool(q_base) and (_normalize_match(gname) == q_base)
        perf_ok = (not q_artists) or bool(q_artists & set(_split_artist_names(gperf)))
        # 聚合每个 group 已知的所有 lyricist/composer，对用户输入做 _person_match
        lyr_vals = ' / '.join(_collect_person_names(g, 'lyricist'))
        com_vals = ' / '.join(_collect_person_names(g, 'composer'))
        # 用户要求「精准匹配」：用户输入的每个作者都必须出现在结果集合里（子集）
        lyr_ok = _person_set_strict(q_lyr, lyr_vals)
        com_ok = _person_set_strict(q_com, com_vals)
        is_orig = name_ok and perf_ok and lyr_ok and com_ok and (not _has_version_suffix(gname))
        entry = _trim_playlist_group(g)
        # v4.3：alt_name 表示「原版歌曲的另一种叫法」。
        # 当 orig.song_name（剥括号后）跟 query.song_name 剥括号后不相等时，说明有别称。
        # 比如：query=「记得你是女人（忘记你是女人）」剥括号=「记得你是女人」，
        #      orig.song_name=「记得你是女人」剥括号=「记得你是女人」——完全相同，alt_name 空。
        # 但若 orig.song_name=「记得你是女人-王佳音版」剥括号=「记得你是女人王佳音版」，
        #      跟 query 不同，则 alt_name=orig.song_name 展示别称。
        if is_orig:
            entry['is_original'] = True
            # 修复 v4.3：原版 tag 一律固定为「原版」字样，
            # 不再用 _extract_version_tag(gname) 抽括号副标题——那会误导
            # 用户把括号里的字面文字当原版标识。括号副标题挪到 alt_name
            # 让前端做旁注（如「原版 [别称：忘记你是女人]」）。
            entry['version_tag'] = '原版'
            entry['alt_name'] = _alt_name_diff(gname, q_name)
            entry['is_cover'] = False
        else:
            entry['is_original'] = False
            # 精细化版本标签：先看是否带版本括号 → 再看表演者 → 再看词曲作者
            if _has_version_suffix(gname):
                tag = _extract_version_tag(gname) or '翻唱版'
            elif not perf_ok:
                tag = '其他演唱'
            elif (has_lyr and not lyr_ok) or (has_com and not com_ok):
                tag = '词曲不一致'
            else:
                tag = '其它版本'
            entry['version_tag'] = tag
            entry['alt_name'] = _alt_name_diff(gname, q_name)
            entry['is_cover'] = bool(_has_version_suffix(gname)) or (not perf_ok) or (has_lyr and not lyr_ok) or (has_com and not com_ok)
        # 暴露给前端展示用：用户输入 vs 该版本上的命中值
        entry['lyricist_query'] = q_lyr
        entry['composer_query'] = q_com
        entry['lyricist_values'] = lyr_vals
        entry['composer_values'] = com_vals
        entry['lyricist_hit'] = lyr_ok
        entry['composer_hit'] = com_ok
        classified.append(entry)
    orig_list = [c for c in classified if c['is_original']]
    if orig_list:
        orig = orig_list[0]
        versions = [c for c in classified if not c['is_original']]
    else:
        orig = max(classified, key=lambda c: c.get('match_score', 0)) if classified else None
        if orig:
            orig['is_original'] = True
            # 修复 v4.3：兜底选 orig 时，如果只是「词曲不一致 / 其它版本 / 翻唱版」
            # 这种由数据缺失触发的标签（而非真正的翻唱/版本括号），强制改回「原版」
            # 跟 is_original 保持一致——避免一张卡同时显示「原版」+「词曲不一致」的矛盾。
            # 真正的翻唱（带版本括号或表演者不同）才保留「翻唱版/其他演唱」label。
            if (orig.get('version_tag') in ('词曲不一致', '其它版本', '翻唱版')
                    and not _has_version_suffix(orig.get('song_name', ''))):
                orig['version_tag'] = '原版'
            else:
                orig['version_tag'] = orig.get('version_tag') or '最佳匹配'
            # 兜底 chosen 也要算命中度
            orig['lyricist_hit'] = (not has_lyr) or _person_set_strict(q_lyr, orig.get('lyricist_values', ''))
            orig['composer_hit'] = (not has_com) or _person_set_strict(q_com, orig.get('composer_values', ''))
            versions = [c for c in classified if c is not orig]
        else:
            versions = []
    return orig, versions


def _playlist_worker(item):
    try:
        core = _search_core(item['song_name'], item.get('lyricist', ''),
                            item.get('composer', ''), item.get('performer', ''), 30)
        grouped = core.get('grouped_results', [])
    except Exception:
        grouped = []
    orig, versions = _classify_playlist_song(item, grouped)
    return {
        'query': {k: item.get(k, '') for k in ('song_name', 'performer', 'lyricist', 'composer')},
        'original': orig,
        'versions': versions,
        'version_count': len(versions),
    }


def _run_playlist_search(songs, concurrency=6):
    """搜全部歌并分类，返回列表（按原始输入顺序）。"""
    return list(_run_playlist_search_iter(songs, concurrency))


def _run_playlist_search_iter(songs, concurrency=6):
    """按完成顺序逐条 yield 分类结果（用于 SSE 边搜边推）。"""
    clean = []
    for s in songs:
        name = (s.get('song_name') or '').strip()
        if not name:
            continue
        clean.append({
            'song_name': name,
            'performer': (s.get('performer') or '').strip(),
            'lyricist': (s.get('lyricist') or '').strip(),
            'composer': (s.get('composer') or '').strip(),
        })
    if not clean:
        return
    max_workers = min(int(concurrency), 10) or 6

    def _job(item):
        try:
            core = _search_core(item['song_name'], item.get('lyricist', ''),
                                item.get('composer', ''), item.get('performer', ''), 30)
            grouped = core.get('grouped_results', [])
        except Exception:
            grouped = []
        orig, versions = _classify_playlist_song(item, grouped)
        return {
            'query': {k: item.get(k, '') for k in ('song_name', 'performer', 'lyricist', 'composer')},
            'original': orig,
            'versions': versions,
            'version_count': len(versions),
        }

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = {ex.submit(_job, item): i for i, item in enumerate(clean)}
        for f in concurrent.futures.as_completed(futs):
            i = futs[f]
            try:
                yield f.result()
            except Exception:
                yield {'query': clean[i], 'original': None, 'versions': [], 'version_count': 0}


def _effective_playlist_paths():
    """返回 (playlist_file, results_file, base_dir)，按当前登录用户隔离；未登录走旧路径。"""
    ud = _user_dir()
    if ud:
        return (os.path.join(ud, 'my_playlist.json'),
                os.path.join(ud, 'my_playlist_results.json'), ud)
    return _PLAYLIST_FILE, _PLAYLIST_RESULTS_FILE, _MARKS_DIR


def _load_playlist():
    pf, _, _ = _effective_playlist_paths()
    try:
        with open(pf, 'r', encoding='utf-8') as f:
            d = json.load(f)
        return d.get('songs', []) if isinstance(d, dict) else []
    except Exception:
        return []


def _save_playlist(songs):
    with _playlist_lock:
        pf, _, pd = _effective_playlist_paths()
        os.makedirs(pd, exist_ok=True)
        with open(pf, 'w', encoding='utf-8') as f:
            json.dump({'updated_at': time.strftime('%Y-%m-%d %H:%M:%S'), 'songs': songs},
                      f, ensure_ascii=False, indent=2)
    _push_blob_async('playlist', {'songs': songs})  # 跨设备同步


def _load_playlist_results():
    _, rf, _ = _effective_playlist_paths()
    try:
        with open(rf, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def _save_playlist_results(items):
    with _playlist_lock:
        _, rf, rd = _effective_playlist_paths()
        os.makedirs(rd, exist_ok=True)
        with open(rf, 'w', encoding='utf-8') as f:
            json.dump({'searched_at': time.strftime('%Y-%m-%d %H:%M:%S'), 'items': items},
                      f, ensure_ascii=False, indent=2)
    _push_blob_async('playlist_results', {'items': items})  # 跨设备同步


# ════════════════════════════════════════════════════════════════════════════
#  v4.6 歌单「删除不要的版本」+「固定监测一批歌」
#   - 删除：存黑名单（my_playlist_deleted.json），重载/重搜都不再出现
#   - 固定：存监测清单（my_playlist_pins.json），以后按平台 ID 直接抓最新数据
# ════════════════════════════════════════════════════════════════════════════
_PINS_FILE = os.path.join(_MARKS_DIR, 'my_playlist_pins.json')
_DELETED_FILE = os.path.join(_MARKS_DIR, 'my_playlist_deleted.json')
_PLAT_ID_FIELDS = {
    'qq': ['_songid', '_songmid'],
    'kugou': ['_hash', '_mixsongid'],
    'netease': ['_song_id'],
    'kuwo': ['_rid', '_album_id'],
    'qishui': ['_song_id', '_hash'],
}


def _platform_ids_of(entry):
    """从某个版本块（original / version）提取各平台歌曲 ID，用于固定后直接抓数据。"""
    if not isinstance(entry, dict):
        return {}
    pd = entry.get('platform_data') or {}
    ids = {}
    if not isinstance(pd, dict):
        return ids
    for code, fields in _PLAT_ID_FIELDS.items():
        p = pd.get(code)
        if not isinstance(p, dict):
            continue
        for fld in fields:
            v = p.get(fld)
            if v:
                ids[code] = str(v)
                break
    return ids


def _version_key(entry):
    """给一个版本块算稳定钥匙（跨次搜索可复现），用于删除黑名单 / 固定匹配。"""
    if not isinstance(entry, dict):
        return 'bad:' + str(id(entry))
    ids = _platform_ids_of(entry)
    # 优先用平台 ID（最稳定）；都没有再用 歌名|歌手|别称
    for code in ('qq', 'kugou', 'netease', 'kuwo', 'qishui'):
        if ids.get(code):
            return code + ':' + ids[code]
    name = (entry.get('song_name') or '').strip()
    perf = (entry.get('performer') or '').strip()
    alt = (entry.get('alt_name') or '').strip()
    return 'name:' + '|'.join([name, perf, alt])


def _effective_pins_deleted_paths():
    """返回 (pins_file, deleted_file, base_dir)，按当前登录用户隔离；未登录走旧路径。"""
    ud = _user_dir()
    if ud:
        return (os.path.join(ud, 'my_playlist_pins.json'),
                os.path.join(ud, 'my_playlist_deleted.json'), ud)
    return _PINS_FILE, _DELETED_FILE, _MARKS_DIR


def _load_pins():
    pf, _, _ = _effective_pins_deleted_paths()
    try:
        with open(pf, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except (OSError, ValueError):
        return []


def _save_pins(pins):
    with _playlist_lock:
        pf, _, pd = _effective_pins_deleted_paths()
        os.makedirs(pd, exist_ok=True)
        with open(pf, 'w', encoding='utf-8') as f:
            json.dump(pins, f, ensure_ascii=False, indent=2)
    _push_blob_async('pins', pins)  # 跨设备同步


def _load_deleted():
    _, df, _ = _effective_pins_deleted_paths()
    try:
        with open(df, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return set(data.get('keys', [])) if isinstance(data, dict) else set()
    except (OSError, ValueError):
        return set()


def _load_deleted_log():
    """回收站可读日志（每条含 歌名/歌手/类型/时间），供前端「回收站」页展示。"""
    _, df, _ = _effective_pins_deleted_paths()
    try:
        with open(df, 'r', encoding='utf-8') as f:
            d = json.load(f)
        return d.get('log', []) if isinstance(d, dict) else []
    except (OSError, ValueError):
        return []


def _save_deleted(keys, log=None):
    with _playlist_lock:
        _, df, dd = _effective_pins_deleted_paths()
        os.makedirs(dd, exist_ok=True)
        if log is None:
            log = _load_deleted_log()
        with open(df, 'w', encoding='utf-8') as f:
            json.dump({'keys': sorted(keys), 'log': log}, f, ensure_ascii=False, indent=2)
    _push_blob_async('deleted', {'keys': sorted(keys), 'log': log})  # 跨设备同步


def _filter_deleted_items(items):
    """从缓存结果里剔除已删除的版本 / 整首。"""
    deleted = _load_deleted()
    if not deleted:
        return items
    out = []
    for it in items:
        if not isinstance(it, dict):
            out.append(it)
            continue
        orig = it.get('original')
        if orig and _version_key(orig) in deleted:
            # 原版被删 → 整首不要
            continue
        vers = it.get('versions') or []
        kept = [v for v in vers if _version_key(v) not in deleted]
        if not orig and not kept:
            continue
        new_it = dict(it)
        new_it['versions'] = kept
        new_it['version_count'] = len(kept)
        out.append(new_it)
    return out


def _patch_results_with_pins(pins):
    """刷新后把最新 platform_data 回写进缓存结果（my_playlist_results.json）。"""
    if not pins:
        return
    results = _load_playlist_results()
    if not results:
        return
    items = results.get('items') or []
    changed = False
    for pin in pins:
        want_ids = pin.get('platform_ids') or {}
        song_name = pin.get('song_name', '')
        performer = pin.get('performer', '')
        for it in items:
            q = it.get('query') or {}
            if q.get('song_name', '').strip() != song_name:
                continue
            if (q.get('performer', '').strip() or '') != (performer or ''):
                continue
            cands = []
            if it.get('original'):
                cands.append(it['original'])
            cands += (it.get('versions') or [])
            for v in cands:
                vids = _platform_ids_of(v)
                if any(want_ids.get(c) and vids.get(c) == want_ids.get(c) for c in want_ids):
                    v['platform_data'] = pin.get('platform_data') or v.get('platform_data')
                    changed = True
                    break
    if changed:
        _save_playlist_results(items)


@app.route('/api/playlist', methods=['GET'])
def api_playlist_get():
    """返回本机歌单与已缓存的查询结果（供『自动展示』直接加载，不重搜）。"""
    songs = _load_playlist()
    results = _load_playlist_results()
    items = (results or {}).get('items') if results else None
    if items is not None:
        items = _filter_deleted_items(items)
    return jsonify({
        'has_playlist': bool(songs),
        'playlist': songs,
        'has_results': results is not None,
        'results': items,
        'searched_at': (results or {}).get('searched_at') if results else None,
        'deleted_count': len(_load_deleted()),
        'pinned_count': len(_load_pins()),
    })


@app.route('/api/playlist/import', methods=['POST'])
def api_playlist_import():
    """导入歌单（设备本地）→ 边搜边推送 SSE → 全部完成后缓存结果。"""
    data = request.get_json(silent=True) or {}
    raw = data.get('songs') or []
    songs = []
    for s in raw:
        if not isinstance(s, dict):
            continue
        name = (s.get('song_name') or s.get('name') or '').strip()
        if not name:
            continue
        songs.append({
            'song_name': name,
            'performer': (s.get('performer') or '').strip(),
            'lyricist': (s.get('lyricist') or '').strip(),
            'composer': (s.get('composer') or '').strip(),
        })
    if not songs:
        return jsonify({'error': '请提供至少一首歌（需包含歌名）'}), 400
    _save_playlist(songs)
    return _stream_playlist_search(songs, data.get('concurrency', 6), label='import')


@app.route('/api/playlist/research', methods=['POST'])
def api_playlist_research():
    """对本机已导入歌单重新搜索（换 Cookie / 补数据时用）。"""
    songs = _load_playlist()
    if not songs:
        return jsonify({'error': '本机尚未导入歌单'}), 400
    return _stream_playlist_search(songs, 6, label='research')


def _stream_playlist_search(songs, concurrency, label='search'):
    """SSE 流式推送：每搜完一首歌就 yield 一条事件，最后再写盘缓存 + done 事件。"""
    def gen():
        items = []
        total = len(songs)
        # 先推一条 start 事件，前端可立刻开始显示空状态/进度
        yield 'data: ' + json.dumps({'event': 'start', 'total': total, 'label': label}, ensure_ascii=False) + '\n\n'
        for item in _run_playlist_search_iter(songs, concurrency):
            items.append(item)
            payload = {'event': 'result', 'total': total, 'done': len(items),
                       'item': item}
            yield 'data: ' + json.dumps(payload, ensure_ascii=False) + '\n\n'
        # 落盘前剔除已删除版本（黑名单），避免重搜又把不要的版本塞回来
        items = _filter_deleted_items(items)
        # 全部搜完再落盘缓存（让「以后打开即展示」拿到完整结果）
        try:
            _save_playlist_results(items)
            saved = True
        except Exception as e:
            logger.error('保存歌单结果缓存失败: %s', e)
            saved = False
        yield 'data: ' + json.dumps({
            'event': 'done', 'total': total, 'saved': saved,
            'searched_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        }, ensure_ascii=False) + '\n\n'

    resp = Response(stream_with_context(gen()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    resp.headers['Connection'] = 'keep-alive'
    return resp


@app.route('/api/playlist/clear', methods=['POST'])
def api_playlist_clear():
    """清除本机歌单与缓存结果（不影响云端标记）。"""
    with _playlist_lock:
        for p in (_PLAYLIST_FILE, _PLAYLIST_RESULTS_FILE):
            try:
                os.remove(p)
            except OSError:
                pass
    return jsonify({'ok': True})


# ── v4.6：版本删除 / 整首删除 / 固定监测 ───────────────────────────────
@app.route('/api/playlist/version/delete', methods=['POST'])
def api_playlist_version_delete():
    """把某个版本（原版/翻唱）加入删除黑名单，重载与重搜都不再出现；同时记入回收站日志。"""
    data = request.get_json(silent=True) or {}
    key = (data.get('key') or '').strip()
    entry = data.get('entry')
    if not key and entry:
        key = _version_key(entry)
    if not key:
        return jsonify({'error': '缺少版本 key 或 entry'}), 400
    deleted = _load_deleted()
    deleted.add(key)
    log = _load_deleted_log()
    ent = entry or {}
    log.insert(0, {
        'type': 'version', 'key': key,
        'song_name': (data.get('song_name') or ent.get('song_name') or '').strip(),
        'performer': (data.get('performer') or ent.get('performer') or '').strip(),
        'label': (data.get('label') or ent.get('label') or '版本'),
        'deleted_at': time.strftime('%Y-%m-%d %H:%M:%S'),
    })
    _save_deleted(deleted, log)
    return jsonify({'ok': True, 'deleted_count': len(deleted)})


@app.route('/api/playlist/version/batch_delete', methods=['POST'])
def api_playlist_version_batch_delete():
    """批量把多个版本加入删除黑名单（可从回收站一键恢复）。"""
    data = request.get_json(silent=True) or {}
    items = data.get('items') or []
    if not isinstance(items, list) or not items:
        return jsonify({'error': '缺少 items 列表'}), 400
    deleted = _load_deleted()
    log = _load_deleted_log()
    inserted = 0
    for it in items:
        if not isinstance(it, dict):
            continue
        key = (it.get('key') or '').strip()
        entry = it.get('entry')
        if not key and entry:
            key = _version_key(entry)
        if not key or key in deleted:
            continue
        deleted.add(key)
        ent = entry or {}
        log.insert(0, {
            'type': 'version', 'key': key,
            'song_name': (it.get('song_name') or ent.get('song_name') or '').strip(),
            'performer': (it.get('performer') or ent.get('performer') or '').strip(),
            'label': (it.get('label') or ent.get('label') or '版本'),
            'deleted_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        })
        inserted += 1
    if inserted:
        _save_deleted(deleted, log)
    return jsonify({'ok': True, 'deleted_count': len(deleted), 'inserted': inserted})


@app.route('/api/playlist/version/restore', methods=['POST'])
def api_playlist_version_restore():
    """从删除黑名单移除某个版本（误删可恢复）。"""
    data = request.get_json(silent=True) or {}
    key = (data.get('key') or '').strip()
    if not key and data.get('entry'):
        key = _version_key(data['entry'])
    if not key:
        return jsonify({'error': '缺少版本 key 或 entry'}), 400
    deleted = _load_deleted()
    if key in deleted:
        deleted.discard(key)
        log = [x for x in _load_deleted_log() if x.get('key') != key]
        _save_deleted(deleted, log)
    return jsonify({'ok': True, 'deleted_count': len(deleted)})


@app.route('/api/playlist/song/delete', methods=['POST'])
def api_playlist_song_delete():
    """删除整首歌：从本机歌单移除，并把它的全部版本key加入黑名单；同时记入回收站日志。"""
    data = request.get_json(silent=True) or {}
    name = (data.get('song_name') or '').strip()
    perf = (data.get('performer') or '').strip()
    if not name:
        return jsonify({'error': '缺少歌名'}), 400
    # 1) 从歌单移除
    songs = _load_playlist()
    songs = [s for s in songs
             if not (s.get('song_name', '').strip() == name
                     and (s.get('performer', '').strip() or '') == perf)]
    _save_playlist(songs)
    # 2) 把它的版本 key 加黑名单（若该首已有搜索结果，key 取自结果）
    results = _load_playlist_results()
    deleted = _load_deleted()
    song_keys = []
    if results:
        for it in results.get('items', []):
            q = it.get('query') or {}
            if q.get('song_name', '').strip() == name and (q.get('performer', '').strip() or '') == perf:
                for v in ([it.get('original')] + (it.get('versions') or [])):
                    if v:
                        k = _version_key(v)
                        deleted.add(k)
                        song_keys.append(k)
    log = _load_deleted_log()
    log.insert(0, {
        'type': 'song', 'key': 'song:' + name + '|' + perf,
        'song_name': name, 'performer': perf, 'label': '整首歌',
        'deleted_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        'keys': song_keys,
    })
    _save_deleted(deleted, log)
    return jsonify({'ok': True, 'playlist_count': len(songs), 'deleted_count': len(deleted)})


@app.route('/api/playlist/song/restore', methods=['POST'])
def api_playlist_song_restore():
    """恢复整首歌：把该首全部版本 key 从黑名单移除（回收站一键恢复）。"""
    data = request.get_json(silent=True) or {}
    name = (data.get('song_name') or '').strip()
    perf = (data.get('performer') or '').strip()
    if not name:
        return jsonify({'error': '缺少歌名'}), 400
    deleted = _load_deleted()
    # 找出该首在日志里的所有 key
    log = _load_deleted_log()
    restore_keys = set()
    for x in log:
        if x.get('type') == 'song' and x.get('song_name', '').strip() == name \
                and (x.get('performer', '').strip() or '') == perf:
            for k in (x.get('keys') or []):
                restore_keys.add(k)
            deleted.discard('song:' + name + '|' + perf)
    for k in restore_keys:
        deleted.discard(k)
    log = [x for x in log if not (x.get('type') == 'song' and x.get('song_name', '').strip() == name
            and (x.get('performer', '').strip() or '') == perf)]
    _save_deleted(deleted, log)
    return jsonify({'ok': True, 'deleted_count': len(deleted)})


@app.route('/api/playlist/deleted', methods=['GET'])
def api_playlist_deleted_list():
    """返回回收站日志（已删版本/整首），供前端展示与一键恢复。"""
    return jsonify({'log': _load_deleted_log(), 'count': len(_load_deleted_log())})


@app.route('/api/playlist/pins', methods=['GET'])
def api_playlist_pins_get():
    """返回固定监测清单。"""
    return jsonify({'pins': _load_pins(), 'count': len(_load_pins())})


@app.route('/api/playlist/pin', methods=['POST'])
def api_playlist_pin():
    """固定一个版本进监测清单。优先用前端传来的 entry 算钥匙 + 平台ID（避免 JS 重复逻辑）。"""
    data = request.get_json(silent=True) or {}
    entry = data.get('entry')
    key = (data.get('version_key') or '').strip()
    if entry:
        key = _version_key(entry)
        platform_ids = _platform_ids_of(entry)
        platform_data = (entry.get('platform_data') or {})
        q = data.get('query') or entry.get('query') or {}
        song_name = (data.get('song_name') or entry.get('song_name') or q.get('song_name') or '').strip()
        performer = (data.get('performer') or entry.get('performer') or q.get('performer') or '').strip()
        lyricist = (data.get('lyricist') or entry.get('lyricist') or q.get('lyricist') or '').strip()
        composer = (data.get('composer') or entry.get('composer') or q.get('composer') or '').strip()
        kind = entry.get('is_original') and 'original' or 'cover'
    else:
        if not key:
            return jsonify({'error': '缺少 version_key 或 entry'}), 400
        platform_ids = data.get('platform_ids') or {}
        platform_data = data.get('platform_data') or {}
        song_name = (data.get('song_name') or '').strip()
        performer = (data.get('performer') or '').strip()
        lyricist = (data.get('lyricist') or '').strip()
        composer = (data.get('composer') or '').strip()
        kind = (data.get('kind') or 'version')
    pins = _load_pins()
    if any(p.get('version_key') == key for p in pins):
        return jsonify({'ok': True, 'already': True, 'count': len(pins)})
    pin = {
        'version_key': key,
        'song_name': song_name,
        'performer': performer,
        'lyricist': lyricist,
        'composer': composer,
        'kind': kind,
        'platform_ids': platform_ids,
        'platform_data': platform_data,
        'pinned_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        'refreshed_at': None,
        'refreshed_missing': False,
    }
    pins.append(pin)
    _save_pins(pins)
    return jsonify({'ok': True, 'count': len(pins)})


@app.route('/api/playlist/unpin', methods=['POST'])
def api_playlist_unpin():
    """取消固定（移出监测清单）。"""
    data = request.get_json(silent=True) or {}
    key = (data.get('version_key') or '').strip()
    if not key and data.get('entry'):
        key = _version_key(data['entry'])
    if not key:
        return jsonify({'error': '缺少 version_key 或 entry'}), 400
    pins = _load_pins()
    before = len(pins)
    pins = [p for p in pins if p.get('version_key') != key]
    _save_pins(pins)
    return jsonify({'ok': True, 'removed': before - len(pins), 'count': len(pins)})


@app.route('/api/playlist/pins/refresh', methods=['POST'])
def api_playlist_pins_refresh():
    """刷新固定清单：对每首固定歌重新查询，按平台 ID 锁定同一版本，回填最新数据。

    不清除未固定的歌、不重排；只更新固定清单里每项的 platform_data / refreshed_at。
    返回更新后的 pins（含 refreshed_missing 标记哪些在平台已下架/搜不到）。
    """
    data = request.get_json(silent=True) or {}
    pins = _load_pins()
    if not pins:
        return jsonify({'ok': True, 'pins': [], 'updated': 0, 'missing': 0})
    conc = min(int(data.get('concurrency', 3)), 6) or 3
    updated = 0
    missing = 0

    def _refresh_one(pin):
        name = pin.get('song_name', '')
        perf = pin.get('performer', '')
        ly = pin.get('lyricist', '')
        co = pin.get('composer', '')
        want_ids = pin.get('platform_ids') or {}
        try:
            core = _search_core(name, ly, co, perf, 30)
            grouped = core.get('grouped_results', [])
        except Exception:
            grouped = []
        # 在分组结果里找「平台 ID 与固定时一致」的版本
        match = None
        for g in grouped:
            gids = _platform_ids_of(g)
            if any(want_ids.get(c) and gids.get(c) == want_ids.get(c) for c in want_ids):
                match = g
                break
        if match:
            # 走 _trim_platform：保留原始 ID（供下次刷新匹配）同时补全展示字段
            _pd = {}
            for code, p in (match.get('platform_data') or {}).items():
                t = _trim_platform(p)
                if t:
                    _pd[code] = t
            pin['platform_data'] = _pd
            pin['refreshed_missing'] = False
            pin['refreshed_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
            return True
        # 没匹配到（可能下架/改名）→ 保留旧数据，标记缺失
        pin['refreshed_missing'] = True
        pin['refreshed_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
        return False

    with concurrent.futures.ThreadPoolExecutor(max_workers=conc) as ex:
        futs = {ex.submit(_refresh_one, p): p for p in pins}
        for f in concurrent.futures.as_completed(futs):
            try:
                ok = f.result()
                updated += 1 if ok else 0
                missing += 0 if ok else 1
            except Exception:
                missing += 1
    _save_pins(pins)
    # 把最新数据也写回缓存结果（my_playlist_results.json），下次打开直接看到新数据
    try:
        _patch_results_with_pins(pins)
    except Exception as e:
        logger.error('刷新后回写缓存结果失败: %s', e)
    return jsonify({'ok': True, 'pins': pins, 'updated': updated, 'missing': missing})


@app.route('/api/verify', methods=['POST'])
def api_verify():
    """授权核对 API"""
    data = request.get_json(silent=True) or {}
    songs = data.get('songs', [])

    if not songs or not isinstance(songs, list):
        return jsonify({'error': '请提供授权歌曲列表'}), 400

    verify_limit = min(int(data.get('limit', 30)), 1000)
    verify_results = []

    for idx, auth_song in enumerate(songs):
        song_name = (auth_song.get('song_name') or auth_song.get('name') or '').strip()
        performer = (auth_song.get('performer') or auth_song.get('artist') or '').strip()
        auth_label = (auth_song.get('record_label') or auth_song.get('label') or '').strip()

        if not song_name:
            verify_results.append({
                'authorized_song': auth_song,
                'status': 'skip',
                'message': '缺少歌曲名称',
                'platforms': [],
            })
            continue

        # 多艺人组合时只用第一个艺人名，避免关键词过长导致搜索分词失败
        if performer:
            _first_artist = re.split(r'[/&,，、;；|]+', performer)[0].strip()
            keyword = f"{song_name} {_first_artist}".strip() if _first_artist else song_name
        else:
            keyword = song_name
        all_results = search_all(keyword, verify_limit)

        by_platform = {}
        for r in all_results:
            code = r['platform_code']
            if code not in by_platform:
                by_platform[code] = []
            by_platform[code].append(r)

        platform_checks = []
        for code in PLATFORM_ORDER:
            platform_results = by_platform.get(code, [])
            matched = []
            for r in platform_results:
                name_ok = _name_match(song_name, r.get('song_name', ''))
                artist_ok = True
                if performer:
                    artist_ok = _performer_match(performer, r.get('performer', ''))
                if name_ok and artist_ok:
                    matched.append(r)

            if matched:
                best = matched[0]
                for m in matched:
                    if m.get('record_label'):
                        best = m
                        break

                label_status = None
                label_note = '该平台不提供唱片公司信息'
                if code in LABEL_PLATFORMS:
                    all_labels = [m.get('record_label') for m in matched if m.get('record_label')]
                    unique_labels = list(dict.fromkeys(all_labels))
                    if auth_label and unique_labels:
                        matched_label = None
                        for lbl in unique_labels:
                            if _label_match(auth_label, lbl):
                                matched_label = lbl
                                break
                        if matched_label:
                            label_status = True
                            label_note = f'平台显示：{matched_label}'
                        else:
                            label_status = False
                            label_note = f'平台显示：{" / ".join(unique_labels[:3])}'
                    elif unique_labels and not auth_label:
                        label_note = f'平台显示：{" / ".join(unique_labels[:3])}（授权名单未提供唱片公司）'
                    elif not unique_labels and auth_label:
                        label_note = '平台未返回该歌曲的唱片公司信息'
                    else:
                        label_note = '双方均无唱片公司信息'

                platform_checks.append({
                    'platform': PLATFORM_NAMES[code],
                    'platform_code': code,
                    'found': True,
                    'match_count': len(matched),
                    'song_url': best.get('song_url', ''),
                    'song_name': best.get('song_name', ''),
                    'performer': best.get('performer', ''),
                    'album': best.get('album', ''),
                    'lyricist': best.get('lyricist'),
                    'composer': best.get('composer'),
                    'label_match': label_status,
                    'label_note': label_note,
                })
            else:
                platform_checks.append({
                    'platform': PLATFORM_NAMES[code],
                    'platform_code': code,
                    'found': False,
                    'match_count': 0,
                    'song_url': '',
                    'song_name': '',
                    'performer': '',
                    'album': '',
                    'lyricist': None,
                    'composer': None,
                    'label_match': None,
                    'label_note': '未找到匹配歌曲',
                })

        found_count = sum(1 for p in platform_checks if p['found'])
        verify_results.append({
            'authorized_song': {
                'song_name': song_name,
                'performer': performer,
                'record_label': auth_label,
            },
            'keyword': keyword,
            'status': 'verified',
            'found_platforms': found_count,
            'total_platforms': len(PLATFORM_ORDER),
            'platforms': platform_checks,
        })

    total_songs = len(verify_results)
    fully_covered = sum(1 for r in verify_results if r.get('found_platforms', 0) == len(PLATFORM_ORDER))
    partial = sum(1 for r in verify_results if 0 < r.get('found_platforms', 0) < len(PLATFORM_ORDER))
    missing = sum(1 for r in verify_results if r.get('found_platforms', 0) == 0)

    return jsonify({
        'total_songs': total_songs,
        'summary': {
            'fully_covered': fully_covered,
            'partial_coverage': partial,
            'missing': missing,
        },
        'results': verify_results,
    })


# ═══════════════════════════════════════════════════
#  Cookie 管理路由
# ═══════════════════════════════════════════════════

@app.route('/api/cookies', methods=['GET'])
def get_cookies():
    """获取所有平台的 Cookie"""
    cookies = load_cookies()
    result = {}
    for code in PLATFORM_ORDER:
        result[code] = {
            'platform': PLATFORM_NAMES[code],
            'platform_code': code,
            'cookie': cookies.get(code, ''),
            'has_cookie': bool(cookies.get(code, '')),
        }
    return jsonify({'cookies': result})


@app.route('/api/cookies', methods=['POST'])
def set_cookies():
    """保存指定平台的 Cookie"""
    data = request.get_json(silent=True) or {}
    platform = data.get('platform_code', '').strip()
    cookie_value = data.get('cookie', '').strip()

    if platform not in PLATFORM_NAMES:
        return jsonify({'error': '无效的平台'}), 400

    cookies = load_cookies()
    if cookie_value:
        cookies[platform] = cookie_value
    else:
        cookies.pop(platform, None)

    if save_cookies(cookies):
        return jsonify({
            'success': True,
            'message': f'{PLATFORM_NAMES[platform]} Cookie 已保存',
            'has_cookie': bool(cookie_value),
        })
    else:
        return jsonify({'error': '保存失败'}), 500


@app.route('/api/cookies/clear', methods=['POST'])
def clear_cookies():
    """清空指定平台的 Cookie（v4.25.10 新增：一键清除按钮后端）"""
    data = request.get_json(silent=True) or {}
    platform = data.get('platform_code', '').strip()

    if platform not in PLATFORM_NAMES:
        return jsonify({'error': '无效的平台'}), 400

    cookies = load_cookies()
    had = bool(cookies.get(platform))
    cookies.pop(platform, None)

    if save_cookies(cookies):
        return jsonify({
            'success': True,
            'message': f"{PLATFORM_NAMES[platform]} Cookie{'已清空' if had else '（本来就没有）'}",
            'had_cookie': had,
            'has_cookie': False,
        })
    else:
        return jsonify({'error': '清空失败'}), 500


@app.route('/api/login/<platform>', methods=['POST'])
def browser_login(platform):
    """立即返回，后台启动浏览器登录（长连接改为前端轮询状态，避免被超时切断）"""
    if platform not in PLATFORM_NAMES:
        return jsonify({'success': False, 'message': '无效平台'})
    start_browser_login(platform)
    return jsonify({'started': True, 'message': '正在打开浏览器，请在弹出的窗口中登录…'})


@app.route('/api/login/status/<platform>', methods=['GET'])
def browser_login_status(platform):
    """前端轮询登录状态"""
    status = LOGIN_STATUS.get(platform)
    if not status:
        return jsonify({'running': False, 'done': True, 'success': False, 'message': '尚未开始', 'cookie': ''})
    return jsonify(status)


@app.route('/api/hearts/login/<platform>', methods=['POST'])
def heart_browser_login(platform):
    """红心管理页：后台启动浏览器登录，登录成功后自动保存到 hearts_cookies.json 并抓取红心歌单。"""
    if platform not in PLATFORM_NAMES:
        return jsonify({'success': False, 'message': '无效平台'})
    start_heart_browser_login(platform)
    return jsonify({'started': True, 'message': '正在打开浏览器，请在弹出的窗口中登录…'})


@app.route('/api/hearts/login/status/<platform>', methods=['GET'])
def heart_browser_login_status(platform):
    """红心管理页：前端轮询浏览器登录+抓取状态"""
    status = HEART_LOGIN_STATUS.get(platform)
    if not status:
        return jsonify({'running': False, 'done': True, 'success': False, 'message': '尚未开始', 'cookie': '', 'count': 0})
    return jsonify(status)


@app.route('/api/open-login/<platform>', methods=['POST'])
def open_browser_login(platform):
    """用系统默认浏览器打开平台登录页（不等待、不阻塞），随后用户手动粘贴 Cookie。
    比 Playwright 弹窗更稳定， especially 在打包后的 .app 里。"""
    if platform not in PLATFORM_NAMES:
        return jsonify({'success': False, 'message': '无效平台'})
    try:
        url = LOGIN_HOME[platform]
        # webbrowser.open 在 macOS .app 里有时需要在新线程调用，避免阻塞 Flask
        threading.Thread(target=webbrowser.open, args=(url,), kwargs={'new': 2}, daemon=True).start()
        return jsonify({
            'success': True,
            'message': f"已打开 {PLATFORM_NAMES[platform]} 登录页，登录后请复制 Cookie 粘贴保存",
            'url': url,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f"打开浏览器失败：{e}"})


@app.route('/api/env')
def api_env():
    """返回前端需要的运行环境信息（是否本地、Playwright 是否可用、各平台登录页）"""
    try:
        from playwright.sync_api import sync_playwright
        playwright_ok = True
    except Exception:
        playwright_ok = False

    # 仅当请求来自本机（浏览器和服务器在同一台机器）时才允许弹出登录相关功能
    remote = request.remote_addr or ''
    is_localhost = remote in ('127.0.0.1', 'localhost', '::1', None, '')

    return jsonify({
        'is_localhost': is_localhost,
        'playwright_available': playwright_ok,
        # 现在「浏览器登录」改为 webbrowser.open，只要本地运行就可用
        'show_browser_login': is_localhost,
        'login_home': LOGIN_HOME,
        # v4：标记同步后端（cloudbase = 全网同步，local = 仅本机）
        'marks_backend': _marks_backend_cfg().get('backend', 'local'),
        'app_version': 'v4',
    })
@app.route('/api/cookies/test', methods=['POST'])
def test_cookies():
    """测试 Cookie 是否有效（简单搜索测试）"""
    data = request.get_json(silent=True) or {}
    platform = data.get('platform_code', '').strip()
    if platform not in PLATFORM_NAMES:
        return jsonify({'error': '无效的平台'}), 400

    cookie_str = get_cookie_string(platform)
    if not cookie_str:
        return jsonify({'valid': False, 'message': '未设置 Cookie'}), 200

    # 简单测试：用关键词搜索看是否返回结果
    try:
        func = PLATFORM_FUNCS.get(platform)
        if func:
            results = func('测试', 5)
            return jsonify({
                'valid': len(results) > 0,
                'message': f'返回 {len(results)} 条结果',
                'count': len(results),
            })
    except Exception as e:
        return jsonify({'valid': False, 'message': str(e)}), 200

    return jsonify({'valid': False, 'message': '未知错误'}), 200


# ═══════════════════════════════════════════════════
#  歌词段落识别（同音字容错）— N1 / Feature2
# ═══════════════════════════════════════════════════

try:
    from pypinyin import lazy_pinyin as _lazy_pinyin
except Exception:
    _lazy_pinyin = None


def _to_pinyin(text):
    """中文转拼音（去声调），用于同音字容错匹配；非中文按原字符处理。"""
    if not text:
        return ''
    if _lazy_pinyin is None:
        return text
    try:
        return ''.join(_lazy_pinyin(text))
    except Exception:
        return text


def _clean_lyric_text(text):
    """清洗歌词：去时间轴、去括号注解、去标点空白，转小写。"""
    if not text:
        return ''
    text = re.sub(r'\[\d{1,2}:\d{2}(?:\.\d+)?\]', ' ', text)  # [00:12.34]
    text = re.sub(r'[\(（\[【][^)\]】]*[\)）\]】]', ' ', text)  # (x2) （合唱）【间奏】
    text = re.sub(r'[\s\-_/\\&＆·•,，、;；:：.。!！?？~～"\'‘’“”()（）\[\]【】]+', '', text)
    return text.lower()


_LYRIC_PLAT_WEIGHT = {'netease': 4, 'qq': 3, 'kugou': 2, 'kuwo': 1, 'qishui': 0}

_CANONICAL_NOISE = ['原唱', '翻唱', 'cover', '伴奏', 'remix', '(live', 'live)', '现场', ' dj', '版)',
                     '试听', '片段', '铃声']


def _canonical_penalty(c):
    """翻唱/ live/ 伴奏等非原版标注轻微降权，让原唱版本优先。"""
    s = ((c.get('song_name') or '') + ' ' + (c.get('performer') or '')).lower()
    return sum(30 for k in _CANONICAL_NOISE if k in s)


# ═══════════════════════════════════════════════════════════════════════════
#  歌词/音频识别 → 找盗版：自动嫌疑标签
# ═══════════════════════════════════════════════════════════════════════════

# 标题/歌手里出现这些词，大概率不是官方原版
_PIRACY_KEYWORDS = [
    '翻唱', 'cover', 'remix', '伴奏', 'instrumental', 'karaoke',
    '(live', 'live)', '现场版', '现场', ' dj', '版)', '改词', '改编',
    '试听', '片段', '铃声', '消音', '变速', '降调', '升调', '合唱版',
    '完整版', '纯音乐', '钢琴版', '吉他版', '小提琴版', '女声版', '男声版',
    '童声版', '戏腔版', '烟嗓版', '古风版', '电音版', 'dj版', '慢摇',
]

# 表演者里出现这些，通常是 AI/虚拟/匿名上传
_ANONYMOUS_ARTIST_HINTS = ['ai ', ' a.i', '人工智能', '虚拟歌手', 'unknown', '未知歌手', '群星']


def _suspect_tags_from_text(song_name, performer):
    """仅基于歌名+歌手的文本，给出初始盗版/翻唱嫌疑标签。"""
    tags = []
    s = ((song_name or '') + ' ' + (performer or '')).lower()
    if any(k in s for k in _PIRACY_KEYWORDS):
        tags.append('疑似翻唱/改编')
    if any(k in (performer or '').lower() for k in _ANONYMOUS_ARTIST_HINTS):
        tags.append('表演者异常')
    return tags


def _max_collection_of_row(row):
    """取聚合行在各平台中的最大收藏量。"""
    pd = row.get('platform_data') or {}
    max_n = None
    for code in PLATFORM_ORDER:
        r = pd.get(code)
        if r and r.get('collection_count') is not None:
            n = int(r['collection_count'])
            if max_n is None or n > max_n:
                max_n = n
    return max_n


def _suspect_tags_from_row(row, canonical_row=None):
    """基于完整平台数据，分析一首歌是否为盗版/翻唱嫌疑对象。

    canonical_row: 同一首歌的「正版参考行」（通常是平台最多/收藏最高的版本）。
    如果没有 canonical_row，仅做绝对值判断。
    """
    tags = _suspect_tags_from_text(row.get('song_name'), row.get('performer'))
    pd = row.get('platform_data') or {}
    has_label = bool(row.get('record_label'))
    max_col = _max_collection_of_row(row)
    plat_count = len(pd)

    # 没有发行公司 + 收藏量低 → 高度可疑
    if not has_label:
        if max_col is None or max_col < 1000:
            tags.append('疑似盗版：无发行公司且热度极低')
        else:
            tags.append('无发行公司信息')

    # 与 canonical_row（同一首歌的正版参考）做相对比较
    if canonical_row:
        can_max = _max_collection_of_row(canonical_row)
        can_label = canonical_row.get('record_label') or ''
        # 收藏量显著偏低
        if can_max and max_col is not None and can_max > 0 and max_col < can_max * 0.05:
            tags.append('热度显著低于正版')
        # 发行公司不同
        row_label = row.get('record_label') or ''
        if can_label and row_label and can_label != row_label:
            tags.append('发行公司与正版不一致')
        # 平台覆盖少
        can_plats = len(canonical_row.get('platform_data') or {})
        if plat_count == 1 and can_plats >= 3:
            tags.append('传播平台单一')
    else:
        # 没有参考时，单平台 + 低热度 → 可疑
        if plat_count == 1 and (max_col is None or max_col < 5000):
            tags.append('单平台低热度')

    # 去重并保持顺序
    seen = set()
    out = []
    for t in tags:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def _pick_canonical_row(rows):
    """在一组同一首歌的候选中，选出最像「正版参考」的行。

    规则：平台覆盖多 > 收藏量高 > 有发行公司。
    """
    best, best_score = None, -1e9
    for r in rows:
        pd = r.get('platform_data') or {}
        plat_score = len(pd) * 1000
        col_score = (_max_collection_of_row(r) or 0)
        label_score = 500 if (r.get('record_label') or (any(p.get('record_label') for p in pd.values()))) else 0
        text_penalty = 0
        s = ((r.get('song_name') or '') + ' ' + (r.get('performer') or '')).lower()
        if any(k in s for k in _PIRACY_KEYWORDS):
            text_penalty -= 800
        score = plat_score + col_score + label_score + text_penalty
        if score > best_score:
            best, best_score = r, score
    return best


def _bigram_containment(a, b):
    """返回 (匹配到的二元组数, 输入二元组总数)，用于片段级模糊匹配。"""
    if len(a) < 2:
        return 0, 0
    ab = set(a[i:i + 2] for i in range(len(a) - 1))
    if not ab:
        return 0, 0
    bb = set(b[i:i + 2] for i in range(len(b) - 1))
    return len(ab & bb), len(ab)


def _lyric_score(in_clean, in_py, cand_clean, cand_py, allow_fuzzy=True):
    """三级匹配打分：精确 > 同音 > 片段。返回 (score, match_type, hit_len)。"""
    if not in_clean or not cand_clean:
        return 0, 'none', 0
    if in_clean in cand_clean:
        return 3 * 1000 + min(len(in_clean), 25) * 10, 'precise', len(in_clean)
    if allow_fuzzy and in_py and cand_py and in_py in cand_py:
        return 2 * 1000 + min(len(in_py), 25) * 10, 'same', len(in_py)
    if allow_fuzzy and in_py and cand_py:
        matched, total = _bigram_containment(in_py, cand_py)
        if total > 0 and matched >= 4 and matched / total >= 0.55:
            return 1 * 1000 + min(matched, 25) * 10, 'fragment', matched
    return 0, 'none', 0


def _clean_lyric_lines(lyr):
    """去掉 LRC 时间轴，过滤哈希/元数据乱码，返回可读歌词行列表。"""
    if not lyr:
        return []
    out = []
    for ln in lyr.splitlines():
        ln = re.sub(r'\[\d{1,2}:\d{2}(?:\.\d+)?\]|\[\w+:[^\]]*\]', '', ln).strip().lstrip('\ufeff')
        if not ln:
            continue
        if len(ln) > 120:
            continue
        low = ln.lower()
        # 过滤酷狗等接口偶尔会混进歌词里的 hash/metadata 行
        if re.search(r'byhash|hash[:=]|signature[:=]|sign[:=]|id3|^h\d{5,}', low):
            continue
        if not re.search(r'[\u4e00-\u9fff]', ln) and ' ' not in ln and len(ln) > 30:
            continue
        out.append(ln)
    return out


def _extract_snippet(lyr, in_clean):
    """从歌词中截取与输入最相关的连续几行，用于结果展示。支持同音字容错。"""
    lines = _clean_lyric_lines(lyr)
    if not lines:
        return ''
    key = in_clean[:12] if len(in_clean) >= 12 else in_clean
    key_py = _to_pinyin(key)

    def matches(ln):
        lclean = re.sub(r'[\s\-_/\\&＆·•,，、;；:：.。!！?？~～"\'‘’“”()（）\[\]【】]+', '', ln.lower())
        if key in lclean:
            return True
        if key_py and key_py in _to_pinyin(lclean):
            return True
        return False

    for i, ln in enumerate(lines):
        if matches(ln):
            start = max(0, i - 1)
            end = min(len(lines), i + 3)
            return ' / '.join(lines[start:end])[:120]
    return lines[0][:100]


def fetch_lyrics(platform_code, meta):
    """按平台拉取歌曲完整歌词文本（原始 lrc）。失败返回 ''。"""
    try:
        if platform_code == 'qq':
            songmid = meta.get('_songmid')
            if not songmid:
                return ''
            url = 'https://c.y.qq.com/lyric/fcgi-bin/fcg_query_lyric_new.fcg'
            params = {'songmid': songmid, 'g_tk': 5381, 'format': 'json',
                      'inCharset': 'utf8', 'outCharset': 'utf-8'}
            h = {'User-Agent': COMMON_UA, 'Referer': 'https://y.qq.com/'}
            ck = get_cookie_string('qq')
            if ck:
                h['Cookie'] = ck
            r = requests.get(url, params=params, headers=h, timeout=6)
            d = r.json()
            b = d.get('lyric', '')
            if b:
                return base64.b64decode(b).decode('utf-8', errors='ignore')
            return ''
        elif platform_code == 'netease':
            sid = meta.get('_song_id')
            if not sid:
                return ''
            url = f'https://music.163.com/api/song/lyric?id={sid}&lv=1&tv=-1'
            h = {'User-Agent': COMMON_UA, 'Referer': 'https://music.163.com/'}
            ck = get_cookie_string('netease')
            if ck:
                h['Cookie'] = ck
            r = requests.get(url, headers=h, timeout=6)
            d = r.json()
            return d.get('lrc', {}).get('lyric', '') or ''
        elif platform_code == 'kugou':
            hv = meta.get('_hash')
            if not hv:
                return ''
            h = {'User-Agent': COMMON_UA, 'Referer': 'https://www.kugou.com/'}
            ck = get_cookie_string('kugou')
            if ck:
                h['Cookie'] = ck
            kw = urllib.parse.quote(meta.get('song_name', '') or '')
            s_url = f'http://lyrics.kugou.com/search?ver=1&man=yes&client=pc&keyword={kw}&hash={hv}'
            r = requests.get(s_url, headers=h, timeout=6)
            cs = r.json().get('candidates', [])
            if not cs:
                return ''
            c = cs[0]
            lid, ak = c.get('id'), c.get('accesskey')
            if not (lid and ak):
                return ''
            dl = f'http://lyrics.kugou.com/download?ver=1&client=pc&id={lid}&accesskey={ak}&fmt=lrc&charset=utf8'
            r2 = requests.get(dl, headers=h, timeout=6)
            content = r2.json().get('content', '')
            if content:
                return base64.b64decode(content).decode('utf-8', errors='ignore')
            return ''
        elif platform_code == 'kuwo':
            rid = meta.get('_rid')
            if not rid:
                return ''
            url = f'http://m.kuwo.cn/newh5/singles/songinfoandlrc?musicId={rid}'
            r = requests.get(url, headers={'User-Agent': COMMON_UA}, timeout=6, verify=False)
            ll = r.json().get('data', {}).get('lrclist', [])
            return '\n'.join(i.get('lineLyric', '') for i in ll)
        else:
            # 汽水音乐（抖音系）暂无公开歌词接口，跳过
            return ''
    except Exception as e:
        print(f"[lyric] {platform_code} fetch error: {e}")
        return ''


def _identify_one_paragraph(text, options):
    """识别单个歌词段落，返回结构化结果。"""
    raw = (text or '').strip()
    cleaned = _clean_lyric_text(raw)
    excerpt = raw[:24].replace('\n', ' ')
    if len(cleaned) < 4:
        return {'found': False, 'reason': '输入过短（少于 4 个有效字符）',
                'input_excerpt': excerpt, 'best': None, 'alternatives': []}
    # 搜索种子：取首行清洗后截断，聚焦最 distinctive 的短语
    lines = [l for l in re.split(r'[\r\n]+', raw) if l.strip()]
    seed = _clean_lyric_text(lines[0] if lines else raw)
    if len(seed) > 20:
        seed = seed[:20]
    if not seed:
        seed = cleaned[:20]
    per_platform = int(options.get('per_platform', 5)) or 5
    candidates = search_all(seed, per_platform_limit=per_platform)
    allow_fuzzy = bool(options.get('same_sound', True))
    in_py = _to_pinyin(cleaned)

    def fetch_and_score(c):
        lyr = fetch_lyrics(c.get('platform_code'), c)
        if not lyr:
            return None
        c_clean = _clean_lyric_text(lyr)
        c_py = _to_pinyin(c_clean)
        sc, mtype, hit = _lyric_score(cleaned, in_py, c_clean, c_py, allow_fuzzy)
        if sc <= 0:
            return None
        return (sc, mtype, c, lyr)

    scored = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
        for res in ex.map(fetch_and_score, candidates):
            if res:
                scored.append(res)
    if not scored:
        return {'found': False,
                'reason': '未找到匹配候选（可能平台未收录该歌词，或候选歌曲无歌词）',
                'input_excerpt': excerpt, 'best': None, 'alternatives': []}
    scored.sort(key=lambda r: (r[0] - _canonical_penalty(r[2]),
                                _LYRIC_PLAT_WEIGHT.get(r[2].get('platform_code'), 0)),
                reverse=True)

    def pack(r):
        sc, mtype, c, lyr = r
        return {
            'song_name': c.get('song_name'), 'performer': c.get('performer'),
            'platform': c.get('platform'), 'platform_code': c.get('platform_code'),
            'song_url': c.get('song_url'), 'match_type': mtype,
            'match_label': {'precise': '精准命中', 'same': '同音命中', 'fragment': '片段匹配'}.get(mtype, mtype),
            'snippet': _extract_snippet(lyr, cleaned),
        }

    return {
        'found': True, 'reason': '', 'input_excerpt': excerpt,
        'best': pack(scored[0]), 'alternatives': [pack(a) for a in scored[1:4]],
    }


def _lyric_search_one(text, options):
    """新版歌词段落搜索：返回所有匹配候选（不只一个 best），用于找盗版。

    options:
      same_sound:     是否开启同音字/片段容错（默认 true）
      per_platform:   每平台搜索候选数（默认 15，最大 30）
      max_candidates: 返回候选总数上限（默认 15，最大 30）
    """
    raw = (text or '').strip()
    cleaned = _clean_lyric_text(raw)
    excerpt = raw[:24].replace('\n', ' ')
    if len(cleaned) < 4:
        return {'found': False, 'reason': '输入过短（少于 4 个有效字符）',
                'input_excerpt': excerpt, 'candidates': []}

    lines = [l for l in re.split(r'[\r\n]+', raw) if l.strip()]
    seed = _clean_lyric_text(lines[0] if lines else raw)
    if len(seed) > 20:
        seed = seed[:20]
    if not seed:
        seed = cleaned[:20]

    per_platform = min(int(options.get('per_platform', 15)), 30)
    max_candidates = min(int(options.get('max_candidates', 15)), 30)
    allow_fuzzy = bool(options.get('same_sound', True))
    in_py = _to_pinyin(cleaned)

    candidates = search_all(seed, per_platform_limit=per_platform)

    def fetch_and_score(c):
        lyr = fetch_lyrics(c.get('platform_code'), c)
        if not lyr:
            return None
        c_clean = _clean_lyric_text(lyr)
        c_py = _to_pinyin(c_clean)
        sc, mtype, hit = _lyric_score(cleaned, in_py, c_clean, c_py, allow_fuzzy)
        if sc <= 0:
            return None
        return (sc, mtype, c, lyr)

    scored = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
        for res in ex.map(fetch_and_score, candidates):
            if res:
                scored.append(res)
    if not scored:
        return {'found': False,
                'reason': '未找到匹配候选（可能平台未收录该歌词，或候选歌曲无歌词）',
                'input_excerpt': excerpt, 'candidates': []}

    # 按匹配分数倒序，不降权翻唱/伴奏，让盗版版本也露出来
    scored.sort(key=lambda r: (r[0],
                               _LYRIC_PLAT_WEIGHT.get(r[2].get('platform_code'), 0)),
                reverse=True)

    def pack(r):
        sc, mtype, c, lyr = r
        item = {
            'song_name': c.get('song_name'),
            'performer': c.get('performer'),
            'platform': c.get('platform'),
            'platform_code': c.get('platform_code'),
            'song_url': c.get('song_url'),
            'match_type': mtype,
            'match_score': sc,
            'match_label': {'precise': '精准命中', 'same': '同音命中', 'fragment': '片段匹配'}.get(mtype, mtype),
            'snippet': _extract_snippet(lyr, cleaned),
            'suspect_tags': _suspect_tags_from_text(c.get('song_name'), c.get('performer')),
            '_needs_enrich': True,
        }
        mk = _find_mark_loose(c.get('song_name'), c.get('performer'))
        if mk:
            item['mark'] = mk
        return item

    return {
        'found': True, 'reason': '', 'input_excerpt': excerpt,
        'candidates': [pack(a) for a in scored[:max_candidates]],
    }


@app.route('/api/lyric-identify', methods=['POST'])
def api_lyric_identify():
    data = request.get_json(silent=True) or {}
    text = (data.get('text') or '').strip()
    options = data.get('options') or {}
    if not text:
        return jsonify({'error': '请输入歌词文本'}), 400
    paragraphs = [p for p in re.split(r'\n\s*\n', text) if p.strip()]
    if not paragraphs:
        return jsonify({'error': '未检测到歌词段落（请用空行分隔每首歌）'}), 400
    if len(paragraphs) > 100:
        return jsonify({'error': '单次最多识别 100 个段落'}), 400
    results = [None] * len(paragraphs)
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(4, len(paragraphs))) as ex:
        futs = {ex.submit(_identify_one_paragraph, p, options): i for i, p in enumerate(paragraphs)}
        for fut in concurrent.futures.as_completed(futs):
            i = futs[fut]
            try:
                results[i] = fut.result()
            except Exception as e:
                results[i] = {'found': False, 'reason': f'识别失败：{e}',
                              'input_excerpt': paragraphs[i][:24].replace('\n', ' '),
                              'best': None, 'alternatives': []}
    # 附加已有的人工标记（打盗版流程：识别出来先看这首歌是不是已经处理过，避免重复打）
    for r in results:
        b = r.get('best') if isinstance(r, dict) else None
        if b:
            mk = _find_mark_loose(b.get('song_name'), b.get('performer'))
            if mk:
                b['mark'] = mk
        for a in (r.get('alternatives') or []) if isinstance(r, dict) else []:
            mk = _find_mark_loose(a.get('song_name'), a.get('performer'))
            if mk:
                a['mark'] = mk
    return jsonify({'count': len(results), 'results': results})


@app.route('/api/lyric-search', methods=['POST'])
def api_lyric_search():
    """歌词片段搜索：把输入的多段歌词视为同一首歌的多个片段，聚合识别结果并自动补全五平台数据。

    body:
      text: 歌词文本（同一首歌的多段歌词用空行分隔）
      options:
        same_sound: true/false（默认 true）
        per_platform: 每平台搜索数（默认 15，最大 30）
        max_candidates: 返回歌曲上限（默认 15，最大 30）
    """
    data = request.get_json(silent=True) or {}
    text = (data.get('text') or '').strip()
    options = data.get('options') or {}
    if not text:
        return jsonify({'error': '请输入歌词文本'}), 400
    paragraphs = [p for p in re.split(r'\n\s*\n', text) if p.strip()]
    if not paragraphs:
        return jsonify({'error': '未检测到歌词段落'}), 400
    if len(paragraphs) > 100:
        return jsonify({'error': '单次最多 100 个歌词段落'}), 400

    per_platform = min(int(options.get('per_platform', 15)), 30)
    max_candidates = min(int(options.get('max_candidates', 15)), 30)
    options['per_platform'] = per_platform
    options['max_candidates'] = max_candidates

    # 1) 每个片段并行搜索原始候选
    raw_results = [None] * len(paragraphs)
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(4, len(paragraphs))) as ex:
        futs = {ex.submit(_lyric_search_one, p, options): i for i, p in enumerate(paragraphs)}
        for fut in concurrent.futures.as_completed(futs):
            i = futs[fut]
            try:
                raw_results[i] = fut.result()
            except Exception as e:
                logger.exception('lyric search paragraph failed')
                raw_results[i] = {'found': False, 'reason': f'识别失败：{e}',
                                  'input_excerpt': paragraphs[i][:24].replace('\n', ' '),
                                  'candidates': []}

    # 2) 把所有候选按「歌名 + 歌手」聚合成「歌曲维度」，并统计命中了几个歌词片段
    groups = {}
    snippet_map = {}
    vote_map = {}
    for pi, res in enumerate(raw_results):
        if not res or not res.get('found'):
            continue
        for c in (res.get('candidates') or []):
            song_key = _normalize_match(c.get('song_name', ''))
            art_key = tuple(sorted(_split_artist_names(c.get('performer', ''))))
            key = (song_key, art_key)
            groups.setdefault(key, []).append(c)
            vote_map.setdefault(key, set()).add(pi)
            snip = c.get('snippet', '')
            if key not in snippet_map or len(snippet_map[key]) < len(snip):
                snippet_map[key] = snip

    if not groups:
        return jsonify({'count': 0, 'results': []})

    # 3) 每个歌曲维度自动补全五平台数据（复用第一页搜索流程）
    def enrich_one(key_cands):
        key, cands = key_cands
        best = max(cands, key=lambda x: x.get('match_score', 0) or 0)
        row = _enrich_row(best.get('song_name', ''), best.get('performer', ''), limit=20)
        if not row:
            row = {
                'song_name': best.get('song_name'),
                'performer': best.get('performer'),
                'album': '',
                'release_date': '',
                'lyricist': '',
                'composer': '',
                'record_label': '',
                'platform_data': dict(best.get('platform_data') or {}),
                'match_score': best.get('match_score', 0),
                'match_label': best.get('match_label', '歌词匹配'),
                'snippet': snippet_map.get(key, ''),
                'suspect_tags': best.get('suspect_tags', []),
                'mark': best.get('mark'),
                'hearts': [],
            }
        else:
            row['_enriched'] = True
        row['_match_source'] = 'lyric'
        row['_input_excerpt'] = best.get('input_excerpt', '')
        row['snippet'] = snippet_map.get(key, row.get('snippet', ''))
        row['_fragment_hits'] = len(vote_map.get(key) or ())
        row['_fragment_total'] = len(paragraphs)
        if not row.get('match_score'):
            row['match_score'] = best.get('match_score', 0)
        # 把歌词搜索原始候选里的平台链接也合并进来（ enrich 没覆盖到的平台）
        pd = row.setdefault('platform_data', {})
        for c in cands:
            code = c.get('platform_code')
            if code and code not in pd:
                pd[code] = {
                    'collection_count': None,
                    'listening_count': None,
                    'comment_count': None,
                    'record_label': '',
                    'release_date': '',
                    'song_url': c.get('song_url', ''),
                }
        return row

    with concurrent.futures.ThreadPoolExecutor(max_workers=min(6, len(groups))) as ex:
        enriched = list(ex.map(enrich_one, groups.items()))

    # 4) 补全后再按「歌名 + 歌手」二次去重（不同歌手写法会指向同一首歌）
    merged = {}
    for row in enriched:
        if not row:
            continue
        dedupe_key = (
            _normalize_match(row.get('song_name', '')),
            tuple(sorted(_split_artist_names(row.get('performer', '')))),
        )
        old = merged.get(dedupe_key)
        if old is None:
            merged[dedupe_key] = row
            continue
        # 合并：片段命中数取并集式相加上限、平台数据互补、分数取高
        old['_fragment_hits'] = max(old.get('_fragment_hits', 0), row.get('_fragment_hits', 0))
        old['match_score'] = max(old.get('match_score', 0) or 0, row.get('match_score', 0) or 0)
        opd = old.setdefault('platform_data', {})
        for code, info in (row.get('platform_data') or {}).items():
            if code not in opd:
                opd[code] = info
        for fld in ('album', 'release_date', 'lyricist', 'composer', 'record_label'):
            if not old.get(fld) and row.get(fld):
                old[fld] = row.get(fld)
        if len(row.get('snippet') or '') > len(old.get('snippet') or ''):
            old['snippet'] = row.get('snippet')

    results = list(merged.values())

    # 5) 单曲收敛：一次只搜一首歌，命中片段最多的才是目标歌曲
    max_hits = max((r.get('_fragment_hits', 0) for r in results), default=0)
    if max_hits >= 2:
        results = [r for r in results if r.get('_fragment_hits', 0) >= max_hits]

    # 6) 排序：命中片段多 > 匹配分高 > 覆盖平台多
    results.sort(key=lambda r: (
        r.get('_fragment_hits', 0),
        r.get('match_score', 0) or 0,
        len(r.get('platform_data') or {})
    ), reverse=True)

    return jsonify({
        'count': len(results),
        'fragment_total': len(paragraphs),
        'results': results[:max_candidates],
    })

def _find_mark_loose(song_name, performer):
    """按「歌名 + 歌手」查已有标记（忽略专辑）。

    第一页标记时 key 含专辑，而歌词识别阶段拿不到专辑，
    因此展示标记状态用宽松匹配；真正写入标记时仍以补全后的完整信息为准。
    """
    if not song_name:
        return None
    tgt_song = _normalize_match(song_name)
    tgt_artists = set(_split_artist_names(performer or ''))
    with _marks_lock:
        marks = _load_marks()
    fallback = None
    for m in marks.values():
        if not isinstance(m, dict):
            continue
        if _normalize_match(m.get('song_name', '')) != tgt_song:
            continue
        arts = set(_split_artist_names(m.get('performer', '')))
        if not tgt_artists or not arts:
            fallback = fallback or m
            continue
        if arts & tgt_artists:
            return m
    return fallback


def _pick_best_group(grouped, song_name, performer):
    """从聚合结果里挑出与目标「歌名 + 歌手」最贴合的一行。"""
    tgt_song = _normalize_match(song_name or '')
    tgt_art = set(_split_artist_names(performer or ''))
    best, best_score = None, -1e9
    for g in grouped:
        gs = _normalize_match(g.get('song_name', ''))
        score = 0.0
        if gs and gs == tgt_song:
            score += 10
        elif tgt_song and (tgt_song in gs or gs in tgt_song):
            score += 4
        arts = set(_split_artist_names(g.get('performer', '')))
        if tgt_art and (arts & tgt_art):
            score += 8
        elif tgt_art and arts:
            score -= 2          # 歌手完全对不上，多半是翻唱/同名歌
        # 覆盖平台越多越可能是原版发行
        score += min(len(g.get('platform_data') or {}), 5) * 0.4
        if score > best_score:
            best, best_score = g, score
    return best


def _enrich_row(song_name, performer='', limit=15):
    """对识别出的「歌名 + 歌手」补全五平台聚合数据（复用第一页搜索流程）。

    返回和第一页表格一致的聚合行 dict（含 platform_data / 发行公司 / 嫌疑标签 / 标记）。
    若五平台均未搜到，返回 None。
    """
    if not song_name:
        return None
    try:
        payload = _search_core(song_name, '', '', performer, max(limit, 50))
    except Exception as e:
        logger.warning('_enrich_row _search_core failed: %s', e)
        return None
    grouped = payload.get('grouped_results') or []
    row = _pick_best_group(grouped, song_name, performer)
    if not row:
        return None

    # 同名同歌手行按平台合并（应对同一首歌在不同平台的专辑/单曲版本）
    tgt_song = _normalize_match(row.get('song_name', ''))
    tgt_art = set(_split_artist_names(row.get('performer', '')))
    pd = row.setdefault('platform_data', {})
    merged_from = []
    for g in grouped:
        if g is row:
            continue
        if _normalize_match(g.get('song_name', '')) != tgt_song:
            continue
        arts = set(_split_artist_names(g.get('performer', '')))
        if tgt_art and arts and not (tgt_art & arts):
            continue
        for code, pr in (g.get('platform_data') or {}).items():
            if code not in pd:
                pd[code] = pr
                merged_from.append(code)
        for h in (g.get('hearts') or []):
            if h not in (row.get('hearts') or []):
                row.setdefault('hearts', []).append(h)
    if merged_from:
        row['_merged_platforms'] = merged_from

    # 顶层发行公司兜底
    if not row.get('record_label'):
        for code in PLATFORM_ORDER:
            if pd.get(code) and pd[code].get('record_label'):
                row['record_label'] = pd[code]['record_label']
                break

    # 找盗版：基于完整平台数据生成/刷新嫌疑标签
    same_name_rows = [g for g in grouped
                      if _normalize_match(g.get('song_name', '')) == tgt_song]
    canonical = _pick_canonical_row(same_name_rows) if same_name_rows else None
    row['suspect_tags'] = _suspect_tags_from_row(row, canonical)

    # 已有标记（宽松匹配）
    mk = _find_mark_loose(row.get('song_name', ''), row.get('performer', ''))
    if mk:
        row['mark'] = mk

    return row


def _enrich_versions(song_name, performer='', limit=20, max_rows=20):
    """按歌名走第一页搜索流程，返回该歌名下「全部版本」的聚合行列表。

    用于音频识别：识别出歌名后，把原唱和各翻唱/DJ 版本一次性归拢出来，
    每行都带完整的五平台数据，不需要用户再点「查平台数据」。
    """
    if not song_name:
        return []
    try:
        payload = _search_core(song_name, '', '', performer, max(limit, 50))
    except Exception as e:
        logger.warning('_enrich_versions _search_core failed: %s', e)
        return []
    grouped = payload.get('grouped_results') or []
    if not grouped:
        return []

    tgt_song = _normalize_match(song_name)
    # 只保留歌名相同/相近的版本，过滤掉搜索引擎带出来的无关歌曲
    same_name_rows = [g for g in grouped
                      if _normalize_match(g.get('song_name', '')) == tgt_song]
    pool = same_name_rows or grouped
    canonical = _pick_canonical_row(same_name_rows) if same_name_rows else None

    best = _pick_best_group(grouped, song_name, performer)
    tgt_art = set(_split_artist_names(performer or ''))

    merged = {}
    for g in pool:
        key = (
            _normalize_match(g.get('song_name', '')),
            tuple(sorted(_split_artist_names(g.get('performer', '')))),
        )
        old = merged.get(key)
        if old is None:
            merged[key] = g
            continue
        opd = old.setdefault('platform_data', {})
        for code, pr in (g.get('platform_data') or {}).items():
            if code not in opd:
                opd[code] = pr
        for fld in ('album', 'release_date', 'lyricist', 'composer', 'record_label'):
            if not old.get(fld) and g.get(fld):
                old[fld] = g.get(fld)

    rows = list(merged.values())
    for row in rows:
        pd = row.setdefault('platform_data', {})
        if not row.get('record_label'):
            for code in PLATFORM_ORDER:
                if pd.get(code) and pd[code].get('record_label'):
                    row['record_label'] = pd[code]['record_label']
                    break
        row['suspect_tags'] = _suspect_tags_from_row(row, canonical)
        mk = _find_mark_loose(row.get('song_name', ''), row.get('performer', ''))
        if mk:
            row['mark'] = mk
        row['_enriched'] = True
        arts = set(_split_artist_names(row.get('performer', '')))
        row['_is_target'] = bool(best is not None and row is best) or bool(tgt_art and arts and (tgt_art & arts))

    rows.sort(key=lambda r: (
        1 if r.get('_is_target') else 0,
        len(r.get('platform_data') or {}),
    ), reverse=True)
    return rows[:max_rows]


@app.route('/api/lyric-enrich', methods=['POST'])
@app.route('/api/enrich', methods=['POST'])
def api_lyric_enrich():
    """歌词/音频识别 → 打盗版数据补全（通用）。

    传入识别出的「歌名 + 歌手」，复用第一页完整搜索流程，
    返回该歌的聚合行（含五平台收藏量 / 发行公司 / 发行时间 / 链接 + 标记 + 红心 + 嫌疑标签）。
    """
    data = request.get_json(silent=True) or {}
    song_name = (data.get('song_name') or '').strip()
    performer = (data.get('performer') or '').strip()
    limit = min(int(data.get('limit', 15)), 50)
    if not song_name:
        return jsonify({'error': '缺少歌名'}), 400
    try:
        row = _enrich_row(song_name, performer, limit)
    except Exception as e:
        logger.exception('lyric enrich failed')
        return jsonify({'error': f'补全失败：{e}'}), 500
    if not row:
        return jsonify({'ok': False, 'reason': '五平台均未搜到这首歌'})
    return jsonify({
        'ok': True,
        'row': row,
        'platform_stats': [],
    })


if __name__ == '__main__':
    _ensure_admin_exists()  # 升级兼容：确保存在管理员账号
    _ensure_perms_all()     # 升级兼容：给老账号补 perms 细粒度权限位
    _load_performer_aliases()  # v4.25.x：启动加载「艺人变体白名单」（同艺人多账号发布用）
    _sync_users_from_cloud()  # v4.20: 启动时从云端同步用户
    _push_local_users_to_cloud()  # v4.27.3: 启动时补推本地用户到云端（修复跨设备登录）
    # 一次性回填：把 note 里「原类型：xxx」提取回 mark_type（v4.13 撤销归一后的反向迁移）
    try:
        _migrate_mark_types_once()
    except Exception as _e:
        print(f'[MusicFinder] mark_type 启动回填失败（不影响主功能）: {_e}')
    # 启动时后台各拉一次云端标记/个人数据（不阻塞、本地优先）。
    # 放在此处而非模块顶层：这两个函数在上方定义，import 时（如测试）不应触发线程。
    try:
        threading.Thread(target=_pull_cloud_marks_once, daemon=True).start()
    except Exception as _e:
        print(f'[MusicFinder] 云端标记同步启动失败（不影响本地）: {_e}')
    try:
        threading.Thread(target=_pull_cloud_blobs_once, daemon=True).start()
    except Exception as _e:
        print(f'[MusicFinder] 云端 blob 同步启动失败（不影响本地）: {_e}')
    print("=" * 50)
    print("  MusicFinder v3 - 多平台音乐搜索聚合")
    print("  支持：QQ音乐 / 酷狗 / 酷我 / 网易云 / 汽水音乐")
    print("  字段：歌名/歌手/专辑/词曲作者/收藏/在听/评论/播放量/唱片公司/链接")
    print("=" * 50)
    print(f"  Cookie 文件: {COOKIE_FILE}")
    print(f"  访问 http://localhost:5050 开始使用")
    print("=" * 50)

    # ── 首次运行把使用说明复制到程序同目录，方便用户查看 ──
    def _copy_usage_doc():
        try:
            _meipass = getattr(sys, '_MEIPASS', None)
            if not _meipass:
                return
            _src = os.path.join(_meipass, '使用说明.md')
            if not os.path.exists(_src):
                return
            import shutil
            _dst = os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), '使用说明.md')
            if not os.path.exists(_dst):
                shutil.copyfile(_src, _dst)
        except Exception:
            pass

    _copy_usage_doc()

    # ── 自定义端口：优先级 命令行 --port > 同目录 port.txt > 环境变量 PORT > 5050 ──
    _arg_parser = argparse.ArgumentParser(description='MusicFinder')
    _arg_parser.add_argument('--port', type=int, default=None, help='自定义监听端口')
    _cli_args, _ = _arg_parser.parse_known_args()

    def _load_port_from_file():
        # 冻结态（PyInstaller 打包后）资源在 sys._MEIPASS，port.txt 已随 bundle 打进该目录；
        # 开发/本地运行时在同目录（BASE_DIR）。两处都试，保证打包版也用 57074 而非回退 5050。
        for _base in (BUNDLE_DIR, os.path.dirname(os.path.abspath(sys.argv[0]))):
            try:
                _cfg = os.path.join(_base, 'port.txt')
                if os.path.exists(_cfg):
                    with open(_cfg, 'r', encoding='utf-8') as _f:
                        return int(str(_f.read()).strip())
            except Exception:
                pass
        return None

    port = (_cli_args.port or _load_port_from_file()
            or int(os.environ.get('PORT', 5050)))

    # reloader 子进程（WERKZEUG_RUN_MAIN=true）由父进程已验证端口，跳过此检测，
    # 否则子进程会误判「端口被占用」而直接退出，导致服务起不来
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        import socket as _sock

        def _parse_etime(_s):
            """把 ps 的 etime（如 12:34:56 / 3-12:34:56 / 00:34）转成秒数。"""
            _s = _s.strip()
            _days = 0
            if '-' in _s:
                _d, _s = _s.split('-', 1)
                _days = int(_d)
            _p = [int(x) for x in _s.split(':')]
            if len(_p) == 3:
                _h, _m, _sec = _p
            elif len(_p) == 2:
                _h, _m, _sec = 0, _p[0], _p[1]
            else:
                _h, _m, _sec = 0, 0, _p[0]
            return _days * 86400 + _h * 3600 + _m * 60 + _sec

        def _resolve_port_conflict(_port):
            """端口被占用时决策：
            - 'yield'：占用者就是当前这版实例（同路径且启动于 exe 更新之后）→ 单实例让位
            - 'take' ：已清理掉旧/异路径占用者，调用方可重试 bind 自己起
            - 'abort'：清理失败，保守退出
            这样「双击新安装包」会自动踢掉旧版本（v2）残留进程，无需用户手动操作。
            """
            import subprocess as _sp
            _myself = os.path.abspath(sys.executable)
            _pid = None
            if sys.platform == 'win32':
                try:
                    _out = _sp.run(['netstat', '-ano'], capture_output=True, text=True).stdout
                    for _ln in _out.splitlines():
                        if f':{_port}' in _ln and 'LISTENING' in _ln:
                            _pid = _ln.split()[-1].strip()
                            break
                except Exception:
                    pass
            else:
                try:
                    _out = _sp.run(['lsof', '-ti', f'tcp:{_port}'], capture_output=True, text=True).stdout.strip()
                    if _out:
                        _pid = _out.split('\n')[0].strip()
                except Exception:
                    pass
            if not _pid:
                return 'take'  # 查不到占用者（可能刚释放），让调用方重试 bind
            # 查占用者 exe 路径
            _other_exe = ''
            try:
                if sys.platform == 'win32':
                    _out = _sp.run(['wmic', 'process', 'where', f'ProcessId={_pid}', 'get', 'ExecutablePath', '/value'],
                                  capture_output=True, text=True).stdout
                    for _ln in _out.splitlines():
                        if _ln.startswith('ExecutablePath='):
                            _other_exe = _ln.split('=', 1)[1].strip()
                else:
                    _other_exe = _sp.run(['ps', '-o', 'comm=', '-p', _pid], capture_output=True, text=True).stdout.strip()
            except Exception:
                pass
            _other_exe = os.path.abspath(_other_exe) if _other_exe else ''
            # 不是自己（旧版本/别的路径）→ 直接踢掉升级
            if _other_exe and _other_exe != _myself:
                pass  # 落到下方 kill 分支
            else:
                # 路径相同：判断是否「覆盖了新二进制却还在跑旧进程」（启动早于 exe 更新）
                try:
                    _exe_mtime = os.path.getmtime(sys.executable)
                    _etime = _sp.run(['ps', '-o', 'etime=', '-p', _pid], capture_output=True, text=True).stdout.strip()
                    _started = time.time() - _parse_etime(_etime)
                    if _started < _exe_mtime - 5:
                        pass  # 占用者跑的是旧二进制 → 踢掉升级
                    else:
                        return 'yield'  # 同版本刚起 → 单实例让位
                except Exception:
                    return 'yield'  # 判断不了就保守让位
            # kill 占用者（跨平台）
            try:
                if sys.platform == 'win32':
                    _sp.run(['taskkill', '/PID', _pid, '/F'], capture_output=True)
                else:
                    _sp.run(['kill', '-9', _pid], capture_output=True)
                time.sleep(1.0)
                return 'take'
            except Exception:
                return 'abort'

        _s = _sock.socket(_sock.AF_INET, _sock.SOCK_STREAM)
        # SO_REUSEADDR：允许在 TIME_WAIT 状态下重建绑定，避免旧进程刚退出时
        # 新进程误判「端口被占用」而退出（重启常见的坑）
        try:
            _s.setsockopt(_sock.SOL_SOCKET, _sock.SO_REUSEADDR, 1)
        except OSError:
            pass
        try:
            _s.bind(('127.0.0.1', port))
            _s.close()
        except OSError:
            _decision = _resolve_port_conflict(port)
            if _decision == 'yield':
                if not os.environ.get('MF_NO_BROWSER'):
                    try:
                        webbrowser.open(f'http://127.0.0.1:{port}/')
                    except Exception:
                        pass
                print(f"[MusicFinder] 端口 {port} 已被同版本实例占用，直接唤起")
                sys.exit(0)
            elif _decision == 'take':
                # 旧/异版本占用者已清理，重试一次 bind
                try:
                    _s2 = _sock.socket(_sock.AF_INET, _sock.SOCK_STREAM)
                    try:
                        _s2.setsockopt(_sock.SOL_SOCKET, _sock.SO_REUSEADDR, 1)
                    except OSError:
                        pass
                    _s2.bind(('127.0.0.1', port))
                    _s2.close()
                    print(f"[MusicFinder] 已清理旧版本实例，独占端口 {port}")
                except OSError:
                    print(f"[MusicFinder] 端口 {port} 仍被占用，无法启动")
                    sys.exit(1)
            else:
                print(f"[MusicFinder] 端口 {port} 冲突无法解决，退出")
                sys.exit(1)

    # 启动后自动打开本地浏览器（仅本机访问，launchd 常驻模式跳过）
    def _open_browser():
        time.sleep(1.5)
        if os.environ.get('MF_NO_BROWSER'):
            return
        try:
            webbrowser.open(f'http://127.0.0.1:{port}/')
        except Exception:
            pass

    threading.Thread(target=_open_browser, daemon=True).start()
    # 记录启动时间戳，供 /api/version 算 uptime
    global _APP_START_TS
    _APP_START_TS = time.time()
    # use_reloader=True: 改动 app.py 保存后服务自动重启，避免「改了代码线上还是旧逻辑」
    # 仅开 reloader，不开交互式调试器（debugger 关闭更安全）
    # 冻结态（PyInstaller 打包后 sys._MEIPASS 存在）下关闭 reloader：
    # 否则 reloader 子进程找不到模块会无限重启，服务起不来
    FROZEN = getattr(sys, '_MEIPASS', None) is not None
    # reloader 默认开（开发期改代码自动重启）；常驻/launchd 托管时用环境变量
    # MF_RELOADER=0 关闭，避免 reloader 父子双进程与 launchd 的 KeepAlive 产生端口竞态
    _USE_RELOADER = (not FROZEN) and os.environ.get('MF_RELOADER', '1') != '0'
    app.run(host='127.0.0.1', port=port, debug=False, use_reloader=_USE_RELOADER, threaded=True)
